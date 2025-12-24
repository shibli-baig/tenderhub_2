"""
I FastAPI backend for TenderHub application - MODULAR VERSION

This is the main application file that has been refactored for better code organization.

MODULAR STRUCTURE:
==================
The codebase has been organized into modules for better maintainability:

1. core/security.py      - Authentication and session management utilities
2. core/dependencies.py  - Shared FastAPI dependencies (get_db, get_current_user, etc.)
3. api/routes/auth.py    - Authentication routes (ready for use)
4. api/routes/certificates.py - Certificate management routes (ready for use)
5. api/routes/tenders.py    - Tender management routes (in progress)
6. api/routes/projects.py   - Project routes (in progress)
7. api/routes/employees.py  - Employee routes (in progress)

CURRENT STATUS:
==============
- ✅ Core modules extracted and working
- ✅ app.py uses core modules for better organization
- ⚠️  Routes are still in this file (can be migrated to routers incrementally)
- 📖 See ROUTER_MIGRATION_GUIDE.md for migration instructions

This provides all API endpoints for tender management, user authentication,
and favorites functionality while maintaining backward compatibility.
"""

from fastapi import FastAPI, HTTPException, Depends, Request, Form, File, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse, Response
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBearer
import os
import csv
import asyncio
import threading
import multiprocessing
import shutil

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

# Hard-coded defaults for the PostgreSQL migration phase.
# Replace these values with environment-specific secrets before deploying.
os.environ.setdefault("DATABASE_URL", os.getenv("DATABASE_URL", "postgresql+psycopg://tenderhub_admin:YOUR_PASSWORD@localhost:5432/tenderhub_db"))
os.environ.setdefault("DB_POOL_SIZE", "5")
os.environ.setdefault("DB_MAX_OVERFLOW", "5")
os.environ.setdefault("DB_POOL_TIMEOUT", "30")
os.environ.setdefault("DB_POOL_RECYCLE", "1800")

from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, or_, desc, String, cast, func, inspect, text, Numeric
from datetime import datetime, timedelta, date
from typing import List, Optional, Dict, Any, Set
import uuid
import re
from pathlib import Path
import secrets
import logging
import json
import ast
import time
from collections import defaultdict
from urllib.parse import unquote, urlencode
from io import BytesIO, StringIO
from openpyxl import load_workbook, Workbook

# Import database models and functions
from database import (
    SessionLocal, TenderDB, TenderDocumentDB, UserDB, FavoriteDB, CustomCardDB, ProjectDB,
    CompanyCodeDB, EmployeeDB, TenderAssignmentDB, TaskDB, TaskCommentDB, TaskFileDB, TaskProgressUpdateDB,
    TenderMessageDB, TaskConcernDB, EmployeeNotificationDB, CertificateDB, CompanyDB, CompanyCertificateDB, ShortlistedTenderDB,
    RejectedTenderDB, DumpedTenderDB, StageDocumentDB, StageTaskTemplateDB, NotificationDB,
    TenderResponseDB, ResponseDocumentDB, ReminderDB, CalendarActivityDB, SeenTenderDB,
    ExpertDB, ExpertProfileDB, ExpertContentDB, ExpertContentCommentDB, ExpertContentLikeDB,
    ExpertServiceRequestDB, ExpertApplicationDB, ExpertCollaborationDB, ExpertReviewDB,
    ExpertPaymentDB, ExpertFavoriteTenderDB, ExpertNotificationDB, EmployeePerformanceRatingDB,
    ExpertHiringRequestDB, ExpertHiringApplicationDB, ExpertProjectMessageDB, ExpertProjectTaskDB,
    create_tables, cleanup_old_tenders, cleanup_expired_tenders, cleanup_orphaned_records, engine,
    ExpertProjectTaskFileDB, ExpertProjectTaskProgressUpdateDB, ExpertProjectTaskQueryDB
)

# Import certificate processor
from certificate_processor import certificate_processor

# Import tender recommendation scorer
from tender_recommendation import TenderRecommendationScorer, TenderEmbeddingManager

# Import WebSocket connection manager
from websocket_manager import manager as ws_manager

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Track whether we've successfully ensured the expert_hiring_requests.tender_location column exists
_expert_request_column_lock = threading.Lock()
_expert_request_column_ready = False

_expert_review_column_lock = threading.Lock()
_expert_review_column_ready = False

def create_expert_notification(db: Session, expert_id: str, notification_type: str, title: str, message: str, link: Optional[str] = None):
    """Utility to log notifications for ExpertVerse."""
    notification = ExpertNotificationDB(
        expert_id=expert_id,
        notification_type=notification_type,
        title=title,
        message=message,
        link=link
    )
    db.add(notification)
    return notification

slug_pattern = re.compile(r'[^a-z0-9]+')

def generate_expert_content_slug(db: Session, title: str) -> str:
    """
    Create a URL-friendly slug from the provided title.
    Ensures uniqueness by appending a numeric suffix when needed.
    """
    base_slug = slug_pattern.sub('-', title.lower()).strip('-') if title else ''
    if not base_slug:
        base_slug = f"post-{uuid.uuid4().hex[:8]}"

    # Gather existing slugs with the same prefix to avoid repeated DB round-trips
    existing_slugs = {
        slug for (slug,) in db.query(ExpertContentDB.slug)
        .filter(ExpertContentDB.slug.like(f"{base_slug}%"))
        .all()
    }

    if base_slug not in existing_slugs:
        return base_slug

    counter = 2
    candidate = f"{base_slug}-{counter}"
    while candidate in existing_slugs:
        counter += 1
        candidate = f"{base_slug}-{counter}"

    return candidate

def ensure_expert_hiring_tender_location_column() -> bool:
    """
    Ensure the expert_hiring_requests table has the tender_location column.
    Automatically runs a lightweight migration if needed.
    """
    global _expert_request_column_ready

    if _expert_request_column_ready:
        return True

    with _expert_request_column_lock:
        if _expert_request_column_ready:
            return True

        try:
            with engine.begin() as conn:
                inspector = inspect(conn)
                if "expert_hiring_requests" not in inspector.get_table_names():
                    logger.warning("Table 'expert_hiring_requests' not found when ensuring tender_location column.")
                    return False

                existing_columns = {col["name"] for col in inspector.get_columns("expert_hiring_requests")}
                if "tender_location" in existing_columns:
                    _expert_request_column_ready = True
                    return True

                backend = getattr(engine.url, "get_backend_name", lambda: engine.dialect.name)()
                ddl = "ALTER TABLE expert_hiring_requests ADD COLUMN tender_location VARCHAR"
                if backend == "postgresql":
                    ddl = "ALTER TABLE expert_hiring_requests ADD COLUMN IF NOT EXISTS tender_location VARCHAR"

                logger.info("Adding tender_location column to expert_hiring_requests table (backend=%s)", backend)
                conn.execute(text(ddl))
                _expert_request_column_ready = True
                return True
        except Exception as exc:
            logger.error("Failed to ensure tender_location column exists: %s", exc, exc_info=True)
            return False

def ensure_expert_review_hiring_column() -> bool:
    """Ensure expert_reviews has hiring_request_id column and index."""
    global _expert_review_column_ready

    if _expert_review_column_ready:
        return True

    with _expert_review_column_lock:
        if _expert_review_column_ready:
            return True

        try:
            with engine.begin() as conn:
                inspector = inspect(conn)
                if "expert_reviews" not in inspector.get_table_names():
                    logger.warning("Table 'expert_reviews' not found when ensuring hiring_request_id column.")
                    return False

                existing_columns = {col["name"] for col in inspector.get_columns("expert_reviews")}
                if "hiring_request_id" not in existing_columns:
                    backend = getattr(engine.url, "get_backend_name", lambda: engine.dialect.name)()
                    ddl = "ALTER TABLE expert_reviews ADD COLUMN hiring_request_id VARCHAR"
                    if backend == "postgresql":
                        ddl = "ALTER TABLE expert_reviews ADD COLUMN IF NOT EXISTS hiring_request_id VARCHAR"
                    conn.execute(text(ddl))

                    if backend == "postgresql":
                        try:
                            conn.execute(text(
                                "ALTER TABLE expert_reviews "
                                "ADD CONSTRAINT IF NOT EXISTS expert_reviews_hiring_request_id_fkey "
                                "FOREIGN KEY (hiring_request_id) REFERENCES expert_hiring_requests(id) "
                                "ON DELETE SET NULL"
                            ))
                        except Exception as exc:
                            logger.warning("Failed to add expert_reviews hiring_request FK: %s", exc)
                    else:
                        try:
                            conn.execute(text(
                                "ALTER TABLE expert_reviews "
                                "ADD CONSTRAINT expert_reviews_hiring_request_id_fkey "
                                "FOREIGN KEY (hiring_request_id) REFERENCES expert_hiring_requests(id) "
                                "ON DELETE SET NULL"
                            ))
                        except Exception as exc:
                            logger.warning("Failed to add expert_reviews hiring_request FK: %s", exc)

                # Ensure unique index exists
                try:
                    conn.execute(text(
                        "CREATE UNIQUE INDEX IF NOT EXISTS idx_expert_company_hiring_request_unique "
                        "ON expert_reviews (expert_id, company_id, hiring_request_id)"
                    ))
                except Exception as exc:
                    logger.warning("Failed to ensure expert_reviews hiring request unique index: %s", exc)

                _expert_review_column_ready = True
                return True
        except Exception as exc:
            logger.error("Failed to ensure expert_reviews hiring_request_id column exists: %s", exc, exc_info=True)
            return False

# ============================================================================
# PROJECT ID GENERATION FUNCTIONS
# ============================================================================

def extract_company_initials(company_name: str) -> str:
    """
    Extract initials from company name by taking first letter of each word.
    Example: "ABC Consultants Pvt. Ltd." -> "ACPL"

    Args:
        company_name: The company name string

    Returns:
        Uppercase initials (2-10 characters), or "PROJ" if name is empty/invalid
    """
    if not company_name or not isinstance(company_name, str):
        return "PROJ"

    # Only filter out common connectors and articles that add no value
    filter_words = {
        "and", "&", "of", "the"
    }

    # Split by spaces and common separators
    words = re.split(r'[\s\-_.,&]+', company_name.strip())

    # Extract first letter of each significant word
    initials = []
    for word in words:
        word_clean = word.strip().lower()
        if word_clean and word_clean not in filter_words and len(word_clean) > 0:
            # Take first character (handles unicode properly)
            first_char = word[0].upper()
            if first_char.isalnum():
                initials.append(first_char)
    
    # Ensure minimum 2 characters, maximum 10
    if len(initials) < 2:
        return "PROJ"
    if len(initials) > 10:
        initials = initials[:10]
    
    return ''.join(initials)


def get_company_name_for_user(user: UserDB, db: Session) -> str:
    """
    Get company name for a user, checking CompanyDB first, then UserDB.company.
    
    Args:
        user: UserDB instance
        db: Database session
        
    Returns:
        Company name string, or empty string if not found
    """
    # Check CompanyDB first (more detailed company info)
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == user.id).first()
    if company_details and company_details.company_name:
        return company_details.company_name.strip()
    
    # Fall back to UserDB.company
    if user.company:
        return user.company.strip()
    
    return ""


def get_next_project_number(user_id: str, company_initials: str, db: Session, in_memory_project_ids: Optional[Set[str]] = None) -> int:
    """
    Get the next project number for a user's company initials.
    Handles overflow beyond 999999 by detecting existing longer numbers.
    
    Args:
        user_id: User ID
        company_initials: Company initials (e.g., "PCPL")
        db: Database session
        in_memory_project_ids: Optional set of project IDs already generated in this session (not yet committed)
        
    Returns:
        Next project number (starting from 1)
    """
    # Query all projects for this user with matching initials pattern
    pattern = f"{company_initials}-%"
    existing_projects = db.query(ProjectDB.project_id).filter(
        ProjectDB.user_id == user_id,
        ProjectDB.project_id.like(pattern)
    ).all()
    
    max_number = 0
    max_digits = 6  # Default to 6 digits
    
    # Extract numbers from existing project_ids in database
    for (project_id,) in existing_projects:
        if not project_id:
            continue
        
        # Match pattern: {INITIALS}-{NUMBER}
        match = re.match(rf'^{re.escape(company_initials)}-(\d+)$', project_id)
        if match:
            number = int(match.group(1))
            max_number = max(max_number, number)
            # Track maximum digits needed
            num_digits = len(match.group(1))
            max_digits = max(max_digits, num_digits)
    
    # Also check in-memory project IDs (from current bulk upload session)
    if in_memory_project_ids:
        for project_id in in_memory_project_ids:
            if not project_id:
                continue
            
            # Match pattern: {INITIALS}-{NUMBER}
            match = re.match(rf'^{re.escape(company_initials)}-(\d+)$', project_id)
            if match:
                number = int(match.group(1))
                max_number = max(max_number, number)
                # Track maximum digits needed
                num_digits = len(match.group(1))
                max_digits = max(max_digits, num_digits)
    
    next_number = max_number + 1
    
    # If next number exceeds 6 digits, we'll format with more digits
    # The formatting will be handled in generate_project_id()
    return next_number


def generate_project_id(user_id: str, db: Session, existing_id: Optional[str] = None, in_memory_project_ids: Optional[Set[str]] = None) -> str:
    """
    Generate a project ID in format {COMPANY_INITIALS}-{NUMBER}.
    
    Args:
        user_id: User ID
        db: Database session
        existing_id: Optional existing project ID to validate/use
        in_memory_project_ids: Optional set of project IDs already generated in this session (not yet committed)
        
    Returns:
        Project ID string (e.g., "PCPL-000001")
    """
    # If existing_id provided, validate format and check uniqueness
    if existing_id:
        existing_id = existing_id.strip()
        # Validate format: {INITIALS}-{NUMBER}
        if re.match(r'^[A-Z]{2,10}-\d{6,}$', existing_id):
            # Check uniqueness in database
            existing = db.query(ProjectDB).filter(
                ProjectDB.user_id == user_id,
                ProjectDB.project_id == existing_id
            ).first()
            # Also check in-memory set
            if not existing and (not in_memory_project_ids or existing_id not in in_memory_project_ids):
                return existing_id
            # If duplicate, fall through to generate new one
            logger.warning(f"Duplicate project_id '{existing_id}' for user {user_id}, generating new one")
    
    # Get user
    user = db.query(UserDB).filter(UserDB.id == user_id).first()
    if not user:
        raise ValueError(f"User {user_id} not found")
    
    # Get company name
    company_name = get_company_name_for_user(user, db)
    if not company_name:
        company_name = "Project"  # Default fallback
    
    # Extract initials
    company_initials = extract_company_initials(company_name)
    
    # Get next project number (passing in_memory_project_ids)
    next_number = get_next_project_number(user_id, company_initials, db, in_memory_project_ids)
    
    # Determine number of digits needed
    # If number >= 999999, use 7 digits; if >= 9999999, use 8, etc.
    if next_number >= 9999999:
        num_digits = 8
    elif next_number >= 999999:
        num_digits = 7
    else:
        num_digits = 6
    
    # Format project ID
    project_id = f"{company_initials}-{next_number:0{num_digits}d}"
    
    # Double-check uniqueness against database and in-memory set
    existing = db.query(ProjectDB).filter(ProjectDB.project_id == project_id).first()
    in_memory_collision = in_memory_project_ids and project_id in in_memory_project_ids
    
    # If collision found, increment and retry until unique
    while existing or in_memory_collision:
        next_number += 1
        if next_number >= 9999999:
            num_digits = 8
        elif next_number >= 999999:
            num_digits = 7
        project_id = f"{company_initials}-{next_number:0{num_digits}d}"
        existing = db.query(ProjectDB).filter(ProjectDB.project_id == project_id).first()
        in_memory_collision = in_memory_project_ids and project_id in in_memory_project_ids
    
    return project_id

# Configure multiprocessing to use 'spawn' start method for safety
# This ensures clean separation between parent and child processes
try:
    multiprocessing.set_start_method('spawn', force=False)
except RuntimeError:
    # Start method already set, which is fine
    pass

# Resolve base directory so static/template paths work in any CWD
BASE_DIR = Path(__file__).resolve().parent

# Create necessary directories if they don't exist
uploads_dir = BASE_DIR / "uploads"
uploads_dir.mkdir(parents=True, exist_ok=True)
(uploads_dir / "projects").mkdir(parents=True, exist_ok=True)
(uploads_dir / "certificates").mkdir(parents=True, exist_ok=True)
(uploads_dir / "stage_documents").mkdir(parents=True, exist_ok=True)
(uploads_dir / "response_documents").mkdir(parents=True, exist_ok=True)
(uploads_dir / "response_pdfs").mkdir(parents=True, exist_ok=True)
(uploads_dir / "signatures").mkdir(parents=True, exist_ok=True)

# Initialize FastAPI app
app = FastAPI(title="TenderHub", description="Government Tender Management System")


# Startup event to initialize database tables
@app.on_event("startup")
async def startup_event():
    """Initialize database tables on application startup."""
    try:
        logger.info("Application startup: Ensuring database tables exist...")
        create_tables()
        logger.info("✓ Database tables verified/created successfully")
    except Exception as e:
        logger.error(f"Failed to initialize database tables on startup: {e}")
        # Don't fail the startup, just log the error
        # Tables might already exist or will be created later


# Import and mount API routers
from api.routes import certificates as certificates_router
from api.routes import certificate_matching as certificate_matching_router
from api.routes import ai_insights as ai_insights_router
# from api.routes import tenders as tenders_router  # COMMENTED OUT - stubs not implemented
app.include_router(certificates_router.router)
app.include_router(certificate_matching_router.router)
app.include_router(ai_insights_router.router)
# app.include_router(tenders_router.router)  # COMMENTED OUT - causes stub override

# Mount static files and templates
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
app.mount("/uploads", StaticFiles(directory=BASE_DIR / "uploads"), name="uploads")
templates = Jinja2Templates(directory=BASE_DIR / "templates")

# ============================================================================
# UI CONFIGURATION
# ============================================================================

# Font Selection System
# Available options: 'inter', 'poppins', 'roboto', 'work-sans', 'plus-jakarta-sans'
# Change this value to test different fonts across the entire application
SELECTED_FONT = "work-sans"  # <-- Change this to switch fonts

FONT_OPTIONS = {
    "inter": {
        "name": "Inter",
        "family": "'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
        "description": "Modern, geometric, excellent readability"
    },
    "poppins": {
        "name": "Poppins",
        "family": "'Poppins', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
        "description": "Friendly, geometric, rounded"
    },
    "roboto": {
        "name": "Roboto",
        "family": "'Roboto', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
        "description": "Neutral, highly legible, Google's flagship"
    },
    "work-sans": {
        "name": "Work Sans",
        "family": "'Work Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
        "description": "Industrial, clean, corporate"
    },
    "plus-jakarta-sans": {
        "name": "Plus Jakarta Sans",
        "family": "'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
        "description": "Contemporary, elegant"
    }
}

def get_active_font():
    """Get the currently selected font configuration"""
    return FONT_OPTIONS.get(SELECTED_FONT, FONT_OPTIONS["inter"])

# ============================================================================

COMPANY_CERTIFICATIONS_DIR = BASE_DIR / "uploads" / "company_certifications"
COMPANY_CERTIFICATIONS_DIR.mkdir(parents=True, exist_ok=True)

PROJECT_SERVICE_KEYS = [
    "design_engineering",
    "dpr_feasibility",
    "gis_data",
    "pmc",
    "pmu",
    "advisory_capacity",
    "survey_investigations",
    "environmental_social"
]

PROJECT_DOCUMENT_TYPES = [
    "tender_documents",
    "technical_proposal",
    "financial_proposal",
    "work_order",
    "deliverables",
    "completion_certificate",
    "invoices_receipts",
    "other_documents"
]


def _ensure_dict(value) -> Dict[str, Any]:
    """Ensure a JSON field is a dictionary."""
    if not value:
        return {}
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            return {}
    return {}


def _from_json_filter(value):
    """Jinja filter to safely parse JSON strings in templates."""
    if value is None:
        return []

    if isinstance(value, (list, dict)):
        return value

    try:
        parsed = json.loads(value)
        return parsed
    except (json.JSONDecodeError, TypeError):
        return value


def _indian_currency_filter(value):
    """Jinja filter to format numbers in Indian currency format (lakhs/crores)."""
    if value is None or value == '':
        return 'N/A'

    try:
        # Convert to float
        num = float(value)

        # Handle negative numbers
        is_negative = num < 0
        num = abs(num)

        # Convert to integer for formatting
        num_str = str(int(num))

        # Handle decimal part
        decimal_part = ''
        if '.' in str(value):
            decimal_part = '.' + str(value).split('.')[1][:2]

        # Indian numbering system formatting
        if len(num_str) <= 3:
            result = num_str
        elif len(num_str) <= 5:
            result = num_str[:-3] + ',' + num_str[-3:]
        elif len(num_str) <= 7:
            result = num_str[:-5] + ',' + num_str[-5:-3] + ',' + num_str[-3:]
        else:
            # For crores and above
            result = num_str[:-7] + ',' + num_str[-7:-5] + ',' + num_str[-5:-3] + ',' + num_str[-3:]
            # Add more commas for larger numbers
            remaining = num_str[:-7]
            if len(remaining) > 2:
                formatted_remaining = ''
                for i, digit in enumerate(reversed(remaining)):
                    if i > 0 and i % 2 == 0:
                        formatted_remaining = ',' + formatted_remaining
                    formatted_remaining = digit + formatted_remaining
                result = formatted_remaining + ',' + num_str[-7:-5] + ',' + num_str[-5:-3] + ',' + num_str[-3:]

        result = result + decimal_part

        if is_negative:
            result = '-' + result

        return result
    except (ValueError, TypeError):
        return str(value)


templates.env.filters["fromjson"] = _from_json_filter
templates.env.filters["indian_currency"] = _indian_currency_filter

CURRENCY_SANITIZE_RE = re.compile(r"[^\d\.\-]")


def _coerce_json_structure(value: Any) -> Any:
    """Convert JSON-like strings into native Python dict/list when possible."""
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            try:
                return ast.literal_eval(value)
            except (ValueError, SyntaxError):
                return value
    return value


def _extract_json_field_value(value: Any, keys: List[str]) -> Any:
    """Fetch the first non-empty value from a JSON-like field using preferred keys."""
    source: Any = _coerce_json_structure(value)

    if isinstance(source, dict):
        for key in keys:
            candidate = source.get(key)
            if candidate not in (None, "", "N/A", "NA"):
                return candidate
        return None

    return value


def _format_currency_display(value: Any) -> str:
    """Format currency values with Indian number formatting, preserving raw text when needed."""
    if value in (None, "", "N/A", "NA"):
        return "N/A"

    if isinstance(value, (int, float)):
        amount = float(value)
    else:
        text = str(value).strip()
        if not text or text.upper() in {"N/A", "NA"}:
            return "N/A"
        if "%" in text:
            return text
        cleaned = CURRENCY_SANITIZE_RE.sub("", text)
        if not cleaned:
            return text
        try:
            amount = float(cleaned)
        except ValueError:
            return text

    formatted = _indian_currency_filter(amount)
    if not formatted or formatted == "N/A":
        return str(value)
    if not formatted.startswith("₹"):
        formatted = f"₹{formatted}"
    return formatted


def _extract_numeric_emd_from_jsonb(emd_fee_details) -> Optional[float]:
    """
    Extract numeric EMD value from JSONB emd_fee_details for filtering.
    Returns None if EMD is not found or cannot be parsed.
    """
    if not emd_fee_details:
        return None
    
    # Try to extract from various possible keys
    emd_keys = [
        "EMD Amount (INR)",
        "EMD Amount in ₹",
        "EMD Amount",
        "Amount",
        "EMD Fee",
        "emd_amount",
    ]
    
    emd_value = _extract_json_field_value(emd_fee_details, emd_keys)
    
    if emd_value is None or emd_value in ("", "N/A", "NA"):
        return None
    
    # Convert to numeric value
    if isinstance(emd_value, (int, float)):
        return float(emd_value)
    
    # Try to parse string value
    text = str(emd_value).strip()
    if not text or text.upper() in {"N/A", "NA"}:
        return None
    
    # Remove currency symbols and non-numeric characters (except decimal point and minus)
    cleaned = CURRENCY_SANITIZE_RE.sub("", text)
    if not cleaned:
        return None
    
    try:
        return float(cleaned)
    except ValueError:
        return None


def _prepare_tender_display_fields(tender: Optional[TenderDB]) -> None:
    """Attach pre-formatted display fields (EMD, tender fee) to a tender instance."""
    if tender is None:
        return

    coerced_emd = _coerce_json_structure(tender.emd_fee_details)
    if isinstance(coerced_emd, dict):
        tender.emd_fee_details = coerced_emd

    coerced_fee = _coerce_json_structure(tender.tender_fee_details)
    if isinstance(coerced_fee, dict):
        tender.tender_fee_details = coerced_fee

    emd_raw = _extract_json_field_value(
        tender.emd_fee_details,
        [
            "EMD Amount (INR)",
            "EMD Amount in ₹",
            "EMD Amount",
            "Amount",
            "EMD Fee",
            "emd_amount",
        ],
    )
    tender.display_emd_amount = _format_currency_display(emd_raw)

    fee_raw = _extract_json_field_value(
        tender.tender_fee_details,
        [
            "Tender Fee (INR)",
            "Tender Fee in ₹",
            "Tender Fee",
            "Fee",
            "Amount",
            "tender_fee",
        ],
    )
    tender.display_tender_fee = _format_currency_display(fee_raw)

# =============================================================================
# MODULAR CODE ORGANIZATION
# Core utilities have been moved to core/ directory for better maintainability
# =============================================================================

# Import core utilities (security, dependencies)
from core.security import (
    hash_password,
    verify_password,
    create_session,
    delete_session,
    user_sessions,
    update_session,
    create_expert_session,
    get_expert_session,
    delete_expert_session,
    expert_sessions,
    SESSION_EXPIRE_DAYS
)
from core.dependencies import (
    get_db,
    get_current_user,
    get_current_employee,
    user_has_complete_company_details,
    require_company_details,
    require_pin_verification,
    get_current_expert,
    expert_has_complete_profile,
    require_expert_login,
    require_expert_profile_complete,
    enforce_test_quarantine
)

# Security
security = HTTPBearer(auto_error=False)

# =============================================================================
# CUSTOM ERROR HANDLERS
# =============================================================================

@app.exception_handler(404)
async def custom_404_handler(request: Request, exc: HTTPException):
    """Custom 404 error page handler."""
    db = SessionLocal()
    try:
        current_user = get_current_user(request, db)
        current_employee = get_current_employee(request, db)
        current_expert = get_current_expert(request, db)

        return templates.TemplateResponse(
            "404.html",
            {
                "request": request,
                "current_user": current_user,
                "current_employee": current_employee,
                "current_expert": current_expert,
                "selected_font": get_active_font()
            },
            status_code=404
        )
    finally:
        db.close()

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """Custom HTTP exception handler for all errors."""
    # For API endpoints, return JSON
    if request.url.path.startswith("/api/"):
        return JSONResponse(
            status_code=exc.status_code,
            content={"detail": exc.detail}
        )

    # For 404 errors, use custom 404 page
    if exc.status_code == 404:
        db = SessionLocal()
        try:
            current_user = get_current_user(request, db)
            current_employee = get_current_employee(request, db)
            current_expert = get_current_expert(request, db)

            return templates.TemplateResponse(
                "404.html",
                {
                    "request": request,
                    "current_user": current_user,
                    "current_employee": current_employee,
                    "current_expert": current_expert,
                    "selected_font": get_active_font()
                },
                status_code=404
            )
        finally:
            db.close()

    # For other HTTP errors, return JSON
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail}
    )

# Helper function for deadline notifications
def create_deadline_notifications(user_id: str, tender_id: str, db: Session):
    """
    Create deadline notifications for a tender at specific intervals.
    Called when a tender is favorited or shortlisted.
    """
    # Get the tender
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender or not tender.deadline:
        return

    # Calculate days until deadline
    now = datetime.utcnow()
    days_until_deadline = (tender.deadline - now).days

    # Define notification thresholds and their types
    notification_thresholds = [
        (10, 'deadline_10d', '10 days'),
        (7, 'deadline_7d', '7 days'),
        (5, 'deadline_5d', '5 days'),
        (2, 'deadline_2d', '2 days'),
        (0, 'deadline_today', 'today')
    ]

    # Create notifications for relevant thresholds
    for threshold_days, notif_type, display_text in notification_thresholds:
        # Only create notification if we're at or past this threshold
        if days_until_deadline <= threshold_days:
            # Check if notification already exists
            existing = db.query(NotificationDB).filter(
                NotificationDB.tender_id == tender_id,
                NotificationDB.notification_type == notif_type
            ).first()

            if not existing:
                # Create notification
                message = f"{tender.title or 'Tender'} - Due in {display_text}"
                if threshold_days == 0:
                    message = f"{tender.title or 'Tender'} - Expires today!"

                notification = NotificationDB(
                    user_id=user_id,
                    tender_id=tender_id,
                    notification_type=notif_type,
                    message=message,
                    tender_title=tender.title,
                    days_remaining=max(0, days_until_deadline),
                    is_read=False
                )
                db.add(notification)

    db.commit()

def check_and_remove_expired_tenders(db: Session):
    """
    Central helper that delegates to the database level cleanup routine so
    that routes/tasks can ensure expired tenders are removed and users are
    notified consistently.
    """
    return cleanup_expired_tenders(db)


# =============================================================================
# TENDER-TO-PROJECT CONVERSION UTILITIES
# =============================================================================

def transfer_tender_documents_to_project(db: Session, tender_id: str, project_id: int, shortlist_id: Optional[int] = None) -> dict:
    """
    Transfer tender documents from database to file system and categorize them.
    Also transfers stage documents uploaded during project tracking if shortlist_id is provided.

    Args:
        db: Database session
        tender_id: Source tender ID
        project_id: Destination project ID
        shortlist_id: Optional shortlist ID for transferring stage documents

    Returns:
        dict with structure: {category: [file_paths], ...} and transfer summary
    """
    # Document type keyword mapping for smart categorization
    DOCUMENT_KEYWORDS = {
        "tender_documents": ["tender", "nit", "notice", "corrigendum", "addendum", "clarification", "rfp", "rfq"],
        "technical_proposal": ["technical", "proposal", "methodology", "approach"],
        "financial_proposal": ["financial", "price", "bid", "quotation", "commercial", "cost"],
        "work_order": ["work order", "loi", "letter of intent", "award", "acceptance", "contract"],
        "deliverables": ["boq", "bill of quantities", "deliverable", "milestone", "report", "submission"],
        "completion_certificate": ["completion", "certificate", "handover", "closing"],
        "invoices_receipts": ["invoice", "receipt", "payment", "bill", "voucher"],
        "other_documents": []  # default category
    }

    # Query all tender documents
    tender_docs = db.query(TenderDocumentDB).filter(
        TenderDocumentDB.tender_id == tender_id
    ).order_by(TenderDocumentDB.display_order).all()

    if not tender_docs:
        logger.info(f"No documents found for tender {tender_id}")
        return {"documents": {}, "count": 0, "categories": []}

    # Create project directory
    project_dir = Path("uploads/projects") / str(project_id)
    project_dir.mkdir(parents=True, exist_ok=True)

    documents_by_category = {}
    transfer_count = 0
    categories_used = set()

    for doc in tender_docs:
        try:
            # Determine category based on filename
            filename_lower = doc.filename.lower() if doc.filename else ""
            category = "other_documents"  # default

            for cat, keywords in DOCUMENT_KEYWORDS.items():
                if any(keyword in filename_lower for keyword in keywords):
                    category = cat
                    break

            # Create category directory
            category_dir = project_dir / category
            category_dir.mkdir(parents=True, exist_ok=True)

            # Generate unique filename if duplicate
            target_file = category_dir / doc.filename
            counter = 1
            while target_file.exists():
                name_parts = doc.filename.rsplit('.', 1)
                if len(name_parts) == 2:
                    target_file = category_dir / f"{name_parts[0]}_{counter}.{name_parts[1]}"
                else:
                    target_file = category_dir / f"{doc.filename}_{counter}"
                counter += 1

            # Write binary data to file
            with open(target_file, 'wb') as f:
                f.write(doc.file_data)

            # Add to category list
            relative_path = str(target_file.relative_to(Path("uploads")))
            if category not in documents_by_category:
                documents_by_category[category] = []
            documents_by_category[category].append(relative_path)

            categories_used.add(category)
            transfer_count += 1
            logger.info(f"Transferred document {doc.filename} to category {category}")

        except Exception as e:
            logger.error(f"Failed to transfer document {doc.id} ({doc.filename}): {e}")
            # Continue with other documents
            continue

    # Transfer stage documents if shortlist_id is provided
    if shortlist_id:
        stage_docs = db.query(StageDocumentDB).filter(
            StageDocumentDB.shortlist_id == shortlist_id
        ).order_by(StageDocumentDB.step_number, StageDocumentDB.uploaded_at).all()

        # Stage-to-category mapping
        # Step 1: Bid Preparation -> technical_proposal
        # Step 2: Bid Submission -> financial_proposal
        # Step 3: Bid Opening -> tender_documents
        # Step 4: Technical Evaluation -> technical_proposal
        # Step 5: Financial Opening -> financial_proposal
        # Step 6: Award/Contract -> work_order
        STAGE_TO_CATEGORY = {
            1: "technical_proposal",   # Bid Preparation
            2: "financial_proposal",   # Bid Submission
            3: "tender_documents",     # Bid Opening
            4: "technical_proposal",   # Technical Evaluation
            5: "financial_proposal",   # Financial Opening
            6: "work_order"            # Award/Contract
        }

        for stage_doc in stage_docs:
            try:
                # Determine category based on step number and title
                category = STAGE_TO_CATEGORY.get(stage_doc.step_number, "other_documents")

                # Check title for additional clues
                title_lower = stage_doc.title.lower() if stage_doc.title else ""
                for cat, keywords in DOCUMENT_KEYWORDS.items():
                    if any(keyword in title_lower for keyword in keywords):
                        category = cat
                        break

                # Create category directory
                category_dir = project_dir / category
                category_dir.mkdir(parents=True, exist_ok=True)

                # Copy file from stage_documents to project directory
                source_path = Path(stage_doc.file_path)
                if source_path.exists():
                    # Generate unique filename if duplicate
                    target_file = category_dir / stage_doc.filename
                    counter = 1
                    while target_file.exists():
                        name_parts = stage_doc.filename.rsplit('.', 1)
                        if len(name_parts) == 2:
                            target_file = category_dir / f"{name_parts[0]}_{counter}.{name_parts[1]}"
                        else:
                            target_file = category_dir / f"{stage_doc.filename}_{counter}"
                        counter += 1

                    # Copy the file
                    shutil.copy2(source_path, target_file)

                    # Add to category list
                    relative_path = str(target_file.relative_to(Path("uploads")))
                    if category not in documents_by_category:
                        documents_by_category[category] = []
                    documents_by_category[category].append(relative_path)

                    categories_used.add(category)
                    transfer_count += 1
                    logger.info(f"Transferred stage document {stage_doc.filename} (step {stage_doc.step_number}) to category {category}")
                else:
                    logger.warning(f"Stage document file not found: {stage_doc.file_path}")

            except Exception as e:
                logger.error(f"Failed to transfer stage document {stage_doc.id} ({stage_doc.filename}): {e}")
                continue

    return {
        "documents": documents_by_category,
        "count": transfer_count,
        "categories": list(categories_used)
    }


def auto_create_project_from_tender(db: Session, tender: TenderDB, user_id: str, shortlist_id: Optional[int] = None) -> Optional[ProjectDB]:
    """
    Automatically create a ProjectDB entry from an awarded tender.

    Args:
        db: Database session
        tender: Awarded tender object
        user_id: User ID who won the tender
        shortlist_id: Optional shortlist ID for transferring stage documents

    Returns:
        Created ProjectDB object or None if failed
    """
    try:
        # Extract client name from tender_inviting_authority or use authority
        client_name = tender.authority or "Unknown Client"
        if tender.tender_inviting_authority and isinstance(tender.tender_inviting_authority, dict):
            client_name = tender.tender_inviting_authority.get("Name", client_name)

        # Build project name from tender title
        project_name = tender.title[:255] if tender.title else f"Project from Tender #{tender.tender_reference_number}"

        # Handle duplicate project names
        original_name = project_name
        counter = 1
        while db.query(ProjectDB).filter(
            ProjectDB.user_id == user_id,
            ProjectDB.project_name == project_name
        ).first():
            if tender.tender_reference_number:
                project_name = f"{original_name} (Tender #{tender.tender_reference_number})"
            else:
                project_name = f"{original_name} ({counter})"
            counter += 1
            if counter > 10:  # Safety limit
                import uuid
                project_name = f"{original_name} ({uuid.uuid4().hex[:8]})"
                break

        # Extract location data
        states_list = []
        cities_list = []
        if tender.state and tender.state != "Unknown":
            states_list = [tender.state]

        if tender.work_item_details and isinstance(tender.work_item_details, dict):
            location = tender.work_item_details.get("Location", "")
            if location and location != "NA":
                # Try to parse location into cities
                cities_list = [loc.strip() for loc in location.split(",")]

        # Build project description
        description_parts = []
        if tender.summary:
            description_parts.append(tender.summary)
        if tender.tender_reference_number:
            description_parts.append(f"\n\nOriginal Tender Reference: {tender.tender_reference_number}")
        if tender.source:
            description_parts.append(f"Source Portal: {tender.source}")

        project_description = "\n".join(description_parts) if description_parts else None

        # Generate project_id
        try:
            project_id_value = generate_project_id(user_id, db)
        except Exception as e:
            logger.error(f"Error generating project_id for auto-created project: {e}")
            # Fallback: generate ID even if there's an error
            project_id_value = generate_project_id(user_id, db)

        # Create project with auto-filled fields
        new_project = ProjectDB(
            user_id=user_id,
            project_name=project_name,
            project_description=project_description,
            client_name=client_name,
            project_cost=tender.estimated_value if tender.estimated_value else None,
            country="India",  # default
            states=states_list if states_list else None,
            cities=cities_list if cities_list else None,
            start_date=tender.awarded_at if tender.awarded_at else datetime.utcnow(),
            # Leave blank for user to fill:
            # consultancy_fee, sector, sub_sector, end_date, project_duration_months,
            # financing_authority, jv_partner, services_rendered
            source_tender_id=tender.id,
            is_auto_generated=True,
            completion_status="incomplete",
            documents={},  # Will be filled by document transfer
            project_id=project_id_value
        )

        db.add(new_project)
        db.flush()  # Get project.id without committing

        # Transfer documents (including stage documents if shortlist_id provided)
        logger.info(f"Transferring documents for project {new_project.id} from tender {tender.id}")
        transfer_result = transfer_tender_documents_to_project(db, tender.id, new_project.id, shortlist_id)

        # Update project documents
        new_project.documents = transfer_result["documents"]

        logger.info(
            f"Auto-created project {new_project.id} ('{new_project.project_name}') "
            f"from tender {tender.id}. Transferred {transfer_result['count']} documents "
            f"across {len(transfer_result['categories'])} categories."
        )

        return new_project

    except Exception as e:
        logger.error(f"Failed to auto-create project from tender {tender.id}: {e}", exc_info=True)
        return None


def finalize_tender_award_for_user(db: Session, tender: TenderDB, awarding_user_id: Optional[str] = None) -> Optional[ProjectDB]:
    """
    Mark a tender as awarded and remove it from all favourites/shortlists.
    Automatically creates a project from the awarded tender.

    Args:
        db: Database session
        tender: Tender to award
        awarding_user_id: User ID who is awarding the tender

    Returns:
        Created ProjectDB object if successful, None otherwise
    """
    now = datetime.utcnow()

    if not tender.awarded:
        tender.awarded = True

    if tender.awarded_at is None:
        tender.awarded_at = now

    if awarding_user_id:
        tender.awarded_by = awarding_user_id

    # Find the user's shortlist record BEFORE deleting (to get stage documents)
    user_shortlist_id = None
    if awarding_user_id:
        user_shortlist = db.query(ShortlistedTenderDB).filter(
            ShortlistedTenderDB.tender_id == tender.id,
            ShortlistedTenderDB.user_id == awarding_user_id
        ).first()
        if user_shortlist:
            user_shortlist_id = user_shortlist.id

    # Auto-create project from tender (with error handling)
    project = None
    if awarding_user_id:
        try:
            project = auto_create_project_from_tender(db, tender, awarding_user_id, user_shortlist_id)
            if project:
                db.commit()  # Commit project creation separately
                logger.info(f"Successfully auto-created project {project.id} from tender {tender.id} (shortlist_id: {user_shortlist_id})")
            else:
                logger.warning(f"Project creation returned None for tender {tender.id}")
        except Exception as e:
            logger.error(f"Failed to auto-create project from tender {tender.id}: {e}", exc_info=True)
            db.rollback()  # Rollback project creation but continue with award
            project = None

    # Remove from favorites, shortlists, and notifications
    db.query(FavoriteDB).filter(FavoriteDB.tender_id == tender.id).delete(synchronize_session=False)

    shortlist_ids = [row[0] for row in db.query(ShortlistedTenderDB.id).filter(ShortlistedTenderDB.tender_id == tender.id).all()]

    if shortlist_ids:
        db.query(StageDocumentDB).filter(StageDocumentDB.shortlist_id.in_(shortlist_ids)).delete(synchronize_session=False)
        db.query(ShortlistedTenderDB).filter(ShortlistedTenderDB.id.in_(shortlist_ids)).delete(synchronize_session=False)

    db.query(NotificationDB).filter(
        NotificationDB.tender_id == tender.id,
        NotificationDB.notification_type.like('deadline_%')
    ).delete(synchronize_session=False)

    return project


def enforce_awarded_tender_state(db: Session) -> int:
    """Ensure awarded tenders no longer appear in favourites/shortlists."""
    awarded_tenders = db.query(TenderDB).filter(TenderDB.awarded == True).all()
    adjustments = 0

    for tender in awarded_tenders:
        # If any favourites or shortlist links remain, remove them.
        has_favorites = db.query(FavoriteDB).filter(FavoriteDB.tender_id == tender.id).count() > 0
        has_shortlisted = db.query(ShortlistedTenderDB).filter(ShortlistedTenderDB.tender_id == tender.id).count() > 0

        if has_favorites or has_shortlisted:
            finalize_tender_award_for_user(db, tender, tender.awarded_by)
            adjustments += 1

    if adjustments:
        db.commit()

    return adjustments

def check_all_deadline_notifications(db: Session):
    """
    Check all favorited and shortlisted tenders for deadline notifications.
    This should be run periodically (e.g., daily cron job).
    """
    # First, remove expired tenders
    expired_count = check_and_remove_expired_tenders(db)
    enforce_awarded_tender_state(db)

    # Get all active favorites
    favorites = db.query(FavoriteDB).all()
    for fav in favorites:
        create_deadline_notifications(fav.user_id, fav.tender_id, db)

    # Get all shortlisted tenders
    shortlisted = db.query(ShortlistedTenderDB).all()
    for shortlist in shortlisted:
        create_deadline_notifications(shortlist.user_id, shortlist.tender_id, db)

    return expired_count

# HTML Routes
@app.get("/", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db)):
    """Home page with tender listings."""
    # Redirect logged-in users to home page
    current_user = get_current_user(request, db)
    if current_user:
        return RedirectResponse(url="/home", status_code=302)

    # Get search and filter parameters
    search = request.query_params.get('search', '')
    category = request.query_params.get('category', '')
    state = request.query_params.get('state', '')
    source = request.query_params.get('source', '')
    min_value = request.query_params.get('min_value', '')
    max_value = request.query_params.get('max_value', '')
    page = int(request.query_params.get('page', 1))
    per_page = 20

    # Check if any filters are applied
    filters_applied = any([search, category, state, source, min_value, max_value])

    tenders = []
    total_tenders = 0
    total_pages = 0
    has_prev = False
    has_next = False

    if filters_applied:
        # Build query only if filters are applied
        query = db.query(TenderDB)

        # Filter out expired tenders (unless awarded)
        # Keep tenders that either:
        # 1. Have no deadline (NULL)
        # 2. Have a future deadline
        # 3. Have been awarded (regardless of deadline)
        now = datetime.utcnow()
        query = query.filter(
            or_(
                TenderDB.deadline == None,
                TenderDB.deadline >= now,
                TenderDB.awarded == True
            )
        )

        # Apply filters (comprehensive search across all tender fields)
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    TenderDB.title.ilike(search_term),
                    TenderDB.summary.ilike(search_term),
                    TenderDB.authority.ilike(search_term),
                    TenderDB.category.ilike(search_term),
                    TenderDB.organisation_chain.ilike(search_term),
                    TenderDB.tender_reference_number.ilike(search_term),
                    TenderDB.tender_id.ilike(search_term),
                    TenderDB.tender_type.ilike(search_term),
                    TenderDB.tender_category.ilike(search_term),
                    TenderDB.state.ilike(search_term),
                    TenderDB.source.ilike(search_term),
                    TenderDB.work_item_details.cast(String).ilike(search_term),
                    TenderDB.tender_inviting_authority.cast(String).ilike(search_term),
                    TenderDB.tags.cast(String).ilike(search_term)
                )
            )

        if category:
            query = query.filter(TenderDB.category == category)

        if state:
            query = query.filter(TenderDB.state == state)

        if source:
            query = query.filter(TenderDB.source == source)

        if min_value:
            try:
                query = query.filter(TenderDB.estimated_value >= float(min_value))
            except ValueError:
                pass

        if max_value:
            try:
                query = query.filter(TenderDB.estimated_value <= float(max_value))
            except ValueError:
                pass

        # Get total count for pagination
        total_tenders = query.count()

        # Apply pagination and ordering
        tenders = query.order_by(desc(TenderDB.published_at)).offset((page - 1) * per_page).limit(per_page).all()

        # Pagination info
        total_pages = (total_tenders + per_page - 1) // per_page
        has_prev = page > 1
        has_next = page < total_pages

    # Get current user
    current_user = get_current_user(request, db)

    # Get unique categories and states for filters (cached)
    from core.cache import filter_cache

    filters = filter_cache.get_filter_options(db)
    categories = sorted(filters['categories'])
    states = sorted(filters['states'])
    sources = sorted(filters['sources'])

    return templates.TemplateResponse("home_2.html", {
        "request": request,
        "tenders": tenders,
        "current_user": current_user,
        "categories": sorted(categories),
        "states": sorted(states),
        "sources": sorted(sources),
        "search": search,
        "category": category,
        "state": state,
        "source": source,
        "min_value": min_value,
        "max_value": max_value,
        "page": page,
        "total_pages": total_pages,
        "has_prev": has_prev,
        "has_next": has_next,
        "total_tenders": total_tenders,
        "selected_font": get_active_font()
    })

@app.get("/robots.txt", response_class=Response)
def robots_txt():
    """Generate robots.txt file."""
    from core.seo_config import SITE_URL, NOINDEX_PAGES
    
    robots_content = f"""User-agent: *
Allow: /
Disallow: /api/
Disallow: /login
Disallow: /employee/login
Disallow: /expert/login
Disallow: /dashboard
Disallow: /employee/dashboard
Disallow: /expert/dashboard
Disallow: /profile
Disallow: /expert/profile

Sitemap: {SITE_URL}/sitemap.xml
"""
    return Response(content=robots_content, media_type="text/plain")

@app.get("/sitemap.xml", response_class=Response)
def sitemap_xml(request: Request, db: Session = Depends(get_db)):
    """Generate dynamic sitemap.xml."""
    from core.seo_config import SITE_URL, SITEMAP_CONFIG
    from datetime import datetime
    
    # Get current time for lastmod
    now = datetime.utcnow().strftime("%Y-%m-%d")
    
    # Start building sitemap
    sitemap_urls = []
    
    # Static pages
    static_pages = [
        {"loc": "/", "changefreq": "daily", "priority": "1.0"},
        {"loc": "/login", "changefreq": "monthly", "priority": "0.3"},
        {"loc": "/employee/login", "changefreq": "monthly", "priority": "0.3"},
        {"loc": "/expert/login", "changefreq": "monthly", "priority": "0.3"},
    ]
    
    for page in static_pages:
        sitemap_urls.append(f"""  <url>
    <loc>{SITE_URL}{page['loc']}</loc>
    <lastmod>{now}</lastmod>
    <changefreq>{page['changefreq']}</changefreq>
    <priority>{page['priority']}</priority>
  </url>""")
    
    # Dynamic tender pages (limit to recent/active tenders for performance)
    try:
        # Get active tenders (not expired, not awarded, or recently published)
        active_tenders = db.query(TenderDB).filter(
            or_(
                TenderDB.deadline.is_(None),
                TenderDB.deadline >= datetime.utcnow() - timedelta(days=30),
                TenderDB.awarded == True
            )
        ).order_by(desc(TenderDB.published_at)).limit(1000).all()
        
        for tender in active_tenders:
            lastmod = tender.published_at.strftime("%Y-%m-%d") if tender.published_at else now
            sitemap_urls.append(f"""  <url>
    <loc>{SITE_URL}/tender/{tender.id}</loc>
    <lastmod>{lastmod}</lastmod>
    <changefreq>weekly</changefreq>
    <priority>0.8</priority>
  </url>""")
    except Exception as e:
        logger.warning(f"Error generating tender sitemap entries: {e}")
    
    # Build complete sitemap XML
    sitemap_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
{''.join(sitemap_urls)}
</urlset>"""
    
    return Response(content=sitemap_xml, media_type="application/xml")

@app.get("/procurement", response_class=HTMLResponse)
@require_company_details
@require_pin_verification
async def procurement(request: Request, db: Session = Depends(get_db)):
    """Procurement page with tender listings (login required)."""
    from core.dependencies import get_current_user_or_bd_employee

    # Check if user or BD employee is logged in
    entity, entity_type = get_current_user_or_bd_employee(request, db)
    if not entity:
        # No user or BD employee logged in
        current_user = None
        return RedirectResponse(url="/login", status_code=302)

    # Set current_user for backward compatibility with template
    current_user = entity if entity_type == 'user' else None
    current_bd_employee = entity if entity_type == 'bd_employee' else None

    # Clean up expired tenders and orphaned records before displaying
    try:
        check_and_remove_expired_tenders(db)
        enforce_awarded_tender_state(db)
        cleanup_orphaned_records(db)
    except Exception as e:
        logger.warning(f"Cleanup of expired tenders and orphaned records failed: {e}")

    # Get search and filter parameters
    search = request.query_params.get('search', '')
    category = request.query_params.get('category', '')
    state = request.query_params.get('state', '')
    source = request.query_params.get('source', '')
    product_category = request.query_params.get('product_category', '').strip()
    
    # Debug logging for Product Category filter
    if product_category:
        logger.info(f"Product Category filter requested: '{product_category}'")
    min_value = request.query_params.get('min_value', '')
    max_value = request.query_params.get('max_value', '')
    min_emd = request.query_params.get('min_emd', '')
    max_emd = request.query_params.get('max_emd', '')
    sort_by = request.query_params.get('sort_by', 'published_desc')
    show_all = request.query_params.get('show_all', '')
    page = int(request.query_params.get('page', 1))
    per_page = 20

    # Check if any filters are applied or show_all is requested
    filters_applied = any([search, category, state, source, product_category, min_value, max_value, min_emd, max_emd, show_all])

    tenders = []
    total_tenders = 0
    total_pages = 0
    has_prev = False
    has_next = False

    if filters_applied:
        # Build query only if filters are applied
        query = db.query(TenderDB)

        # Filter out expired tenders and awarded tenders
        # Awarded tenders should only appear in tender_management page
        # Keep tenders that:
        # 1. Are not awarded AND
        # 2. Either have no deadline (NULL) OR have a future deadline
        now = datetime.utcnow()
        query = query.filter(
            TenderDB.awarded == False,  # Exclude awarded tenders from search
            or_(
                TenderDB.deadline == None,
                TenderDB.deadline >= now
            )
        )

        # Apply filters (comprehensive search across all tender fields)
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    TenderDB.title.ilike(search_term),
                    TenderDB.summary.ilike(search_term),
                    TenderDB.authority.ilike(search_term),
                    TenderDB.category.ilike(search_term),
                    TenderDB.organisation_chain.ilike(search_term),
                    TenderDB.tender_reference_number.ilike(search_term),
                    TenderDB.tender_id.ilike(search_term),
                    TenderDB.tender_type.ilike(search_term),
                    TenderDB.tender_category.ilike(search_term),
                    TenderDB.state.ilike(search_term),
                    TenderDB.source.ilike(search_term),
                    TenderDB.work_item_details.cast(String).ilike(search_term),
                    TenderDB.tender_inviting_authority.cast(String).ilike(search_term),
                    TenderDB.tags.cast(String).ilike(search_term)
                )
            )

        if category:
            query = query.filter(TenderDB.category == category)

        if state:
            query = query.filter(TenderDB.state == state)

        if source:
            query = query.filter(TenderDB.source == source)

        if product_category:
            # Filter by Product Category from work_item_details JSONB
            # Handle both "Product Category" and "PRODUCT CATEGORY" key variations
            # Use case-insensitive matching to handle variations in stored values
            product_category_lower = product_category.lower().strip()
            logger.info(f"Filtering by Product Category: '{product_category}' (normalized: '{product_category_lower}')")
            
            # Use PostgreSQL TRIM and LOWER functions for case-insensitive, whitespace-tolerant matching
            # func.trim() maps to PostgreSQL's TRIM() function which removes leading/trailing whitespace
            query = query.filter(
                or_(
                    func.lower(func.trim(TenderDB.work_item_details['Product Category'].astext)) == product_category_lower,
                    func.lower(func.trim(TenderDB.work_item_details['PRODUCT CATEGORY'].astext)) == product_category_lower
                )
            )
            
            # Debug: Log count before pagination to verify filter is working
            count_before_pagination = query.count()
            logger.info(f"Product Category filter '{product_category}' matched {count_before_pagination} tenders")

        if min_value:
            try:
                query = query.filter(TenderDB.estimated_value >= float(min_value))
            except ValueError:
                pass

        if max_value:
            try:
                query = query.filter(TenderDB.estimated_value <= float(max_value))
            except ValueError:
                pass

        # EMD filtering - only show tenders with EMD when EMD filters are applied
        if min_emd or max_emd:
            # Extract numeric EMD value from JSONB using PostgreSQL functions
            # Try multiple possible keys in order of preference
            emd_value_expr = None
            emd_keys = [
                "EMD Amount (INR)",
                "EMD Amount in ₹",
                "EMD Amount",
                "Amount",
                "EMD Fee",
                "emd_amount",
            ]
            
            # Build COALESCE chain to try each key
            for i, key in enumerate(emd_keys):
                # Extract value and clean it (remove non-numeric chars except decimal and minus)
                key_expr = func.regexp_replace(
                    func.regexp_replace(
                        TenderDB.emd_fee_details[key].astext,
                        r'[^\d\.\-]', '', 'g'
                    ),
                    r'^$', '', 'g'
                )
                
                if emd_value_expr is None:
                    emd_value_expr = func.nullif(key_expr, '')
                else:
                    emd_value_expr = func.coalesce(emd_value_expr, func.nullif(key_expr, ''))
            
            # Filter out tenders without EMD (where emd_value_expr is NULL or empty)
            query = query.filter(emd_value_expr.isnot(None))
            query = query.filter(emd_value_expr != '')
            
            # Apply min_emd filter
            if min_emd:
                try:
                    min_emd_float = float(min_emd)
                    # Cast the cleaned EMD value to numeric for comparison
                    emd_numeric = cast(emd_value_expr, Numeric)
                    query = query.filter(emd_numeric >= min_emd_float)
                except (ValueError, Exception) as e:
                    logger.warning(f"Error parsing min_emd: {e}")
            
            # Apply max_emd filter
            if max_emd:
                try:
                    max_emd_float = float(max_emd)
                    # Cast the cleaned EMD value to numeric for comparison
                    emd_numeric = cast(emd_value_expr, Numeric)
                    query = query.filter(emd_numeric <= max_emd_float)
                except (ValueError, Exception) as e:
                    logger.warning(f"Error parsing max_emd: {e}")

        # Get total count for pagination
        total_tenders = query.count()

        # Apply sorting based on sort_by parameter
        if sort_by == 'published_asc':
            query = query.order_by(TenderDB.published_at.asc())
        elif sort_by == 'deadline_asc':
            query = query.order_by(TenderDB.deadline.asc().nullslast())
        elif sort_by == 'deadline_desc':
            query = query.order_by(TenderDB.deadline.desc().nullslast())
        elif sort_by == 'value_desc':
            query = query.order_by(TenderDB.estimated_value.desc().nullslast())
        elif sort_by == 'value_asc':
            query = query.order_by(TenderDB.estimated_value.asc().nullslast())
        elif sort_by == 'title_asc':
            query = query.order_by(TenderDB.title.asc())
        elif sort_by == 'title_desc':
            query = query.order_by(TenderDB.title.desc())
        else:  # Default: published_desc
            query = query.order_by(TenderDB.published_at.desc())

        # Apply pagination
        tenders = query.offset((page - 1) * per_page).limit(per_page).all()

        # Pagination info
        total_pages = (total_tenders + per_page - 1) // per_page
        has_prev = page > 1
        has_next = page < total_pages

    # Get unique categories and states for filters (cached)
    from core.cache import filter_cache

    filters = filter_cache.get_filter_options(db)
    categories = sorted(filters['categories'])
    states = sorted(filters['states'])
    sources = sorted(filters['sources'])
    product_categories = filters.get('product_categories', [])

    # Get company sectors (for user or BD employee)
    company_details = None
    user_sectors_data = []

    if current_user:
        # Regular user
        company_details = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()
    elif current_bd_employee:
        # BD employee - get company via company_code
        company_code = db.query(CompanyCodeDB).filter(CompanyCodeDB.id == current_bd_employee.company_code_id).first()
        if company_code:
            company_details = db.query(CompanyDB).filter(CompanyDB.user_id == company_code.user_id).first()

    if company_details and company_details.industry_sector:
        if company_details.industry_sector.startswith('[{'):
            try:
                user_sectors_data = json.loads(company_details.industry_sector)
            except json.JSONDecodeError:
                user_sectors_data = []

    # Get seen tender IDs for the current entity
    seen_tender_ids = set()
    if current_user:
        seen_records = db.query(SeenTenderDB).filter(SeenTenderDB.user_id == current_user.id).all()
        seen_tender_ids = {record.tender_id for record in seen_records}
    elif current_bd_employee:
        seen_records = db.query(SeenTenderDB).filter(SeenTenderDB.employee_id == current_bd_employee.id).all()
        seen_tender_ids = {record.tender_id for record in seen_records}

    return templates.TemplateResponse("tender_searching.html", {
        "request": request,
        "tenders": tenders,
        "current_user": current_user,
        "current_employee": current_bd_employee,  # Pass BD employee separately
        "categories": sorted(categories),
        "states": sorted(states),
        "sources": sorted(sources),
        "product_categories": product_categories,
        "user_sectors_data": user_sectors_data,
        "search": search,
        "category": category,
        "state": state,
        "source": source,
        "product_category": product_category,
        "min_value": min_value,
        "max_value": max_value,
        "min_emd": min_emd,
        "max_emd": max_emd,
        "sort_by": sort_by,
        "show_all": show_all,
        "page": page,
        "total_pages": total_pages,
        "has_prev": has_prev,
        "has_next": has_next,
        "total_tenders": total_tenders,
        "seen_tender_ids": seen_tender_ids,
        "selected_font": get_active_font()
    })


@app.get("/show-all-tneders", response_class=HTMLResponse)
async def show_all_tneders(
    request: Request,
    search: str = "",
    db: Session = Depends(get_db)
):
    """Temporary troubleshooting view that lists every tender without filters."""
    query = db.query(TenderDB)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                TenderDB.title.ilike(search_term),
                TenderDB.summary.ilike(search_term),
                TenderDB.tender_reference_number.ilike(search_term),
                TenderDB.tender_id.ilike(search_term),
                TenderDB.organisation_chain.ilike(search_term),
                TenderDB.authority.ilike(search_term)
            )
        )

    tenders = query.order_by(desc(TenderDB.scraped_at), desc(TenderDB.published_at)).all()

    return templates.TemplateResponse("show_all_tneders.html", {
        "request": request,
        "tenders": tenders,
        "search": search,
        "total_tenders": len(tenders),
        "selected_font": get_active_font()
    })


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Login/signup page."""
    current_user = get_current_user(request, next(get_db()))
    if current_user:
        return RedirectResponse(url="/home", status_code=302)

    # Handle error messages from failed login/signup attempts
    error_message = None
    signup_error_message = None
    error_param = request.query_params.get('error')

    if error_param == 'invalid_credentials':
        error_message = "Invalid email or password. Please check your credentials and try again."
    elif error_param == 'email_exists':
        signup_error_message = "Email already registered. Please use a different email or try logging in."

    return templates.TemplateResponse("login.html", {
        "request": request,
        "error": error_message,
        "signup_error": signup_error_message,
        "selected_font": get_active_font()
    })

@app.get("/employee/login", response_class=HTMLResponse)
async def employee_login_page(request: Request):
    """Employee login page."""
    current_employee = get_current_employee(request, next(get_db()))
    if current_employee:
        return RedirectResponse(url="/employee/dashboard", status_code=302)

    # Handle error messages from failed login attempts
    error_message = None
    error_param = request.query_params.get('error')

    if error_param == 'invalid_credentials':
        error_message = "Invalid email or password. Please contact your administrator if you've forgotten your credentials."

    return templates.TemplateResponse("employee_login.html", {
        "request": request,
        "error": error_message,
        "selected_font": get_active_font()
    })

@app.get("/dashboard", response_class=HTMLResponse)
@enforce_test_quarantine
@require_company_details
async def dashboard(request: Request, db: Session = Depends(get_db)):
    """User dashboard with favorites."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Clean up expired tenders before displaying
    try:
        check_and_remove_expired_tenders(db)
        enforce_awarded_tender_state(db)
        cleanup_orphaned_records(db)
    except Exception as e:
        logger.warning(f"Cleanup of expired tenders and orphaned records failed: {e}")
        # Roll back the failed transaction to restore session to valid state
        db.rollback()

    # Get user's favorite tenders with user_filled_data (using eager loading to prevent N+1 queries)
    from sqlalchemy.orm import joinedload

    now = datetime.utcnow()
    try:
        favorites = db.query(FavoriteDB).options(
            joinedload(FavoriteDB.tender)  # Eagerly load tender relationship
        ).filter(FavoriteDB.user_id == current_user.id).all()
    except Exception as e:
        logger.error(f"Error fetching favorites: {e}")
        db.rollback()
        favorites = []
    favorite_tenders = []

    for fav in favorites:
        if fav.tender:
            # Skip expired tenders (unless awarded)
            if fav.tender.deadline and fav.tender.deadline < now and not fav.tender.awarded:
                continue

            _prepare_tender_display_fields(fav.tender)

            # Merge tender data with user filled data
            tender_data = {
                'id': fav.tender.id,
                'favorite_id': fav.id,
                'created_at': fav.created_at,
                'status': fav.status,
                'user_filled_data': fav.user_filled_data or {},
                'notes': fav.notes,
                'tender': fav.tender
            }
            favorite_tenders.append(tender_data)

    # Get user's company details
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()

    # Get counts for tender management stats
    favorite_count = len(favorites)
    shortlisted_count = db.query(ShortlistedTenderDB).filter(
        ShortlistedTenderDB.user_id == current_user.id
    ).count()

    # Get project statistics
    all_projects = db.query(ProjectDB).filter(ProjectDB.user_id == current_user.id).all()
    total_projects = len(all_projects)
    completed_projects = len([p for p in all_projects if p.end_date is not None])

    # Calculate total project value (sum of consultancy fees)
    total_project_value = sum(
        p.consultancy_fee for p in all_projects
        if p.consultancy_fee is not None
    )

    # Get total tender count (excluding expired non-awarded tenders)
    now = datetime.utcnow()
    total_tender_count = db.query(TenderDB).filter(
        or_(
            TenderDB.deadline == None,
            TenderDB.deadline >= now,
            TenderDB.awarded == True
        )
    ).count()

    # Format tender count with k/l notation
    def format_tender_count(count):
        if count >= 100000:  # 1 lakh or more
            return f"{count / 100000:.1f}L".rstrip('0').rstrip('.')
        elif count >= 1000:  # 1k or more
            return f"{count / 1000:.1f}k".rstrip('0').rstrip('.')
        else:
            return str(count)

    formatted_tender_count = format_tender_count(total_tender_count)

    # Prepare calendar data - activities and deadlines
    from datetime import timedelta
    import calendar as cal_module

    current_date = datetime.now()
    current_month = current_date.month
    current_year = current_date.year

    # Get activities from persistent calendar storage
    from database import CalendarActivityDB, ReminderDB, save_calendar_activity

    # First, save any new activities from current favorites/projects/reminders
    # This ensures new activities are captured

    # Save favorite tender deadlines
    for fav in favorites:
        if fav.tender and fav.tender.deadline:
            try:
                save_calendar_activity(
                    db=db,
                    user_id=current_user.id,
                    activity_date=fav.tender.deadline,
                    activity_type='deadline',
                    title=f"Deadline: {fav.tender.title[:50]}..." if fav.tender.title else "Tender Deadline",
                    description=f"Tender deadline for {fav.tender.title}" if fav.tender.title else "Tender deadline",
                    tender_id=fav.tender.id
                )
            except Exception as e:
                logger.debug(f"Error saving calendar activity for tender {fav.tender.id}: {e}")

    # Save project activities
    for project in all_projects:
        if project.start_date:
            try:
                start_date = project.start_date if isinstance(project.start_date, datetime) else datetime.strptime(str(project.start_date), '%Y-%m-%d')
                save_calendar_activity(
                    db=db,
                    user_id=current_user.id,
                    activity_date=start_date,
                    activity_type='activity',
                    title=f"Project Started: {project.project_name[:40]}",
                    description=f"Started project: {project.project_name}",
                    project_id=project.id
                )
            except Exception as e:
                logger.debug(f"Error saving calendar activity for project start {project.id}: {e}")

        if project.end_date:
            try:
                end_date = project.end_date if isinstance(project.end_date, datetime) else datetime.strptime(str(project.end_date), '%Y-%m-%d')
                save_calendar_activity(
                    db=db,
                    user_id=current_user.id,
                    activity_date=end_date,
                    activity_type='activity',
                    title=f"Project Completed: {project.project_name[:40]}",
                    description=f"Completed project: {project.project_name}",
                    project_id=project.id
                )
            except Exception as e:
                logger.debug(f"Error saving calendar activity for project end {project.id}: {e}")

    # Save reminder activities
    reminders = db.query(ReminderDB).filter(
        ReminderDB.user_id == current_user.id,
        ReminderDB.is_dismissed == False
    ).all()

    for reminder in reminders:
        try:
            save_calendar_activity(
                db=db,
                user_id=current_user.id,
                activity_date=reminder.reminder_datetime,
                activity_type='reminder',
                title=f"Reminder: {reminder.title[:50]}..." if len(reminder.title) > 50 else f"Reminder: {reminder.title}",
                description=reminder.note or f"Reminder for {reminder.title}",
                tender_id=reminder.tender_id,
                reminder_id=reminder.id
            )
        except Exception as e:
            logger.debug(f"Error saving calendar activity for reminder {reminder.id}: {e}")

    # Now load ALL calendar activities from persistent storage (including historical ones)
    calendar_activities = db.query(CalendarActivityDB).filter(
        CalendarActivityDB.user_id == current_user.id,
        CalendarActivityDB.is_active == True  # Only active activities
    ).all()

    # Convert to calendar events format
    calendar_events = []
    for activity in calendar_activities:
        try:
            event = {
                'date': activity.activity_date.strftime('%Y-%m-%d'),
                'type': activity.activity_type,
                'title': activity.title,
                'description': activity.description,
                'source_deleted': activity.source_deleted  # Include deletion status
            }

            # Add source IDs if present
            if activity.tender_id:
                event['tender_id'] = activity.tender_id
            if activity.project_id:
                event['project_id'] = activity.project_id
            if activity.reminder_id:
                event['reminder_id'] = activity.reminder_id

            calendar_events.append(event)
        except Exception as e:
            logger.debug(f"Error formatting calendar activity {activity.id}: {e}")

    def _normalize_json(value, default):
        if value is None:
            return default
        if isinstance(value, (dict, list)):
            return value
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                return default
            try:
                return json.loads(stripped)
            except json.JSONDecodeError:
                return default
        return default

    if company_details:
        company_details.legal_details = _normalize_json(company_details.legal_details, {})
        company_details.financial_details = _normalize_json(company_details.financial_details, {})
        company_details.company_certifications = company_details.company_certifications or []


    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "current_user": current_user,
        "company_details": company_details,
        "favorite_tenders": favorite_tenders,
        "favorites": favorites,
        "favorite_count": favorite_count,
        "shortlisted_count": shortlisted_count,
        "total_projects": total_projects,
        "completed_projects": completed_projects,
        "total_project_value": total_project_value,
        "total_tender_count": total_tender_count,
        "formatted_tender_count": formatted_tender_count,
        "now": datetime.utcnow,  # Add current datetime function to template context
        "calendar_events": calendar_events,
        "current_month": current_month,
        "current_year": current_year,
        "selected_font": get_active_font()
    })

@app.get("/company-details", response_class=HTMLResponse)
async def company_details_page(request: Request, db: Session = Depends(get_db)):
    """Company details form page."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Get existing company details if any
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()

    return templates.TemplateResponse("company_details.html", {
        "request": request,
        "current_user": current_user,
        "company_details": company_details,
        "selected_font": get_active_font()
    })

@app.get("/home", response_class=HTMLResponse)
@require_company_details
async def home_login_page(request: Request, db: Session = Depends(get_db)):
    """Home page for logged-in users with navigation cards."""
    from core.dependencies import get_current_bd_employee

    # Check if BD employee is logged in
    bd_employee = get_current_bd_employee(request, db)
    if bd_employee:
        # Redirect BD employees to their home page
        return RedirectResponse(url="/bd/home", status_code=302)

    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Get company details
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()

    # Parse industry sectors for template
    sector_list = []
    if company_details and company_details.industry_sector:
        if isinstance(company_details.industry_sector, str):
            try:
                # Try to parse as JSON array
                parsed_data = json.loads(company_details.industry_sector)

                # Check if it's a list of dictionaries with "sector" key
                if isinstance(parsed_data, list) and len(parsed_data) > 0:
                    if isinstance(parsed_data[0], dict) and 'sector' in parsed_data[0]:
                        # Extract sector names from dictionaries
                        sector_list = [item['sector'] for item in parsed_data if 'sector' in item]
                    else:
                        # Simple list of sector names
                        sector_list = parsed_data
                else:
                    sector_list = parsed_data if isinstance(parsed_data, list) else [parsed_data]
            except json.JSONDecodeError:
                # If not JSON, treat as single sector
                sector_list = [company_details.industry_sector]
        elif isinstance(company_details.industry_sector, list):
            # Handle list of dictionaries
            if len(company_details.industry_sector) > 0 and isinstance(company_details.industry_sector[0], dict):
                sector_list = [item['sector'] for item in company_details.industry_sector if 'sector' in item]
            else:
                sector_list = company_details.industry_sector
        else:
            sector_list = []

    return templates.TemplateResponse("home_login.html", {
        "request": request,
        "current_user": current_user,
        "company_details": company_details,
        "sector_list": sector_list,
        "selected_font": get_active_font()
    })

@app.get("/analytics", response_class=HTMLResponse)
@require_company_details
async def analytics_page(request: Request, db: Session = Depends(get_db)):
    """
    Analytics Dashboard - Comprehensive tender management analytics.

    Provides 3 main sections:
    1. Dates & Deadlines - Upcoming deadlines and reminders
    2. Tenders Analytics - Status distribution, trends, and performance
    3. Stage Analytics - 6-stage workflow breakdown and employee workload
    """
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    from database import CalendarActivityDB, ReminderDB
    from datetime import datetime, timedelta
    from sqlalchemy import func, and_, or_, case
    import calendar as cal_module

    try:
        # =============================================================================
        # CALENDAR DATA PREPARATION
        # =============================================================================
        
        current_date = datetime.now()
        current_month = current_date.month
        current_year = current_date.year

        # Get all favorites for calendar
        favorites_for_calendar = db.query(FavoriteDB).filter(
            FavoriteDB.user_id == current_user.id
        ).all()

        # Get all projects for calendar
        all_projects = db.query(ProjectDB).filter(ProjectDB.user_id == current_user.id).all()

        # Save favorite tender deadlines to calendar
        for fav in favorites_for_calendar:
            if fav.tender and fav.tender.deadline:
                try:
                    from database import save_calendar_activity
                    save_calendar_activity(
                        db=db,
                        user_id=current_user.id,
                        activity_date=fav.tender.deadline,
                        activity_type='deadline',
                        title=f"Deadline: {fav.tender.title[:50]}..." if fav.tender.title else "Tender Deadline",
                        description=f"Tender deadline for {fav.tender.title}" if fav.tender.title else "Tender deadline",
                        tender_id=fav.tender.id
                    )
                except Exception as e:
                    logger.debug(f"Error saving calendar activity for tender {fav.tender.id}: {e}")

        # Save project activities to calendar
        for project in all_projects:
            if project.start_date:
                try:
                    from database import save_calendar_activity
                    start_date = project.start_date if isinstance(project.start_date, datetime) else datetime.strptime(str(project.start_date), '%Y-%m-%d')
                    save_calendar_activity(
                        db=db,
                        user_id=current_user.id,
                        activity_date=start_date,
                        activity_type='activity',
                        title=f"Project Started: {project.project_name[:40]}",
                        description=f"Started project: {project.project_name}",
                        project_id=project.id
                    )
                except Exception as e:
                    logger.debug(f"Error saving calendar activity for project start {project.id}: {e}")

            if project.end_date:
                try:
                    from database import save_calendar_activity
                    end_date = project.end_date if isinstance(project.end_date, datetime) else datetime.strptime(str(project.end_date), '%Y-%m-%d')
                    save_calendar_activity(
                        db=db,
                        user_id=current_user.id,
                        activity_date=end_date,
                        activity_type='activity',
                        title=f"Project Completed: {project.project_name[:40]}",
                        description=f"Completed project: {project.project_name}",
                        project_id=project.id
                    )
                except Exception as e:
                    logger.debug(f"Error saving calendar activity for project end {project.id}: {e}")

        # Save reminder activities to calendar
        reminders_for_calendar = db.query(ReminderDB).filter(
            ReminderDB.user_id == current_user.id,
            ReminderDB.is_dismissed == False
        ).all()

        for reminder in reminders_for_calendar:
            try:
                from database import save_calendar_activity
                save_calendar_activity(
                    db=db,
                    user_id=current_user.id,
                    activity_date=reminder.reminder_datetime,
                    activity_type='reminder',
                    title=f"Reminder: {reminder.title[:50]}..." if len(reminder.title) > 50 else f"Reminder: {reminder.title}",
                    description=reminder.note or f"Reminder for {reminder.title}",
                    tender_id=reminder.tender_id,
                    reminder_id=reminder.id
                )
            except Exception as e:
                logger.debug(f"Error saving calendar activity for reminder {reminder.id}: {e}")

        # Load ALL calendar activities from persistent storage
        calendar_activities = db.query(CalendarActivityDB).filter(
            CalendarActivityDB.user_id == current_user.id,
            CalendarActivityDB.is_active == True
        ).all()

        # Convert to calendar events format
        calendar_events = []
        for activity in calendar_activities:
            try:
                event = {
                    'date': activity.activity_date.strftime('%Y-%m-%d'),
                    'type': activity.activity_type,
                    'title': activity.title,
                    'description': activity.description,
                    'source_deleted': activity.source_deleted
                }

                # Add source IDs if present
                if activity.tender_id:
                    event['tender_id'] = activity.tender_id
                if activity.project_id:
                    event['project_id'] = activity.project_id
                if activity.reminder_id:
                    event['reminder_id'] = activity.reminder_id

                calendar_events.append(event)
            except Exception as e:
                logger.debug(f"Error formatting calendar activity {activity.id}: {e}")

        # =============================================================================
        # SECTION 1: DATES & DEADLINES DATA
        # =============================================================================

        # Get all shortlisted tenders with their progress data
        shortlisted_tenders = db.query(ShortlistedTenderDB).filter(
            ShortlistedTenderDB.user_id == current_user.id
        ).all()

        deadlines = []
        today = datetime.utcnow().date()

        # Extract ALL deadlines from shortlisted tenders (regardless of completion status)
        for st in shortlisted_tenders:
            tender = db.query(TenderDB).filter(TenderDB.id == st.tender_id).first()
            if not tender:
                continue

            progress = st.progress_data or {}
            critical_dates = tender.critical_dates or {}

            # Get assigned employees for display
            assignments = db.query(TenderAssignmentDB).filter(
                TenderAssignmentDB.tender_id == tender.id
            ).all()
            assigned_employees = []
            if assignments:
                emp_ids = [a.employee_id for a in assignments]
                employees = db.query(EmployeeDB).filter(EmployeeDB.id.in_(emp_ids)).all()
                assigned_employees = [{"id": e.id, "name": e.name, "email": e.email} for e in employees]

            # Step 1: Pre-Bid Meeting / Clarification Deadline
            # Try 'Clarification End Date' key (matches tender_management.html)
            clarification_end = critical_dates.get('Clarification End Date')
            if clarification_end:
                try:
                    deadline_date = datetime.fromisoformat(clarification_end).date()
                    deadlines.append({
                        'tender_id': tender.id,
                        'tender_title': tender.title,
                        'deadline_date': deadline_date,
                        'deadline_type': 'Pre-Bid Meeting',
                        'step_number': 1,
                        'tender_status': 'Shortlisted',
                        'assigned_employees': assigned_employees,
                        'is_manual': False
                    })
                except:
                    pass

            # Step 2: Bid Submission Deadline
            # Priority: tender.deadline, fallback: critical_dates['Bid Submission End Date']
            submission_deadline = None
            if tender.deadline:
                submission_deadline = tender.deadline.date() if isinstance(tender.deadline, datetime) else tender.deadline
            elif 'Bid Submission End Date' in critical_dates:
                try:
                    submission_deadline = datetime.fromisoformat(critical_dates['Bid Submission End Date']).date()
                except:
                    pass

            if submission_deadline:
                deadlines.append({
                    'tender_id': tender.id,
                    'tender_title': tender.title,
                    'deadline_date': submission_deadline,
                    'deadline_type': 'Bid Submission',
                    'step_number': 2,
                    'tender_status': 'Shortlisted',
                    'assigned_employees': assigned_employees,
                    'is_manual': False
                })

            # Step 3: Bid Opening Date
            bid_opening = critical_dates.get('Bid Opening Date')
            if bid_opening:
                try:
                    deadline_date = datetime.fromisoformat(bid_opening).date()
                    deadlines.append({
                        'tender_id': tender.id,
                        'tender_title': tender.title,
                        'deadline_date': deadline_date,
                        'deadline_type': 'Bid Opening',
                        'step_number': 3,
                        'tender_status': 'Shortlisted',
                        'assigned_employees': assigned_employees,
                        'is_manual': False
                    })
                except:
                    pass

            # Step 4: Financial Proposal Deadline (manual)
            step4_deadline = progress.get('step4_deadline')
            if step4_deadline:
                try:
                    deadline_date = datetime.fromisoformat(step4_deadline).date() if isinstance(step4_deadline, str) else step4_deadline
                    deadlines.append({
                        'tender_id': tender.id,
                        'tender_title': tender.title,
                        'deadline_date': deadline_date,
                        'deadline_type': 'Financial Proposal',
                        'step_number': 4,
                        'tender_status': 'Shortlisted',
                        'assigned_employees': assigned_employees,
                        'is_manual': True
                    })
                except:
                    pass

            # Step 5: Negotiation Deadline (manual)
            step5_deadline = progress.get('step5_deadline')
            if step5_deadline:
                try:
                    deadline_date = datetime.fromisoformat(step5_deadline).date() if isinstance(step5_deadline, str) else step5_deadline
                    deadlines.append({
                        'tender_id': tender.id,
                        'tender_title': tender.title,
                        'deadline_date': deadline_date,
                        'deadline_type': 'Negotiation',
                        'step_number': 5,
                        'tender_status': 'Shortlisted',
                        'assigned_employees': assigned_employees,
                        'is_manual': True
                    })
                except:
                    pass

        # Also get ALL deadlines from favourited tenders (not yet shortlisted)
        favourited = db.query(FavoriteDB).filter(
            FavoriteDB.user_id == current_user.id
        ).all()

        for fav in favourited:
            tender = db.query(TenderDB).filter(TenderDB.id == fav.tender_id).first()
            if not tender:
                continue

            critical_dates = tender.critical_dates or {}

            # Bid Submission Deadline (most important for favourited tenders)
            # Priority: tender.deadline, fallback: critical_dates['Bid Submission End Date']
            submission_deadline = None
            if tender.deadline:
                submission_deadline = tender.deadline.date() if isinstance(tender.deadline, datetime) else tender.deadline
            elif 'Bid Submission End Date' in critical_dates:
                try:
                    submission_deadline = datetime.fromisoformat(critical_dates['Bid Submission End Date']).date()
                except:
                    pass

            if submission_deadline:
                deadlines.append({
                    'tender_id': tender.id,
                    'tender_title': tender.title,
                    'deadline_date': submission_deadline,
                    'deadline_type': 'Bid Submission',
                    'step_number': None,
                    'tender_status': 'Favourited',
                    'assigned_employees': [],
                    'is_manual': False
                })

            # Pre-Bid Meeting Deadline
            # Try both 'Pre Bid Meeting Date' and 'Clarification End Date'
            pre_bid_deadline = None
            if 'Pre Bid Meeting Date' in critical_dates:
                try:
                    pre_bid_deadline = datetime.fromisoformat(critical_dates['Pre Bid Meeting Date']).date()
                except:
                    pass
            elif 'Clarification End Date' in critical_dates:
                try:
                    pre_bid_deadline = datetime.fromisoformat(critical_dates['Clarification End Date']).date()
                except:
                    pass

            if pre_bid_deadline:
                deadlines.append({
                    'tender_id': tender.id,
                    'tender_title': tender.title,
                    'deadline_date': pre_bid_deadline,
                    'deadline_type': 'Pre-Bid Meeting',
                    'step_number': None,
                    'tender_status': 'Favourited',
                    'assigned_employees': [],
                    'is_manual': False
                })

            # Bid Opening Deadline
            bid_opening = critical_dates.get('Bid Opening Date')
            if bid_opening:
                try:
                    deadline_date = datetime.fromisoformat(bid_opening).date()
                    deadlines.append({
                        'tender_id': tender.id,
                        'tender_title': tender.title,
                        'deadline_date': deadline_date,
                        'deadline_type': 'Bid Opening',
                        'step_number': None,
                        'tender_status': 'Favourited',
                        'assigned_employees': [],
                        'is_manual': False
                    })
                except:
                    pass

        # Categorize deadlines by urgency
        overdue = []
        today_deadlines = []
        this_week = []
        next_two_weeks = []
        later = []

        for d in deadlines:
            days_diff = (d['deadline_date'] - today).days
            d['days_remaining'] = days_diff

            if days_diff < 0:
                d['urgency'] = 'overdue'
                overdue.append(d)
            elif days_diff == 0:
                d['urgency'] = 'today'
                today_deadlines.append(d)
            elif 1 <= days_diff <= 7:
                d['urgency'] = 'this_week'
                this_week.append(d)
            elif 8 <= days_diff <= 14:
                d['urgency'] = 'next_two_weeks'
                next_two_weeks.append(d)
            else:
                d['urgency'] = 'later'
                later.append(d)

        # Get active reminders
        reminders = db.query(ReminderDB).filter(
            ReminderDB.user_id == current_user.id,
            ReminderDB.is_triggered == False,
            ReminderDB.is_dismissed == False,
            ReminderDB.reminder_datetime >= datetime.utcnow()
        ).order_by(ReminderDB.reminder_datetime).all()

        reminders_data = []
        for r in reminders:
            tender = db.query(TenderDB).filter(TenderDB.id == r.tender_id).first()
            reminders_data.append({
                'id': r.id,
                'tender_id': r.tender_id,
                'tender_title': r.title or (tender.title if tender else 'Unknown'),
                'reminder_datetime': r.reminder_datetime,
                'note': r.note,
                'days_until': (r.reminder_datetime.date() - today).days if r.reminder_datetime else 0
            })

        # =============================================================================
        # SECTION 2: TENDERS ANALYTICS DATA
        # =============================================================================

        # Count by status
        favorites_count = db.query(func.count(FavoriteDB.id)).filter(
            FavoriteDB.user_id == current_user.id
        ).scalar() or 0

        shortlisted_count = db.query(func.count(ShortlistedTenderDB.id)).filter(
            ShortlistedTenderDB.user_id == current_user.id
        ).scalar() or 0

        rejected_count = db.query(func.count(RejectedTenderDB.id)).filter(
            RejectedTenderDB.user_id == current_user.id
        ).scalar() or 0

        dumped_count = db.query(func.count(DumpedTenderDB.id)).filter(
            DumpedTenderDB.user_id == current_user.id
        ).scalar() or 0

        awarded_count = db.query(func.count(TenderDB.id)).filter(
            TenderDB.awarded == True,
            TenderDB.awarded_by == current_user.id
        ).scalar() or 0

        # Calculate total value by status
        shortlisted_value = db.query(func.sum(TenderDB.estimated_value)).select_from(
            ShortlistedTenderDB
        ).join(
            TenderDB, ShortlistedTenderDB.tender_id == TenderDB.id
        ).filter(
            ShortlistedTenderDB.user_id == current_user.id
        ).scalar() or 0

        awarded_value = db.query(func.sum(TenderDB.estimated_value)).filter(
            TenderDB.awarded == True,
            TenderDB.awarded_by == current_user.id
        ).scalar() or 0

        dumped_value = db.query(func.sum(TenderDB.estimated_value)).select_from(
            DumpedTenderDB
        ).join(
            TenderDB, DumpedTenderDB.tender_id == TenderDB.id
        ).filter(
            DumpedTenderDB.user_id == current_user.id
        ).scalar() or 0

        # Calculate metrics
        # Win Rate: Awarded tenders divided by total ever shortlisted tenders
        # IMPORTANT: When tenders are awarded, they are DELETED from ShortlistedTenderDB (see finalize_tender_award_for_user)
        # Similarly, killed/cancelled tenders are moved to DumpedTenderDB
        # Therefore: Total Ever Shortlisted = Current Shortlisted + Awarded + Dumped
        total_ever_shortlisted = shortlisted_count + awarded_count + dumped_count
        win_rate = round((awarded_count / total_ever_shortlisted * 100), 1) if total_ever_shortlisted > 0 else 0

        # Rejection Rate: Rejected tenders divided by total favourited tenders
        # This represents tenders that were favourited but rejected before shortlisting
        rejection_rate = round((rejected_count / favorites_count * 100), 1) if favorites_count > 0 else 0

        # Time series data (last 30 days for chart)
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)

        favorites_timeline = db.query(
            func.date(FavoriteDB.created_at).label('date'),
            func.count(FavoriteDB.id).label('count')
        ).filter(
            FavoriteDB.user_id == current_user.id,
            FavoriteDB.created_at >= thirty_days_ago
        ).group_by(func.date(FavoriteDB.created_at)).all()

        shortlisted_timeline = db.query(
            func.date(ShortlistedTenderDB.created_at).label('date'),
            func.count(ShortlistedTenderDB.id).label('count')
        ).filter(
            ShortlistedTenderDB.user_id == current_user.id,
            ShortlistedTenderDB.created_at >= thirty_days_ago
        ).group_by(func.date(ShortlistedTenderDB.created_at)).all()

        rejected_timeline = db.query(
            func.date(RejectedTenderDB.created_at).label('date'),
            func.count(RejectedTenderDB.id).label('count')
        ).filter(
            RejectedTenderDB.user_id == current_user.id,
            RejectedTenderDB.created_at >= thirty_days_ago
        ).group_by(func.date(RejectedTenderDB.created_at)).all()

        dumped_timeline = db.query(
            func.date(DumpedTenderDB.created_at).label('date'),
            func.count(DumpedTenderDB.id).label('count')
        ).filter(
            DumpedTenderDB.user_id == current_user.id,
            DumpedTenderDB.created_at >= thirty_days_ago
        ).group_by(func.date(DumpedTenderDB.created_at)).all()

        # Awarded timeline - use updated_at if awarded_date doesn't exist
        awarded_timeline = db.query(
            func.date(TenderDB.updated_at).label('date'),
            func.count(TenderDB.id).label('count')
        ).filter(
            TenderDB.awarded == True,
            TenderDB.awarded_by == current_user.id,
            TenderDB.updated_at >= thirty_days_ago
        ).group_by(func.date(TenderDB.updated_at)).all()

        # Build combined timeline_data for charts
        # Create a date range for last 30 days
        date_range = [(datetime.utcnow() - timedelta(days=i)).date() for i in range(29, -1, -1)]
        
        # Convert timeline queries to dictionaries for easy lookup
        favorites_dict = {str(d.date): d.count for d in favorites_timeline}
        shortlisted_dict = {str(d.date): d.count for d in shortlisted_timeline}
        rejected_dict = {str(d.date): d.count for d in rejected_timeline}
        dumped_dict = {str(d.date): d.count for d in dumped_timeline}
        awarded_dict = {str(d.date): d.count for d in awarded_timeline}
        
        # Build timeline_data array
        timeline_data = []
        for date in date_range:
            date_str = str(date)
            timeline_data.append({
                'date': date_str,
                'favorites': favorites_dict.get(date_str, 0),
                'shortlisted': shortlisted_dict.get(date_str, 0),
                'rejected': rejected_dict.get(date_str, 0),
                'dumped': dumped_dict.get(date_str, 0),
                'awarded': awarded_dict.get(date_str, 0)
            })

        # =============================================================================
        # SECTION 3: STAGE ANALYTICS DATA
        # =============================================================================

        # Analyze 6-stage workflow
        # NOTE: "Pending" status is excluded from analytics - only count tenders with actual progress
        stage_breakdown = {
            'step1': {'Attended': 0, 'Not Attended': 0, 'Not Applicable': 0, 'Cancelled': 0},
            'step2': {'Submitted': 0, 'Not Submitted': 0, 'Tender Cancelled': 0},
            'step3': {'Not Opened': 0, 'Opened & Qualified': 0, 'Opened & Not Qualified': 0, 'Tender Cancelled': 0},
            'step4': {'Not Opened': 0, 'Opened & Won': 0, 'Opened & Lost': 0, 'Opened & Not Eligible': 0, 'Tender Cancelled': 0},
            'step5': {'Applicable': 0, 'Not Applicable': 0, 'Tender Cancelled': 0},
            'step6': {'Yes': 0, 'No': 0}
        }

        stage_tenders = {f'step{i}': [] for i in range(1, 7)}

        for st in shortlisted_tenders:
            tender = db.query(TenderDB).filter(TenderDB.id == st.tender_id).first()
            if not tender:
                continue

            progress = st.progress_data or {}

            # FILTER: Only include tenders where at least one stage has action taken
            # If all stages are "Pending", skip this tender from stage analytics
            has_progress = False
            for step_num in range(1, 7):
                step_key = f'step{step_num}'
                status = progress.get(step_key, 'Pending')
                if status != 'Pending':
                    has_progress = True
                    break

            if not has_progress:
                continue  # Skip tenders with no progress at any stage

            # Get assigned employees
            assignments = db.query(TenderAssignmentDB).filter(
                TenderAssignmentDB.tender_id == tender.id
            ).all()
            assigned_employees = []
            if assignments:
                emp_ids = [a.employee_id for a in assignments]
                employees = db.query(EmployeeDB).filter(EmployeeDB.id.in_(emp_ids)).all()
                assigned_employees = [{"id": e.id, "name": e.name} for e in employees]

            tender_info = {
                'id': tender.id,
                'title': tender.title,
                'assigned_employees': assigned_employees,
                'created_at': st.created_at
            }

            for step_num in range(1, 7):
                step_key = f'step{step_num}'
                status = progress.get(step_key, 'Pending')

                # Skip "Pending" status - only count and track tenders with actual progress
                if status == 'Pending':
                    continue

                # Count status in breakdown (only non-Pending statuses)
                if status in stage_breakdown[step_key]:
                    stage_breakdown[step_key][status] += 1

                # Track which tenders are at each stage (only non-Pending statuses)
                tender_step_info = {
                    **tender_info,
                    'status': status,
                    'step_employees': progress.get(f'{step_key}_employees', [])
                }
                stage_tenders[step_key].append(tender_step_info)

        # Employee workload analysis
        all_employees = db.query(EmployeeDB).filter(
            EmployeeDB.company_code_id == (current_user.company_code_id if hasattr(current_user, 'company_code_id') else None)
        ).all()

        employee_workload = []
        for emp in all_employees:
            # Count tenders assigned to this employee
            assigned_tenders = db.query(TenderAssignmentDB).filter(
                TenderAssignmentDB.employee_id == emp.id
            ).count()

            # Analyze stage distribution
            stage_dist = {f'step{i}': 0 for i in range(1, 7)}
            for st in shortlisted_tenders:
                progress = st.progress_data or {}
                for step_num in range(1, 7):
                    step_key = f'step{step_num}'
                    step_emps = progress.get(f'{step_key}_employees', [])
                    if emp.id in step_emps:
                        stage_dist[step_key] += 1

            # Determine workload level
            if assigned_tenders == 0:
                workload_level = 'none'
            elif assigned_tenders <= 3:
                workload_level = 'light'
            elif assigned_tenders <= 7:
                workload_level = 'medium'
            else:
                workload_level = 'heavy'

            employee_workload.append({
                'id': emp.id,
                'name': emp.name,
                'email': emp.email,
                'total_tenders': assigned_tenders,
                'stage_distribution': stage_dist,
                'workload_level': workload_level
            })

        # =============================================================================
        # RETURN ANALYTICS DATA
        # =============================================================================

        return templates.TemplateResponse("analytics.html", {
            "request": request,
            "current_user": current_user,

            # Section 1: Dates & Deadlines
            "overdue_deadlines": sorted(overdue, key=lambda x: x['deadline_date']),
            "today_deadlines": sorted(today_deadlines, key=lambda x: x['deadline_date']),
            "this_week_deadlines": sorted(this_week, key=lambda x: x['deadline_date']),
            "next_two_weeks_deadlines": sorted(next_two_weeks, key=lambda x: x['deadline_date']),
            "later_deadlines": sorted(later, key=lambda x: x['deadline_date']),
            "active_reminders": reminders_data,

            # Section 2: Tenders Analytics
            "favorites_count": favorites_count,
            "shortlisted_count": shortlisted_count,
            "rejected_count": rejected_count,
            "dumped_count": dumped_count,
            "awarded_count": awarded_count,
            "shortlisted_value": shortlisted_value,
            "awarded_value": awarded_value,
            "dumped_value": dumped_value,
            "win_rate": win_rate,
            "rejection_rate": rejection_rate,
            "favorites_timeline": [(str(d.date), d.count) for d in favorites_timeline],
            "shortlisted_timeline": [(str(d.date), d.count) for d in shortlisted_timeline],
            "timeline_data": timeline_data,  # Combined timeline data for charts

            # Section 3: Stage Analytics
            "stage_breakdown": stage_breakdown,
            "stage_tenders": stage_tenders,
            "employee_workload": employee_workload,

            # Summary stats for header cards
            "total_deadlines": len(deadlines),
            "urgent_deadlines": len(deadlines) + len(reminders_data),  # Total deadlines + reminders
            "total_tenders": favorites_count + shortlisted_count + rejected_count + dumped_count + awarded_count,

            # Calendar data
            "calendar_events": calendar_events,
            "current_month": current_month,
            "current_year": current_year,
            "selected_font": get_active_font()
        })

    except Exception as e:
        logger.error(f"Error in analytics page: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error loading analytics: {str(e)}")

@app.get("/profile", response_class=HTMLResponse)
@require_company_details
async def profile_page(request: Request, db: Session = Depends(get_db)):
    """User profile page for managing personal information and settings."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Get company details for additional context
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()

    return templates.TemplateResponse("profile.html", {
        "request": request,
        "current_user": current_user,
        "company_details": company_details,
        "selected_font": get_active_font()
    })

@app.post("/api/profile/update-basic")
@require_company_details
async def update_basic_profile(request: Request, db: Session = Depends(get_db)):
    """Update basic profile information (name, phone, job title, department, bio)."""
    current_user = get_current_user(request, db)
    if not current_user:
        return {"success": False, "error": "User not authenticated"}

    try:
        data = await request.json()

        # Update fields if provided
        if "name" in data and data["name"].strip():
            current_user.name = data["name"].strip()

        if "phone_number" in data:
            current_user.phone_number = data["phone_number"].strip() if data["phone_number"] else None

        if "job_title" in data:
            current_user.job_title = data["job_title"].strip() if data["job_title"] else None

        if "department" in data:
            current_user.department = data["department"].strip() if data["department"] else None

        if "bio" in data:
            current_user.bio = data["bio"].strip() if data["bio"] else None

        db.commit()
        db.refresh(current_user)

        logger.info(f"✅ Profile updated successfully for user {current_user.email}")
        return {
            "success": True,
            "message": "Profile updated successfully",
            "user": {
                "name": current_user.name,
                "phone_number": current_user.phone_number,
                "job_title": current_user.job_title,
                "department": current_user.department,
                "bio": current_user.bio
            }
        }

    except Exception as e:
        logger.error(f"❌ Error updating profile: {e}")
        db.rollback()
        return {"success": False, "error": str(e)}

@app.post("/api/profile/upload-image")
@require_company_details
async def upload_profile_image(request: Request, db: Session = Depends(get_db)):
    """Upload profile image."""
    current_user = get_current_user(request, db)
    if not current_user:
        return {"success": False, "error": "User not authenticated"}

    try:
        form = await request.form()
        image_file = form.get("profile_image")

        if not image_file:
            return {"success": False, "error": "No image file provided"}

        # Read file content
        file_content = await image_file.read()

        # Validate file size (max 5MB)
        max_size = 5 * 1024 * 1024  # 5MB
        if len(file_content) > max_size:
            return {"success": False, "error": "Image size must be less than 5MB"}

        # Validate file type
        allowed_types = ["image/jpeg", "image/jpg", "image/png", "image/gif", "image/webp"]
        if image_file.content_type not in allowed_types:
            return {"success": False, "error": "Invalid file type. Only JPG, PNG, GIF, and WebP are allowed"}

        # Create uploads directory if it doesn't exist
        uploads_dir = Path("static/uploads/profiles")
        uploads_dir.mkdir(parents=True, exist_ok=True)

        # Generate unique filename
        file_extension = Path(image_file.filename).suffix
        unique_filename = f"{current_user.id}_{int(time.time())}{file_extension}"
        file_path = uploads_dir / unique_filename

        # Save file
        with open(file_path, "wb") as f:
            f.write(file_content)

        # Update user profile
        old_image = current_user.profile_image
        current_user.profile_image = f"/static/uploads/profiles/{unique_filename}"
        db.commit()
        db.refresh(current_user)

        # Delete old image if exists
        if old_image and old_image.startswith("/static/uploads/profiles/"):
            old_file_path = Path(old_image.lstrip("/"))
            if old_file_path.exists():
                old_file_path.unlink()

        logger.info(f"✅ Profile image uploaded successfully for user {current_user.email}")
        return {
            "success": True,
            "message": "Profile image uploaded successfully",
            "image_url": current_user.profile_image
        }

    except Exception as e:
        logger.error(f"❌ Error uploading profile image: {e}")
        db.rollback()
        return {"success": False, "error": str(e)}

@app.post("/api/profile/change-password")
@require_company_details
async def change_password(request: Request, db: Session = Depends(get_db)):
    """Change user password."""
    current_user = get_current_user(request, db)
    if not current_user:
        return {"success": False, "error": "User not authenticated"}

    try:
        data = await request.json()
        current_password = data.get("current_password", "")
        new_password = data.get("new_password", "")
        confirm_password = data.get("confirm_password", "")

        # Validate inputs
        if not current_password or not new_password or not confirm_password:
            return {"success": False, "error": "All fields are required"}

        # Verify current password
        if not pwd_context.verify(current_password, current_user.password_hash):
            return {"success": False, "error": "Current password is incorrect"}

        # Validate new password
        if len(new_password) < 8:
            return {"success": False, "error": "New password must be at least 8 characters long"}

        if new_password != confirm_password:
            return {"success": False, "error": "New passwords do not match"}

        if current_password == new_password:
            return {"success": False, "error": "New password must be different from current password"}

        # Update password
        current_user.password_hash = pwd_context.hash(new_password)
        db.commit()

        logger.info(f"✅ Password changed successfully for user {current_user.email}")
        return {
            "success": True,
            "message": "Password changed successfully"
        }

    except Exception as e:
        logger.error(f"❌ Error changing password: {e}")
        db.rollback()
        return {"success": False, "error": str(e)}

@app.post("/api/profile/update-preferences")
@require_company_details
async def update_preferences(request: Request, db: Session = Depends(get_db)):
    """Update user notification preferences."""
    current_user = get_current_user(request, db)
    if not current_user:
        return {"success": False, "error": "User not authenticated"}

    try:
        data = await request.json()
        preferences = data.get("preferences", {})

        # Update notification preferences
        current_user.notification_preferences = preferences
        db.commit()
        db.refresh(current_user)

        logger.info(f"✅ Preferences updated successfully for user {current_user.email}")
        return {
            "success": True,
            "message": "Preferences updated successfully",
            "preferences": current_user.notification_preferences
        }

    except Exception as e:
        logger.error(f"❌ Error updating preferences: {e}")
        db.rollback()
        return {"success": False, "error": str(e)}


# =============================================================================
# FEATURE LOCK API ENDPOINTS
# PIN-based feature locking system for sharing limited account access
# =============================================================================

@app.post("/api/feature-lock/setup")
@require_company_details
async def setup_feature_lock(request: Request, db: Session = Depends(get_db)):
    """
    Setup or update PIN and locked features.
    First time: No password required
    Updates: Requires current password verification
    """
    current_user = get_current_user(request, db)
    if not current_user:
        return {"success": False, "error": "User not authenticated"}

    try:
        data = await request.json()
        pin = data.get("pin", "").strip()
        locked_features = data.get("locked_features", [])
        password = data.get("password", "").strip()  # Required for updates

        # Validate PIN format (6-digit numeric)
        if not pin or len(pin) != 6 or not pin.isdigit():
            return {
                "success": False,
                "error": "PIN must be exactly 6 digits"
            }

        # Validate locked features (must be valid route paths)
        valid_features = ["/procurement", "/tender-management", "/projects", "/project-management"]
        for feature in locked_features:
            if feature not in valid_features:
                return {
                    "success": False,
                    "error": f"Invalid feature: {feature}"
                }

        # Check if this is an update (system already enabled)
        is_update = current_user.feature_lock_enabled

        # If updating, require password verification
        if is_update:
            if not password:
                return {
                    "success": False,
                    "error": "Password required to update feature lock settings"
                }

            if not verify_password(password, current_user.password_hash):
                return {
                    "success": False,
                    "error": "Incorrect password"
                }

        # Hash the PIN
        hashed_pin = hash_password(pin)

        # Update user settings
        current_user.feature_lock_pin = hashed_pin
        current_user.feature_lock_enabled = True
        current_user.locked_features = locked_features

        db.commit()
        db.refresh(current_user)

        logger.info(f"✅ Feature lock {'updated' if is_update else 'setup'} for user {current_user.email}")

        return {
            "success": True,
            "message": f"Feature lock {'updated' if is_update else 'enabled'} successfully",
            "locked_features": locked_features
        }

    except Exception as e:
        logger.error(f"❌ Error setting up feature lock: {e}")
        db.rollback()
        return {"success": False, "error": str(e)}


@app.post("/api/feature-lock/verify")
async def verify_feature_lock_pin(request: Request, db: Session = Depends(get_db)):
    """
    Verify PIN and update session to mark PIN as verified.
    Once verified, remains valid until logout.
    """
    current_user = get_current_user(request, db)
    if not current_user:
        return {"success": False, "error": "User not authenticated"}

    try:
        data = await request.json()
        pin = data.get("pin", "").strip()

        # Check if feature lock is enabled
        if not current_user.feature_lock_enabled or not current_user.feature_lock_pin:
            return {
                "success": False,
                "error": "Feature lock is not enabled"
            }

        # Verify PIN
        if not verify_password(pin, current_user.feature_lock_pin):
            logger.warning(f"⚠️ Failed PIN verification attempt for user {current_user.email}")
            return {
                "success": False,
                "error": "Incorrect PIN"
            }

        # Update session to mark PIN as verified
        session_token = request.cookies.get("session_token")
        if session_token:
            update_session(session_token, {
                'pin_verified': True,
                'pin_verified_at': datetime.utcnow().isoformat()
            })

        logger.info(f"✅ PIN verified successfully for user {current_user.email}")

        return {
            "success": True,
            "message": "PIN verified successfully"
        }

    except Exception as e:
        logger.error(f"❌ Error verifying PIN: {e}")
        return {"success": False, "error": str(e)}


@app.get("/api/feature-lock/status")
@require_company_details
async def get_feature_lock_status(request: Request, db: Session = Depends(get_db)):
    """
    Get current feature lock status and PIN verification state.
    Returns: enabled status, locked features, and whether PIN is verified in current session.
    """
    current_user = get_current_user(request, db)
    if not current_user:
        return {"success": False, "error": "User not authenticated"}

    try:
        # Check if PIN is verified in current session
        pin_verified = False
        session_token = request.cookies.get("session_token")
        if session_token:
            from core.security import get_session
            session_data = get_session(session_token)
            if session_data:
                pin_verified = session_data.get('pin_verified', False)

        return {
            "success": True,
            "enabled": current_user.feature_lock_enabled or False,
            "locked_features": current_user.locked_features or [],
            "pin_verified": pin_verified
        }

    except Exception as e:
        logger.error(f"❌ Error getting feature lock status: {e}")
        return {"success": False, "error": str(e)}


@app.post("/api/feature-lock/reset")
@require_company_details
async def reset_feature_lock(request: Request, db: Session = Depends(get_db)):
    """
    Disable feature lock system entirely.
    Requires password verification.
    """
    current_user = get_current_user(request, db)
    if not current_user:
        return {"success": False, "error": "User not authenticated"}

    try:
        data = await request.json()
        password = data.get("password", "").strip()

        if not password:
            return {
                "success": False,
                "error": "Password required to disable feature lock"
            }

        # Verify password
        if not verify_password(password, current_user.password_hash):
            return {
                "success": False,
                "error": "Incorrect password"
            }

        # Disable feature lock
        current_user.feature_lock_pin = None
        current_user.feature_lock_enabled = False
        current_user.locked_features = []

        db.commit()
        db.refresh(current_user)

        logger.info(f"✅ Feature lock disabled for user {current_user.email}")

        return {
            "success": True,
            "message": "Feature lock disabled successfully"
        }

    except Exception as e:
        logger.error(f"❌ Error disabling feature lock: {e}")
        db.rollback()
        return {"success": False, "error": str(e)}


@app.get("/subscription/manage", response_class=HTMLResponse)
async def manage_subscription_page(request: Request, db: Session = Depends(get_db)):
    """Manage subscription page with pricing plans."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse("manage_subscription.html", {
        "request": request,
        "current_user": current_user,
        "selected_font": get_active_font()
    })

@app.get("/subscription/checkout", response_class=HTMLResponse)
async def checkout_page(request: Request, plan: str = "professional", db: Session = Depends(get_db)):
    """Checkout page for subscription payment."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Define plan details
    plans = {
        "starter": {
            "name": "Starter",
            "price": 50000,
            "display_price": "₹50,000"
        },
        "professional": {
            "name": "Professional",
            "price": 100000,
            "display_price": "₹1,00,000"
        },
        "premium": {
            "name": "Premium",
            "price": 200000,
            "display_price": "₹2,00,000"
        }
    }

    plan_details = plans.get(plan, plans["professional"])
    gst_amount = int(plan_details["price"] * 0.18)
    total_amount = plan_details["price"] + gst_amount

    return templates.TemplateResponse("checkout.html", {
        "request": request,
        "current_user": current_user,
        "plan_name": plan_details["name"],
        "plan_price": plan_details["display_price"],
        "gst": f"₹{gst_amount:,}",
        "total": f"₹{total_amount:,}",
        "selected_font": get_active_font()
    })

@app.get("/manage_certificates", response_class=HTMLResponse)
@enforce_test_quarantine
@require_company_details
async def manage_certificates_page(request: Request, db: Session = Depends(get_db)):
    """Manage certificates page (upload and search)."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse("manage_certificates.html", {
        "request": request,
        "current_user": current_user,
        "selected_font": get_active_font()
    })

@app.get("/certificates/search", response_class=HTMLResponse)
@enforce_test_quarantine
@require_company_details
async def certificates_search_page(request: Request, db: Session = Depends(get_db)):
    """Certificate search page."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse("certificate_search.html", {
        "request": request,
        "current_user": current_user,
        "selected_font": get_active_font()
    })

@app.get("/certificate/{certificate_id}", response_class=HTMLResponse)
@enforce_test_quarantine
@require_company_details
async def certificate_detail_page(request: Request, certificate_id: str, db: Session = Depends(get_db)):
    """Certificate detail page."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Get the certificate
    certificate = db.query(CertificateDB).filter(
        and_(CertificateDB.id == certificate_id, CertificateDB.user_id == current_user.id)
    ).first()

    if not certificate:
        raise HTTPException(status_code=404, detail="Certificate not found")

    # Ensure services_rendered is always a list (not a JSON string)
    if certificate.services_rendered:
        if isinstance(certificate.services_rendered, str):
            try:
                import json
                certificate.services_rendered = json.loads(certificate.services_rendered)
            except (json.JSONDecodeError, ValueError):
                certificate.services_rendered = []
        elif not isinstance(certificate.services_rendered, list):
            certificate.services_rendered = []
    else:
        certificate.services_rendered = []

    def _normalize_list(value):
        if not value:
            return []
        if isinstance(value, list):
            return [str(item).strip() for item in value if isinstance(item, str) and item and str(item).strip()]
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
                if isinstance(parsed, list):
                    return [str(item).strip() for item in parsed if isinstance(item, str) and item.strip()]
            except (json.JSONDecodeError, ValueError):
                pass
            return [part.strip() for part in re.split(r"[;,/\n]+", value) if part and part.strip()]
        return []

    def _normalize_metrics(value):
        if not value:
            return []
        if isinstance(value, str):
            try:
                value = json.loads(value)
            except (json.JSONDecodeError, ValueError):
                return []
        if not isinstance(value, list):
            return []

        cleaned = []
        for entry in value:
            if not isinstance(entry, dict):
                continue
            metric_record = {
                "metric_name": (entry.get("metric_name") or "").strip(),
                "value": (entry.get("value") or "").strip(),
                "unit": (entry.get("unit") or "").strip(),
                "notes": (entry.get("notes") or "").strip()
            }
            if any(metric_record.values()):
                cleaned.append(metric_record)
        return cleaned

    certificate.services_rendered = _normalize_list(certificate.services_rendered)
    certificate.sectors = _normalize_list(getattr(certificate, "sectors", []))
    certificate.sub_sectors = _normalize_list(getattr(certificate, "sub_sectors", []))
    certificate.jv_partners = _normalize_list(getattr(certificate, "jv_partners", []))
    certificate.metrics = _normalize_metrics(getattr(certificate, "metrics", []))

    if getattr(certificate, "consultancy_fee_inr", None) and isinstance(certificate.consultancy_fee_inr, str):
        certificate.consultancy_fee_inr = certificate.consultancy_fee_inr.strip()
    if getattr(certificate, "project_value_inr", None) and isinstance(certificate.project_value_inr, str):
        certificate.project_value_inr = certificate.project_value_inr.strip()
    if getattr(certificate, "scope_of_work", None) and isinstance(certificate.scope_of_work, str):
        certificate.scope_of_work = certificate.scope_of_work.strip()

    confidence_percentage = None
    if getattr(certificate, "confidence_score", None) is not None:
        try:
            confidence_percentage = int(round(float(certificate.confidence_score) * 100))
        except (ValueError, TypeError):
            confidence_percentage = None

    return templates.TemplateResponse("certificate_detail.html", {
        "request": request,
        "current_user": current_user,
        "certificate": certificate,
        "confidence_percentage": confidence_percentage,
        "selected_font": get_active_font()
    })

@app.get("/projects", response_class=HTMLResponse)
@require_company_details
@require_pin_verification
async def projects_list(request: Request, db: Session = Depends(get_db)):
    """List all user projects with search functionality."""
    from core.dependencies import get_current_user_or_bd_employee

    # Check if user or BD employee is logged in
    entity, entity_type = get_current_user_or_bd_employee(request, db)
    if not entity:
        return RedirectResponse(url="/login", status_code=302)

    # Set current_user for backward compatibility
    current_user = entity if entity_type == 'user' else None
    current_bd_employee = entity if entity_type == 'bd_employee' else None

    # Detect quarantined test session (allows access but tracks status)
    is_quarantined = (
        current_user and
        hasattr(current_user, '_quarantined') and
        current_user._quarantined
    )

    # Get the user_id for queries
    user_id_for_queries = None
    if current_user:
        user_id_for_queries = current_user.id
    elif current_bd_employee:
        company_code = db.query(CompanyCodeDB).filter(CompanyCodeDB.id == current_bd_employee.company_code_id).first()
        if company_code:
            user_id_for_queries = company_code.user_id

    if not user_id_for_queries:
        return RedirectResponse(url="/login", status_code=302)

    # Get company details to retrieve industry sectors
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == user_id_for_queries).first()
    user_industry_sectors = []
    if company_details and company_details.industry_sector:
        if company_details.industry_sector.startswith('[{'):
            # New format: [{"sector": "...", "subsectors": [...]}]
            try:
                sectors_data = json.loads(company_details.industry_sector)
                user_industry_sectors = [s['sector'] for s in sectors_data if 'sector' in s]
            except (json.JSONDecodeError, KeyError):
                user_industry_sectors = [company_details.industry_sector]
        elif company_details.industry_sector.startswith('['):
            try:
                user_industry_sectors = json.loads(company_details.industry_sector)
            except json.JSONDecodeError:
                user_industry_sectors = [company_details.industry_sector]
        else:
            user_industry_sectors = [company_details.industry_sector]

    # Get search and filter parameters
    search = request.query_params.get('search', '')
    sector_filter = request.query_params.get('sector_filter', '')
    project_name = request.query_params.get('project_name', '')
    client = request.query_params.get('client', '')
    sub_sector = request.query_params.get('sub_sector', '')
    start_date = request.query_params.get('start_date', '')
    end_date = request.query_params.get('end_date', '')
    state = request.query_params.get('state', '')
    show_all = request.query_params.get('show_all', '')
    page = int(request.query_params.get('page', 1))
    per_page = 10

    # Get total project count for cards
    total_project_count = db.query(ProjectDB).filter(ProjectDB.user_id == user_id_for_queries).count()

    # Only show projects if search parameters are provided OR show_all is true
    projects = []
    total_projects = 0
    total_pages = 0
    has_prev = False
    has_next = False

    # Check if any search/filter parameter is provided or show_all flag is set
    search_performed = any([search, sector_filter, project_name, client, sub_sector, start_date, end_date, state, show_all])

    if search_performed:
        # Build query
        query = db.query(ProjectDB).filter(ProjectDB.user_id == user_id_for_queries)

        # Apply filters
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    ProjectDB.project_name.ilike(search_term),
                    ProjectDB.project_description.ilike(search_term),
                    ProjectDB.client_name.ilike(search_term)
                )
            )

        if project_name:
            query = query.filter(ProjectDB.project_name.ilike(f"%{project_name}%"))

        if sector_filter:
            query = query.filter(ProjectDB.sector == sector_filter)

        if client:
            query = query.filter(ProjectDB.client_name.ilike(f"%{client}%"))

        if sub_sector:
            query = query.filter(ProjectDB.sub_sector.ilike(f"%{sub_sector}%"))

        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(ProjectDB.start_date >= start_date_obj)
            except ValueError:
                pass

        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                query = query.filter(ProjectDB.end_date <= end_date_obj)
            except ValueError:
                pass

        if state:
            # Filter projects where the state is in the JSON states list
            query = query.filter(ProjectDB.states.contains([state]))

        # Get total count for pagination
        total_projects = query.count()

        # Apply pagination and ordering
        projects = query.order_by(desc(ProjectDB.created_at)).offset((page - 1) * per_page).limit(per_page).all()

        # Pagination info
        total_pages = (total_projects + per_page - 1) // per_page
        has_prev = page > 1
        has_next = page < total_pages

    # Get unique values for search dropdowns
    all_projects = db.query(ProjectDB).filter(ProjectDB.user_id == user_id_for_queries).all()
    clients = list(set([p.client_name for p in all_projects if p.client_name]))

    # Build sectors and subsectors from actual user projects (not company profile)
    sectors = []
    sectors_subsectors_map = {}

    # Build the sectors and subsectors map from actual project data
    for project in all_projects:
        if project.sector:
            # Add sector to list if not already present
            if project.sector not in sectors:
                sectors.append(project.sector)

            # Initialize subsector list for this sector if needed
            if project.sector not in sectors_subsectors_map:
                sectors_subsectors_map[project.sector] = []

            # Add subsector if it exists and isn't already in the list
            if project.sub_sector and project.sub_sector not in sectors_subsectors_map[project.sector]:
                sectors_subsectors_map[project.sector].append(project.sub_sector)

    # Sort for consistent display
    sectors = sorted(sectors)
    for sector in sectors_subsectors_map:
        sectors_subsectors_map[sector] = sorted(sectors_subsectors_map[sector])

    # Get subsectors based on selected sector filter, or all subsectors if no filter
    sub_sectors = []
    if sector_filter and sector_filter in sectors_subsectors_map:
        sub_sectors = sectors_subsectors_map[sector_filter]
    elif not sector_filter and sectors_subsectors_map:
        # Show all subsectors from all sectors
        all_subsectors = set()
        for subsector_list in sectors_subsectors_map.values():
            all_subsectors.update(subsector_list)
        sub_sectors = sorted(list(all_subsectors))

    # Extract all unique states from JSON states field
    states_set = set()
    for p in all_projects:
        if p.states and isinstance(p.states, list):
            states_set.update(p.states)
    states = sorted(list(states_set))

    # Calculate sector analytics (total value and average value per sector)
    sector_analytics = {}
    for sector in user_industry_sectors:
        sector_projects = [p for p in all_projects if p.sector == sector and p.project_cost is not None]
        if sector_projects:
            total_value = sum(p.project_cost for p in sector_projects)
            average_value = total_value / len(sector_projects)
            count = len(sector_projects)
            sector_analytics[sector] = {
                'total_value': total_value,
                'average_value': average_value,
                'count': count
            }
        else:
            sector_analytics[sector] = {
                'total_value': 0,
                'average_value': 0,
                'count': 0
            }

    # Calculate sector fee analytics (total consultancy fee and average consultancy fee per sector)
    sector_fee_analytics = {}
    for sector in user_industry_sectors:
        sector_projects = [p for p in all_projects if p.sector == sector and p.consultancy_fee is not None]
        if sector_projects:
            total_fee = sum(p.consultancy_fee for p in sector_projects)
            average_fee = total_fee / len(sector_projects)
            count = len(sector_projects)
            sector_fee_analytics[sector] = {
                'total_fee': total_fee,
                'average_fee': average_fee,
                'count': count
            }
        else:
            sector_fee_analytics[sector] = {
                'total_fee': 0,
                'average_fee': 0,
                'count': 0
            }

    return templates.TemplateResponse("past_projects.html", {
        "request": request,
        "current_user": current_user,
        "current_employee": current_bd_employee,  # Pass BD employee separately
        "projects": projects,
        "search": search,
        "sector_filter": sector_filter,
        "project_name": project_name,
        "client": client,
        "sub_sector": sub_sector,
        "start_date": start_date,
        "end_date": end_date,
        "state": state,
        "show_all": show_all,
        "page": page,
        "total_pages": total_pages,
        "has_prev": has_prev,
        "has_next": has_next,
        "total_projects": total_projects,
        "total_project_count": total_project_count,
        "search_performed": search_performed,
        "user_industry_sectors": user_industry_sectors,
        "sectors": sorted(sectors),
        "sub_sectors": sorted(sub_sectors),
        "clients": sorted(clients),
        "states": states,
        "sectors_subsectors_map": json.dumps(sectors_subsectors_map),
        "sector_analytics": sector_analytics,
        "sector_fee_analytics": sector_fee_analytics,
        "selected_font": get_active_font()
    })

# ==================== Test Endpoint Helper Functions ====================

def validate_test_token_and_get_user(test_token: Optional[str], db: Session) -> UserDB:
    """
    Validate test token and return test user.
    NO session creation, NO login - just token validation.
    """
    SECURE_TEST_TOKEN = os.getenv('TEST_TOKEN', 'pratyaksh_secure_test_2024_xyz')
    TEST_USER_EMAIL = 'nkbpl.pratyaksh@gmail.com'

    if not test_token or test_token != SECURE_TEST_TOKEN:
        raise HTTPException(
            status_code=403,
            detail="Invalid or missing test token. Access denied."
        )

    user = db.query(UserDB).filter(UserDB.email == TEST_USER_EMAIL).first()
    if not user:
        raise HTTPException(
            status_code=404,
            detail=f"Test user {TEST_USER_EMAIL} not found in database"
        )

    return user


# ==================== Test Endpoints (Complete Isolation) ====================

@app.get("/public-projects/nkbpl.pratyaksh", response_class=HTMLResponse)
async def public_test_projects(
    request: Request,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    ISOLATED testing endpoint for nkbpl.pratyaksh@gmail.com

    Complete functionality with STRICT segregation:
    - NO session creation
    - NO login cookies
    - Token validation on every request
    - All filtering, search, pagination supported
    - URLs never leave /public-projects/nkbpl.pratyaksh base

    Usage: /public-projects/nkbpl.pratyaksh?test_token=X&sector_filter=Water&search=...
    """
    # Validate token and get user (NO session)
    test_user = validate_test_token_and_get_user(test_token, db)

    # Get user_id for queries
    user_id_for_queries = test_user.id

    # Get company details for industry sectors
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == user_id_for_queries).first()
    user_industry_sectors = []
    if company_details and company_details.industry_sector:
        if company_details.industry_sector.startswith('[{'):
            try:
                sectors_data = json.loads(company_details.industry_sector)
                user_industry_sectors = [s['sector'] for s in sectors_data if 'sector' in s]
            except (json.JSONDecodeError, KeyError):
                user_industry_sectors = [company_details.industry_sector]
        elif company_details.industry_sector.startswith('['):
            try:
                user_industry_sectors = json.loads(company_details.industry_sector)
            except json.JSONDecodeError:
                user_industry_sectors = [company_details.industry_sector]
        else:
            user_industry_sectors = [company_details.industry_sector]

    # Get search and filter parameters (same as /projects endpoint)
    search = request.query_params.get('search', '')
    sector_filter = request.query_params.get('sector_filter', '')
    project_name = request.query_params.get('project_name', '')
    client = request.query_params.get('client', '')
    sub_sector = request.query_params.get('sub_sector', '')
    start_date = request.query_params.get('start_date', '')
    end_date = request.query_params.get('end_date', '')
    state = request.query_params.get('state', '')
    show_all = request.query_params.get('show_all', '')
    page = int(request.query_params.get('page', 1))
    per_page = 10

    # Get total project count
    total_project_count = db.query(ProjectDB).filter(ProjectDB.user_id == user_id_for_queries).count()

    # Initialize variables
    projects = []
    total_projects = 0
    total_pages = 0
    has_prev = False
    has_next = False

    # Check if search is performed
    search_performed = any([search, sector_filter, project_name, client, sub_sector, start_date, end_date, state, show_all])

    if search_performed:
        # Build query (same logic as /projects endpoint)
        query = db.query(ProjectDB).filter(ProjectDB.user_id == user_id_for_queries)

        # Apply filters
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    ProjectDB.project_name.ilike(search_term),
                    ProjectDB.project_description.ilike(search_term),
                    ProjectDB.client_name.ilike(search_term)
                )
            )

        if project_name:
            query = query.filter(ProjectDB.project_name.ilike(f"%{project_name}%"))

        if client:
            query = query.filter(ProjectDB.client_name.ilike(f"%{client}%"))

        if sector_filter:
            query = query.filter(ProjectDB.sector == sector_filter)

        if sub_sector:
            query = query.filter(ProjectDB.sub_sector.ilike(f"%{sub_sector}%"))

        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(ProjectDB.start_date >= start_date_obj)
            except ValueError:
                pass

        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                query = query.filter(ProjectDB.end_date <= end_date_obj)
            except ValueError:
                pass

        if state:
            query = query.filter(ProjectDB.states.contains([state]))

        # Get total count
        total_projects = query.count()

        # Apply pagination and ordering
        projects = query.order_by(desc(ProjectDB.created_at)).offset((page - 1) * per_page).limit(per_page).all()

        # Pagination info
        total_pages = (total_projects + per_page - 1) // per_page
        has_prev = page > 1
        has_next = page < total_pages

    # Get all projects for dropdowns
    all_projects = db.query(ProjectDB).filter(ProjectDB.user_id == user_id_for_queries).all()
    clients_list = list(set([p.client_name for p in all_projects if p.client_name]))

    # Build sectors and subsectors from actual user projects
    sectors = []
    sectors_subsectors_map = {}

    for project in all_projects:
        if project.sector:
            if project.sector not in sectors:
                sectors.append(project.sector)

            if project.sector not in sectors_subsectors_map:
                sectors_subsectors_map[project.sector] = []

            if project.sub_sector and project.sub_sector not in sectors_subsectors_map[project.sector]:
                sectors_subsectors_map[project.sector].append(project.sub_sector)

    # Sort for consistent display
    sectors = sorted(sectors)
    for sector in sectors_subsectors_map:
        sectors_subsectors_map[sector] = sorted(sectors_subsectors_map[sector])

    # Get subsectors based on selected sector filter
    sub_sectors = []
    if sector_filter and sector_filter in sectors_subsectors_map:
        sub_sectors = sectors_subsectors_map[sector_filter]
    elif not sector_filter and sectors_subsectors_map:
        all_subsectors = set()
        for subsector_list in sectors_subsectors_map.values():
            all_subsectors.update(subsector_list)
        sub_sectors = sorted(list(all_subsectors))

    # Extract all unique states from JSON states field
    states_set = set()
    for p in all_projects:
        if p.states and isinstance(p.states, list):
            states_set.update(p.states)
    states = sorted(list(states_set))

    # Calculate sector analytics
    sector_analytics = {}
    for sector in user_industry_sectors:
        sector_projects = [p for p in all_projects if p.sector == sector and p.project_cost is not None]
        if sector_projects:
            total_value = sum(p.project_cost for p in sector_projects)
            average_value = total_value / len(sector_projects)
            count = len(sector_projects)
            sector_analytics[sector] = {
                'total_value': total_value,
                'average_value': average_value,
                'count': count
            }
        else:
            sector_analytics[sector] = {
                'total_value': 0,
                'average_value': 0,
                'count': 0
            }

    # Calculate sector fee analytics
    sector_fee_analytics = {}
    for sector in user_industry_sectors:
        sector_projects = [p for p in all_projects if p.sector == sector and p.consultancy_fee is not None]
        if sector_projects:
            total_fee = sum(p.consultancy_fee for p in sector_projects)
            average_fee = total_fee / len(sector_projects)
            count = len(sector_projects)
            sector_fee_analytics[sector] = {
                'total_fee': total_fee,
                'average_fee': average_fee,
                'count': count
            }
        else:
            sector_fee_analytics[sector] = {
                'total_fee': 0,
                'average_fee': 0,
                'count': 0
            }

    # Render template - NO session, NO cookies, STRICT test mode
    return templates.TemplateResponse("past_projects.html", {
        "request": request,
        "current_user": test_user,
        "current_employee": None,
        "projects": projects,
        "search": search,
        "sector_filter": sector_filter,
        "project_name": project_name,
        "client": client,
        "sub_sector": sub_sector,
        "start_date": start_date,
        "end_date": end_date,
        "state": state,
        "show_all": show_all,
        "page": page,
        "total_pages": total_pages,
        "has_prev": has_prev,
        "has_next": has_next,
        "total_projects": total_projects,
        "total_project_count": total_project_count,
        "search_performed": search_performed,
        "user_industry_sectors": user_industry_sectors,
        "sectors": sorted(sectors),
        "sub_sectors": sorted(sub_sectors),
        "clients": sorted(clients_list),
        "states": states,
        "sectors_subsectors_map": json.dumps(sectors_subsectors_map),
        "sector_analytics": sector_analytics,
        "sector_fee_analytics": sector_fee_analytics,
        "is_test_mode": True,  # CRITICAL: Flag for template to use test URLs
        "test_token": test_token,  # CRITICAL: Pass token for URL building
        "test_base_url": "/public-projects/nkbpl.pratyaksh",  # CRITICAL: Base URL
        "selected_font": get_active_font()
    })


@app.get("/public-projects/nkbpl.pratyaksh/project/{project_id}", response_class=HTMLResponse)
async def test_project_detail(
    project_id: int,
    test_token: Optional[str] = None,
    return_url: Optional[str] = None,
    request: Request = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Project detail page."""
    test_user = validate_test_token_and_get_user(test_token, db)

    # Get project and verify ownership
    project = db.query(ProjectDB).filter(
        ProjectDB.id == project_id,
        ProjectDB.user_id == test_user.id
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Log what we retrieved from database
    logger.info(f"[TEST DETAIL] Loading project {project_id}: {project.project_name}")
    if project.documents:
        doc_count = sum(len(v) if isinstance(v, list) else 1 for v in project.documents.values())
        logger.info(f"[TEST DETAIL] Project has {doc_count} documents across {len(project.documents)} categories")
        for doc_type, docs in project.documents.items():
            count = len(docs) if isinstance(docs, list) else 1
            logger.info(f"[TEST DETAIL]   - {doc_type}: {count} files")
    else:
        logger.warning(f"[TEST DETAIL] ⚠ Project {project_id} has NO documents in database!")

    # Calculate file sizes for all documents (same logic as regular endpoint)
    documents_with_sizes = {}
    if project.documents and isinstance(project.documents, dict):
        for doc_type, file_entries in project.documents.items():
            documents_with_sizes[doc_type] = []
            # Ensure file_entries is a list
            if not isinstance(file_entries, list):
                file_entries = [file_entries] if file_entries else []

            for file_entry in file_entries:
                # Handle two formats:
                # 1. Old format: simple string file path
                # 2. New format: dict with file_path, metadata (for task deliverables)
                if isinstance(file_entry, dict):
                    # New format - deliverable with metadata
                    file_path = file_entry.get("file_path", "")
                    original_filename = file_entry.get("original_filename", os.path.basename(file_path))
                    uploaded_by = file_entry.get("uploaded_by")
                    uploaded_at = file_entry.get("uploaded_at")
                    task_title = file_entry.get("task_title")
                    task_file_id = file_entry.get("task_file_id")
                    stored_size = file_entry.get("file_size")
                    description = file_entry.get("description")

                    try:
                        # Get file size from disk or use stored size
                        if os.path.exists(file_path):
                            file_size = os.path.getsize(file_path)
                        else:
                            file_size = stored_size

                        documents_with_sizes[doc_type].append({
                            'path': file_path,
                            'size': file_size,
                            'original_filename': original_filename,
                            'uploaded_by': uploaded_by,
                            'uploaded_at': uploaded_at,
                            'task_title': task_title,
                            'task_file_id': task_file_id,
                            'description': description,
                            'is_deliverable': True
                        })
                    except Exception:
                        documents_with_sizes[doc_type].append({
                            'path': file_path,
                            'size': stored_size,
                            'original_filename': original_filename,
                            'uploaded_by': uploaded_by,
                            'uploaded_at': uploaded_at,
                            'task_title': task_title,
                            'description': description,
                            'is_deliverable': True
                        })
                else:
                    # Old format - simple file path string
                    file_path = file_entry
                    try:
                        # Get file size in bytes
                        file_size = os.path.getsize(file_path)
                        documents_with_sizes[doc_type].append({
                            'path': file_path,
                            'size': file_size,
                            'original_filename': os.path.basename(file_path),
                            'is_deliverable': False
                        })
                    except (OSError, IOError):
                        # Handle missing or inaccessible files
                        documents_with_sizes[doc_type].append({
                            'path': file_path,
                            'size': None,  # Will display as 'Unknown size'
                            'original_filename': os.path.basename(file_path),
                            'is_deliverable': False
                        })

    # Log what will be displayed
    total_display_docs = sum(len(v) for v in documents_with_sizes.values())
    logger.info(f"[TEST DETAIL] Rendering template with {total_display_docs} documents for display")
    for doc_type, docs in documents_with_sizes.items():
        logger.info(f"[TEST DETAIL]   - {doc_type}: {len(docs)} files for display")

    # Use return_url if provided, otherwise fallback to default
    if return_url:
        from urllib.parse import unquote
        decoded_return_url = unquote(return_url)
        back_url = decoded_return_url
    else:
        back_url = f"/public-projects/nkbpl.pratyaksh?test_token={test_token}"

    return templates.TemplateResponse("project_detail.html", {
        "request": request,
        "current_user": test_user,
        "project": project,
        "documents_with_sizes": documents_with_sizes,
        "is_test_mode": True,
        "test_token": test_token,
        "test_base_url": "/public-projects/nkbpl.pratyaksh",
        "test_redirect_url": back_url,
        "selected_font": get_active_font()
    })


@app.get("/public-projects/nkbpl.pratyaksh/edit/{project_id}", response_class=HTMLResponse)
async def test_edit_project_page(
    project_id: int,
    test_token: Optional[str] = None,
    return_url: Optional[str] = None,
    request: Request = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Show edit project form."""
    test_user = validate_test_token_and_get_user(test_token, db)

    # Get the project and verify ownership
    project = db.query(ProjectDB).filter(
        ProjectDB.id == project_id,
        ProjectDB.user_id == test_user.id
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get user's company details to retrieve industry sectors
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == test_user.id).first()
    user_industry_sectors = []
    user_sectors_data = []
    if company_details and company_details.industry_sector:
        if company_details.industry_sector.startswith('[{'):
            # New format: [{"sector": "...", "subsectors": [...]}]
            try:
                sectors_data = json.loads(company_details.industry_sector)
                user_industry_sectors = [s['sector'] for s in sectors_data if 'sector' in s]
                user_sectors_data = sectors_data
            except (json.JSONDecodeError, KeyError):
                user_industry_sectors = [company_details.industry_sector]
        elif company_details.industry_sector.startswith('['):
            try:
                user_industry_sectors = json.loads(company_details.industry_sector)
            except json.JSONDecodeError:
                user_industry_sectors = [company_details.industry_sector]
        else:
            user_industry_sectors = [company_details.industry_sector]

    return templates.TemplateResponse("edit_project.html", {
        "request": request,
        "current_user": test_user,
        "project": project,
        "user_industry_sectors": user_industry_sectors,
        "user_sectors_data": user_sectors_data,
        "now": datetime.now,
        "is_test_mode": True,
        "test_token": test_token,
        "test_base_url": "/public-projects/nkbpl.pratyaksh",
        "return_url": return_url,
        "selected_font": get_active_font()
    })


@app.post("/public-projects/nkbpl.pratyaksh/edit/{project_id}")
async def test_update_project(
    request: Request,
    project_id: int,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Handle project update with full document upload support."""
    test_user = validate_test_token_and_get_user(test_token, db)
    
    # Get return_url from form data
    form_data = await request.form()
    return_url = form_data.get('return_url', '')

    # Get the project
    project = db.query(ProjectDB).filter(
        ProjectDB.id == project_id,
        ProjectDB.user_id == test_user.id
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Parse form data manually to handle arrays and complex data
    form_data = await request.form()

    # Basic project fields
    project.project_name = form_data.get('project_name', '')
    project.project_description = form_data.get('project_description', '')
    project.complete_scope_of_work = form_data.get('complete_scope_of_work', '')
    project.client_name = form_data.get('client_name', '')
    project.sector = form_data.get('sector', '')
    project.sub_sector = form_data.get('sub_sector', '')
    project.jv_partner = form_data.get('jv_partner', '')
    project.country = form_data.get('country', 'India')

    # Financing authority handling
    financing_authority = form_data.get('financing_authority', '')
    financing_authority_entries = []
    if financing_authority:
        financing_authority_entries = [entry.strip() for entry in financing_authority.split(',') if entry.strip()]
    project.financing_authority = ", ".join(financing_authority_entries) if financing_authority_entries else "Financing Not Required"

    # Convert numeric fields
    consultancy_fee = form_data.get('consultancy_fee', '')
    if consultancy_fee and consultancy_fee.strip():
        try:
            project.consultancy_fee = float(consultancy_fee)
        except ValueError:
            project.consultancy_fee = None
    else:
        project.consultancy_fee = None

    project_cost = form_data.get('project_cost', '')
    if project_cost and project_cost.strip():
        try:
            project.project_cost = float(project_cost)
        except ValueError:
            project.project_cost = None
    else:
        project.project_cost = None

    # Date handling
    start_date = form_data.get('start_date')
    if start_date:
        try:
            project.start_date = datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            pass

    ongoing = form_data.get('ongoing', 'false')
    if ongoing == 'true':
        project.end_date = None
    else:
        end_date = form_data.get('end_date')
        if end_date:
            try:
                parsed_end_date = datetime.strptime(end_date, '%Y-%m-%d')
                # Validate that end_date is not in the future
                today = datetime.utcnow().date()
                if parsed_end_date.date() > today:
                    raise HTTPException(
                        status_code=400,
                        detail="Project completion date cannot be in the future. Please select today's date or earlier."
                    )
                project.end_date = parsed_end_date
            except ValueError:
                pass

    # Calculate project duration in months
    if project.start_date:
        end_for_calc = project.end_date if project.end_date else datetime.utcnow()
        if project.start_date <= end_for_calc:
            diff = end_for_calc - project.start_date
            project.project_duration_months = round(diff.days / 30.44)

    # Handle states (comma-separated string)
    states = form_data.get('states', '')
    project.states = [s.strip() for s in states.split(',') if s.strip()] if states else []

    # Handle cities (comma-separated string)
    cities = form_data.get('cities', '')
    project.cities = [c.strip() for c in cities.split(',') if c.strip()] if cities else []

    # Parse services rendered
    services_rendered = {}
    for key in PROJECT_SERVICE_KEYS:
        service_value = form_data.get(f'services[{key}]', '')
        if service_value:
            services_rendered[key] = service_value
    project.services_rendered = services_rendered

    # Handle document uploads (append to existing documents)
    existing_documents = project.documents if project.documents else {}

    # Ensure uploads directory exists
    os.makedirs("uploads/projects", exist_ok=True)

    logger.info(f"[TEST EDIT] Processing document uploads for project {project_id}")
    logger.info(f"[TEST EDIT] Current documents in DB: {sum(len(v) if isinstance(v, list) else 1 for v in existing_documents.values()) if existing_documents else 0} files across {len(existing_documents)} categories")

    # Log current document structure
    for doc_type, docs in existing_documents.items():
        doc_count = len(docs) if isinstance(docs, list) else 1
        logger.info(f"[TEST EDIT]   - {doc_type}: {doc_count} files")

    for doc_type in PROJECT_DOCUMENT_TYPES:
        doc_files = form_data.getlist(f'documents[{doc_type}][]')
        file_entries = []

        # Check each file to see if it's actually a file upload or just an empty input
        actual_files = []
        for file in doc_files:
            if hasattr(file, 'filename') and file.filename:
                # Check if file has content
                file_content = await file.read()
                if file_content:
                    actual_files.append((file.filename, file_content))

        logger.info(f"[TEST EDIT] Document type '{doc_type}': {len(doc_files)} inputs, {len(actual_files)} actual files")

        for filename, content in actual_files:
            try:
                # Generate unique filename
                file_extension = os.path.splitext(filename)[1]
                unique_filename = f"{uuid.uuid4()}{file_extension}"
                file_path = f"uploads/projects/{unique_filename}"

                with open(file_path, "wb") as buffer:
                    buffer.write(content)

                # Store both file path and original filename
                file_entries.append({
                    "file_path": file_path,
                    "original_filename": filename
                })
                logger.info(f"[TEST EDIT] ✓ Successfully uploaded: {filename} -> {file_path}")
            except Exception as e:
                logger.error(f"[TEST EDIT] ✗ Error uploading {doc_type} file {filename}: {e}")
                continue

        # Append new files to existing documents for this category
        if file_entries:
            if doc_type not in existing_documents:
                existing_documents[doc_type] = []
                logger.info(f"[TEST EDIT] Created new category: {doc_type}")

            before_count = len(existing_documents[doc_type]) if isinstance(existing_documents[doc_type], list) else 0
            existing_documents[doc_type].extend(file_entries)
            after_count = len(existing_documents[doc_type])
            logger.info(f"[TEST EDIT] ✓ Category '{doc_type}': {before_count} → {after_count} files (+{len(file_entries)} new)")

    # Update documents in project
    project.documents = existing_documents

    # CRITICAL: Flag the JSONB column as modified so SQLAlchemy tracks the change
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(project, "documents")

    total_files = sum(len(v) if isinstance(v, list) else 1 for v in existing_documents.values()) if existing_documents else 0
    logger.info(f"[TEST EDIT] === FINAL DOCUMENT COUNT: {total_files} files across {len(existing_documents)} categories ===")
    logger.info(f"[TEST EDIT] ✓ Flagged 'documents' field as modified for SQLAlchemy tracking")

    # If project was auto-generated and is now being edited, mark as completed_by_user
    if project.is_auto_generated and project.completion_status == "incomplete":
        project.completion_status = "completed_by_user"

    # Commit changes
    try:
        db.commit()
        db.refresh(project)
        logger.info(f"[TEST EDIT] ✓ Database commit successful for project {project_id}")

        # Verify documents were actually saved
        verify_project = db.query(ProjectDB).filter(ProjectDB.id == project_id).first()
        if verify_project and verify_project.documents:
            verified_count = sum(len(v) if isinstance(v, list) else 1 for v in verify_project.documents.values())
            logger.info(f"[TEST EDIT] ✓ VERIFICATION: Database shows {verified_count} documents after commit")

            # Log what's actually in the database
            for doc_type, docs in verify_project.documents.items():
                doc_count = len(docs) if isinstance(docs, list) else 1
                logger.info(f"[TEST EDIT]   ✓ DB has {doc_type}: {doc_count} files")
        else:
            logger.warning(f"[TEST EDIT] ⚠ VERIFICATION FAILED: No documents found in database after commit!")

    except Exception as e:
        db.rollback()
        logger.error(f"[TEST EDIT] ✗ Failed to update project {project_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update project: {str(e)}")

    # Redirect back to project detail in test mode, preserving return_url
    logger.info(f"[TEST EDIT] Redirecting to project detail page")
    redirect_url = f"/public-projects/nkbpl.pratyaksh/project/{project_id}?test_token={test_token}"
    if return_url:
        redirect_url += f"&return_url={return_url}"
    return RedirectResponse(url=redirect_url, status_code=302)


@app.delete("/public-projects/nkbpl.pratyaksh/project/{project_id}/document/{doc_type}/{doc_index}")
async def test_delete_project_document(
    project_id: int,
    doc_type: str,
    doc_index: int,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Delete a specific document from a project."""
    test_user = validate_test_token_and_get_user(test_token, db)

    logger.info(f"[TEST DELETE] Deleting document: project={project_id}, type={doc_type}, index={doc_index}")

    # Get project and verify ownership
    project = db.query(ProjectDB).filter(
        ProjectDB.id == project_id,
        ProjectDB.user_id == test_user.id
    ).first()

    if not project:
        logger.error(f"[TEST DELETE] Project {project_id} not found")
        raise HTTPException(status_code=404, detail="Project not found")

    # Check if project has documents
    if not project.documents or doc_type not in project.documents:
        logger.error(f"[TEST DELETE] Document type '{doc_type}' not found in project {project_id}")
        raise HTTPException(status_code=404, detail=f"Document type '{doc_type}' not found")

    # Get the documents for this type
    docs = project.documents[doc_type]
    if not isinstance(docs, list):
        docs = [docs]

    # Check if index is valid
    if doc_index < 0 or doc_index >= len(docs):
        logger.error(f"[TEST DELETE] Invalid document index {doc_index} for type '{doc_type}'")
        raise HTTPException(status_code=404, detail="Document not found")

    # Get the document to delete
    doc_to_delete = docs[doc_index]

    # Extract file path
    if isinstance(doc_to_delete, dict):
        file_path = doc_to_delete.get('file_path', '')
        filename = doc_to_delete.get('original_filename', 'unknown')
    else:
        file_path = doc_to_delete
        filename = os.path.basename(file_path)

    logger.info(f"[TEST DELETE] Deleting file: {filename} at {file_path}")

    # Delete file from filesystem
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"[TEST DELETE] ✓ File deleted from disk: {file_path}")
        else:
            logger.warning(f"[TEST DELETE] ⚠ File not found on disk: {file_path}")
    except Exception as e:
        logger.error(f"[TEST DELETE] Error deleting file from disk: {e}")
        # Continue with database deletion even if file deletion fails

    # Remove document from list
    docs.pop(doc_index)

    # Update project documents
    if len(docs) == 0:
        # Remove the category if no documents left
        del project.documents[doc_type]
        logger.info(f"[TEST DELETE] ✓ Removed empty category '{doc_type}'")
    else:
        # Update the category with remaining documents
        project.documents[doc_type] = docs

    # CRITICAL: Flag as modified for SQLAlchemy
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(project, "documents")

    logger.info(f"[TEST DELETE] ✓ Flagged 'documents' field as modified")

    # Commit changes
    try:
        db.commit()
        db.refresh(project)
        logger.info(f"[TEST DELETE] ✓ Successfully deleted document from project {project_id}")

        # Verify deletion
        verify_project = db.query(ProjectDB).filter(ProjectDB.id == project_id).first()
        if verify_project and verify_project.documents:
            remaining_count = sum(len(v) if isinstance(v, list) else 1 for v in verify_project.documents.values())
            logger.info(f"[TEST DELETE] ✓ VERIFICATION: {remaining_count} documents remaining in database")
        else:
            logger.info(f"[TEST DELETE] ✓ VERIFICATION: No documents remaining in project")

        return {"success": True, "message": f"Document '{filename}' deleted successfully"}

    except Exception as e:
        db.rollback()
        logger.error(f"[TEST DELETE] ✗ Failed to delete document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete document: {str(e)}")


@app.get("/public-projects/nkbpl.pratyaksh/api/sector-count", response_class=JSONResponse)
async def test_sector_count(
    test_token: Optional[str] = None,
    sector: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Get project count for a sector."""
    test_user = validate_test_token_and_get_user(test_token, db)

    if not sector:
        return {"count": 0}

    count = db.query(ProjectDB).filter(
        ProjectDB.user_id == test_user.id,
        ProjectDB.sector == sector
    ).count()

    return {"count": count}


@app.get("/public-projects/nkbpl.pratyaksh/project/{project_id}/download/{doc_type}/{doc_index}")
async def test_download_project_document(
    project_id: int,
    doc_type: str,
    doc_index: int,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Download project document."""
    test_user = validate_test_token_and_get_user(test_token, db)

    # Get project and verify ownership
    project = db.query(ProjectDB).filter(
        ProjectDB.id == project_id,
        ProjectDB.user_id == test_user.id
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get document path
    if not project.documents or doc_type not in project.documents:
        raise HTTPException(status_code=404, detail="Document type not found")

    docs_list = project.documents[doc_type]
    if doc_index >= len(docs_list):
        raise HTTPException(status_code=404, detail="Document index out of range")

    file_entry = docs_list[doc_index]
    
    # Handle both old format (string) and new format (dict with metadata)
    if isinstance(file_entry, dict):
        # New format - deliverable with metadata
        file_path = file_entry.get("file_path", "")
        filename = file_entry.get("original_filename", os.path.basename(file_path) if file_path else "document")
    else:
        # Old format - simple file path string
        file_path = file_entry
        filename = os.path.basename(file_path)

    # Normalize file path - remove leading slash if present (paths stored as "/uploads/..." should be "uploads/...")
    if file_path.startswith('/') and not os.path.exists(file_path):
        file_path = file_path.lstrip('/')
    
    # Check if file exists
    if not os.path.exists(file_path):
        logger.error(f"[TEST DOWNLOAD] File not found: path='{file_path}', filename='{filename}', file_entry={file_entry}")
        raise HTTPException(status_code=404, detail=f"File not found on server: {filename}")

    from fastapi.responses import FileResponse
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type='application/octet-stream'
    )


@app.get("/public-projects/nkbpl.pratyaksh/project/{project_id}/download/pdf")
async def test_download_project_pdf(
    project_id: int,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Generate and download a comprehensive PDF of the project."""
    test_user = validate_test_token_and_get_user(test_token, db)

    # Get the project
    project = db.query(ProjectDB).filter(
        and_(ProjectDB.id == project_id, ProjectDB.user_id == test_user.id)
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Generate comprehensive PDF (same logic as regular endpoint)
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        import io
        import os
        from pathlib import Path
        from PyPDF2 import PdfReader
        from pdf2image import convert_from_path
        from PIL import Image as PILImage
        import tempfile
        import html

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.75*inch, bottomMargin=0.75*inch)
        styles = getSampleStyleSheet()
        story = []

        # Helper function to escape HTML and convert to plain text for safe rendering
        def safe_paragraph(text, style=styles['Normal']):
            if not text:
                return Paragraph("", style)
            # Escape HTML entities and convert newlines
            text = html.escape(str(text))
            text = text.replace('\n', '<br/>')
            return Paragraph(text, style)

        # Title Page
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=28,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1e293b')
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=14,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#64748b')
        )
        
        story.append(Spacer(1, 2*inch))
        story.append(Paragraph(project.project_name, title_style))
        story.append(Paragraph("Project Comprehensive Documentation", subtitle_style))
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(f"Generated on {datetime.utcnow().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(PageBreak())

        # Table of Contents
        toc_style = ParagraphStyle(
            'TOCTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        story.append(Paragraph("Table of Contents", toc_style))
        story.append(Spacer(1, 20))
        
        # Build TOC items
        toc_items = ["1. Project Metadata"]
        if project.consultancy_fee or project.project_cost:
            toc_items.append("   1.1 Basic Information")
            toc_items.append("   1.2 Commercial Information")
        else:
            toc_items.append("   1.1 Basic Information")
        toc_items.append("   1.3 Timeline & Geography")
        if project.project_description:
            toc_items.append("2. Project Description")
        if project.complete_scope_of_work:
            toc_items.append("3. Complete Scope of Work")
        if project.services_rendered:
            toc_items.append("4. Services Delivered")
        if project.documents:
            toc_items.append("5. Supporting Documents")
        
        for item in toc_items:
            story.append(Paragraph(item, styles['Normal']))
        
        story.append(PageBreak())
        
        # 1. Project Metadata Section
        story.append(Paragraph("1. Project Metadata", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        # Basic Information
        story.append(Paragraph("1.1 Basic Information", styles['Heading2']))
        story.append(Spacer(1, 12))
        basic_data = []
        basic_data.append(["Project Name:", project.project_name])
        if project.client_name:
            basic_data.append(["Client:", project.client_name])
        if project.sector:
            sector_info = project.sector
            if project.sub_sector:
                sector_info += f" - {project.sub_sector}"
            basic_data.append(["Sector:", sector_info])
        basic_data.append(["Financing Authority:", project.financing_authority or "Financing Not Required"])
        if project.country:
            basic_data.append(["Country:", project.country])
        if project.jv_partner:
            basic_data.append(["JV Partner:", project.jv_partner])

        if basic_data:
            table = Table(basic_data, colWidths=[2.5*inch, 4.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
            ]))
            story.append(table)
            story.append(Spacer(1, 12))

        # Commercial Information
        if project.consultancy_fee or project.project_cost:
            story.append(Paragraph("1.2 Commercial Information", styles['Heading2']))
            story.append(Spacer(1, 12))
            commercial_data = []
            if project.consultancy_fee:
                commercial_data.append(["Consultancy Fee:", f"₹{project.consultancy_fee:,.0f}"])
            if project.project_cost:
                commercial_data.append(["Project Cost:", f"₹{project.project_cost:,.0f}"])

            if commercial_data:
                table = Table(commercial_data, colWidths=[2.5*inch, 4.5*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0fdf4')),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
                ]))
                story.append(table)
                story.append(Spacer(1, 12))

        # Timeline & Geography
        story.append(Paragraph("1.3 Timeline & Geography", styles['Heading2']))
        story.append(Spacer(1, 12))
        timeline_geo_data = []
        if project.start_date:
            timeline_geo_data.append(["Start Date:", project.start_date.strftime('%B %d, %Y')])
        timeline_geo_data.append(["Completion Date:", 
                                  project.end_date.strftime('%B %d, %Y') if project.end_date else "Ongoing"])
        if project.project_duration_months:
            timeline_geo_data.append(["Duration:", f"{project.project_duration_months} months"])
        if project.states:
            states_str = ", ".join(project.states) if isinstance(project.states, list) else str(project.states)
            timeline_geo_data.append(["States:", states_str])
        if project.cities:
            cities_str = ", ".join(project.cities) if isinstance(project.cities, list) else str(project.cities)
            timeline_geo_data.append(["Cities:", cities_str])

        if timeline_geo_data:
            table = Table(timeline_geo_data, colWidths=[2.5*inch, 4.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fefce8')),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
            ]))
            story.append(table)
            story.append(Spacer(1, 12))

        # 2. Project Description
        if project.project_description:
            story.append(PageBreak())
            story.append(Paragraph("2. Project Description", styles['Heading1']))
            story.append(Spacer(1, 12))
            story.append(safe_paragraph(project.project_description))
            story.append(Spacer(1, 12))

        # 3. Complete Scope of Work
        if project.complete_scope_of_work:
            story.append(PageBreak())
            story.append(Paragraph("3. Complete Scope of Work", styles['Heading1']))
            story.append(Spacer(1, 12))
            story.append(safe_paragraph(project.complete_scope_of_work))
            story.append(Spacer(1, 12))

        # 4. Services Delivered
        if project.services_rendered:
            story.append(PageBreak())
            story.append(Paragraph("4. Services Delivered", styles['Heading1']))
            story.append(Spacer(1, 12))
            services_data = []
            service_labels = {
                'survey_investigations': 'Survey & investigations',
                'design_engineering': 'Design & engineering',
                'dpr_feasibility': 'DPR & feasibility studies',
                'gis_data': 'GIS & data services',
                'pmc': 'Project management consultancy (PMC)',
                'pmu': 'Project management unit (PMU)',
                'advisory_capacity': 'Advisory & capacity building',
                'supervision': 'Construction supervision',
                'quality_control': 'Quality control & assurance',
                'environmental': 'Environmental & social impact',
                'financial_advisory': 'Financial advisory services'
            }
            if isinstance(project.services_rendered, dict):
                for key, value in project.services_rendered.items():
                    if value:
                        label = service_labels.get(key, key.replace('_', ' ').title())
                        services_data.append([label, "✓"])
            elif isinstance(project.services_rendered, list):
                for service in project.services_rendered:
                    if isinstance(service, dict):
                        label = service_labels.get(service.get('key', ''), service.get('label', ''))
                        services_data.append([label, "✓"])
                    else:
                        services_data.append([str(service), "✓"])

            if services_data:
                table = Table(services_data, colWidths=[5*inch, 2*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
                ]))
                story.append(table)
                story.append(Spacer(1, 12))

        # 5. Supporting Documents
        if project.documents:
            story.append(PageBreak())
            story.append(Paragraph("5. Supporting Documents", styles['Heading1']))
            story.append(Spacer(1, 12))
            
            doc_type_labels = {
                'technical_documents': 'Technical Documents',
                'financial_documents': 'Financial Documents',
                'legal_documents': 'Legal Documents',
                'other_documents': 'Other Documents'
            }
            
            for doc_type, file_entries in project.documents.items():
                if not file_entries:
                    continue
                    
                doc_label = doc_type_labels.get(doc_type, doc_type.replace('_', ' ').title())
                story.append(Paragraph(doc_label, styles['Heading2']))
                story.append(Spacer(1, 8))
                
                # Ensure file_entries is a list
                if not isinstance(file_entries, list):
                    file_entries = [file_entries] if file_entries else []
                
                for idx, file_entry in enumerate(file_entries):
                    # Handle both old format (string) and new format (dict)
                    if isinstance(file_entry, dict):
                        file_path = file_entry.get("file_path", "")
                        original_filename = file_entry.get("original_filename", os.path.basename(file_path))
                    else:
                        file_path = file_entry
                        original_filename = os.path.basename(file_path)
                    
                    story.append(Paragraph(f"• {original_filename}", styles['Normal']))
                
                story.append(Spacer(1, 12))

        # Build PDF
        doc.build(story)

        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{project.project_name.replace(" ", "_")}_comprehensive.pdf"'}
        )

    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"PDF generation library not available: {str(e)}")
    except Exception as e:
        logger.error(f"Error generating comprehensive PDF: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")


@app.get("/project-management", response_class=HTMLResponse)
@require_company_details
@require_pin_verification
async def project_management_page(request: Request, db: Session = Depends(get_db)):
    """Project Management page - Shows awarded tenders as project cards."""
    from core.dependencies import get_current_bd_employee

    # Block BD employees from accessing project management
    bd_employee = get_current_bd_employee(request, db)
    if bd_employee:
        return RedirectResponse(url="/bd/home", status_code=302)

    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Get all awarded tenders for this user
    awarded_tenders = db.query(TenderDB).filter(
        TenderDB.awarded_by == current_user.id,
        TenderDB.awarded == True
    ).options(
        joinedload(TenderDB.documents)
    ).order_by(TenderDB.awarded_at.desc()).all()

    # Build project data with team and task statistics
    projects = []
    for tender in awarded_tenders:
        # Get team assignments
        assignments = db.query(TenderAssignmentDB).filter(
            TenderAssignmentDB.tender_id == tender.id
        ).options(
            joinedload(TenderAssignmentDB.employee)
        ).all()

        # Get task statistics
        tasks = db.query(TaskDB).join(
            TenderAssignmentDB, TaskDB.assignment_id == TenderAssignmentDB.id
        ).filter(
            TenderAssignmentDB.tender_id == tender.id
        ).all()

        total_tasks = len(tasks)
        completed_tasks = sum(1 for t in tasks if t.status == 'completed')
        in_progress_tasks = sum(1 for t in tasks if t.status == 'in_progress')
        pending_tasks = sum(1 for t in tasks if t.status == 'pending')

        # Calculate progress percentage
        progress = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0

        # Determine project status
        if total_tasks == 0:
            project_status = "Planning"
            status_color = "neutral"
        elif completed_tasks == total_tasks:
            project_status = "Completed"
            status_color = "success"
        elif in_progress_tasks > 0 or completed_tasks > 0:
            project_status = "In Progress"
            status_color = "primary"
        else:
            project_status = "Not Started"
            status_color = "warning"

        # Get the actual ProjectDB entry if it exists
        project_entry = db.query(ProjectDB).filter(
            ProjectDB.source_tender_id == tender.id
        ).first()

        projects.append({
            'id': project_entry.id if project_entry else None,
            'project_name': project_entry.project_name if project_entry else tender.title,
            'tender': tender,
            'team_members': [a.employee for a in assignments if a.employee],
            'team_count': len([a for a in assignments if a.employee]),
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'in_progress_tasks': in_progress_tasks,
            'pending_tasks': pending_tasks,
            'progress': round(progress, 1),
            'status': project_status,
            'status_color': status_color
        })

    return templates.TemplateResponse("project_management.html", {
        "request": request,
        "current_user": current_user,
        "projects": projects,
        "total_projects": len(projects),
        "active_projects": len([p for p in projects if p['status'] == 'In Progress']),
        "completed_projects": len([p for p in projects if p['status'] == 'Completed']),
        "selected_font": get_active_font()
    })

@app.get("/api/expert-requests")
@require_company_details
async def get_manager_expert_requests(request: Request, db: Session = Depends(get_db)):
    """Get all expert hiring requests created by the current manager."""
    current_user = get_current_user(request, db)
    if not current_user:
        logger.error("[get_manager_expert_requests] No authenticated user")
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    logger.info(f"[get_manager_expert_requests] Fetching requests for user_id={current_user.id}")
    
    # Get all requests created by this user
    requests = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.company_id == current_user.id
    ).order_by(ExpertHiringRequestDB.created_at.desc()).all()
    
    logger.info(f"[get_manager_expert_requests] Found {len(requests)} hiring requests for user_id={current_user.id}")
    
    requests_data = []
    for req in requests:
        # Get application counts by status
        total_applications = db.query(ExpertHiringApplicationDB).filter(
            ExpertHiringApplicationDB.request_id == req.id
        ).count()
        
        pending_applications = db.query(ExpertHiringApplicationDB).filter(
            ExpertHiringApplicationDB.request_id == req.id,
            ExpertHiringApplicationDB.status == 'pending'
        ).count()
        
        accepted_applications = db.query(ExpertHiringApplicationDB).filter(
            ExpertHiringApplicationDB.request_id == req.id,
            ExpertHiringApplicationDB.status == 'accepted'
        ).count()
        
        if total_applications > 0:
            logger.info(f"[get_manager_expert_requests] Request {req.id} has {total_applications} applications (pending: {pending_applications}, accepted: {accepted_applications})")
        
        # Format budget display
        if req.budget_type == 'fixed':
            budget_display = f"₹{req.budget_amount:,.0f}" if req.budget_amount else "Not specified"
        else:
            if req.budget_min and req.budget_max:
                budget_display = f"₹{req.budget_min:,.0f} - ₹{req.budget_max:,.0f}"
            elif req.budget_min:
                budget_display = f"₹{req.budget_min:,.0f}+"
            else:
                budget_display = "Negotiable"
        
        requests_data.append({
            "id": req.id,
            "request_name": req.request_name,
            "description": req.description,
            "company_name": req.company_name,
            "company_location": req.company_location,
            "company_details": {
                "name": req.company_name,
                "location": req.company_location
            },
            "tender_title": req.tender_title,
            "tender_sector": req.tender_sector,
            "tender_state": req.tender_state,  # Legacy field
            "tender_location": getattr(req, 'tender_location', None) or req.tender_state,  # Use location if available, fallback to state
            "budget_type": req.budget_type,
            "budget_display": budget_display,
            "status": req.status,
            "project_id": req.project_id,
            "tender_id": req.tender_id,
            "created_at": req.created_at.isoformat() if req.created_at else None,
            "updated_at": req.updated_at.isoformat() if req.updated_at else None,
            "closed_at": req.closed_at.isoformat() if req.closed_at else None,
            "application_stats": {
                "total": total_applications,
                "pending": pending_applications,
                "accepted": accepted_applications
            }
        })
    
    total_apps_count = sum(r["application_stats"]["total"] for r in requests_data)
    logger.info(f"[get_manager_expert_requests] Returning {len(requests_data)} requests with {total_apps_count} total applications")
    
    return {
        "requests": requests_data,
        "total_requests": len(requests_data)
    }

@app.get("/api/expert-request/{request_id}/applications")
@require_company_details
async def get_request_applications(
    request: Request,
    request_id: str,
    db: Session = Depends(get_db)
):
    """Get all applications for a specific hiring request."""
    current_user = get_current_user(request, db)
    if not current_user:
        logger.error(f"[get_request_applications] No authenticated user for request_id={request_id}")
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    logger.info(f"[get_request_applications] User {current_user.id} fetching applications for request_id={request_id}")
    
    # Get the hiring request
    hiring_request = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.id == request_id
    ).first()
    
    if not hiring_request:
        logger.error(f"[get_request_applications] Hiring request not found: request_id={request_id}")
        raise HTTPException(status_code=404, detail="Hiring request not found")
    
    logger.info(f"[get_request_applications] Found hiring request: request_id={request_id}, company_id={hiring_request.company_id}, request_name={hiring_request.request_name}")
    
    # Verify ownership
    if hiring_request.company_id != current_user.id:
        logger.error(f"[get_request_applications] Ownership mismatch: request.company_id={hiring_request.company_id} != current_user.id={current_user.id} for request_id={request_id}")
        raise HTTPException(status_code=403, detail="Not authorized to view applications for this request")

    # Ensure expert_reviews table has necessary column (with error handling)
    try:
        ensure_expert_review_hiring_column()
    except Exception as e:
        logger.warning(f"[get_request_applications] Failed to ensure expert_review column (non-fatal): {e}")
        # Continue anyway - this is not critical for fetching applications
    
    # Get all applications
    applications = db.query(ExpertHiringApplicationDB).filter(
        ExpertHiringApplicationDB.request_id == request_id
    ).order_by(ExpertHiringApplicationDB.applied_at.desc()).all()
    
    logger.info(f"[get_request_applications] Found {len(applications)} applications for request_id={request_id}")
    
    applications_data = []
    for app in applications:
        # Get expert details
        expert = db.query(ExpertDB).filter(ExpertDB.id == app.expert_id).first()
        expert_profile = db.query(ExpertProfileDB).filter(
            ExpertProfileDB.expert_id == app.expert_id
        ).first() if expert else None

        # Try to get review - handle case where hiring_request_id column might not exist
        review = None
        try:
            review = db.query(ExpertReviewDB).filter(
                ExpertReviewDB.hiring_request_id == app.request_id,
                ExpertReviewDB.expert_id == app.expert_id,
                ExpertReviewDB.company_id == current_user.id
            ).first()
        except Exception as e:
            logger.warning(f"[get_request_applications] Could not query expert review for app {app.id}: {e}")
            # Review is optional, so we can continue without it
        
        applications_data.append({
            "id": app.id,
            "expert_id": app.expert_id,
            "expert_name": expert.name if expert else "Unknown Expert",
            "expert_email": expert.email if expert else None,
            "expert_rating": expert.rating_average if expert else None,
            "expert_location": expert_profile.location if expert_profile else None,
            "expert_experience_years": expert_profile.experience_years if expert_profile else None,
            "expert_expertise_areas": expert_profile.expertise_areas if expert_profile else [],
            "cover_letter": app.cover_letter,
            "proposed_rate": app.proposed_rate,
            "estimated_timeline": app.estimated_timeline,
            "relevant_experience": app.relevant_experience,
            "status": app.status,
            "manager_notes": app.manager_notes,
            "applied_at": app.applied_at.isoformat() if app.applied_at else None,
            "reviewed_at": app.reviewed_at.isoformat() if app.reviewed_at else None,
            "rating_value": review.rating if review else None,
            "rating_review": review.review_text if review else None
        })
    
    logger.info(f"[get_request_applications] Returning {len(applications_data)} applications for request_id={request_id}")
    
    return {
        "request_id": request_id,
        "request_name": hiring_request.request_name,
        "applications": applications_data,
        "total_applications": len(applications_data)
    }

@app.put("/api/expert-application/{application_id}/status")
@require_company_details
async def update_application_status(
    request: Request,
    application_id: str,
    db: Session = Depends(get_db)
):
    """Update the status of an expert application."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Get the application
    application = db.query(ExpertHiringApplicationDB).filter(
        ExpertHiringApplicationDB.id == application_id
    ).first()
    
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    
    # Get the hiring request to verify ownership
    hiring_request = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.id == application.request_id
    ).first()
    
    if not hiring_request or hiring_request.company_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to update this application")
    
    # Get update data
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON data")
    
    new_status = data.get("status", "").lower()
    manager_notes = data.get("manager_notes", "").strip()
    
    # Validate status
    valid_statuses = ['pending', 'shortlisted', 'accepted', 'rejected', 'completed']
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Status must be one of: {', '.join(valid_statuses)}")
    
    # Update application
    application.status = new_status
    application.reviewed_at = datetime.utcnow()
    
    if manager_notes:
        application.manager_notes = manager_notes
    
    # If accepted, optionally close the request
    if new_status == 'accepted' and data.get("close_request", False):
        hiring_request.status = 'filled'
        hiring_request.closed_at = datetime.utcnow()

    if new_status != 'pending':
        status_label = new_status.capitalize()
        base_message = f"Your application for \"{hiring_request.request_name}\" was {status_label}."
        if manager_notes:
            base_message += f" Notes: {manager_notes}"
        create_expert_notification(
            db,
            application.expert_id,
            f"application_{new_status}",
            f"Application {status_label}",
            base_message,
            link=f"/expert/opportunities?source=requests&request_id={hiring_request.id}"
        )
    
    db.commit()
    db.refresh(application)
    
    return {
        "success": True,
        "message": "Application status updated successfully",
        "application_id": application.id,
        "status": application.status
    }

@app.post("/api/expert-application/{application_id}/close-project")
@require_company_details
async def close_expert_application_project(
    request: Request,
    application_id: str,
    db: Session = Depends(get_db)
):
    """Close an accepted expert application/project."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    application = db.query(ExpertHiringApplicationDB).filter(
        ExpertHiringApplicationDB.id == application_id
    ).first()

    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    hiring_request = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.id == application.request_id
    ).first()

    if not hiring_request or hiring_request.company_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to close this project")

    if application.status != 'accepted':
        raise HTTPException(status_code=400, detail="Only accepted applications can be closed")

    try:
        data = await request.json()
    except Exception:
        data = {}

    closing_notes = (data.get("closing_notes") or "").strip()

    application.status = 'completed'
    application.reviewed_at = datetime.utcnow()

    if closing_notes:
        application.manager_notes = closing_notes

    create_expert_notification(
        db,
        application.expert_id,
        "project_closed",
        "Project Closed",
        f"Your engagement for \"{hiring_request.request_name}\" has been closed by the manager.{' Notes: ' + closing_notes if closing_notes else ''}",
        link=f"/expert/opportunities?source=requests&request_id={hiring_request.id}"
    )

    db.commit()
    db.refresh(application)

    return {
        "success": True,
        "message": "Project closed successfully",
        "application_id": application.id,
        "status": application.status
    }

@app.post("/api/expert-application/{application_id}/rate")
@require_company_details
async def rate_expert_application(
    request: Request,
    application_id: str,
    db: Session = Depends(get_db)
):
    """Submit a rating and review for a completed expert application."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    application = db.query(ExpertHiringApplicationDB).filter(
        ExpertHiringApplicationDB.id == application_id
    ).first()

    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    hiring_request = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.id == application.request_id
    ).first()

    if not hiring_request or hiring_request.company_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to rate this project")

    if not ensure_expert_review_hiring_column():
        raise HTTPException(
            status_code=500,
            detail="Unable to submit rating because the database schema is missing required columns."
        )

    if application.status != 'completed':
        raise HTTPException(status_code=400, detail="You can only rate experts after closing the project")

    existing_review = db.query(ExpertReviewDB).filter(
        ExpertReviewDB.hiring_request_id == application.request_id,
        ExpertReviewDB.expert_id == application.expert_id,
        ExpertReviewDB.company_id == current_user.id
    ).first()

    if existing_review:
        raise HTTPException(status_code=400, detail="You have already rated this expert for this project")

    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    try:
        rating_value = int(data.get("rating"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Rating must be an integer between 1 and 10")

    if rating_value < 1 or rating_value > 10:
        raise HTTPException(status_code=400, detail="Rating must be between 1 and 10")

    review_text = (data.get("review") or "").strip()

    review = ExpertReviewDB(
        expert_id=application.expert_id,
        company_id=current_user.id,
        request_id=None,
        hiring_request_id=application.request_id,
        rating=rating_value,
        review_text=review_text
    )
    db.add(review)

    # Update expert rating average
    expert = db.query(ExpertDB).filter(ExpertDB.id == application.expert_id).first()
    if expert:
        db.flush()
        avg_rating = db.query(func.avg(ExpertReviewDB.rating)).filter(
            ExpertReviewDB.expert_id == expert.id
        ).scalar()
        expert.rating_average = float(avg_rating) if avg_rating else 0.0

    create_expert_notification(
        db,
        application.expert_id,
        "rating_received",
        "New Rating Received",
        f"You received a {rating_value}/10 rating for \"{hiring_request.request_name}\".",
        link="/expert/dashboard"
    )

    db.commit()
    db.refresh(review)

    return {
        "success": True,
        "message": "Rating submitted successfully",
        "application_id": application.id,
        "rating": review.rating,
        "review_text": review.review_text
    }

@app.put("/api/expert-request/{request_id}/close")
@require_company_details
async def close_hiring_request(
    request: Request,
    request_id: str,
    db: Session = Depends(get_db)
):
    """Close an expert hiring request."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Get the hiring request
    hiring_request = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.id == request_id
    ).first()
    
    if not hiring_request:
        raise HTTPException(status_code=404, detail="Hiring request not found")
    
    # Verify ownership
    if hiring_request.company_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to close this request")
    
    # Update status
    hiring_request.status = 'closed'
    hiring_request.closed_at = datetime.utcnow()
    
    db.commit()
    db.refresh(hiring_request)
    
    return {
        "success": True,
        "message": "Hiring request closed successfully",
        "request_id": hiring_request.id,
        "status": hiring_request.status
    }

@app.get("/manage-experts", response_class=HTMLResponse)
@require_company_details
async def manage_experts_page(request: Request, db: Session = Depends(get_db)):
    """Manage Experts page - View and manage expert hiring requests and applications."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("manage_experts.html", {
        "request": request,
        "current_user": current_user,
        "selected_font": get_active_font()
    })

@app.get("/employee-performance", response_class=HTMLResponse)
@require_company_details
async def employee_performance_page(request: Request, db: Session = Depends(get_db)):
    """Employee Performance Analytics page - Manager only."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)
    
    # Get all employees for this company via CompanyCodeDB (same logic as API endpoint)
    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
    
    teams = []
    has_unassigned = False
    
    if company_codes:
        # Get company code IDs for this user
        company_code_ids = [cc.id for cc in company_codes]
        
        # Get all unique teams from employees belonging to this user's company codes only
        team_results = db.query(EmployeeDB.team).filter(
            EmployeeDB.company_code_id.in_(company_code_ids),
            EmployeeDB.is_active == True,
            EmployeeDB.team.isnot(None),
            EmployeeDB.team != ''
        ).distinct().all()
        
        teams = [t[0] for t in team_results if t[0]]
        
        # Check if there are any employees without teams (unassigned)
        unassigned_count = db.query(EmployeeDB).filter(
            EmployeeDB.company_code_id.in_(company_code_ids),
            EmployeeDB.is_active == True,
            or_(
                EmployeeDB.team.is_(None),
                EmployeeDB.team == ''
            )
        ).count()
        
        if unassigned_count > 0:
            has_unassigned = True
    
    # Add "Unassigned" to teams list if there are employees without teams
    if has_unassigned:
        teams.append("Unassigned")
    
    return templates.TemplateResponse("employee_performance.html", {
        "request": request,
        "current_user": current_user,
        "teams": sorted(teams),
        "selected_font": get_active_font()
    })

@app.get("/api/employee-performance")
async def get_employee_performance_data(request: Request, days: int = 30, db: Session = Depends(get_db)):
    """Get performance analytics for all employees."""
    from datetime import datetime, timedelta
    from sqlalchemy import func, case
    
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Calculate date range
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)
    
    # Get all employees for this company via CompanyCodeDB
    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
    if not company_codes:
        return {"employees": []}
    
    # Get employees from all company codes associated with this user
    company_code_ids = [cc.id for cc in company_codes]
    employees = db.query(EmployeeDB).filter(
        EmployeeDB.company_code_id.in_(company_code_ids),
        EmployeeDB.is_active == True
    ).order_by(EmployeeDB.name).all()
    
    employee_data = []
    
    for emp in employees:
        # Get all tasks assigned to this employee
        assignments = db.query(TenderAssignmentDB).filter(
            TenderAssignmentDB.employee_id == emp.id
        ).all()
        
        assignment_ids = [a.id for a in assignments]
        
        if not assignment_ids:
            # Employee has no tasks yet
            employee_data.append({
                "id": emp.id,
                "name": emp.name,
                "email": emp.email,
                "team": emp.team,
                "performance_score": 0,
                "tasks_completed": 0,
                "tasks_pending": 0,
                "tasks_in_progress": 0,
                "completion_rate": 0,
                "avg_completion_time": "N/A",
                "avg_first_response_time": "N/A",
                "total_progress_logs": 0,
                "avg_updates_per_task": 0,
                "total_concerns": 0,
                "concerns_resolved": 0,
                "deliverables_uploaded": 0,
                "tenders_participated": 0,
                "active_tenders": 0,
                "total_tender_value": "$0",
                "current_rating": 0,
                "rating_date": None,
                "weekly_tasks": []
            })
            continue
        
        # Get tasks in date range
        tasks_all = db.query(TaskDB).filter(
            TaskDB.assignment_id.in_(assignment_ids)
        ).all()
        
        tasks_in_range = [t for t in tasks_all if t.created_at and t.created_at >= start_date]
        
        # Task metrics
        total_tasks = len(tasks_all)
        tasks_completed = len([t for t in tasks_all if t.status == 'completed'])
        tasks_pending = len([t for t in tasks_all if t.status == 'pending'])
        tasks_in_progress = len([t for t in tasks_all if t.status == 'in_progress'])
        completion_rate = round((tasks_completed / total_tasks * 100) if total_tasks > 0 else 0, 1)
        
        # Average completion time (for completed tasks)
        completed_tasks = [t for t in tasks_all if t.status == 'completed' and t.created_at and t.completed_at]
        if completed_tasks:
            total_time = sum((t.completed_at - t.created_at).total_seconds() for t in completed_tasks)
            avg_completion_days = round(total_time / len(completed_tasks) / 86400, 1)
            avg_completion_time = str(avg_completion_days)
        else:
            avg_completion_time = "N/A"
        
        # Average first response time (time from task creation to first progress log)
        task_ids = [t.id for t in tasks_in_range]
        first_responses = []
        for task_id in task_ids:
            task = next((t for t in tasks_in_range if t.id == task_id), None)
            if not task or not task.created_at:
                continue
            
            first_log = db.query(TaskProgressUpdateDB).filter(
                TaskProgressUpdateDB.task_id == task_id
            ).order_by(TaskProgressUpdateDB.created_at).first()
            
            if first_log and first_log.created_at:
                response_time = (first_log.created_at - task.created_at).total_seconds() / 3600
                first_responses.append(response_time)
        
        if first_responses:
            avg_first_response = round(sum(first_responses) / len(first_responses), 1)
            avg_first_response_time = str(avg_first_response)
        else:
            avg_first_response_time = "N/A"
        
        # Progress logs
        total_progress_logs = db.query(TaskProgressUpdateDB).join(
            TaskDB, TaskProgressUpdateDB.task_id == TaskDB.id
        ).filter(
            TaskDB.assignment_id.in_(assignment_ids)
        ).count()
        
        avg_updates_per_task = round(total_progress_logs / total_tasks, 1) if total_tasks > 0 else 0
        
        # Concerns
        total_concerns = db.query(TaskConcernDB).join(
            TaskDB, TaskConcernDB.task_id == TaskDB.id
        ).filter(
            TaskDB.assignment_id.in_(assignment_ids)
        ).count()
        
        concerns_resolved = db.query(TaskConcernDB).join(
            TaskDB, TaskConcernDB.task_id == TaskDB.id
        ).filter(
            TaskDB.assignment_id.in_(assignment_ids),
            TaskConcernDB.status == 'resolved'
        ).count()
        
        # Deliverables
        deliverables_uploaded = db.query(TaskFileDB).join(
            TaskDB, TaskFileDB.task_id == TaskDB.id
        ).filter(
            TaskDB.assignment_id.in_(assignment_ids)
        ).count()
        
        # Tender participation
        unique_tender_ids = list(set([a.tender_id for a in assignments]))
        tenders_participated = len(unique_tender_ids)
        
        # Active tenders (with pending or in-progress tasks)
        active_tender_ids = list(set([
            t.assignment_id for t in tasks_all 
            if t.status in ['pending', 'in_progress']
        ]))
        active_assignments = [a for a in assignments if a.id in active_tender_ids]
        active_tenders = len(set([a.tender_id for a in active_assignments]))
        
        # Total tender value
        tenders = db.query(TenderDB).filter(TenderDB.id.in_(unique_tender_ids)).all()
        total_value = sum(t.estimated_value or 0 for t in tenders)
        total_tender_value = f"${total_value:,.0f}" if total_value else "$0"
        
        # Get current rating
        current_rating = 0
        rating_date = None
        latest_rating = db.query(EmployeePerformanceRatingDB).filter(
            EmployeePerformanceRatingDB.employee_id == emp.id
        ).order_by(EmployeePerformanceRatingDB.week_start_date.desc()).first()
        
        if latest_rating:
            current_rating = latest_rating.rating
            rating_date = latest_rating.week_start_date.strftime('%b %d, %Y')
        
        # Calculate performance score (0-100)
        # Weighted average of different metrics
        performance_score = round(
            (completion_rate * 0.4) +  # 40% weight on completion rate
            (min(avg_updates_per_task / 3 * 100, 100) * 0.2) +  # 20% on communication
            (max(100 - (total_concerns * 5), 0) * 0.2) +  # 20% on quality (fewer concerns)
            (current_rating / 5 * 100 * 0.2)  # 20% on manager rating
        )
        
        # Weekly task data for chart
        weekly_tasks = []
        for i in range(min(days // 7, 12)):  # Last 12 weeks max
            week_start = end_date - timedelta(days=(i+1)*7)
            week_end = end_date - timedelta(days=i*7)
            
            week_tasks = [t for t in tasks_all if t.created_at and week_start <= t.created_at < week_end]
            week_completed = len([t for t in week_tasks if t.status == 'completed'])
            
            weekly_tasks.insert(0, {
                "week": f"W{i+1}",
                "assigned": len(week_tasks),
                "completed": week_completed
            })
        
        employee_data.append({
            "id": emp.id,
            "name": emp.name,
            "email": emp.email,
            "team": emp.team,
            "performance_score": performance_score,
            "tasks_completed": tasks_completed,
            "tasks_pending": tasks_pending,
            "tasks_in_progress": tasks_in_progress,
            "completion_rate": completion_rate,
            "avg_completion_time": avg_completion_time,
            "avg_first_response_time": avg_first_response_time,
            "total_progress_logs": total_progress_logs,
            "avg_updates_per_task": avg_updates_per_task,
            "total_concerns": total_concerns,
            "concerns_resolved": concerns_resolved,
            "deliverables_uploaded": deliverables_uploaded,
            "tenders_participated": tenders_participated,
            "active_tenders": active_tenders,
            "total_tender_value": total_tender_value,
            "current_rating": current_rating,
            "rating_date": rating_date,
            "weekly_tasks": weekly_tasks[:8]  # Show last 8 weeks
        })
    
    return {"employees": employee_data}

@app.post("/api/employee-performance/rating")
async def save_employee_rating(request: Request, db: Session = Depends(get_db)):
    """Save manager's rating for an employee (Friday only)."""
    from datetime import datetime, timedelta
    
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Check if today is Friday (4 = Friday in Python's weekday())
    if datetime.now().weekday() != 4:
        raise HTTPException(status_code=403, detail="Ratings can only be submitted on Fridays")
    
    data = await request.json()
    employee_id = data.get('employee_id')
    rating = data.get('rating')
    feedback = data.get('feedback', '')
    
    if not employee_id or rating is None:
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    if rating < 0 or rating > 5:
        raise HTTPException(status_code=400, detail="Rating must be between 0 and 5")
    
    # Verify employee exists
    employee = db.query(EmployeeDB).filter(EmployeeDB.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Calculate week start (Monday) and end (Sunday)
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())  # Monday
    week_end = week_start + timedelta(days=6)  # Sunday
    
    # Check if rating already exists for this week
    existing_rating = db.query(EmployeePerformanceRatingDB).filter(
        EmployeePerformanceRatingDB.employee_id == employee_id,
        EmployeePerformanceRatingDB.manager_id == current_user.id,
        EmployeePerformanceRatingDB.week_start_date == week_start
    ).first()
    
    if existing_rating:
        # Update existing rating
        existing_rating.rating = rating
        existing_rating.feedback = feedback
        existing_rating.updated_at = datetime.utcnow()
    else:
        # Create new rating
        new_rating = EmployeePerformanceRatingDB(
            employee_id=employee_id,
            manager_id=current_user.id,
            rating=rating,
            feedback=feedback,
            week_start_date=week_start,
            week_end_date=week_end
        )
        db.add(new_rating)
    
    db.commit()
    
    return {"message": "Rating saved successfully"}

@app.get("/api/project/{project_id}/deliverables")
async def get_project_deliverables(request: Request, project_id: int, db: Session = Depends(get_db)):
    """Get all deliverables uploaded for a project."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Get the project
    project = db.query(ProjectDB).filter(ProjectDB.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Verify ownership - check if user owns the project via tender
    if project.source_tender_id:
        tender = db.query(TenderDB).filter(TenderDB.id == project.source_tender_id).first()
        if tender and tender.awarded_by != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized to view this project")
    
    deliverables = []
    
    # Get deliverables from project.documents JSONB field
    if project.documents and isinstance(project.documents, dict):
        task_deliverables = project.documents.get("Task Deliverables", [])
        
        for item in task_deliverables:
            if isinstance(item, dict):
                deliverables.append({
                    "filename": item.get("original_filename", "Unknown File"),
                    "uploader": item.get("uploaded_by", "Unknown"),
                    "uploaded_at": item.get("uploaded_at"),
                    "task": item.get("task_title", ""),
                    "file_size": item.get("file_size"),
                    "description": item.get("description", ""),
                    "download_url": f"/api/project/{project_id}/download/Task Deliverables/{task_deliverables.index(item)}"
                })
    
    # Sort by upload time (newest first)
    deliverables.sort(key=lambda x: x.get("uploaded_at") or "", reverse=True)
    
    return {
        "project_id": project_id,
        "project_name": project.project_name,
        "deliverables": deliverables,
        "total_count": len(deliverables)
    }

@app.post("/api/project/{project_id}/expert-request")
@require_company_details
async def create_expert_request(
    request: Request,
    project_id: int,
    db: Session = Depends(get_db)
):
    """Create an expert hiring request for a project."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Get the project
    project = db.query(ProjectDB).filter(ProjectDB.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Verify ownership
    if project.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to create requests for this project")
    
    # Get request data
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON data")
    
    request_name = data.get("request_name", "").strip()
    description = data.get("description", "").strip()
    budget_type = data.get("budget_type", "").lower()
    budget_amount = data.get("budget_amount")
    budget_min = data.get("budget_min")
    budget_max = data.get("budget_max")
    
    # Validation
    if not request_name:
        raise HTTPException(status_code=400, detail="Request name is required")
    if not description:
        raise HTTPException(status_code=400, detail="Description is required")
    if budget_type not in ["fixed", "negotiable"]:
        raise HTTPException(status_code=400, detail="Budget type must be 'fixed' or 'negotiable'")
    
    if budget_type == "fixed":
        if not budget_amount or budget_amount <= 0:
            raise HTTPException(status_code=400, detail="Budget amount is required for fixed budget")
        budget_min = None
        budget_max = None
    else:  # negotiable
        if not budget_min or budget_min <= 0:
            raise HTTPException(status_code=400, detail="Minimum budget is required for negotiable budget")
        if budget_max and budget_max < budget_min:
            raise HTTPException(status_code=400, detail="Maximum budget must be greater than minimum budget")
        budget_amount = None
    
    # Get company details
    company = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()
    if not company:
        raise HTTPException(status_code=400, detail="Company details not found. Please complete company profile.")
    
    company_name = company.company_name or "Unknown Company"
    # Safely get company location - CompanyDB has city, not location
    try:
        company_location = getattr(company, 'city', None) or getattr(company, 'operational_address', None) or None
        # Extract city from address if available
        if not company_location and hasattr(company, 'operational_address') and company.operational_address:
            # Try to extract city from address string
            address_parts = company.operational_address.split(',')
            if len(address_parts) > 0:
                company_location = address_parts[-1].strip()
    except Exception:
        company_location = None
    
    # Get tender details if project has source_tender_id
    tender_id = None
    tender_title = None
    tender_sector = None
    tender_state = None
    tender_location = None
    
    if project.source_tender_id:
        tender = db.query(TenderDB).filter(TenderDB.id == project.source_tender_id).first()
        if tender:
            tender_id = tender.id
            tender_title = tender.title
            tender_sector = tender.category
            tender_state = tender.state
            
            # Extract location from work_item_details JSON
            try:
                if tender.work_item_details and isinstance(tender.work_item_details, dict):
                    # Try different possible keys for location
                    tender_location = (
                        tender.work_item_details.get('Location') or
                        tender.work_item_details.get('location') or
                        tender.work_item_details.get('Work Location') or
                        tender.work_item_details.get('Project Location') or
                        None
                    )
                    # Clean up location string if found
                    if tender_location:
                        tender_location = str(tender_location).strip()
                        if not tender_location or tender_location.lower() in ['unknown', 'n/a', 'not specified', '']:
                            tender_location = None
            except Exception as e:
                logger.warning(f"Error extracting location from work_item_details: {e}")
                tender_location = None
    
    # Ensure schema supports tender_location to avoid ORM insert failures
    if not ensure_expert_hiring_tender_location_column():
        raise HTTPException(
            status_code=500,
            detail="Unable to create expert request because the database schema is missing the tender_location column. "
                   "Please run the migration script or try again after the system updates."
        )

    # Create expert hiring request
    request_kwargs = {
        "company_id": current_user.id,
        "project_id": project_id,
        "tender_id": tender_id,
        "request_name": request_name,
        "description": description,
        "budget_type": budget_type,
        "budget_amount": budget_amount,
        "budget_min": budget_min,
        "budget_max": budget_max,
        "company_name": company_name,
        "company_location": company_location,
        "tender_title": tender_title,
        "tender_sector": tender_sector,
        "tender_state": tender_state,
        "status": "open"
    }
    
    # Column exists now; safe to set tender_location if we have data
    request_kwargs["tender_location"] = tender_location
    
    hiring_request = ExpertHiringRequestDB(**request_kwargs)
    
    db.add(hiring_request)
    db.commit()
    db.refresh(hiring_request)
    
    return {
        "success": True,
        "message": "Expert hiring request created successfully",
        "request_id": hiring_request.id
    }

@app.get("/api/projects/sector-count")
@require_company_details
async def get_projects_sector_count(request: Request, sector: str, db: Session = Depends(get_db)):
    """Get count of projects for a specific sector."""
    current_user = get_current_user(request, db)
    if not current_user:
        return {"error": "User not found", "count": 0}

    count = db.query(ProjectDB).filter(
        ProjectDB.user_id == current_user.id,
        ProjectDB.sector == sector
    ).count()

    return {"count": count}

@app.get("/project/{project_id}", response_class=HTMLResponse)
@require_company_details
async def project_detail(
    request: Request, 
    project_id: int, 
    return_url: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Project detail page with document download options."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Check if this is a quarantined test session
    is_test_session = hasattr(current_user, '_is_test_session') and current_user._is_test_session
    is_quarantined = hasattr(current_user, '_quarantined') and current_user._quarantined
    test_redirect_url = None

    # Use return_url if provided, otherwise build default redirect URL
    if return_url:
        from urllib.parse import unquote
        decoded_return_url = unquote(return_url)
        back_url = decoded_return_url
    elif is_quarantined:
        # Build redirect URL back to test endpoint
        test_base_url = getattr(current_user, '_test_base_url', '/public-projects/nkbpl.pratyaksh')
        test_token = getattr(current_user, '_test_token', '')
        if test_token:
            back_url = f"{test_base_url}?test_token={test_token}"
        else:
            back_url = test_base_url
    else:
        back_url = "/projects"
    
    test_redirect_url = back_url

    # Get the project
    project = db.query(ProjectDB).filter(
        and_(ProjectDB.id == project_id, ProjectDB.user_id == current_user.id)
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Generate project_id if it doesn't exist (for legacy projects)
    if not project.project_id:
        try:
            project.project_id = generate_project_id(current_user.id, db)
            db.commit()
            db.refresh(project)
            logger.info(f"[PROJECT DETAIL] Generated project_id: {project.project_id} for project {project_id}")
        except Exception as e:
            logger.error(f"[PROJECT DETAIL] Error generating project_id: {e}")
            db.rollback()
            # Continue without project_id - not critical for viewing

    # Deserialize JSONB fields if they're strings
    import json

    # Fix services_rendered if it's a string
    if project.services_rendered and isinstance(project.services_rendered, str):
        try:
            project.services_rendered = json.loads(project.services_rendered)
        except:
            pass

    # Fix states if it's a string
    if project.states and isinstance(project.states, str):
        try:
            project.states = json.loads(project.states)
        except:
            pass

    # Fix cities if it's a string
    if project.cities and isinstance(project.cities, str):
        try:
            project.cities = json.loads(project.cities)
        except:
            pass

    # Fix documents if it's a string
    if project.documents and isinstance(project.documents, str):
        try:
            project.documents = json.loads(project.documents)
        except:
            pass

    # Calculate file sizes for all documents
    documents_with_sizes = {}
    if project.documents and isinstance(project.documents, dict):
        for doc_type, file_entries in project.documents.items():
            documents_with_sizes[doc_type] = []
            # Ensure file_entries is a list
            if not isinstance(file_entries, list):
                file_entries = [file_entries] if file_entries else []

            for file_entry in file_entries:
                # Handle two formats:
                # 1. Old format: simple string file path
                # 2. New format: dict with file_path, metadata (for task deliverables)
                if isinstance(file_entry, dict):
                    # New format - deliverable with metadata
                    file_path = file_entry.get("file_path", "")
                    original_filename = file_entry.get("original_filename", os.path.basename(file_path))
                    uploaded_by = file_entry.get("uploaded_by")
                    uploaded_at = file_entry.get("uploaded_at")
                    task_title = file_entry.get("task_title")
                    task_file_id = file_entry.get("task_file_id")
                    stored_size = file_entry.get("file_size")
                    description = file_entry.get("description")
                    
                    try:
                        # Get file size from disk or use stored size
                        if os.path.exists(file_path):
                            file_size = os.path.getsize(file_path)
                        else:
                            file_size = stored_size
                        
                        documents_with_sizes[doc_type].append({
                            'path': file_path,
                            'size': file_size,
                            'original_filename': original_filename,
                            'uploaded_by': uploaded_by,
                            'uploaded_at': uploaded_at,
                            'task_title': task_title,
                            'task_file_id': task_file_id,
                            'description': description,
                            'is_deliverable': True
                        })
                    except Exception as e:
                        logger.debug(f"Error processing deliverable metadata: {e}")
                        documents_with_sizes[doc_type].append({
                            'path': file_path,
                            'size': stored_size,
                            'original_filename': original_filename,
                            'uploaded_by': uploaded_by,
                            'uploaded_at': uploaded_at,
                            'task_title': task_title,
                            'description': description,
                            'is_deliverable': True
                        })
                else:
                    # Old format - simple file path string
                    file_path = file_entry
                    try:
                        # Get file size in bytes
                        file_size = os.path.getsize(file_path)
                        documents_with_sizes[doc_type].append({
                            'path': file_path,
                            'size': file_size,
                            'original_filename': os.path.basename(file_path),
                            'is_deliverable': False
                        })
                    except (OSError, IOError):
                        # Handle missing or inaccessible files
                        documents_with_sizes[doc_type].append({
                            'path': file_path,
                            'size': None,  # Will display as 'Unknown size'
                            'original_filename': os.path.basename(file_path),
                            'is_deliverable': False
                        })

    # Debug logging for description and scope
    logger.info(f"Project {project.id} - Description: {bool(project.project_description)} ({len(project.project_description) if project.project_description else 0} chars)")
    logger.info(f"Project {project.id} - Scope: {bool(project.complete_scope_of_work)} ({len(project.complete_scope_of_work) if project.complete_scope_of_work else 0} chars)")

    return templates.TemplateResponse("project_detail.html", {
        "request": request,
        "current_user": current_user,
        "project": project,
        "documents_with_sizes": documents_with_sizes,
        "is_test_session": is_test_session,
        "is_quarantined": is_quarantined,
        "test_redirect_url": test_redirect_url,
        "selected_font": get_active_font()
    })

@app.post("/api/project/{project_id}/mark-finished")
async def mark_project_finished(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Mark a project as finished by setting end_date to today."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get the project
    project = db.query(ProjectDB).filter(
        and_(ProjectDB.id == project_id, ProjectDB.user_id == current_user.id)
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Check if project is already finished
    if project.end_date:
        raise HTTPException(status_code=400, detail="Project is already marked as finished")

    # Set end_date to today
    project.end_date = datetime.utcnow().date()  # type: ignore

    # Recalculate duration if start_date exists
    if project.start_date:
        # Convert both dates to date objects to ensure compatibility
        start = project.start_date.date() if isinstance(project.start_date, datetime) else project.start_date
        end = project.end_date.date() if isinstance(project.end_date, datetime) else project.end_date
        duration = (end - start).days
        project.project_duration_months = round(duration / 30)  # type: ignore

    db.commit()
    db.refresh(project)

    return {
        "message": "Project marked as finished",
        "end_date": project.end_date.strftime('%d %b %Y') if project.end_date else None,
        "duration_months": project.project_duration_months
    }

@app.get("/api/project/{project_id}/download/{doc_type}/{file_index}")
async def download_project_document(
    project_id: int,
    doc_type: str,
    file_index: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Download a specific document from a project."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get the project
    project = db.query(ProjectDB).filter(
        and_(ProjectDB.id == project_id, ProjectDB.user_id == current_user.id)
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Check if document type exists and has files
    if not project.documents or doc_type not in project.documents:
        raise HTTPException(status_code=404, detail="Document type not found")

    files = project.documents[doc_type]
    if file_index >= len(files):
        raise HTTPException(status_code=404, detail="File not found")

    file_entry = files[file_index]
    
    # Handle both old format (string) and new format (dict with metadata)
    if isinstance(file_entry, dict):
        # New format - deliverable with metadata
        file_path = file_entry.get("file_path", "")
        filename = file_entry.get("original_filename", os.path.basename(file_path))
    else:
        # Old format - simple file path string
        file_path = file_entry
        filename = os.path.basename(file_path)

    # Check if file exists
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found on server")

    # Return file
    from fastapi.responses import FileResponse

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type='application/octet-stream'
    )

@app.get("/api/project/{project_id}/download/pdf")
async def download_project_pdf(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Generate and download a comprehensive PDF of the project with all information and documents."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get the project
    project = db.query(ProjectDB).filter(
        and_(ProjectDB.id == project_id, ProjectDB.user_id == current_user.id)
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Generate comprehensive PDF
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        import io
        import os
        from pathlib import Path
        from PyPDF2 import PdfReader
        from pdf2image import convert_from_path
        from PIL import Image as PILImage
        import tempfile
        import html

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.75*inch, bottomMargin=0.75*inch)
        styles = getSampleStyleSheet()
        story = []

        # Helper function to escape HTML and convert to plain text for safe rendering
        def safe_paragraph(text, style=styles['Normal']):
            if not text:
                return Paragraph("", style)
            # Escape HTML entities and convert newlines
            text = html.escape(str(text))
            text = text.replace('\n', '<br/>')
            return Paragraph(text, style)

        # Title Page
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=28,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1e293b')
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=14,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#64748b')
        )
        
        story.append(Spacer(1, 2*inch))
        story.append(Paragraph(project.project_name, title_style))
        story.append(Paragraph("Project Comprehensive Documentation", subtitle_style))
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(f"Generated on {datetime.utcnow().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(PageBreak())

        # Table of Contents
        toc_style = ParagraphStyle(
            'TOCTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        story.append(Paragraph("Table of Contents", toc_style))
        story.append(Spacer(1, 20))
        
        # Build TOC items
        toc_items = ["1. Project Metadata"]
        if project.consultancy_fee or project.project_cost:
            toc_items.append("   1.1 Basic Information")
            toc_items.append("   1.2 Commercial Information")
        else:
            toc_items.append("   1.1 Basic Information")
        toc_items.append("   1.3 Timeline & Geography")
        if project.project_description:
            toc_items.append("2. Project Description")
        if project.complete_scope_of_work:
            toc_items.append("3. Complete Scope of Work")
        if project.services_rendered:
            toc_items.append("4. Services Delivered")
        if project.documents:
            toc_items.append("5. Supporting Documents")
        
        for item in toc_items:
            story.append(Paragraph(item, styles['Normal']))
        
        story.append(PageBreak())
        
        # 1. Project Metadata Section
        story.append(Paragraph("1. Project Metadata", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        # Basic Information
        story.append(Paragraph("1.1 Basic Information", styles['Heading2']))
        story.append(Spacer(1, 12))
        basic_data = []
        basic_data.append(["Project Name:", project.project_name])
        if project.client_name:
            basic_data.append(["Client:", project.client_name])
        if project.sector:
            sector_info = project.sector
            if project.sub_sector:
                sector_info += f" - {project.sub_sector}"
            basic_data.append(["Sector:", sector_info])
        basic_data.append(["Financing Authority:", project.financing_authority or "Financing Not Required"])
        if project.country:
            basic_data.append(["Country:", project.country])
        if project.jv_partner:
            basic_data.append(["JV Partner:", project.jv_partner])

        if basic_data:
            table = Table(basic_data, colWidths=[2.5*inch, 4.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
            ]))
            story.append(table)
            story.append(Spacer(1, 12))

        # Commercial Information
        if project.consultancy_fee or project.project_cost:
            story.append(Paragraph("1.2 Commercial Information", styles['Heading2']))
            story.append(Spacer(1, 12))
            commercial_data = []
            if project.consultancy_fee:
                commercial_data.append(["Consultancy Fee:", f"₹{project.consultancy_fee:,.0f}"])
            if project.project_cost:
                commercial_data.append(["Project Cost:", f"₹{project.project_cost:,.0f}"])

            if commercial_data:
                table = Table(commercial_data, colWidths=[2.5*inch, 4.5*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0fdf4')),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
                ]))
                story.append(table)
                story.append(Spacer(1, 12))

        # Timeline & Geography
        story.append(Paragraph("1.3 Timeline & Geography", styles['Heading2']))
        story.append(Spacer(1, 12))
        timeline_geo_data = []
        if project.start_date:
            timeline_geo_data.append(["Start Date:", project.start_date.strftime('%B %d, %Y')])
        timeline_geo_data.append(["Completion Date:", 
                                  project.end_date.strftime('%B %d, %Y') if project.end_date else "Ongoing"])
        if project.project_duration_months:
            timeline_geo_data.append(["Duration:", f"{project.project_duration_months} months"])
        if project.states:
            states_str = ", ".join(project.states) if isinstance(project.states, list) else str(project.states)
            timeline_geo_data.append(["States:", states_str])
        if project.cities:
            cities_str = ", ".join(project.cities) if isinstance(project.cities, list) else str(project.cities)
            timeline_geo_data.append(["Cities:", cities_str])

        if timeline_geo_data:
            table = Table(timeline_geo_data, colWidths=[2.5*inch, 4.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fefce8')),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
                ]))
            story.append(table)
            story.append(Spacer(1, 12))

        story.append(PageBreak())

        # 2. Project Description
        if project.project_description:
            story.append(Paragraph("2. Project Description", styles['Heading1']))
            story.append(Spacer(1, 12))
            story.append(safe_paragraph(project.project_description, styles['Normal']))
            story.append(Spacer(1, 12))
            story.append(PageBreak())

        # 3. Complete Scope of Work
        if project.complete_scope_of_work:
            story.append(Paragraph("3. Complete Scope of Work", styles['Heading1']))
            story.append(Spacer(1, 12))
            story.append(safe_paragraph(project.complete_scope_of_work, styles['Normal']))
            story.append(Spacer(1, 12))
            story.append(PageBreak())

        # 4. Services Rendered
        if project.services_rendered:
            story.append(Paragraph("4. Services Delivered", styles['Heading1']))
            story.append(Spacer(1, 12))
            service_labels = {
                'survey_investigations': 'Survey & Investigations',
                'design_engineering': 'Design & Engineering',
                'dpr_feasibility': 'DPR & Feasibility Studies',
                'gis_data': 'GIS & Data Services',
                'pmc': 'Project Management Consultancy (PMC)',
                'pmu': 'Project Management Unit (PMU)',
                'advisory_capacity': 'Advisory & Capacity Building',
                'supervision': 'Construction Supervision',
                'quality_control': 'Quality Control & Assurance',
                'environmental': 'Environmental & Social Impact',
                'financial_advisory': 'Financial Advisory Services'
            }
            
            if isinstance(project.services_rendered, dict):
                for service_key, service_value in project.services_rendered.items():
                    if service_value:
                            service_name = service_labels.get(service_key, service_key.replace('_', ' ').title())
                            story.append(Paragraph(f"<b>{service_name}</b>", styles['Heading3']))
                            story.append(safe_paragraph(service_value, styles['Normal']))
                    story.append(Spacer(1, 12))
            else:
                story.append(safe_paragraph(str(project.services_rendered), styles['Normal']))

            story.append(PageBreak())

        # 5. Supporting Documents
        doc_type_labels = {
            'Task Deliverables': 'Task Deliverables (Employee Uploads)',
                'tender_documents': 'Tender Documents',
                'technical_proposal': 'Technical Proposal',
                'financial_proposal': 'Financial Proposal',
                'work_order': 'Work Order',
                'deliverables': 'Deliverables',
                'completion_certificate': 'Completion Certificate',
                'invoices_receipts': 'Invoices & Receipts',
                'other_documents': 'Other Documents'
            }

        if project.documents:
            story.append(Paragraph("5. Supporting Documents", styles['Heading1']))
            story.append(Spacer(1, 12))
            
            # Create temporary directory for image conversion
            with tempfile.TemporaryDirectory() as temp_dir:
                # Process all document types in project.documents
                section_num = 1
                processed_doc_types = []
                
                # First handle Task Deliverables if they exist
                if 'Task Deliverables' in project.documents:
                    display_name = 'Task Deliverables (Employee Uploads)'
                    file_paths = project.documents['Task Deliverables']
                    if file_paths and len(file_paths) > 0:
                        processed_doc_types.append(('Task Deliverables', display_name, file_paths))
                
                # Then handle other document types
                for doc_type, display_name in doc_type_labels.items():
                    if doc_type in project.documents:
                        file_paths = project.documents[doc_type]
                        if file_paths and len(file_paths) > 0:
                            processed_doc_types.append((doc_type, display_name, file_paths))
                
                # Process each document type
                for doc_type_key, display_name, file_paths in processed_doc_types:
                    
                    story.append(Paragraph(f"5.{section_num} {display_name}", styles['Heading2']))
                    story.append(Spacer(1, 12))
                    section_num += 1
                    
                    for idx, file_path in enumerate(file_paths):
                        if not file_path or not os.path.exists(file_path):
                            continue
                        
                        filename = os.path.basename(file_path)
                        file_ext = os.path.splitext(filename)[1].lower()
                        
                        story.append(Paragraph(f"<b>Document {idx + 1}: {filename}</b>", styles['Heading3']))
                        story.append(Spacer(1, 6))

                        try:
                            if file_ext == '.pdf':
                                # Convert PDF to images and embed
                                try:
                                    images = convert_from_path(file_path, dpi=150)
                                    for page_num, img in enumerate(images):
                                        # Save image temporarily
                                        img_path = os.path.join(temp_dir, f"{filename}_page_{page_num}.png")
                                        img.save(img_path, 'PNG')
                                        
                                        # Get image dimensions
                                        img_width, img_height = img.size
                                        # Scale to fit page width (A4 width - margins)
                                        available_width = A4[0] - 1.5*inch
                                        available_height = A4[1] - 1.5*inch
                                        
                                        scale_w = available_width / img_width
                                        scale_h = available_height / img_height
                                        scale = min(scale_w, scale_h, 1.0)  # Don't scale up
                                        
                                        img_width_pdf = img_width * scale
                                        img_height_pdf = img_height * scale
                                        
                                        story.append(Image(img_path, width=img_width_pdf, height=img_height_pdf))
                                        if page_num < len(images) - 1:
                                            story.append(Spacer(1, 6))
                                    
                                    story.append(Spacer(1, 12))
                                except Exception as e:
                                    logger.warning(f"Could not convert PDF {filename} to images: {e}")
                                    story.append(Paragraph(f"<i>Note: PDF document could not be embedded. File: {filename}</i>", styles['Normal']))
                                    story.append(Spacer(1, 12))
                            
                            elif file_ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp']:
                                # Embed image directly
                                try:
                                    img = PILImage.open(file_path)
                                    img_width, img_height = img.size
                                    
                                    # Scale to fit page
                                    available_width = A4[0] - 1.5*inch
                                    available_height = A4[1] - 1.5*inch
                                    
                                    scale_w = available_width / img_width
                                    scale_h = available_height / img_height
                                    scale = min(scale_w, scale_h, 1.0)
                                    
                                    img_width_pdf = img_width * scale
                                    img_height_pdf = img_height * scale
                                    
                                    story.append(Image(file_path, width=img_width_pdf, height=img_height_pdf))
                                    story.append(Spacer(1, 12))
                                except Exception as e:
                                    logger.warning(f"Could not embed image {filename}: {e}")
                                    story.append(Paragraph(f"<i>Note: Image could not be embedded. File: {filename}</i>", styles['Normal']))
                                    story.append(Spacer(1, 12))
                            
                            else:
                                # Other file types - just list
                                story.append(Paragraph(f"<i>File type not supported for embedding: {filename} ({file_ext})</i>", styles['Normal']))
                                story.append(Spacer(1, 12))
                        
                        except Exception as e:
                            logger.error(f"Error processing document {filename}: {e}")
                            story.append(Paragraph(f"<i>Error processing document: {filename}</i>", styles['Normal']))
                            story.append(Spacer(1, 12))
                    
                    story.append(PageBreak())

        # Build PDF
        doc.build(story)
        buffer.seek(0)

        # Return comprehensive PDF
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            io.BytesIO(buffer.getvalue()),
            media_type='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{project.project_name.replace(" ", "_")}_comprehensive.pdf"'}
        )

    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"PDF generation library not available: {str(e)}")
    except Exception as e:
        logger.error(f"Error generating comprehensive PDF: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@app.get("/employee/dashboard", response_class=HTMLResponse)
async def employee_dashboard(request: Request, db: Session = Depends(get_db)):
    """Employee dashboard with assigned tenders."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        return RedirectResponse(url="/employee/login", status_code=302)

    # Block BD employees from accessing regular employee dashboard
    if current_employee.is_bd:
        return RedirectResponse(url="/bd/home", status_code=302)

    # Get employee's assignments (using eager loading to prevent N+1 queries)
    from sqlalchemy.orm import joinedload

    assignments = db.query(TenderAssignmentDB).options(
        joinedload(TenderAssignmentDB.tender),  # Eagerly load tender
        joinedload(TenderAssignmentDB.tasks)     # Eagerly load tasks
    ).filter(
        TenderAssignmentDB.employee_id == current_employee.id
    ).all()

    # Calculate statistics
    active_tenders = 0
    pending_tasks = 0
    overdue_tasks = 0
    due_today = 0
    now = datetime.utcnow()

    tender_summaries = []
    for assignment in assignments:
        if not assignment.tender:
            continue

        # Skip expired tenders (unless awarded)
        if assignment.tender.deadline and assignment.tender.deadline < now and not assignment.tender.awarded:
            continue

        # Check if tender is active
        if assignment.tender.deadline is not None and assignment.tender.deadline > now:
            active_tenders += 1

        # Use pre-loaded tasks (no additional query)
        tasks = assignment.tasks
        completed_tasks = len([t for t in tasks if t.status == 'completed']) # type: ignore
        total_tasks = len(tasks)
        pending_tasks += total_tasks - completed_tasks

        # Check for overdue tasks
        for task in tasks:
            if task.deadline is not None and task.status != 'completed':  # type: ignore
                if task.deadline < now:  # type: ignore
                    overdue_tasks += 1
                elif task.deadline.date() == now.date():  # type: ignore
                    due_today += 1

        # Find nearest deadline
        nearest_deadline = None
        if tasks:
            active_tasks = [t for t in tasks if t.status != 'completed' and t.deadline is not None]  # type: ignore
            if active_tasks:
                deadlines = [t.deadline for t in active_tasks if t.deadline is not None]
                if deadlines:
                    nearest_deadline = min(deadlines)  # type: ignore

        tender_summaries.append({
            'id': assignment.tender.id,
            'title': assignment.tender.title,
            'authority': assignment.tender.authority,
            'reference_id': assignment.tender.tender_id,
            'tasks_left': total_tasks - completed_tasks,
            'total_tasks': total_tasks,
            'nearest_deadline': nearest_deadline,
            'priority': assignment.priority,
            'role': assignment.role
        })

    return templates.TemplateResponse("employee_dashboard.html", {
        "request": request,
        "current_employee": current_employee,
        "tender_summaries": tender_summaries,
        "active_tenders": active_tenders,
        "pending_tasks": pending_tasks,
        "overdue_tasks": overdue_tasks,
        "due_today": due_today,
        "now": datetime.utcnow(),
        "selected_font": get_active_font()
    })

@app.get("/employee/tender/{tender_id}", response_class=HTMLResponse)
async def employee_tender_detail(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Employee tender detail page with tasks, deliverables, and chat."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        return RedirectResponse(url="/employee/login", status_code=302)

    # Get assignment for this employee and tender
    assignment = db.query(TenderAssignmentDB).filter(
        and_(TenderAssignmentDB.tender_id == tender_id, TenderAssignmentDB.employee_id == current_employee.id)
    ).first()

    if not assignment or not assignment.tender:
        raise HTTPException(status_code=404, detail="Tender assignment not found")

    # Get tasks for this assignment (only parent tasks, load subtasks separately)
    parent_tasks = db.query(TaskDB).filter(
        and_(TaskDB.assignment_id == assignment.id, TaskDB.is_subtask == False)
    ).order_by(TaskDB.created_at.desc()).all()

    # Load subtasks for each parent task
    tasks = []
    for task in parent_tasks:
        task.subtasks_list = db.query(TaskDB).filter(
            TaskDB.parent_task_id == task.id
        ).order_by(TaskDB.subtask_order).all()
        tasks.append(task)

    # Get task comments, files, and concerns for each task
    task_comment_payloads = {}
    task_comment_counts = {}
    task_files_map = {}
    task_concerns_map = {}
    
    all_task_ids = [t.id for t in tasks]
    for subtask_list in [t.subtasks_list for t in tasks]:
        all_task_ids.extend([st.id for st in subtask_list])
    
    for task_id in all_task_ids:
        task_obj = db.query(TaskDB).filter(TaskDB.id == task_id).first()
        if not task_obj:
            continue
            
        # Comments
        comment_payload = []
        if task_obj.comments:
            for comment in task_obj.comments:
                author_name = None
                if comment.employee:
                    author_name = comment.employee.name
                comment_payload.append({
                    "id": comment.id,
                    "author": author_name,
                    "body": comment.comment,
                    "timestamp": comment.created_at.isoformat() if comment.created_at else None
                })
        task_comment_payloads[task_id] = comment_payload
        task_comment_counts[task_id] = len(comment_payload)
        
        # Files (deliverables)
        files_list = []
        if task_obj.files:
            for file_obj in task_obj.files:
                files_list.append({
                    "id": file_obj.id,
                    "filename": file_obj.filename,
                    "mime_type": file_obj.mime_type,
                    "file_size": file_obj.file_size,
                    "description": file_obj.description,
                    "created_at": file_obj.created_at.isoformat() if file_obj.created_at else None,
                    "uploaded_by": file_obj.employee.name if file_obj.employee else "Unknown"
                })
        task_files_map[task_id] = files_list
        
        # Concerns (only for employee's own tasks)
        concerns_list = []
        if task_obj.concerns:
            for concern in task_obj.concerns:
                concerns_list.append({
                    "id": concern.id,
                    "concern_type": concern.concern_type,
                    "title": concern.title,
                    "description": concern.description,
                    "status": concern.status,
                    "priority": concern.priority,
                    "created_at": concern.created_at.isoformat() if concern.created_at else None,
                    "resolved_at": concern.resolved_at.isoformat() if concern.resolved_at else None,
                    "resolution_notes": concern.resolution_notes
                })
        task_concerns_map[task_id] = concerns_list

    # Create progress map for all tasks
    task_progress_map = {}
    for task_id in all_task_ids:
        task_obj = db.query(TaskDB).filter(TaskDB.id == task_id).first()
        if not task_obj:
            continue

        progress_list = []
        if task_obj.progress_updates:
            for progress in task_obj.progress_updates:
                progress_list.append({
                    "id": progress.id,
                    "employee_id": progress.employee_id,
                    "employee_name": progress.employee.name if progress.employee else "Unknown",
                    "employee_avatar": progress.employee.name[0].upper() if progress.employee and progress.employee.name else "?",
                    "update_text": progress.update_text,
                    "created_at": progress.created_at.isoformat() if progress.created_at else None,
                    "formatted_date": progress.created_at.strftime('%d %b %Y, %I:%M %p') if progress.created_at else None,
                    "is_edited": progress.is_edited,
                    "is_current_user": progress.employee_id == current_employee.id
                })
        task_progress_map[task_id] = progress_list

    # Get team members (other employees assigned to this tender)
    team_assignments = db.query(TenderAssignmentDB).filter(
        and_(TenderAssignmentDB.tender_id == tender_id, TenderAssignmentDB.employee_id != current_employee.id)
    ).all()
    team_members = [ta.employee for ta in team_assignments if ta.employee]

    # Get manager (user who assigned the tender)
    manager = None
    if assignment.assigned_by: # type: ignore
        manager = db.query(UserDB).filter(UserDB.id == assignment.assigned_by).first()

    # Get messages for this assignment with robust ordering
    messages = db.query(TenderMessageDB).filter(
        TenderMessageDB.assignment_id == assignment.id
    ).order_by(TenderMessageDB.created_at.asc(), TenderMessageDB.id.asc()).all()

    # Calculate statistics
    total_tasks = len([t for t in tasks if not t.is_subtask])
    completed_tasks = len([t for t in tasks if t.status == 'completed' and not t.is_subtask])
    pending_tasks = total_tasks - completed_tasks
    overdue_tasks = 0
    due_today = 0
    now = datetime.utcnow()
    
    for task in tasks:
        if task.deadline and task.status != 'completed':
            if task.deadline < now:
                overdue_tasks += 1
            elif task.deadline.date() == now.date():
                due_today += 1

    return templates.TemplateResponse("employee_task_detail.html", {
        "request": request,
        "current_employee": current_employee,
        "assignment": assignment,
        "tender": assignment.tender,
        "tasks": tasks,
        "manager": manager,
        "team_members": team_members,
        "messages": messages,
        "task_comment_payloads": task_comment_payloads,
        "task_comment_counts": task_comment_counts,
        "task_files_map": task_files_map,
        "task_concerns_map": task_concerns_map,
        "task_progress_map": task_progress_map,
        "stats": {
            "total_tasks": total_tasks,
            "completed_tasks": completed_tasks,
            "pending_tasks": pending_tasks,
            "overdue_tasks": overdue_tasks,
            "due_today": due_today
        },
        "now": datetime.utcnow(),
        "selected_font": get_active_font()
    })

@app.get("/tender-management", response_class=HTMLResponse)
@enforce_test_quarantine
@require_pin_verification
async def tender_management_page(request: Request, db: Session = Depends(get_db)):
    """Tender management page showing favorites, shortlisted, and awarded tenders."""
    from core.dependencies import get_id_for_tender_management, get_all_bd_employee_ids_for_company
    from sqlalchemy.orm import joinedload

    # Get authenticated entity (user or BD employee)
    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        return RedirectResponse(url="/login", status_code=302)

    # Set context variables for templates
    current_user = entity if entity_type == 'user' else None
    current_bd_employee = entity if entity_type == 'bd_employee' else None
    is_admin = entity_type == 'user'

    # Clean up expired tenders and orphaned records before displaying
    try:
        check_and_remove_expired_tenders(db)
        enforce_awarded_tender_state(db)
        cleanup_orphaned_records(db)
    except Exception as e:
        logger.warning(f"Cleanup of expired tenders and orphaned records failed: {e}")

    # Determine IDs to query based on user type (for shortlisted and awarded)
    if entity_type == 'user':
        # Admin: Get all BD employee IDs + their own ID
        bd_employee_ids = get_all_bd_employee_ids_for_company(entity_id, db)
        all_ids = [entity_id] + bd_employee_ids
    else:
        # BD Employee: Only their own ID
        all_ids = [entity_id]

    # Get user_id for database queries (company owner's user_id for BD employees)
    from core.dependencies import get_user_id_for_queries
    user_id_for_query, _ = get_user_id_for_queries(request, db)
    if not user_id_for_query:
        return RedirectResponse(url="/login", status_code=302)

    now = datetime.utcnow()

    # Get favorite tenders with tender details
    # For admins: Get all favorites (admin's own + BD employees')
    # For BD employees: Get only their own favorites (filtered by worked_by fields)
    if entity_type == 'user':
        # Admin: Get all favorites for this user_id (includes both admin's and BD employees')
        all_favorites = db.query(FavoriteDB).options(
            joinedload(FavoriteDB.tender)
        ).filter(
            FavoriteDB.user_id == user_id_for_query
        ).order_by(desc(FavoriteDB.created_at)).all()
    else:
        # BD Employee: Get only their own favorites
        all_favorites = db.query(FavoriteDB).options(
            joinedload(FavoriteDB.tender)
        ).filter(
            and_(
                FavoriteDB.user_id == user_id_for_query,
                FavoriteDB.worked_by_type == 'bd_employee',
                FavoriteDB.worked_by_name == entity.name
            )
        ).order_by(desc(FavoriteDB.created_at)).all()

    # Filter out expired tenders (unless awarded)
    favorites = []
    for fav in all_favorites:
        if fav.tender:
            if fav.tender.awarded:
                continue
            # Skip expired tenders (unless awarded)
            if fav.tender.deadline and fav.tender.deadline < now and not fav.tender.awarded:
                continue
            _prepare_tender_display_fields(fav.tender)
            favorites.append(fav)

    # Get shortlisted tenders with tender details
    all_shortlisted = db.query(ShortlistedTenderDB).options(
        joinedload(ShortlistedTenderDB.tender)
    ).filter(
        ShortlistedTenderDB.user_id.in_(all_ids)
    ).order_by(desc(ShortlistedTenderDB.created_at)).all()

    # Get corresponding favorites for shortlisted tenders
    shortlisted_with_favorites = []
    for shortlisted in all_shortlisted:
        if shortlisted.tender and shortlisted.tender.awarded:
            continue
        # Skip expired tenders (unless awarded)
        if shortlisted.tender and shortlisted.tender.deadline and shortlisted.tender.deadline < now and not shortlisted.tender.awarded:
            continue
        if shortlisted.tender:
            _prepare_tender_display_fields(shortlisted.tender)

        # Try to find the corresponding favorite
        # Note: shortlisted.user_id might be a BD employee ID, but favorites are stored with company owner's user_id
        # So we need to check by user_id_for_query and worked_by fields if it's a BD employee
        if entity_type == 'user' and shortlisted.user_id != user_id_for_query:
            # This is a BD employee's shortlist (admin viewing), find favorite by worked_by fields
            # We need to get the BD employee's name from their ID
            from database import EmployeeDB
            bd_employee = db.query(EmployeeDB).filter(EmployeeDB.id == shortlisted.user_id).first()
            if bd_employee:
                favorite = db.query(FavoriteDB).filter(
                    and_(
                        FavoriteDB.user_id == user_id_for_query,
                        FavoriteDB.tender_id == shortlisted.tender_id,
                        FavoriteDB.worked_by_type == 'bd_employee',
                        FavoriteDB.worked_by_name == bd_employee.name
                    )
                ).first()
            else:
                favorite = None
        elif entity_type == 'bd_employee':
            # BD employee viewing their own shortlist, find their favorite
            favorite = db.query(FavoriteDB).filter(
                and_(
                    FavoriteDB.user_id == user_id_for_query,
                    FavoriteDB.tender_id == shortlisted.tender_id,
                    FavoriteDB.worked_by_type == 'bd_employee',
                    FavoriteDB.worked_by_name == entity.name
                )
            ).first()
        else:
            # Admin viewing their own shortlist
            favorite = db.query(FavoriteDB).filter(
                and_(
                    FavoriteDB.user_id == user_id_for_query,
                    FavoriteDB.tender_id == shortlisted.tender_id
                )
            ).first()
        shortlisted_with_favorites.append({
            'shortlisted': shortlisted,
            'favorite': favorite
        })

    # Get awarded tenders
    awarded_tenders = db.query(TenderDB).filter(
        and_(TenderDB.awarded == True, TenderDB.awarded_by.in_(all_ids))
    ).order_by(desc(TenderDB.published_at)).all()

    # Create awarded tender objects with additional metadata
    awarded_list = []
    for tender in awarded_tenders:
        _prepare_tender_display_fields(tender)
        awarded_list.append({
            'tender': tender,
            'awarded_date': tender.published_at,
            'notes': None
        })

    # Get counts
    favorite_count = len(favorites)
    shortlisted_count = len(shortlisted_with_favorites)
    awarded_count = len(awarded_tenders)

    return templates.TemplateResponse("tender_management.html", {
        "request": request,
        "current_user": current_user,
        "current_employee": current_bd_employee,
        "is_admin": is_admin,  # NEW: Flag for conditional template rendering
        "favorites": favorites,
        "shortlisted_tenders": shortlisted_with_favorites,
        "awarded_tenders": awarded_list,
        "favorite_count": favorite_count,
        "shortlisted_count": shortlisted_count,
        "awarded_count": awarded_count,
        "selected_font": get_active_font()
    })

@app.get("/shortlisted-tenders", response_class=HTMLResponse)
@enforce_test_quarantine
async def shortlisted_tenders_page(request: Request, db: Session = Depends(get_db)):
    """Shortlisted tenders page showing all shortlisted tenders."""
    from core.dependencies import get_id_for_tender_management, get_all_bd_employee_ids_for_company
    from sqlalchemy.orm import joinedload

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        return RedirectResponse(url="/login", status_code=302)

    current_user = entity if entity_type == 'user' else None
    current_bd_employee = entity if entity_type == 'bd_employee' else None
    is_admin = entity_type == 'user'

    # Clean up expired tenders before displaying
    try:
        check_and_remove_expired_tenders(db)
        enforce_awarded_tender_state(db)
        cleanup_orphaned_records(db)
    except Exception as e:
        logger.warning(f"Cleanup of expired tenders and orphaned records failed: {e}")

    # Determine IDs to query
    if entity_type == 'user':
        bd_employee_ids = get_all_bd_employee_ids_for_company(entity_id, db)
        all_ids = [entity_id] + bd_employee_ids
    else:
        all_ids = [entity_id]

    shortlisted_tenders = db.query(ShortlistedTenderDB).options(
        joinedload(ShortlistedTenderDB.tender)
    ).filter(
        ShortlistedTenderDB.user_id.in_(all_ids)
    ).order_by(desc(ShortlistedTenderDB.created_at)).all()

    # Get corresponding favorites
    now = datetime.utcnow()
    shortlisted_with_favorites = []
    for shortlisted in shortlisted_tenders:
        if shortlisted.tender and shortlisted.tender.awarded:
            continue
        if shortlisted.tender and shortlisted.tender.deadline and shortlisted.tender.deadline < now and not shortlisted.tender.awarded:
            continue
        if shortlisted.tender:
            _prepare_tender_display_fields(shortlisted.tender)

        favorite = db.query(FavoriteDB).filter(
            and_(
                FavoriteDB.user_id == shortlisted.user_id,
                FavoriteDB.tender_id == shortlisted.tender_id
            )
        ).first()
        shortlisted_with_favorites.append({
            'shortlisted': shortlisted,
            'favorite': favorite
        })

    shortlisted_count = len(shortlisted_with_favorites)

    return templates.TemplateResponse("shortlisted_tenders.html", {
        "request": request,
        "current_user": current_user,
        "current_employee": current_bd_employee,
        "is_admin": is_admin,
        "shortlisted_tenders": shortlisted_with_favorites,
        "shortlisted_count": shortlisted_count,
        "selected_font": get_active_font()
    })

@app.get("/awarded-tenders", response_class=HTMLResponse)
@enforce_test_quarantine
async def awarded_tenders_page(request: Request, db: Session = Depends(get_db)):
    """Awarded tenders page showing all tenders won by the user or BD employee."""
    from core.dependencies import get_id_for_tender_management, get_all_bd_employee_ids_for_company

    # Get authenticated entity (user or BD employee)
    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        return RedirectResponse(url="/login", status_code=302)

    # Set context variables for templates
    current_user = entity if entity_type == 'user' else None
    current_bd_employee = entity if entity_type == 'bd_employee' else None
    is_admin = entity_type == 'user'

    # Determine IDs to query based on user type
    if entity_type == 'user':
        # Admin: Get all BD employee IDs + their own ID
        bd_employee_ids = get_all_bd_employee_ids_for_company(entity_id, db)
        all_ids = [entity_id] + bd_employee_ids
    else:
        # BD Employee: Only their own ID
        all_ids = [entity_id]

    # Get awarded tenders with admin aggregation
    awarded_tenders = db.query(TenderDB).filter(
        and_(TenderDB.awarded == True, TenderDB.awarded_by.in_(all_ids))
    ).order_by(desc(TenderDB.published_at)).all()

    # Create awarded tender objects with additional metadata
    awarded_list = []
    for tender in awarded_tenders:
        _prepare_tender_display_fields(tender)
        awarded_list.append({
            'tender': tender,
            'awarded_date': tender.published_at,  # You can add a specific awarded_date field later
            'notes': None  # Placeholder for future notes functionality
        })

    return templates.TemplateResponse("awarded_tender.html", {
        "request": request,
        "current_user": current_user,
        "current_employee": current_bd_employee,
        "is_admin": is_admin,
        "awarded_tenders": awarded_list,
        "selected_font": get_active_font()
    })

@app.get("/employee/task-assignment", response_class=HTMLResponse)
async def employee_task_assignment_page(request: Request, db: Session = Depends(get_db)):
    """Employee task assignment page for managers."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Get all tenders (for assignment)
    tenders = db.query(TenderDB).filter(TenderDB.awarded == True).order_by(desc(TenderDB.published_at)).limit(50).all()

    # Get user's employees
    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
    employees = []
    for code in company_codes:
        employees.extend(db.query(EmployeeDB).filter(EmployeeDB.company_code_id == code.id).all())

    return templates.TemplateResponse("work_assignment.html", {
        "request": request,
        "current_user": current_user,
        "tenders": tenders,
        "employees": employees,
        "selected_font": get_active_font()
    })

@app.get("/team/{tender_id}", response_class=HTMLResponse)
@require_company_details
async def team_management_page(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Team management page for a specific tender."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Get the tender
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    _prepare_tender_display_fields(tender)

    # Check if tender is awarded and belongs to user's company
    if tender.awarded != True:
        raise HTTPException(status_code=403, detail="Tender is not awarded")

    # Get user's company codes to filter assignments
    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
    company_code_ids = [code.id for code in company_codes]

    # Get assignments for this tender where employees belong to user's company
    assignments = db.query(TenderAssignmentDB).join(EmployeeDB).filter(
        and_(TenderAssignmentDB.tender_id == tender_id, EmployeeDB.company_code_id.in_(company_code_ids))
    ).all()

    # Format assignments with tasks
    formatted_assignments = []
    for assignment in assignments:
        if not assignment.employee:
            continue

        # Get tasks for this assignment
        tasks = db.query(TaskDB).filter(TaskDB.assignment_id == assignment.id).all()

        formatted_assignments.append({
            'id': assignment.id,
            'employee': assignment.employee,
            'role': assignment.role,
            'priority': assignment.priority,
            'tasks': tasks
        })

    # Get comments for this tender (from all assignments for this tender)
    assignment_ids = [a.id for a in assignments]
    comments = []
    if assignment_ids:
        comments = db.query(TaskCommentDB).join(TaskDB).filter(
            TaskDB.assignment_id.in_(assignment_ids)
        ).order_by(desc(TaskCommentDB.created_at)).limit(50).all()

    # Get messages for this tender (from all assignments for this tender)
    messages = []
    if assignment_ids:
        messages = db.query(TenderMessageDB).filter(
            TenderMessageDB.assignment_id.in_(assignment_ids)
        ).order_by(TenderMessageDB.created_at.asc()).limit(100).all()

    return templates.TemplateResponse("team.html", {
        "request": request,
        "current_user": current_user,
        "tender": tender,
        "assignments": formatted_assignments,
        "comments": comments,
        "messages": messages,
        "now": datetime.utcnow(),
        "selected_font": get_active_font()
    })

@app.get("/tender/{tender_id}", response_class=HTMLResponse)
@enforce_test_quarantine
@require_company_details
async def tender_detail(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Tender detail page."""
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    def _copy_default(value):
        if isinstance(value, dict):
            return dict(value)
        if isinstance(value, list):
            return list(value)
        return value

    def _normalize_json(value, default):
        if value is None:
            return _copy_default(default)
        if isinstance(value, (dict, list)):
            return value
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                return _copy_default(default)
            try:
                parsed = json.loads(stripped)
            except json.JSONDecodeError:
                return _copy_default(default)
            if isinstance(default, dict) and isinstance(parsed, dict):
                return parsed
            if isinstance(default, list) and isinstance(parsed, list):
                return parsed
            return _copy_default(default)
        return _copy_default(default)

    json_defaults = {
        "work_item_details": {},
        "critical_dates": {},
        "tender_documents": {},
        "tender_fee_details": {},
        "emd_fee_details": {},
        "payment_instruments": {},
        "tender_inviting_authority": {},
        "additional_fields": {},
        "covers_information": [],
        "tags": [],
    }

    for field, default in json_defaults.items():
        if hasattr(tender, field):
            setattr(tender, field, _normalize_json(getattr(tender, field), default))

    _prepare_tender_display_fields(tender)

    # Get authenticated entity (user or BD employee)
    from core.dependencies import get_id_for_tender_management, get_user_id_for_queries
    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    
    # Set context variables for templates
    current_user = entity if entity_type == 'user' else None
    current_employee = entity if entity_type == 'bd_employee' else None

    # Get user_id for database queries (company owner's user_id for BD employees)
    user_id_for_query, _ = get_user_id_for_queries(request, db) if entity_id else (None, None)

    # Mark tender as seen (independent of other checks)
    if entity_id:
        try:
            if entity_type == 'user':
                # Check if already seen
                seen_record = db.query(SeenTenderDB).filter(
                    and_(SeenTenderDB.user_id == entity_id, SeenTenderDB.tender_id == tender.id)
                ).first()
                if not seen_record:
                    seen_record = SeenTenderDB(
                        user_id=entity_id,
                        tender_id=tender.id
                    )
                    db.add(seen_record)
                    db.commit()
                    logger.info(f"Marked tender {tender.id} as seen for user {entity_id}")
            elif entity_type == 'bd_employee':
                # Check if already seen
                seen_record = db.query(SeenTenderDB).filter(
                    and_(SeenTenderDB.employee_id == entity_id, SeenTenderDB.tender_id == tender.id)
                ).first()
                if not seen_record:
                    seen_record = SeenTenderDB(
                        employee_id=entity_id,
                        tender_id=tender.id
                    )
                    db.add(seen_record)
                    db.commit()
                    logger.info(f"Marked tender {tender.id} as seen for BD employee {entity_id}")
            else:
                logger.warning(f"Unknown entity_type '{entity_type}' for entity_id {entity_id}, cannot mark tender as seen")
        except Exception as e:
            logger.error(f"Failed to mark tender {tender.id} as seen for {entity_type} {entity_id}: {e}", exc_info=True)
    else:
        logger.debug(f"No entity_id found, cannot mark tender {tender.id} as seen")

    # Check if entity has favorited this tender
    is_favorited = False
    is_shortlisted = False
    is_rejected = False
    is_expired = False
    is_awarded = tender.awarded if tender.awarded else False

    if entity_id and user_id_for_query:
        # Check favorite - for BD employees, check by worked_by fields
        if entity_type == 'bd_employee':
            favorite = db.query(FavoriteDB).filter(
                and_(
                    FavoriteDB.user_id == user_id_for_query,
                    FavoriteDB.tender_id == tender.id,
                    FavoriteDB.worked_by_type == 'bd_employee',
                    FavoriteDB.worked_by_name == entity.name
                )
            ).first()
        else:
            favorite = db.query(FavoriteDB).filter(
                and_(FavoriteDB.user_id == user_id_for_query, FavoriteDB.tender_id == tender.id)
            ).first()
        is_favorited = favorite is not None

        # Check if shortlisted (shortlisted still uses entity_id directly)
        shortlisted = db.query(ShortlistedTenderDB).filter(
            and_(ShortlistedTenderDB.user_id == entity_id, ShortlistedTenderDB.tender_id == tender.id)
        ).first()
        is_shortlisted = shortlisted is not None

        # Check if rejected (rejected still uses entity_id directly)
        rejected = db.query(RejectedTenderDB).filter(
            and_(RejectedTenderDB.user_id == entity_id, RejectedTenderDB.tender_id == tender.id)
        ).first()
        is_rejected = rejected is not None

        # Check if tender has expired and entity had it favorited/shortlisted
        if tender.deadline:
            now = datetime.utcnow()
            if tender.deadline < now and (is_favorited or is_shortlisted):
                is_expired = True

    return_to = request.query_params.get('return_to')
    if return_to:
        return_to = unquote(return_to)
        if not return_to.startswith('/'):
            return_to = None

    return templates.TemplateResponse("tender_detail.html", {
        "request": request,
        "tender": tender,
        "current_user": current_user,
        "current_employee": current_employee,
        "is_favorited": is_favorited,
        "is_shortlisted": is_shortlisted,
        "is_rejected": is_rejected,
        "is_expired": is_expired,
        "is_awarded": is_awarded,
        "return_to": return_to,
        "selected_font": get_active_font()
    })

@app.get("/tender/{tender_id}/tasks", response_class=HTMLResponse)
@require_company_details
async def tender_task_workspace(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """
    Task workspace page for managing tasks, team members, and communication for a specific tender.
    Displays all assignments, tasks, subtasks, concerns, and team chat for the tender.
    """
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get tender details
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Verify tender is awarded
    if not tender.awarded:
        raise HTTPException(status_code=400, detail="Tasks can only be managed for awarded tenders")

    # Verify tender was awarded by current user
    if tender.awarded_by != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    # Get all assignments for this tender with employee details
    assignments = db.query(TenderAssignmentDB).filter(
        TenderAssignmentDB.tender_id == tender_id
    ).order_by(TenderAssignmentDB.assigned_at.desc()).all()

    # Get all tasks for these assignments (only parent tasks, not subtasks)
    assignment_ids = [a.id for a in assignments]
    tasks = []
    if assignment_ids:
        tasks = db.query(TaskDB).filter(
            and_(
                TaskDB.assignment_id.in_(assignment_ids),
                TaskDB.is_subtask == False
            )
        ).order_by(TaskDB.created_at.desc()).all()

        # For each task, load subtasks explicitly
        for task in tasks:
            task.subtasks_list = db.query(TaskDB).filter(
                TaskDB.parent_task_id == task.id
            ).order_by(TaskDB.subtask_order).all()

    # Get all concerns for this tender's tasks
    task_ids = [t.id for t in tasks]
    concerns = []
    if task_ids:
        concerns = db.query(TaskConcernDB).filter(
            TaskConcernDB.task_id.in_(task_ids)
        ).order_by(TaskConcernDB.created_at.desc()).all()

    # Get suggested employees from shortlisted phase
    shortlisted_tender = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.tender_id == tender_id,
            ShortlistedTenderDB.user_id == current_user.id
        )
    ).first()

    suggested_employee_ids = set()
    if shortlisted_tender and shortlisted_tender.progress_data:
        progress_data = shortlisted_tender.progress_data
        if isinstance(progress_data, str):
            try:
                progress_data = json.loads(progress_data)
            except:
                progress_data = {}

        # Extract employee IDs from all steps
        for step_key in ['step1_employees', 'step2_employees', 'step3_employees',
                         'step4_employees', 'step5_employees', 'step6_employees']:
            if step_key in progress_data:
                step_employees = progress_data[step_key]
                if isinstance(step_employees, list):
                    suggested_employee_ids.update(step_employees)

    # Get employee details for suggestions (excluding already assigned)
    assigned_employee_ids = set([a.employee_id for a in assignments])
    available_suggested_ids = suggested_employee_ids - assigned_employee_ids

    # Get user's company codes to filter employees
    user_company_codes = db.query(CompanyCodeDB).filter(
        CompanyCodeDB.user_id == current_user.id
    ).all()
    company_code_ids = [cc.id for cc in user_company_codes]

    suggested_employees = []
    if available_suggested_ids and company_code_ids:
        suggested_employees = db.query(EmployeeDB).filter(
            and_(
                EmployeeDB.id.in_(list(available_suggested_ids)),
                EmployeeDB.company_code_id.in_(company_code_ids)
            )
        ).all()

    # Get all employees for user (for manual selection)
    all_employees = []
    if company_code_ids:
        all_employees = db.query(EmployeeDB).filter(
            EmployeeDB.company_code_id.in_(company_code_ids)
        ).order_by(EmployeeDB.name).all()

    # Get team chat messages for all assignments
    messages = []
    if assignment_ids:
        messages = db.query(TenderMessageDB).filter(
            TenderMessageDB.assignment_id.in_(assignment_ids)
        ).order_by(TenderMessageDB.created_at.asc()).all()

    # Calculate task statistics
    total_tasks = len(tasks)
    completed_tasks = sum(1 for t in tasks if t.status == 'completed')
    in_progress_tasks = sum(1 for t in tasks if t.status == 'in_progress')
    pending_tasks = sum(1 for t in tasks if t.status == 'pending')
    open_concerns = sum(1 for c in concerns if c.status == 'open')

    return templates.TemplateResponse("employee_task_page.html", {
        "request": request,
        "current_user": current_user,
        "tender": tender,
        "assignments": assignments,
        "assigned_employee_ids": assigned_employee_ids,
        "tasks": tasks,
        "concerns": concerns,
        "suggested_employees": suggested_employees,
        "all_employees": all_employees,
        "messages": messages,
        "stats": {
            "total_tasks": total_tasks,
            "completed_tasks": completed_tasks,
            "in_progress_tasks": in_progress_tasks,
            "pending_tasks": pending_tasks,
            "open_concerns": open_concerns,
            "team_size": len(assignments)
        },
        "selected_font": get_active_font()
    })

@app.get("/api/tender/{tender_id}/document")
async def serve_tender_document(tender_id: str, db: Session = Depends(get_db)):
    """
    Serve the tender document (PDF/ZIP) from database for viewing or download.
    Extracts PDF from ZIP if necessary.
    """
    import zipfile
    import io

    try:
        # Get tender from database
        tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
        if not tender:
            raise HTTPException(status_code=404, detail="Tender not found")

        # Query PDF or ZIP documents from TenderDocumentDB
        # Prioritize PDF, then ZIP
        pdf_doc = db.query(TenderDocumentDB).filter(
            TenderDocumentDB.tender_id == tender_id,
            TenderDocumentDB.document_type == 'pdf'
        ).first()

        if pdf_doc:
            # Serve PDF directly from database
            return Response(
                content=pdf_doc.file_data,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f'inline; filename="{pdf_doc.filename}"',
                    "Cache-Control": "public, max-age=3600"
                }
            )

        # If no PDF, try ZIP file
        zip_doc = db.query(TenderDocumentDB).filter(
            TenderDocumentDB.tender_id == tender_id,
            TenderDocumentDB.document_type == 'zip'
        ).first()

        if zip_doc:
            # Extract PDF from ZIP stored in database
            try:
                with zipfile.ZipFile(io.BytesIO(zip_doc.file_data), 'r') as zip_ref:
                    # Find all PDF files in the ZIP
                    pdf_files = [f for f in zip_ref.namelist() if f.lower().endswith('.pdf')]

                    if not pdf_files:
                        raise HTTPException(status_code=404, detail="No PDF found in ZIP file")

                    # Extract the first PDF to memory
                    pdf_data = zip_ref.read(pdf_files[0])

                    # Return PDF from memory
                    return Response(
                        content=pdf_data,
                        media_type="application/pdf",
                        headers={
                            "Content-Disposition": f'inline; filename="tender_{tender_id}.pdf"',
                            "Cache-Control": "public, max-age=3600"
                        }
                    )
            except zipfile.BadZipFile:
                raise HTTPException(status_code=500, detail="Invalid ZIP file")

        # No documents found
        raise HTTPException(status_code=404, detail="No document available for this tender")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving tender document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error serving document: {str(e)}")

@app.get("/api/tender/{tender_id}/documents")
async def list_tender_documents(tender_id: str, db: Session = Depends(get_db)):
    """
    List all documents (non-screenshot) for a tender.
    Returns metadata about each document for display in UI.
    """
    try:
        # Get tender from database
        tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
        if not tender:
            raise HTTPException(status_code=404, detail="Tender not found")

        # Query all non-screenshot documents
        documents = db.query(TenderDocumentDB).filter(
            TenderDocumentDB.tender_id == tender_id,
            TenderDocumentDB.document_type != 'screenshot'
        ).order_by(TenderDocumentDB.display_order).all()

        # Format document list for frontend
        from s3_utils import get_presigned_url
        
        doc_list = []
        for doc in documents:
            # Always prioritize s3_key to generate presigned URL (direct S3 URLs don't work without public access)
            if doc.s3_key:
                presigned_url = get_presigned_url(doc.s3_key, expiration=3600)
                download_url = presigned_url if presigned_url else f"/api/tender/document/{doc.id}"
            elif doc.s3_url and '?' in doc.s3_url:
                # s3_url is already a presigned URL (contains query params)
                download_url = doc.s3_url
            elif doc.s3_url:
                # s3_url is a direct S3 URL, which won't work - fall back to endpoint
                download_url = f"/api/tender/document/{doc.id}"
            else:
                download_url = f"/api/tender/document/{doc.id}"
            
            doc_list.append({
                'id': doc.id,
                'filename': doc.filename,
                'document_type': doc.document_type,
                'file_size': doc.file_size,
                'display_order': doc.display_order,
                'created_at': doc.created_at.isoformat() if doc.created_at else None,
                'url': download_url,  # Add URL for direct access
                's3_url': doc.s3_url,
                's3_key': doc.s3_key
            })

        return {"success": True, "documents": doc_list, "count": len(doc_list)}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing tender documents: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error listing documents: {str(e)}")

@app.get("/api/tender/document/{document_id}")
async def download_tender_document_by_id(document_id: int, db: Session = Depends(get_db)):
    """
    Download a specific document by its database ID.
    Serves the binary content with appropriate headers for viewing or download.
    """
    try:
        # Get document from database
        document = db.query(TenderDocumentDB).filter(TenderDocumentDB.id == document_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")

        # Priority 1: Generate presigned URL from s3_key (direct S3 URLs don't work without public access)
        if document.s3_key:
            from s3_utils import get_presigned_url
            presigned_url = get_presigned_url(document.s3_key, expiration=3600)
            if presigned_url:
                return RedirectResponse(url=presigned_url, status_code=302)
        
        # Priority 2: Use s3_url only if it's already a presigned URL (contains query params)
        if document.s3_url and '?' in document.s3_url:
            return RedirectResponse(url=document.s3_url, status_code=302)
        
        # Priority 3: Fall back to file_data
        if document.file_data:
            return Response(
                content=document.file_data,
                media_type=document.mime_type or "application/octet-stream",
                headers={
                    "Content-Disposition": f'inline; filename="{document.filename}"',
                    "Cache-Control": "public, max-age=3600"
                }
            )
        
        raise HTTPException(status_code=404, detail="Document not available")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading document {document_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error downloading document: {str(e)}")

@app.get("/api/tender/{tender_id}/screenshots")
async def serve_tender_screenshots(tender_id: str, db: Session = Depends(get_db)):
    """
    Serve the tender screenshots as a JSON list from database.
    """
    try:
        # Get tender from database
        tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
        if not tender:
            raise HTTPException(status_code=404, detail="Tender not found")

        # Query screenshot documents from TenderDocumentDB
        screenshots = db.query(TenderDocumentDB).filter(
            TenderDocumentDB.tender_id == tender_id,
            TenderDocumentDB.document_type == 'screenshot'
        ).order_by(TenderDocumentDB.display_order).all()

        if not screenshots:
            raise HTTPException(status_code=404, detail="No screenshots available for this tender")

        # Return list of screenshot metadata
        from s3_utils import get_presigned_url
        
        available_screenshots = []
        for screenshot in screenshots:
            # Always prioritize s3_key to generate presigned URL (direct S3 URLs don't work without public access)
            if screenshot.s3_key:
                presigned_url = get_presigned_url(screenshot.s3_key, expiration=3600)
                url = presigned_url if presigned_url else f"/api/tender/{tender_id}/screenshot/{screenshot.id}"
            elif screenshot.s3_url and '?' in screenshot.s3_url:
                # s3_url is already a presigned URL (contains query params)
                url = screenshot.s3_url
            elif screenshot.s3_url:
                # s3_url is a direct S3 URL, which won't work - fall back to endpoint
                url = f"/api/tender/{tender_id}/screenshot/{screenshot.id}"
            else:
                url = f"/api/tender/{tender_id}/screenshot/{screenshot.id}"
            
            available_screenshots.append({
                'id': screenshot.id,
                'filename': screenshot.filename,
                'file_size': screenshot.file_size,
                'mime_type': screenshot.mime_type,
                'url': url,
                's3_url': screenshot.s3_url,  # Include for frontend reference
                's3_key': screenshot.s3_key
            })

        return {"screenshots": available_screenshots}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting tender screenshots: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting screenshots: {str(e)}")


@app.get("/api/tender/{tender_id}/screenshot/{screenshot_id}")
async def serve_tender_screenshot_file(tender_id: str, screenshot_id: int, db: Session = Depends(get_db)):
    """
    Serve a specific screenshot file from database as binary data.
    """
    try:
        # Get screenshot document from database
        screenshot = db.query(TenderDocumentDB).filter(
            TenderDocumentDB.id == screenshot_id,
            TenderDocumentDB.tender_id == tender_id,
            TenderDocumentDB.document_type == 'screenshot'
        ).first()

        if not screenshot:
            raise HTTPException(status_code=404, detail="Screenshot not found")

        # Priority 1: Generate presigned URL from s3_key (direct S3 URLs don't work without public access)
        if screenshot.s3_key:
            from s3_utils import get_presigned_url
            presigned_url = get_presigned_url(screenshot.s3_key, expiration=3600)
            if presigned_url:
                return RedirectResponse(url=presigned_url, status_code=302)
        
        # Priority 2: Use s3_url only if it's already a presigned URL (contains query params)
        if screenshot.s3_url and '?' in screenshot.s3_url:
            return RedirectResponse(url=screenshot.s3_url, status_code=302)
        
        # Priority 3: Fall back to file_data (backwards compatibility)
        if screenshot.file_data:
            return Response(
                content=screenshot.file_data,
                media_type=screenshot.mime_type,
                headers={
                    "Content-Disposition": f'inline; filename="{screenshot.filename}"',
                    "Cache-Control": "public, max-age=3600"
                }
            )
        
        # If nothing available, return 404
        raise HTTPException(status_code=404, detail="Screenshot not available")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving screenshot: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error serving screenshot: {str(e)}")


@app.get("/custom-card/{card_id}", response_class=HTMLResponse)
@require_company_details
async def custom_card_tenders(request: Request, card_id: int, db: Session = Depends(get_db)):
    """Custom card tenders page."""
    from core.dependencies import get_user_id_for_queries
    
    user_id_for_query, _ = get_user_id_for_queries(request, db)
    if not user_id_for_query:
        return RedirectResponse(url="/login", status_code=302)

    # Find the card - query by user_id_for_query
    card = db.query(CustomCardDB).filter(
        and_(CustomCardDB.id == card_id, CustomCardDB.user_id == user_id_for_query)
    ).first()

    if not card:
        raise HTTPException(status_code=404, detail="Custom card not found")

    # Get page parameters
    page = int(request.query_params.get('page', 1))
    per_page = 20
    skip = (page - 1) * per_page

    # Get tenders for this card
    result = get_custom_card_tenders(card_id, skip, per_page, request, db)
    tenders = result["tenders"]
    total_tenders = result["total"]

    # Pagination info
    total_pages = (total_tenders + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages

    return templates.TemplateResponse("custom_card_tenders.html", {
        "request": request,
        "current_user": current_user,
        "card": result["card"],
        "tenders": tenders,
        "page": page,
        "total_pages": total_pages,
        "has_prev": has_prev,
        "has_next": has_next,
        "total_tenders": total_tenders,
        "selected_font": get_active_font()
    })



# Authentication API endpoints
@app.post("/api/auth/signup")
def signup(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    company: str = Form(""),
    role: str = Form(""),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """User registration."""
    # Check if user already exists
    existing_user = db.query(UserDB).filter(UserDB.email == email).first()
    if existing_user:
        # Redirect to login page with signup error message
        return RedirectResponse(url=f"/login?error=email_exists&email={email}", status_code=302)

    # Create new user
    user_id = str(uuid.uuid4())
    hashed_password = hash_password(password)

    new_user = UserDB(
        id=user_id,
        email=email,
        name=name,
        company=company,
        role=role,
        password_hash=hashed_password,
        created_at=datetime.utcnow()
    )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    # Create session
    session_token = create_session(user_id)

    # Create response with redirect to company details
    response = RedirectResponse(url="/company-details", status_code=302)
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        max_age=7*24*60*60  # 7 days
    )

    return response

@app.post("/api/auth/login")
def login(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """User login."""
    # Find user
    user = db.query(UserDB).filter(UserDB.email == email).first()
    if not user or not verify_password(password, user.password_hash): # type: ignore
        # Redirect to login page with error message
        return RedirectResponse(url=f"/login?error=invalid_credentials&email={email}", status_code=302)

    # Update last login
    user.last_login = datetime.utcnow() # type: ignore
    db.commit()

    # Create session
    session_token = create_session(user.id) # type: ignore

    # Check if user has complete company details
    redirect_url = "/home" if user_has_complete_company_details(user.id, db) else "/company-details" # type: ignore

    # Create response with redirect
    response = RedirectResponse(url=redirect_url, status_code=302)
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        max_age=7*24*60*60  # 7 days
    )

    return response

@app.post("/api/auth/logout")
async def logout(request: Request):
    """User logout."""
    session_token = request.cookies.get('session_token')
    if session_token and session_token in user_sessions:
        del user_sessions[session_token]

    response = RedirectResponse(url="/", status_code=302)
    response.delete_cookie(key="session_token")
    return response

# Employee authentication API endpoints
# Employee signup route removed - employees are now added by administrators only
# Employees receive their login credentials from their company admin

@app.post("/api/auth/employee/login")
async def employee_login(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """Employee login."""
    # Find employee
    employee = db.query(EmployeeDB).filter(EmployeeDB.email == email).first()
    if not employee or not verify_password(password, employee.password_hash): # type: ignore
        return RedirectResponse(url=f"/employee/login?error=invalid_credentials&email={email}", status_code=302)

    # Update last login
    employee.last_login = datetime.utcnow() # type: ignore
    db.commit()

    # Create session
    session_token = create_session(employee.id) # type: ignore

    # Redirect based on employee type
    if employee.is_bd:  # type: ignore
        redirect_url = "/bd/home"
    else:
        redirect_url = "/employee/dashboard"

    # Create response with redirect
    response = RedirectResponse(url=redirect_url, status_code=302)
    response.set_cookie(
        key="employee_session_token",
        value=session_token,
        httponly=True,
        max_age=7*24*60*60  # 7 days
    )

    return response

@app.post("/api/auth/employee/logout")
async def employee_logout(request: Request):
    """Employee logout."""
    session_token = request.cookies.get('employee_session_token')
    if session_token and session_token in user_sessions:
        del user_sessions[session_token]

    response = RedirectResponse(url="/", status_code=302)
    response.delete_cookie(key="employee_session_token")
    return response


# ==================== BD Employee Routes ====================

@app.get("/bd/home", response_class=HTMLResponse)
async def bd_employee_home(request: Request, db: Session = Depends(get_db)):
    """BD Employee home page."""
    from core.dependencies import get_current_bd_employee

    current_employee = get_current_bd_employee(request, db)
    if not current_employee:
        return RedirectResponse(url="/employee/login", status_code=302)

    # Get company details via company_code
    company_details = None
    sector_list = []

    if current_employee.company_code:
        # Get the user who owns this company code
        company_code = db.query(CompanyCodeDB).filter(CompanyCodeDB.id == current_employee.company_code_id).first()
        if company_code:
            company_details = db.query(CompanyDB).filter(CompanyDB.user_id == company_code.user_id).first()

            # Parse industry sectors
            if company_details and company_details.industry_sector:
                import json as json_module
                if company_details.industry_sector.startswith('[{'):
                    try:
                        sectors_data = json_module.loads(company_details.industry_sector)
                        sector_list = [s['sector'] for s in sectors_data if 'sector' in s]
                    except (json_module.JSONDecodeError, KeyError):
                        sector_list = [company_details.industry_sector]
                elif company_details.industry_sector.startswith('['):
                    try:
                        sector_list = json_module.loads(company_details.industry_sector)
                    except json_module.JSONDecodeError:
                        sector_list = [company_details.industry_sector]
                else:
                    sector_list = [company_details.industry_sector]

    return templates.TemplateResponse("bd_employee_home.html", {
        "request": request,
        "current_employee": current_employee,
        "company_details": company_details,
        "sector_list": sector_list,
        "selected_font": get_active_font()
    })


# ==================== Expert-Verse Authentication Routes ====================


@app.get("/expert/login", response_class=HTMLResponse)
async def expert_login_page(request: Request):
    """Expert login/signup page."""
    return templates.TemplateResponse("expert_login.html", {
        "request": request,
        "selected_font": get_active_font()
    })


@app.post("/api/expert/auth/signup")
async def expert_signup(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db)
):
    """Expert registration - simplified signup, full profile completed after."""
    try:
        # Server-side password confirmation validation
        if password != confirm_password:
            return JSONResponse(
                status_code=400,
                content={"detail": "Passwords do not match. Please ensure both password fields are identical."}
            )

        # Validate password strength
        if len(password) < 8:
            return JSONResponse(
                status_code=400,
                content={"detail": "Password must be at least 8 characters long."}
            )

        # Check if email already exists
        existing_expert = db.query(ExpertDB).filter(ExpertDB.email == email).first()
        if existing_expert:
            return JSONResponse(
                status_code=400,
                content={"detail": "An expert account with this email already exists. Please login instead."}
            )

        # Check if email is used by a manager or employee
        existing_user = db.query(UserDB).filter(UserDB.email == email).first()
        existing_employee = db.query(EmployeeDB).filter(EmployeeDB.email == email).first()
        if existing_user or existing_employee:
            return JSONResponse(
                status_code=400,
                content={"detail": "This email is already registered as a company/employee account. Please use a different email or login to the appropriate account."}
            )

        # Create new expert
        expert_id = str(uuid.uuid4())
        hashed_password = hash_password(password)

        new_expert = ExpertDB(
            id=expert_id,
            email=email.lower().strip(),
            name=name.strip(),
            password_hash=hashed_password,
            created_at=datetime.utcnow(),
            profile_completed=False  # Must complete profile in next step
        )

        db.add(new_expert)
        db.commit()
        db.refresh(new_expert)

        logger.info(f"✓ New expert created: {new_expert.email} (ID: {new_expert.id})")

        # Create session
        session_token = create_expert_session(expert_id)

        # Create response redirecting to profile setup
        response = JSONResponse(
            content={
                "success": True,
                "message": "Expert account created successfully",
                "redirect": "/expert/profile-setup"
            }
        )
        response.set_cookie(
            key="expert_session_token",
            value=session_token,
            httponly=True,
            max_age=7*24*60*60,  # 7 days
            samesite="lax"
        )

        return response

    except Exception as e:
        logger.error(f"Expert signup error: {e}", exc_info=True)
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={"detail": f"An error occurred during signup: {str(e)}"}
        )


@app.post("/api/expert/auth/login")
async def expert_login(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """Expert login."""
    try:
        # Find expert
        expert = db.query(ExpertDB).filter(ExpertDB.email == email.lower().strip()).first()
        if not expert or not verify_password(password, expert.password_hash):
            return JSONResponse(
                status_code=401,
                content={"detail": "Invalid email or password. Please check your credentials and try again."}
            )

        # Check if account is active
        if not expert.is_active:
            return JSONResponse(
                status_code=403,
                content={"detail": "Your expert account has been deactivated. Please contact support."}
            )

        # Update last login
        expert.last_login = datetime.utcnow()
        db.commit()

        # Create session
        session_token = create_expert_session(expert.id)

        # Determine redirect based on profile completion
        if expert.profile_completed and expert_has_complete_profile(expert.id, db):
            redirect_url = "/expert/dashboard"
        else:
            redirect_url = "/expert/profile-setup"

        # Create response
        response = JSONResponse(
            content={
                "success": True,
                "message": "Login successful",
                "redirect": redirect_url
            }
        )
        response.set_cookie(
            key="expert_session_token",
            value=session_token,
            httponly=True,
            max_age=7*24*60*60,  # 7 days
            samesite="lax"
        )

        return response

    except Exception as e:
        logger.error(f"Expert login error: {e}")
        return JSONResponse(
            status_code=500,
            content={"detail": "An error occurred during login. Please try again."}
        )


@app.get("/expert/logout")
async def expert_logout(request: Request):
    """Expert logout."""
    session_token = request.cookies.get('expert_session_token')
    if session_token:
        delete_expert_session(session_token)

    response = RedirectResponse(url="/", status_code=302)
    response.delete_cookie(key="expert_session_token")
    return response


# ==================== Expert Profile Setup Routes ====================


@app.get("/expert/profile-setup", response_class=HTMLResponse)
@require_expert_login
async def expert_profile_setup_page(request: Request, db: Session = Depends(get_db)):
    """Expert profile setup page - multi-step form."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Get existing profile data if any
    existing_profile = db.query(ExpertProfileDB).filter(
        ExpertProfileDB.expert_id == current_expert.id
    ).first()

    return templates.TemplateResponse("expert_profile_setup.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "profile": existing_profile,
        "selected_font": get_active_font()
    })


@app.post("/api/expert/profile-setup")
@require_expert_login
async def save_expert_profile(
    request: Request,
    db: Session = Depends(get_db),
    # Step 1 - Mandatory fields
    expertise_areas: list[str] = Form(default=[]),
    services_offered: list[str] = Form(default=[]),
    # Step 2 - Optional experience fields
    employment_type: Optional[str] = Form(default=None),
    experience_years: Optional[str] = Form(default=None),
    qualification_degree: list[str] = Form(default=[]),
    qualification_institution: list[str] = Form(default=[]),
    qualification_year: list[str] = Form(default=[]),
    certification_name: list[str] = Form(default=[]),
    certification_issuer: list[str] = Form(default=[]),
    # Step 3 - Optional additional fields
    bio: Optional[str] = Form(default=None),
    phone_number: Optional[str] = Form(default=None),
    location: Optional[str] = Form(default=None),
    hourly_rate: Optional[str] = Form(default=None),
    availability_status: str = Form(default="available"),
    linkedin_url: Optional[str] = Form(default=None),
    portfolio_url: Optional[str] = Form(default=None),
    languages: Optional[str] = Form(default=None),
    willing_to_travel: bool = Form(default=False)
):
    """Save expert profile data from multi-step form."""
    try:
        current_expert = get_current_expert(request, db)
        if not current_expert:
            return JSONResponse(
                status_code=401,
                content={"detail": "Not authenticated"}
            )

        # Validate mandatory fields (Step 1)
        if not expertise_areas or len(expertise_areas) == 0:
            return JSONResponse(
                status_code=400,
                content={"detail": "Please select at least one area of expertise"}
            )

        if not services_offered or len(services_offered) == 0:
            return JSONResponse(
                status_code=400,
                content={"detail": "Please select at least one service offering"}
            )

        # Convert numeric fields from strings to proper types
        experience_years_int = None
        if experience_years and experience_years.strip():
            try:
                experience_years_int = int(experience_years)
            except (ValueError, AttributeError):
                experience_years_int = None

        hourly_rate_float = None
        if hourly_rate and hourly_rate.strip():
            try:
                hourly_rate_float = float(hourly_rate)
            except (ValueError, AttributeError):
                hourly_rate_float = None

        # Process qualifications into array of objects
        qualifications = []
        if qualification_degree:
            for i in range(len(qualification_degree)):
                if qualification_degree[i].strip():  # Only add if degree is provided
                    qual = {
                        "degree": qualification_degree[i].strip(),
                        "institution": qualification_institution[i].strip() if i < len(qualification_institution) else "",
                        "year": qualification_year[i].strip() if i < len(qualification_year) else ""
                    }
                    qualifications.append(qual)

        # Handle certification file uploads
        certifications = []
        form_data = await request.form()
        certification_files = form_data.getlist("certification_file")

        if certification_name:
            for i in range(len(certification_name)):
                if certification_name[i].strip():  # Only add if name is provided
                    cert = {
                        "name": certification_name[i].strip(),
                        "issuer": certification_issuer[i].strip() if i < len(certification_issuer) else "",
                        "file_path": None
                    }

                    # Handle file upload for this certification
                    if i < len(certification_files) and certification_files[i].filename:
                        file = certification_files[i]
                        # Validate file type and size
                        allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png']
                        file_ext = os.path.splitext(file.filename)[1].lower()

                        if file_ext not in allowed_extensions:
                            return JSONResponse(
                                status_code=400,
                                content={"detail": f"Invalid file type for certification {i+1}. Allowed: PDF, JPG, PNG"}
                            )

                        # Check file size (max 5MB)
                        file_content = await file.read()
                        if len(file_content) > 5 * 1024 * 1024:  # 5MB
                            return JSONResponse(
                                status_code=400,
                                content={"detail": f"File size exceeds 5MB for certification {i+1}"}
                            )

                        # Create upload directory if it doesn't exist
                        upload_dir = "static/uploads/expert_certifications"
                        os.makedirs(upload_dir, exist_ok=True)

                        # Generate unique filename
                        unique_filename = f"{current_expert.id}_{int(datetime.utcnow().timestamp())}_{i}{file_ext}"
                        file_path = os.path.join(upload_dir, unique_filename)

                        # Save file
                        with open(file_path, "wb") as f:
                            f.write(file_content)

                        cert["file_path"] = file_path

                    certifications.append(cert)

        # Check if profile exists
        existing_profile = db.query(ExpertProfileDB).filter(
            ExpertProfileDB.expert_id == current_expert.id
        ).first()

        if existing_profile:
            # Update existing profile
            existing_profile.expertise_areas = expertise_areas
            existing_profile.services_offered = services_offered
            existing_profile.employment_type = employment_type.strip() if employment_type else None
            existing_profile.experience_years = experience_years_int
            existing_profile.qualifications = qualifications
            existing_profile.certifications = certifications
            existing_profile.location = location.strip() if location else None
            existing_profile.linkedin_url = linkedin_url.strip() if linkedin_url else None
            existing_profile.portfolio_url = portfolio_url.strip() if portfolio_url else None
            existing_profile.willing_to_travel = willing_to_travel
            existing_profile.updated_at = datetime.utcnow()
        else:
            # Create new profile
            new_profile = ExpertProfileDB(
                id=str(uuid.uuid4()),
                expert_id=current_expert.id,
                expertise_areas=expertise_areas,
                services_offered=services_offered,
                employment_type=employment_type.strip() if employment_type else None,
                experience_years=experience_years_int,
                qualifications=qualifications,
                certifications=certifications,
                location=location.strip() if location else None,
                linkedin_url=linkedin_url.strip() if linkedin_url else None,
                portfolio_url=portfolio_url.strip() if portfolio_url else None,
                willing_to_travel=willing_to_travel
            )
            db.add(new_profile)

        # Update expert base record fields (these are in ExpertDB, not ExpertProfileDB)
        current_expert.profile_completed = True
        current_expert.bio = bio.strip() if bio else None
        current_expert.phone_number = phone_number.strip() if phone_number else None
        current_expert.hourly_rate = hourly_rate_float
        current_expert.availability_status = availability_status
        if languages and languages.strip():
            current_expert.languages = [lang.strip() for lang in languages.split(',') if lang.strip()]
        else:
            current_expert.languages = []

        db.commit()

        return JSONResponse(content={
            "success": True,
            "message": "Profile completed successfully!",
            "redirect": "/expert/dashboard"
        })

    except Exception as e:
        db.rollback()
        logger.error(f"Error saving expert profile: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"detail": f"Error saving profile: {str(e)}"}
        )


# ==================== Expert Dashboard Routes ====================


@app.get("/expert/dashboard", response_class=HTMLResponse)
@require_expert_login
@require_expert_profile_complete
async def expert_dashboard(request: Request, db: Session = Depends(get_db)):
    """Expert dashboard - main landing page after profile completion."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Get expert profile
    profile = db.query(ExpertProfileDB).filter(
        ExpertProfileDB.expert_id == current_expert.id
    ).first()

    # Calculate statistics
    service_requests_count = db.query(ExpertServiceRequestDB).filter(
        ExpertServiceRequestDB.expert_id == current_expert.id
    ).count()

    proactive_applications_count = db.query(ExpertApplicationDB).filter(
        ExpertApplicationDB.expert_id == current_expert.id
    ).count()

    hiring_applications_count = db.query(ExpertHiringApplicationDB).filter(
        ExpertHiringApplicationDB.expert_id == current_expert.id
    ).count()

    total_applications = proactive_applications_count + hiring_applications_count

    active_collaboration_projects = db.query(ExpertCollaborationDB).filter(
        or_(
            ExpertCollaborationDB.primary_expert_id == current_expert.id,
            ExpertCollaborationDB.collaborator_expert_id == current_expert.id
        ),
        ExpertCollaborationDB.status == 'active'
    ).count()

    active_hiring_projects = db.query(ExpertHiringApplicationDB).join(ExpertHiringRequestDB).filter(
        ExpertHiringApplicationDB.expert_id == current_expert.id,
        ExpertHiringApplicationDB.status == 'accepted',
        ExpertHiringRequestDB.status.in_(["open", "filled"])
    ).count()

    active_projects_count = active_collaboration_projects + active_hiring_projects

    stats = {
        'service_requests': service_requests_count,
        'applications': total_applications,
        'proactive_applications': proactive_applications_count,
        'hiring_applications': hiring_applications_count,
        'active_projects': active_projects_count,
        'active_collaboration_projects': active_collaboration_projects,
        'active_hiring_projects': active_hiring_projects
    }

    # Get recent service requests
    service_requests = db.query(ExpertServiceRequestDB).filter(
        ExpertServiceRequestDB.expert_id == current_expert.id
    ).order_by(ExpertServiceRequestDB.created_at.desc()).limit(5).all()

    # Fetch company and tender details for service requests
    for request_item in service_requests:
        # Get company details
        company = db.query(CompanyDB).filter(
            CompanyDB.user_id == request_item.company_id
        ).first()
        request_item.company_name = company.company_name if company else "Unknown Company"

        # Get tender details
        tender = db.query(TenderDB).filter(
            TenderDB.id == request_item.tender_id
        ).first()
        request_item.tender_title = tender.title if tender else "Unknown Tender"

    notifications = db.query(ExpertNotificationDB).filter(
        ExpertNotificationDB.expert_id == current_expert.id
    ).order_by(desc(ExpertNotificationDB.created_at)).limit(10).all()

    unread_notifications = sum(1 for notif in notifications if not notif.is_read)

    return templates.TemplateResponse("expert_dashboard.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "profile": profile,
        "stats": stats,
        "service_requests": service_requests,
        "notifications": notifications,
        "unread_notifications": unread_notifications,
        "selected_font": get_active_font()
    })


@app.get("/expert/opportunities", response_class=HTMLResponse)
@require_expert_login
@require_expert_profile_complete
async def expert_opportunities(request: Request, db: Session = Depends(get_db)):
    """Browse tender opportunities for experts."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Get expert profile for matching
    profile = db.query(ExpertProfileDB).filter(
        ExpertProfileDB.expert_id == current_expert.id
    ).first()

    # Get search and filter parameters
    search = request.query_params.get('search', '')
    sector = request.query_params.get('sector', '')
    page = int(request.query_params.get('page', 1))
    per_page = 20

    # Only show opportunities that have been specifically requested for this expert
    # Check for service requests that include tender opportunities
    service_requests = db.query(ExpertServiceRequestDB).filter(
        ExpertServiceRequestDB.expert_id == current_expert.id,
        ExpertServiceRequestDB.status.in_(['pending', 'accepted'])
    ).all()
    
    # Get tender IDs from service requests
    requested_tender_ids = {req.tender_id for req in service_requests if req.tender_id}
    
    # If no service requests, return empty list
    if not requested_tender_ids:
        open_request_count = db.query(ExpertHiringRequestDB).filter(
            ExpertHiringRequestDB.status == 'open'
        ).count()
        return templates.TemplateResponse("expert_opportunities.html", {
            "request": request,
            "current_expert": current_expert,
            "expert": current_expert,
            "profile": profile,
            "tenders": [],
            "total_tenders": 0,
            "page": 1,
            "total_pages": 0,
            "has_prev": False,
            "has_next": False,
            "search": search,
            "selected_sector": sector,
            "sectors": [],
            "favorite_tender_ids": set(),
            "applied_tender_ids": set(),
            "open_request_count": open_request_count,
            "selected_font": get_active_font()
        })

    # Determine if there are open hiring requests to highlight in Expert-Verse
    open_request_count = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.status == 'open'
    ).count()

    # Build query for active tenders that have been requested for this expert
    now = datetime.utcnow()
    query = db.query(TenderDB).filter(
        TenderDB.id.in_(requested_tender_ids),
        or_(
            TenderDB.deadline > now,
            TenderDB.deadline == None
        )
    )

    # Apply search filter
    if search:
        query = query.filter(
            or_(
                TenderDB.title.ilike(f'%{search}%'),
                TenderDB.authority.ilike(f'%{search}%'),
                TenderDB.category.ilike(f'%{search}%')
            )
        )

    # Apply sector filter
    if sector:
        query = query.filter(TenderDB.category == sector)

    # Get total count
    total_tenders = query.count()

    # Get paginated tenders
    tenders = query.order_by(TenderDB.scraped_at.desc()).offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    # Check which tenders the expert has favorited or applied to
    expert_favorites = db.query(ExpertFavoriteTenderDB).filter(
        ExpertFavoriteTenderDB.expert_id == current_expert.id
    ).all()
    favorite_tender_ids = {fav.tender_id for fav in expert_favorites}

    expert_applications = db.query(ExpertApplicationDB).filter(
        ExpertApplicationDB.expert_id == current_expert.id
    ).all()
    applied_tender_ids = {app.tender_id for app in expert_applications}

    # Get unique sectors for filter (only from requested tenders)
    sectors = db.query(TenderDB.category).filter(
        TenderDB.id.in_(requested_tender_ids),
        or_(
            TenderDB.deadline > now,
            TenderDB.deadline == None
        )
    ).distinct().all()
    sector_list = sorted([s[0] for s in sectors if s[0]])

    # Calculate pagination
    total_pages = (total_tenders + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages

    return templates.TemplateResponse("expert_opportunities.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "profile": profile,
        "tenders": tenders,
        "favorite_tender_ids": favorite_tender_ids,
        "applied_tender_ids": applied_tender_ids,
        "sectors": sector_list,
        "search": search,
        "selected_sector": sector,
        "page": page,
        "total_pages": total_pages,
        "has_prev": has_prev,
        "has_next": has_next,
        "total_tenders": total_tenders,
        "open_request_count": open_request_count,
        "selected_font": get_active_font()
    })

@app.get("/api/expert/opportunities/requests")
@require_expert_login
async def get_expert_hiring_requests(
    request: Request,
    db: Session = Depends(get_db),
    search: str = None,
    location: str = None,
    sector: str = None,
    budget_min: float = None,
    budget_max: float = None,
    page: int = 1,
    per_page: int = 20
):
    """Get open expert hiring requests with filtering options."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Build query for open requests
    query = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.status == 'open'
    )
    
    # Apply search filter
    if search:
        search_term = f'%{search}%'
        query = query.filter(
            or_(
                ExpertHiringRequestDB.request_name.ilike(search_term),
                ExpertHiringRequestDB.description.ilike(search_term),
                ExpertHiringRequestDB.company_name.ilike(search_term)
            )
        )
    
    # Apply location filter
    if location:
        query = query.filter(
            ExpertHiringRequestDB.company_location.ilike(f'%{location}%')
        )
    
    # Apply sector filter
    if sector:
        query = query.filter(
            ExpertHiringRequestDB.tender_sector.ilike(f'%{sector}%')
        )
    
    # Apply budget filter
    if budget_min is not None or budget_max is not None:
        budget_conditions = []
        if budget_min is not None:
            # Request budget should be >= budget_min
            budget_conditions.append(
                or_(
                    ExpertHiringRequestDB.budget_amount >= budget_min,
                    ExpertHiringRequestDB.budget_max >= budget_min,
                    and_(
                        ExpertHiringRequestDB.budget_type == 'negotiable',
                        ExpertHiringRequestDB.budget_min >= budget_min
                    )
                )
            )
        if budget_max is not None:
            # Request budget should be <= budget_max
            budget_conditions.append(
                or_(
                    ExpertHiringRequestDB.budget_amount <= budget_max,
                    ExpertHiringRequestDB.budget_min <= budget_max,
                    and_(
                        ExpertHiringRequestDB.budget_type == 'negotiable',
                        ExpertHiringRequestDB.budget_max <= budget_max
                    )
                )
            )
        if budget_conditions:
            query = query.filter(and_(*budget_conditions))
    
    # Get total count
    total_requests = query.count()
    
    # Get paginated requests
    requests = query.order_by(ExpertHiringRequestDB.created_at.desc()).offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    # Prefetch related tenders for richer detail views
    tender_details_map: Dict[str, Dict[str, Any]] = {}
    tender_ids = [req.tender_id for req in requests if req.tender_id]
    if tender_ids:
        tender_records = db.query(TenderDB).filter(TenderDB.id.in_(tender_ids)).all()
        for tender in tender_records:
            tender_details_map[tender.id] = {
                "title": tender.title,
                "authority": tender.authority,
                "state": tender.state,
                "category": tender.category,
                "estimated_value": tender.estimated_value,
                "currency": tender.currency,
                "deadline": tender.deadline.isoformat() if tender.deadline else None,
                "published_at": tender.published_at.isoformat() if tender.published_at else None,
                "summary": tender.summary,
                "tender_reference_number": tender.tender_reference_number,
                "tender_id": tender.tender_id,
                "organisation_chain": tender.organisation_chain,
                "tender_type": tender.tender_type,
                "form_of_contract": tender.form_of_contract,
                "tender_fee_details": tender.tender_fee_details,
                "emd_fee_details": tender.emd_fee_details,
                "work_item_details": tender.work_item_details,
                "critical_dates": tender.critical_dates,
                "tender_documents": tender.tender_documents,
                "tender_inviting_authority": tender.tender_inviting_authority,
                "source_url": tender.source_url,
                "pdf_url": tender.pdf_url
            }
    
    # Check which requests the expert has already applied to
    expert_applications = db.query(ExpertHiringApplicationDB).filter(
        ExpertHiringApplicationDB.expert_id == current_expert.id
    ).all()
    applied_request_ids = {app.request_id for app in expert_applications}
    
    # Format response
    requests_data = []
    for req in requests:
        # Get application count
        application_count = db.query(ExpertHiringApplicationDB).filter(
            ExpertHiringApplicationDB.request_id == req.id
        ).count()
        
        # Format budget display
        if req.budget_type == 'fixed':
            budget_display = f"₹{req.budget_amount:,.0f}" if req.budget_amount else "Not specified"
        else:
            if req.budget_min and req.budget_max:
                budget_display = f"₹{req.budget_min:,.0f} - ₹{req.budget_max:,.0f}"
            elif req.budget_min:
                budget_display = f"₹{req.budget_min:,.0f}+"
            else:
                budget_display = "Negotiable"
        
        requests_data.append({
            "id": req.id,
            "request_name": req.request_name,
            "description": req.description,
            "company_name": req.company_name,
            "company_location": req.company_location,
            "tender_title": req.tender_title,
            "tender_sector": req.tender_sector,
            "tender_state": req.tender_state,  # Legacy field
            "tender_location": getattr(req, 'tender_location', None) or req.tender_state,  # Use location if available, fallback to state
            "budget_type": req.budget_type,
            "budget_display": budget_display,
            "budget_amount": req.budget_amount,
            "budget_min": req.budget_min,
            "budget_max": req.budget_max,
            "project_id": req.project_id,
            "tender_id": req.tender_id,
            "created_at": req.created_at.isoformat() if req.created_at else None,
            "tender_details": tender_details_map.get(req.tender_id) if req.tender_id else None,
            "application_count": application_count,
            "has_applied": req.id in applied_request_ids
        })
    
    # Get unique locations and sectors for filters
    all_requests = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.status == 'open'
    ).all()
    
    locations = sorted(list(set([
        req.company_location for req in all_requests 
        if req.company_location
    ])))
    
    sectors = sorted(list(set([
        req.tender_sector for req in all_requests 
        if req.tender_sector
    ])))
    
    total_pages = (total_requests + per_page - 1) // per_page
    
    return {
        "requests": requests_data,
        "total_requests": total_requests,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "filter_options": {
            "locations": locations,
            "sectors": sectors
        }
    }

@app.post("/api/expert/request/{request_id}/apply")
@require_expert_login
async def apply_to_hiring_request(
    request: Request,
    request_id: str,
    db: Session = Depends(get_db)
):
    """Expert applies to a hiring request."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Get the hiring request
    hiring_request = db.query(ExpertHiringRequestDB).filter(
        ExpertHiringRequestDB.id == request_id,
        ExpertHiringRequestDB.status == 'open'
    ).first()
    
    if not hiring_request:
        raise HTTPException(status_code=404, detail="Hiring request not found or closed")
    
    # Check if already applied
    existing_application = db.query(ExpertHiringApplicationDB).filter(
        ExpertHiringApplicationDB.request_id == request_id,
        ExpertHiringApplicationDB.expert_id == current_expert.id
    ).first()
    
    if existing_application:
        raise HTTPException(status_code=400, detail="You have already applied to this request")
    
    # Get application data
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON data")
    
    def safe_strip(value):
        if value is None:
            return ""
        return str(value).strip()

    cover_letter = safe_strip(data.get("cover_letter", ""))
    proposed_rate = data.get("proposed_rate")
    estimated_timeline = safe_strip(data.get("estimated_timeline", ""))
    relevant_experience = safe_strip(data.get("relevant_experience", ""))
    
    # Validation
    if not cover_letter:
        raise HTTPException(status_code=400, detail="Cover letter is required")
    
    # Create application
    application = ExpertHiringApplicationDB(
        request_id=request_id,
        expert_id=current_expert.id,
        cover_letter=cover_letter,
        proposed_rate=proposed_rate if proposed_rate else None,
        estimated_timeline=estimated_timeline if estimated_timeline else None,
        relevant_experience=relevant_experience if relevant_experience else None,
        status='pending'
    )
    
    db.add(application)
    db.commit()
    db.refresh(application)
    
    return {
        "success": True,
        "message": "Application submitted successfully",
        "application_id": application.id
    }

@app.get("/expert/wall", response_class=HTMLResponse)
@require_expert_login
@require_expert_profile_complete
async def expert_wall(request: Request, db: Session = Depends(get_db)):
    """Expert Wall - publish and view articles, insights, case studies."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Get expert profile
    profile = db.query(ExpertProfileDB).filter(
        ExpertProfileDB.expert_id == current_expert.id
    ).first()

    # Get filter parameters
    content_type = request.query_params.get('type', 'all')
    page = int(request.query_params.get('page', 1))
    per_page = 15

    # Build query for published content
    query = db.query(ExpertContentDB).filter(ExpertContentDB.status == 'published')

    # Apply content type filter
    if content_type != 'all':
        query = query.filter(ExpertContentDB.content_type == content_type)

    # Get total count
    total_content = query.count()

    # Get paginated content with expert details
    content_items = query.options(
        joinedload(ExpertContentDB.expert)
    ).order_by(ExpertContentDB.published_at.desc()).offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    # Get like counts and user's likes
    for item in content_items:
        item.like_count = db.query(ExpertContentLikeDB).filter(
            ExpertContentLikeDB.content_id == item.id
        ).count()
        item.comment_count = db.query(ExpertContentCommentDB).filter(
            ExpertContentCommentDB.content_id == item.id
        ).count()
        item.is_liked = db.query(ExpertContentLikeDB).filter(
            ExpertContentLikeDB.content_id == item.id,
            ExpertContentLikeDB.expert_id == current_expert.id
        ).first() is not None

    # Get expert's own content
    my_content = db.query(ExpertContentDB).filter(
        ExpertContentDB.expert_id == current_expert.id
    ).order_by(ExpertContentDB.created_at.desc()).limit(5).all()

    # Calculate pagination
    total_pages = (total_content + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages

    return templates.TemplateResponse("expert_wall.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "profile": profile,
        "content_items": content_items,
        "my_content": my_content,
        "content_type": content_type,
        "page": page,
        "total_pages": total_pages,
        "has_prev": has_prev,
        "has_next": has_next,
        "total_content": total_content,
        "selected_font": get_active_font()
    })


@app.get("/expert/profile", response_class=HTMLResponse)
@require_expert_login
async def expert_profile_edit(request: Request, db: Session = Depends(get_db)):
    """Edit expert profile page."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Get existing profile data
    profile = db.query(ExpertProfileDB).filter(
        ExpertProfileDB.expert_id == current_expert.id
    ).first()

    return templates.TemplateResponse("expert_profile_edit.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "profile": profile,
        "selected_font": get_active_font()
    })


@app.get("/expert/requests", response_class=HTMLResponse)
@require_expert_login
@require_expert_profile_complete
async def expert_all_requests(request: Request, db: Session = Depends(get_db)):
    """View all service requests for the expert."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Get filter parameters
    status_filter = request.query_params.get('status', 'all')
    page = int(request.query_params.get('page', 1))
    per_page = 20

    # Build query
    query = db.query(ExpertServiceRequestDB).filter(
        ExpertServiceRequestDB.expert_id == current_expert.id
    )

    # Apply status filter
    if status_filter != 'all':
        query = query.filter(ExpertServiceRequestDB.status == status_filter)

    # Get total count
    total_requests = query.count()

    # Get paginated requests
    service_requests = query.order_by(
        ExpertServiceRequestDB.created_at.desc()
    ).offset((page - 1) * per_page).limit(per_page).all()

    # Fetch company and tender details for service requests
    for request_item in service_requests:
        # Get company details
        company = db.query(CompanyDB).filter(
            CompanyDB.user_id == request_item.company_id
        ).first()
        request_item.company_name = company.company_name if company else "Unknown Company"

        # Get tender details
        tender = db.query(TenderDB).filter(
            TenderDB.id == request_item.tender_id
        ).first()
        request_item.tender_title = tender.title if tender else "Unknown Tender"

    # Calculate pagination
    total_pages = (total_requests + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages

    return templates.TemplateResponse("expert_requests.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "service_requests": service_requests,
        "status_filter": status_filter,
        "page": page,
        "total_pages": total_pages,
        "has_prev": has_prev,
        "has_next": has_next,
        "total_requests": total_requests,
        "selected_font": get_active_font()
    })


@app.get("/expert/request/{request_id}", response_class=HTMLResponse)
@require_expert_login
@require_expert_profile_complete
async def expert_request_detail(request: Request, request_id: str, db: Session = Depends(get_db)):
    """View detailed service request."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Get the service request
    service_request = db.query(ExpertServiceRequestDB).filter(
        ExpertServiceRequestDB.id == request_id,
        ExpertServiceRequestDB.expert_id == current_expert.id
    ).first()

    if not service_request:
        raise HTTPException(status_code=404, detail="Service request not found")

    # Get company details
    company = db.query(CompanyDB).filter(
        CompanyDB.user_id == service_request.company_id
    ).first()

    # Get tender details
    tender = db.query(TenderDB).filter(
        TenderDB.id == service_request.tender_id
    ).first()

    return templates.TemplateResponse("expert_request_detail.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "service_request": service_request,
        "company": company,
        "tender": tender,
        "selected_font": get_active_font()
    })


@app.get("/expert/tender/{tender_id}", response_class=HTMLResponse)
@require_expert_login
@require_expert_profile_complete
async def expert_tender_detail(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """View tender details for experts."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Get tender
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if expert has favorited or applied to this tender
    is_favorited = db.query(ExpertFavoriteTenderDB).filter(
        ExpertFavoriteTenderDB.expert_id == current_expert.id,
        ExpertFavoriteTenderDB.tender_id == tender_id
    ).first() is not None

    has_applied = db.query(ExpertApplicationDB).filter(
        ExpertApplicationDB.expert_id == current_expert.id,
        ExpertApplicationDB.tender_id == tender_id
    ).first() is not None

    return templates.TemplateResponse("expert_tender_detail.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "tender": tender,
        "is_favorited": is_favorited,
        "has_applied": has_applied,
        "selected_font": get_active_font()
    })


@app.post("/expert/apply/{tender_id}")
@require_expert_login
@require_expert_profile_complete
async def expert_apply_to_tender(
    request: Request,
    tender_id: str,
    proposal: str = Form(...),
    proposed_fee: float = Form(...),
    estimated_duration: str = Form(...),
    db: Session = Depends(get_db)
):
    """Apply to a tender as an expert."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if tender exists
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if already applied
    existing_application = db.query(ExpertApplicationDB).filter(
        ExpertApplicationDB.expert_id == current_expert.id,
        ExpertApplicationDB.tender_id == tender_id
    ).first()

    if existing_application:
        return JSONResponse(
            status_code=400,
            content={"detail": "You have already applied to this tender"}
        )

    # Create new application
    application = ExpertApplicationDB(
        id=str(uuid.uuid4()),
        expert_id=current_expert.id,
        tender_id=tender_id,
        proposal=proposal,
        proposed_fee=proposed_fee,
        estimated_duration=estimated_duration,
        status='pending',
        created_at=datetime.utcnow()
    )

    db.add(application)
    db.commit()
    db.refresh(application)

    logger.info(f"Expert {current_expert.email} applied to tender {tender_id}")

    return JSONResponse(
        content={
            "success": True,
            "message": "Application submitted successfully",
            "redirect": f"/expert/tender/{tender_id}"
        }
    )


@app.get("/expert/collaborations", response_class=HTMLResponse)
@require_expert_login
@require_expert_profile_complete
async def expert_collaborations(request: Request, db: Session = Depends(get_db)):
    """View and manage collaborations with other experts."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Get active collaborations
    active_collaborations = db.query(ExpertCollaborationDB).filter(
        or_(
            ExpertCollaborationDB.primary_expert_id == current_expert.id,
            ExpertCollaborationDB.collaborator_expert_id == current_expert.id
        ),
        ExpertCollaborationDB.status == 'active'
    ).all()

    # Get pending collaboration requests
    pending_collaborations = db.query(ExpertCollaborationDB).filter(
        ExpertCollaborationDB.collaborator_expert_id == current_expert.id,
        ExpertCollaborationDB.status == 'pending'
    ).all()

    # Get completed collaborations
    completed_collaborations = db.query(ExpertCollaborationDB).filter(
        or_(
            ExpertCollaborationDB.primary_expert_id == current_expert.id,
            ExpertCollaborationDB.collaborator_expert_id == current_expert.id
        ),
        ExpertCollaborationDB.status == 'completed'
    ).order_by(ExpertCollaborationDB.invited_at.desc()).limit(10).all()

    return templates.TemplateResponse("expert_collaborations.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "active_collaborations": active_collaborations,
        "pending_collaborations": pending_collaborations,
        "completed_collaborations": completed_collaborations,
        "selected_font": get_active_font()
    })


@app.get("/expert/projects", response_class=HTMLResponse)
@require_expert_login
@require_expert_profile_complete
async def expert_projects_page(request: Request, db: Session = Depends(get_db)):
    """Show current projects (accepted hiring requests and collaborations) for the expert."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        return RedirectResponse(url="/expert/login", status_code=302)

    # Accepted hiring-request projects
    hiring_projects = []
    accepted_apps = db.query(
        ExpertHiringApplicationDB,
        ExpertHiringRequestDB,
        ProjectDB,
        TenderDB
    ).join(
        ExpertHiringRequestDB, ExpertHiringApplicationDB.request_id == ExpertHiringRequestDB.id
    ).outerjoin(
        ProjectDB, ExpertHiringRequestDB.project_id == ProjectDB.id
    ).outerjoin(
        TenderDB, ExpertHiringRequestDB.tender_id == TenderDB.id
    ).filter(
        ExpertHiringApplicationDB.expert_id == current_expert.id,
        ExpertHiringApplicationDB.status == 'accepted'
    ).all()

    for app_entry, hiring_request, project, tender in accepted_apps:
        other_experts = db.query(ExpertHiringApplicationDB).filter(
            ExpertHiringApplicationDB.request_id == hiring_request.id,
            ExpertHiringApplicationDB.status == 'accepted',
            ExpertHiringApplicationDB.expert_id != current_expert.id
        ).count()

        hiring_projects.append({
            "id": hiring_request.id,
            "type": "hiring",
            "channel_type": "hiring",
            "channel_id": hiring_request.id,
            "title": hiring_request.request_name,
            "company": hiring_request.company_name,
            "location": hiring_request.company_location,
            "sector": hiring_request.tender_sector,
            "status": hiring_request.status,
            "project_name": project.project_name if project else None,
            "tender_title": tender.title if tender else hiring_request.tender_title,
            "tender_id": tender.id if tender else hiring_request.tender_id,
            "participants": other_experts + 1,
            "budget_display": hiring_request.budget_type == 'fixed' and hiring_request.budget_amount,
            "detail_url": f"/expert/projects/hiring/{hiring_request.id}"
        })

    # Active collaborations
    collaboration_projects = []
    collabs = db.query(ExpertCollaborationDB, TenderDB).outerjoin(
        TenderDB, ExpertCollaborationDB.tender_id == TenderDB.id
    ).filter(
        or_(
            ExpertCollaborationDB.primary_expert_id == current_expert.id,
            ExpertCollaborationDB.collaborator_expert_id == current_expert.id
        ),
        ExpertCollaborationDB.status == 'active'
    ).all()

    for collab, tender in collabs:
        other_expert_id = collab.collaborator_expert_id if collab.primary_expert_id == current_expert.id else collab.primary_expert_id
        other_expert = db.query(ExpertDB).filter(ExpertDB.id == other_expert_id).first()

        collaboration_projects.append({
            "id": collab.id,
            "type": "collaboration",
            "channel_type": "collaboration",
            "channel_id": collab.id,
            "title": tender.title if tender else "Active Collaboration",
            "tender_id": tender.id if tender else collab.tender_id,
            "tender_state": tender.state if tender else None,
            "role": collab.role,
            "participants": 2,
            "other_expert": other_expert.name if other_expert else "Expert teammate",
            "detail_url": f"/expert/projects/collaboration/{collab.id}"
        })

    projects = hiring_projects + collaboration_projects

    return templates.TemplateResponse("expert_projects.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "projects": projects,
        "selected_font": get_active_font()
    })


def _fetch_project_channel(db: Session, channel_type: str, channel_id: str):
    channel_type = channel_type.lower()
    if channel_type == "hiring":
        hiring_request = db.query(ExpertHiringRequestDB).filter(
            ExpertHiringRequestDB.id == channel_id
        ).first()
        if not hiring_request:
            raise HTTPException(status_code=404, detail="Hiring request not found")
        tender = db.query(TenderDB).filter(TenderDB.id == hiring_request.tender_id).first() if hiring_request.tender_id else None
        project = db.query(ProjectDB).filter(ProjectDB.id == hiring_request.project_id).first() if hiring_request.project_id else None
        accepted_experts = [
            app.expert_id for app in db.query(ExpertHiringApplicationDB).filter(
                ExpertHiringApplicationDB.request_id == channel_id,
                ExpertHiringApplicationDB.status == 'accepted'
            ).all()
        ]
        return {
            "kind": "hiring",
            "channel_type": "hiring",
            "channel_id": channel_id,
            "hiring_request": hiring_request,
            "tender": tender,
            "project": project,
            "accepted_experts": accepted_experts,
            "owner_user_id": hiring_request.company_id
        }
    elif channel_type == "collaboration":
        collaboration = db.query(ExpertCollaborationDB).filter(
            ExpertCollaborationDB.id == channel_id
        ).first()
        if not collaboration:
            raise HTTPException(status_code=404, detail="Collaboration not found")
        tender = db.query(TenderDB).filter(TenderDB.id == collaboration.tender_id).first() if collaboration.tender_id else None
        return {
            "kind": "collaboration",
            "channel_type": "collaboration",
            "channel_id": channel_id,
            "collaboration": collaboration,
            "tender": tender,
            "participants": [collaboration.primary_expert_id, collaboration.collaborator_expert_id]
        }
    else:
        raise HTTPException(status_code=400, detail="Invalid project type")

def _expert_has_project_access(channel: Dict[str, Any], expert_id: str) -> bool:
    if channel["kind"] == "hiring":
        return expert_id in channel["accepted_experts"]
    elif channel["kind"] == "collaboration":
        collab: ExpertCollaborationDB = channel["collaboration"]
        return collab.status == 'active' and expert_id in channel["participants"]
    return False

def _manager_has_project_access(channel: Dict[str, Any], user_id: str) -> bool:
    if channel["kind"] == "hiring":
        return channel.get("owner_user_id") == user_id
    return False

def _project_is_active(channel: Dict[str, Any]) -> bool:
    if channel["kind"] == "hiring":
        return channel["hiring_request"].status in ["open", "filled"]
    elif channel["kind"] == "collaboration":
        return channel["collaboration"].status == 'active'
    return False

def _project_participants(channel: Dict[str, Any], db: Session) -> List[Dict[str, str]]:
    participants = []
    if channel["kind"] == "hiring":
        applications = db.query(ExpertHiringApplicationDB).filter(
            ExpertHiringApplicationDB.request_id == channel["hiring_request"].id,
            ExpertHiringApplicationDB.status == 'accepted'
        ).all()
        for app in applications:
            expert = db.query(ExpertDB).filter(ExpertDB.id == app.expert_id).first()
            if expert:
                participants.append({"id": expert.id, "name": expert.name})
    elif channel["kind"] == "collaboration":
        collab: ExpertCollaborationDB = channel["collaboration"]
        for expert_id in channel["participants"]:
            expert = db.query(ExpertDB).filter(ExpertDB.id == expert_id).first()
            if expert:
                participants.append({"id": expert.id, "name": expert.name})
    return participants

@app.get("/expert/projects/{channel_type}/{channel_id}", response_class=HTMLResponse)
async def expert_project_detail(
    request: Request,
    channel_type: str,
    channel_id: str,
    db: Session = Depends(get_db)
):
    """Project workspace for experts (and managers)."""
    current_expert = get_current_expert(request, db)
    current_user = get_current_user(request, db)
    if not current_expert and not current_user:
        return RedirectResponse(url="/expert/login", status_code=302)

    channel = _fetch_project_channel(db, channel_type, channel_id)
    expert_allowed = bool(current_expert and _expert_has_project_access(channel, current_expert.id))
    manager_allowed = bool(current_user and _manager_has_project_access(channel, current_user.id))
    if not (expert_allowed or manager_allowed):
        raise HTTPException(status_code=403, detail="You do not have access to this project")

    project_active = _project_is_active(channel)

    if channel["kind"] == "hiring":
        hr: ExpertHiringRequestDB = channel["hiring_request"]
        tender = channel.get("tender")
        project = channel.get("project")
        project_info = {
            "type": "Hiring Request",
            "title": hr.request_name,
            "company": hr.company_name,
            "location": hr.company_location,
            "sector": hr.tender_sector,
            "status": hr.status,
            "tender_title": tender.title if tender else hr.tender_title,
            "tender_id": tender.id if tender else hr.tender_id,
            "project_name": project.project_name if project else None,
            "budget_type": hr.budget_type,
            "budget_amount": hr.budget_amount,
            "deadline": tender.deadline if tender else None
        }
    else:
        collab: ExpertCollaborationDB = channel["collaboration"]
        tender = channel.get("tender")
        project_info = {
            "type": "Collaboration",
            "title": tender.title if tender else "Collaboration Workspace",
            "company": None,
            "location": tender.state if tender else None,
            "sector": tender.category if tender else None,
            "status": collab.status,
            "tender_title": tender.title if tender else None,
            "tender_id": tender.id if tender else collab.tender_id,
            "project_name": None,
            "budget_type": None,
            "budget_amount": None,
            "deadline": tender.deadline if tender else None
        }

    participants = _project_participants(channel, db)

    return templates.TemplateResponse("expert_project_detail.html", {
        "request": request,
        "current_expert": current_expert,
        "expert": current_expert,
        "current_user": current_user,
        "project": project_info,
        "channel_type": channel["channel_type"],
        "channel_id": channel["channel_id"],
        "participants": participants,
        "can_manage_tasks": manager_allowed,
        "can_close_project": expert_allowed and project_active,
        "project_active": project_active,
        "selected_font": get_active_font()
    })

@app.get("/api/expert/projects/{channel_type}/{channel_id}/messages")
async def get_expert_project_messages(
    request: Request,
    channel_type: str,
    channel_id: str,
    db: Session = Depends(get_db)
):
    """Return chat history for a project."""
    current_expert = get_current_expert(request, db)
    current_user = get_current_user(request, db)
    if not current_expert and not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    channel = _fetch_project_channel(db, channel_type, channel_id)
    expert_allowed = bool(current_expert and _expert_has_project_access(channel, current_expert.id))
    manager_allowed = bool(current_user and _manager_has_project_access(channel, current_user.id))
    if not (expert_allowed or manager_allowed):
        raise HTTPException(status_code=403, detail="Not authorized to view this chat")

    messages = db.query(ExpertProjectMessageDB).filter(
        ExpertProjectMessageDB.channel_type == channel["channel_type"],
        ExpertProjectMessageDB.channel_id == channel["channel_id"]
    ).order_by(ExpertProjectMessageDB.created_at.asc()).all()

    formatted = []
    for message in messages:
        sender_name = None
        sender_id = None
        if message.sender_expert:
            sender_name = message.sender_expert.name
            sender_id = message.sender_expert_id
        elif message.sender_user:
            sender_name = message.sender_user.name
            sender_id = message.sender_user_id
        formatted.append({
            "id": message.id,
            "sender": sender_name or "Team",
            "sender_id": sender_id,
            "message": message.message,
            "created_at": message.created_at.isoformat()
        })

    return {"messages": formatted}

@app.post("/api/expert/projects/{channel_type}/{channel_id}/messages")
async def send_expert_project_message(
    request: Request,
    channel_type: str,
    channel_id: str,
    db: Session = Depends(get_db)
):
    """Send a chat message in a project workspace."""
    current_expert = get_current_expert(request, db)
    current_user = get_current_user(request, db)
    if not current_expert and not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    channel = _fetch_project_channel(db, channel_type, channel_id)
    expert_allowed = bool(current_expert and _expert_has_project_access(channel, current_expert.id))
    manager_allowed = bool(current_user and _manager_has_project_access(channel, current_user.id))
    if not (expert_allowed or manager_allowed):
        raise HTTPException(status_code=403, detail="Not authorized to message in this project")

    payload = await request.json()
    message_text = (payload.get("message") or "").strip()
    if not message_text:
        raise HTTPException(status_code=400, detail="Message cannot be empty")

    new_message = ExpertProjectMessageDB(
        channel_type=channel["channel_type"],
        channel_id=channel["channel_id"],
        hiring_request_id=channel["hiring_request"].id if channel["kind"] == "hiring" else None,
        collaboration_id=channel["collaboration"].id if channel["kind"] == "collaboration" else None,
        project_id=channel.get("project").id if channel.get("project") else None,
        tender_id=channel.get("tender").id if channel.get("tender") else None,
        sender_expert_id=current_expert.id if expert_allowed and current_expert else None,
        sender_user_id=current_user.id if manager_allowed and current_user else None,
        message=message_text
    )

    db.add(new_message)
    db.commit()
    db.refresh(new_message)

    sender_name = current_expert.name if expert_allowed and current_expert else current_user.name if current_user else "Team"

    return {
        "message": {
            "id": new_message.id,
            "sender": sender_name,
            "sender_id": current_expert.id if expert_allowed and current_expert else current_user.id if current_user else None,
            "message": new_message.message,
            "created_at": new_message.created_at.isoformat()
        }
    }

@app.get("/api/expert/projects/{channel_type}/{channel_id}/tasks")
async def get_project_tasks(
    request: Request,
    channel_type: str,
    channel_id: str,
    db: Session = Depends(get_db)
):
    current_expert = get_current_expert(request, db)
    current_user = get_current_user(request, db)
    if not current_expert and not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    channel = _fetch_project_channel(db, channel_type, channel_id)
    expert_allowed = bool(current_expert and _expert_has_project_access(channel, current_expert.id))
    manager_allowed = bool(current_user and _manager_has_project_access(channel, current_user.id))
    if not (expert_allowed or manager_allowed):
        raise HTTPException(status_code=403, detail="Not authorized to view tasks")

    tasks = db.query(ExpertProjectTaskDB).filter(
        ExpertProjectTaskDB.channel_type == channel["channel_type"],
        ExpertProjectTaskDB.channel_id == channel["channel_id"]
    ).order_by(ExpertProjectTaskDB.created_at.desc()).all()

    results = []
    for task in tasks:
        results.append({
            "id": task.id,
            "title": task.title,
            "description": task.description,
            "priority": task.priority,
            "status": task.status,
            "deadline": task.deadline.isoformat() if task.deadline else None,
            "assignee_expert_id": task.assignee_expert_id,
            "assignee_name": task.assignee_expert.name if task.assignee_expert else None,
            "created_at": task.created_at.isoformat() if task.created_at else None
        })

    return {
        "tasks": results,
        "can_manage": manager_allowed
    }

@app.post("/api/expert/projects/{channel_type}/{channel_id}/tasks")
async def create_project_task(
    request: Request,
    channel_type: str,
    channel_id: str,
    db: Session = Depends(get_db)
):
    current_user = get_current_user(request, db)
    current_expert = get_current_expert(request, db)
    if not current_user and not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")

    channel = _fetch_project_channel(db, channel_type, channel_id)
    manager_allowed = bool(current_user and _manager_has_project_access(channel, current_user.id))
    if not manager_allowed:
        raise HTTPException(status_code=403, detail="Only project owners can add tasks")

    payload = await request.json()
    title = (payload.get("title") or "").strip()
    description = (payload.get("description") or "").strip()
    priority = payload.get("priority", "medium")
    deadline = payload.get("deadline")
    assignee_id = payload.get("assignee_expert_id")

    if not title:
        raise HTTPException(status_code=400, detail="Task title is required")

    deadline_dt = None
    if deadline:
        try:
            deadline_dt = datetime.fromisoformat(deadline)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid deadline format")

    task = ExpertProjectTaskDB(
        channel_type=channel["channel_type"],
        channel_id=channel["channel_id"],
        hiring_request_id=channel["hiring_request"].id if channel["kind"] == "hiring" else None,
        collaboration_id=channel["collaboration"].id if channel["kind"] == "collaboration" else None,
        title=title,
        description=description,
        priority=priority,
        deadline=deadline_dt,
        created_by_user_id=current_user.id if current_user else None,
        assignee_expert_id=assignee_id if assignee_id else None
    )

    db.add(task)
    db.commit()
    db.refresh(task)

    return {
        "task": {
            "id": task.id,
            "title": task.title,
            "description": task.description,
            "priority": task.priority,
            "status": task.status,
            "deadline": task.deadline.isoformat() if task.deadline else None,
            "assignee_expert_id": task.assignee_expert_id,
            "assignee_name": task.assignee_expert.name if task.assignee_expert else None
        }
    }

@app.post("/api/expert/projects/{channel_type}/{channel_id}/tasks/{task_id}/status")
async def update_project_task_status(
    request: Request,
    channel_type: str,
    channel_id: str,
    task_id: str,
    db: Session = Depends(get_db)
):
    current_expert = get_current_expert(request, db)
    current_user = get_current_user(request, db)
    if not current_expert and not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    channel = _fetch_project_channel(db, channel_type, channel_id)
    expert_allowed = bool(current_expert and _expert_has_project_access(channel, current_expert.id))
    manager_allowed = bool(current_user and _manager_has_project_access(channel, current_user.id))
    if not (expert_allowed or manager_allowed):
        raise HTTPException(status_code=403, detail="Not authorized to update tasks")

    task = db.query(ExpertProjectTaskDB).filter(
        ExpertProjectTaskDB.id == task_id,
        ExpertProjectTaskDB.channel_type == channel["channel_type"],
        ExpertProjectTaskDB.channel_id == channel["channel_id"]
    ).first()

    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    payload = await request.json()
    new_status = payload.get("status", "").lower()
    if new_status not in ["pending", "in_progress", "completed"]:
        raise HTTPException(status_code=400, detail="Invalid status value")

    task.status = new_status
    db.commit()
    db.refresh(task)

    return {"success": True, "status": task.status}

@app.post("/api/expert/projects/{channel_type}/{channel_id}/close")
async def close_expert_project(
    request: Request,
    channel_type: str,
    channel_id: str,
    db: Session = Depends(get_db)
):
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")

    channel = _fetch_project_channel(db, channel_type, channel_id)
    if not _expert_has_project_access(channel, current_expert.id):
        raise HTTPException(status_code=403, detail="You are not an active participant")

    if channel["kind"] == "hiring":
        app = db.query(ExpertHiringApplicationDB).filter(
            ExpertHiringApplicationDB.request_id == channel["hiring_request"].id,
            ExpertHiringApplicationDB.expert_id == current_expert.id,
            ExpertHiringApplicationDB.status == 'accepted'
        ).first()
        if not app:
            raise HTTPException(status_code=404, detail="Active participation not found")
        app.status = 'completed'
        app.reviewed_at = datetime.utcnow()
        remaining = db.query(ExpertHiringApplicationDB).filter(
            ExpertHiringApplicationDB.request_id == channel["hiring_request"].id,
            ExpertHiringApplicationDB.status == 'accepted'
        ).count()
        if remaining == 0:
            channel["hiring_request"].status = 'closed'
            channel["hiring_request"].closed_at = datetime.utcnow()
    else:
        collab: ExpertCollaborationDB = channel["collaboration"]
        collab.status = 'completed'
        collab.responded_at = datetime.utcnow()

    db.commit()

    return {"success": True, "message": "Project closed"}


# ==================== Expert Project Task Files ====================

@app.post("/api/expert/projects/{channel_type}/{channel_id}/tasks/{task_id}/files")
async def upload_expert_task_file(
    request: Request,
    channel_type: str,
    channel_id: str,
    task_id: str,
    file: UploadFile = File(...),
    description: str = Form(None),
    db: Session = Depends(get_db)
):
    """Upload a file/deliverable for an expert project task."""
    current_expert = get_current_expert(request, db)
    current_user = get_current_user(request, db)
    
    if not current_expert and not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    channel = _fetch_project_channel(db, channel_type, channel_id)
    expert_allowed = bool(current_expert and _expert_has_project_access(channel, current_expert.id))
    manager_allowed = bool(current_user and _manager_has_project_access(channel, current_user.id))
    
    if not (expert_allowed or manager_allowed):
        raise HTTPException(status_code=403, detail="Not authorized to upload files")
    
    # Verify task exists
    task = db.query(ExpertProjectTaskDB).filter(
        ExpertProjectTaskDB.id == task_id,
        ExpertProjectTaskDB.channel_type == channel["channel_type"],
        ExpertProjectTaskDB.channel_id == channel["channel_id"]
    ).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    # Validate file size (max 50MB)
    file_size = 0
    file_content = await file.read()
    file_size = len(file_content)
    
    if file_size > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size exceeds 50MB limit")
    
    # Create uploads directory if it doesn't exist
    upload_dir = "static/uploads/expert_project_files"
    os.makedirs(upload_dir, exist_ok=True)
    
    # Generate unique filename
    file_ext = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(upload_dir, unique_filename)
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(file_content)
    
    # Create database record
    file_record = ExpertProjectTaskFileDB(
        task_id=task_id,
        channel_type=channel["channel_type"],
        channel_id=channel["channel_id"],
        filename=file.filename,
        file_path=file_path,
        file_size=file_size,
        file_type=file.content_type,
        description=description,
        uploaded_by_expert_id=current_expert.id if expert_allowed and current_expert else None,
        uploaded_by_user_id=current_user.id if manager_allowed and current_user else None
    )
    
    db.add(file_record)
    db.commit()
    db.refresh(file_record)
    
    return {
        "message": "File uploaded successfully",
        "file": {
            "id": file_record.id,
            "filename": file_record.filename,
            "file_size": file_record.file_size,
            "created_at": file_record.created_at.isoformat()
        }
    }


@app.get("/api/expert/projects/{channel_type}/{channel_id}/tasks/{task_id}/files/{file_id}/download")
async def download_expert_task_file(
    request: Request,
    channel_type: str,
    channel_id: str,
    task_id: str,
    file_id: str,
    db: Session = Depends(get_db)
):
    """Download a file from an expert project task."""
    current_expert = get_current_expert(request, db)
    current_user = get_current_user(request, db)
    
    if not current_expert and not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    channel = _fetch_project_channel(db, channel_type, channel_id)
    expert_allowed = bool(current_expert and _expert_has_project_access(channel, current_expert.id))
    manager_allowed = bool(current_user and _manager_has_project_access(channel, current_user.id))
    
    if not (expert_allowed or manager_allowed):
        raise HTTPException(status_code=403, detail="Not authorized to download files")
    
    # Get file record
    file_record = db.query(ExpertProjectTaskFileDB).filter(
        ExpertProjectTaskFileDB.id == file_id,
        ExpertProjectTaskFileDB.task_id == task_id
    ).first()
    
    if not file_record:
        raise HTTPException(status_code=404, detail="File not found")
    
    if not os.path.exists(file_record.file_path):
        raise HTTPException(status_code=404, detail="File not found on server")
    
    return FileResponse(
        path=file_record.file_path,
        filename=file_record.filename,
        media_type=file_record.file_type or 'application/octet-stream'
    )


# ==================== Expert Project Task Progress Updates ====================

@app.post("/api/expert/projects/{channel_type}/{channel_id}/tasks/{task_id}/progress")
async def add_expert_task_progress_update(
    request: Request,
    channel_type: str,
    channel_id: str,
    task_id: str,
    db: Session = Depends(get_db)
):
    """Add a progress update to an expert project task."""
    current_expert = get_current_expert(request, db)
    current_user = get_current_user(request, db)
    
    if not current_expert and not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    channel = _fetch_project_channel(db, channel_type, channel_id)
    expert_allowed = bool(current_expert and _expert_has_project_access(channel, current_expert.id))
    manager_allowed = bool(current_user and _manager_has_project_access(channel, current_user.id))
    
    if not (expert_allowed or manager_allowed):
        raise HTTPException(status_code=403, detail="Not authorized to add progress updates")
    
    # Verify task exists
    task = db.query(ExpertProjectTaskDB).filter(
        ExpertProjectTaskDB.id == task_id,
        ExpertProjectTaskDB.channel_type == channel["channel_type"],
        ExpertProjectTaskDB.channel_id == channel["channel_id"]
    ).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    payload = await request.json()
    update_text = (payload.get("update_text") or "").strip()
    
    if not update_text:
        raise HTTPException(status_code=400, detail="Update text is required")
    
    if len(update_text) > 2000:
        raise HTTPException(status_code=400, detail="Update text is too long (max 2000 characters)")
    
    # Create progress update
    progress_update = ExpertProjectTaskProgressUpdateDB(
        task_id=task_id,
        channel_type=channel["channel_type"],
        channel_id=channel["channel_id"],
        update_text=update_text,
        expert_id=current_expert.id if expert_allowed and current_expert else None,
        user_id=current_user.id if manager_allowed and current_user else None
    )
    
    db.add(progress_update)
    db.commit()
    db.refresh(progress_update)
    
    # Format response
    author_name = None
    if progress_update.expert:
        author_name = progress_update.expert.name
    elif progress_update.user:
        author_name = progress_update.user.name
    
    return {
        "message": "Progress update added successfully",
        "update": {
            "id": progress_update.id,
            "task_id": progress_update.task_id,
            "update_text": progress_update.update_text,
            "author_name": author_name,
            "expert_id": progress_update.expert_id,
            "user_id": progress_update.user_id,
            "created_at": progress_update.created_at.isoformat()
        }
    }


@app.delete("/api/expert/projects/{channel_type}/{channel_id}/tasks/{task_id}/progress/{progress_id}")
async def delete_expert_task_progress_update(
    request: Request,
    channel_type: str,
    channel_id: str,
    task_id: str,
    progress_id: str,
    db: Session = Depends(get_db)
):
    """Delete a progress update from an expert project task."""
    current_expert = get_current_expert(request, db)
    current_user = get_current_user(request, db)
    
    if not current_expert and not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Get progress update
    progress_update = db.query(ExpertProjectTaskProgressUpdateDB).filter(
        ExpertProjectTaskProgressUpdateDB.id == progress_id,
        ExpertProjectTaskProgressUpdateDB.task_id == task_id
    ).first()
    
    if not progress_update:
        raise HTTPException(status_code=404, detail="Progress update not found")
    
    # Check authorization - only the author can delete their own updates
    is_author = (current_expert and progress_update.expert_id == current_expert.id) or \
                (current_user and progress_update.user_id == current_user.id)
    
    if not is_author:
        raise HTTPException(status_code=403, detail="You can only delete your own progress updates")
    
    db.delete(progress_update)
    db.commit()
    
    return {"message": "Progress update deleted successfully"}


# ==================== Expert Project Task Queries ====================

@app.post("/api/expert/projects/{channel_type}/{channel_id}/tasks/{task_id}/queries")
async def raise_expert_task_query(
    request: Request,
    channel_type: str,
    channel_id: str,
    task_id: str,
    db: Session = Depends(get_db)
):
    """Raise a query/concern for an expert project task."""
    current_expert = get_current_expert(request, db)
    
    if not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    channel = _fetch_project_channel(db, channel_type, channel_id)
    
    if not _expert_has_project_access(channel, current_expert.id):
        raise HTTPException(status_code=403, detail="Not authorized to raise queries")
    
    # Verify task exists
    task = db.query(ExpertProjectTaskDB).filter(
        ExpertProjectTaskDB.id == task_id,
        ExpertProjectTaskDB.channel_type == channel["channel_type"],
        ExpertProjectTaskDB.channel_id == channel["channel_id"]
    ).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    payload = await request.json()
    query_type = payload.get("query_type", "").strip()
    title = payload.get("title", "").strip()
    description = payload.get("description", "").strip()
    priority = payload.get("priority", "medium")
    
    if not query_type or query_type not in ["clarification", "timeline", "resources", "blocker", "other"]:
        raise HTTPException(status_code=400, detail="Invalid query type")
    
    if not title:
        raise HTTPException(status_code=400, detail="Query title is required")
    
    if not description:
        raise HTTPException(status_code=400, detail="Query description is required")
    
    # Create query
    query = ExpertProjectTaskQueryDB(
        task_id=task_id,
        channel_type=channel["channel_type"],
        channel_id=channel["channel_id"],
        query_type=query_type,
        title=title,
        description=description,
        priority=priority,
        raised_by_expert_id=current_expert.id
    )
    
    db.add(query)
    db.commit()
    db.refresh(query)
    
    return {
        "message": "Query submitted successfully",
        "query": {
            "id": query.id,
            "task_id": query.task_id,
            "query_type": query.query_type,
            "title": query.title,
            "description": query.description,
            "priority": query.priority,
            "status": query.status,
            "created_at": query.created_at.isoformat()
        }
    }


@app.post("/api/expert/favorite/{tender_id}")
@require_expert_login
async def toggle_expert_favorite_tender(
    request: Request,
    tender_id: str,
    db: Session = Depends(get_db)
):
    """Toggle favorite status for a tender."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if already favorited
    favorite = db.query(ExpertFavoriteTenderDB).filter(
        ExpertFavoriteTenderDB.expert_id == current_expert.id,
        ExpertFavoriteTenderDB.tender_id == tender_id
    ).first()

    if favorite:
        # Remove favorite
        db.delete(favorite)
        db.commit()
        return JSONResponse(content={"success": True, "action": "removed", "is_favorited": False})
    else:
        # Add favorite
        new_favorite = ExpertFavoriteTenderDB(
            id=str(uuid.uuid4()),
            expert_id=current_expert.id,
            tender_id=tender_id,
            created_at=datetime.utcnow()
        )
        db.add(new_favorite)
        db.commit()
        return JSONResponse(content={"success": True, "action": "added", "is_favorited": True})


@app.post("/api/expert/content/publish")
@require_expert_login
@require_expert_profile_complete
async def publish_expert_content(
    request: Request,
    title: str = Form(...),
    content_type: str = Form(...),
    content: str = Form(...),
    tags: str = Form(""),
    db: Session = Depends(get_db)
):
    """Publish new content on Expert Wall."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Parse tags
    tag_list = [tag.strip() for tag in tags.split(',') if tag.strip()] if tags else []

    slug = generate_expert_content_slug(db, title)

    # Create new content
    new_content = ExpertContentDB(
        id=str(uuid.uuid4()),
        expert_id=current_expert.id,
        title=title,
        slug=slug,
        content_type=content_type,
        content=content,
        tags=tag_list,
        status='published',
        published_at=datetime.utcnow(),
        created_at=datetime.utcnow()
    )

    db.add(new_content)
    db.commit()
    db.refresh(new_content)

    logger.info(f"Expert {current_expert.email} published new content: {title}")

    return JSONResponse(
        content={
            "success": True,
            "message": "Content published successfully",
            "redirect": "/expert/wall"
        }
    )


@app.post("/api/expert/content/{content_id}/like")
@require_expert_login
async def toggle_content_like(
    request: Request,
    content_id: str,
    db: Session = Depends(get_db)
):
    """Toggle like on Expert Wall content."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if already liked
    like = db.query(ExpertContentLikeDB).filter(
        ExpertContentLikeDB.expert_id == current_expert.id,
        ExpertContentLikeDB.content_id == content_id
    ).first()

    if like:
        # Remove like
        db.delete(like)
        db.commit()
        action = "unliked"
    else:
        # Add like
        new_like = ExpertContentLikeDB(
            id=str(uuid.uuid4()),
            expert_id=current_expert.id,
            content_id=content_id,
            created_at=datetime.utcnow()
        )
        db.add(new_like)
        db.commit()
        action = "liked"

    # Get updated like count
    like_count = db.query(ExpertContentLikeDB).filter(
        ExpertContentLikeDB.content_id == content_id
    ).count()

    return JSONResponse(content={
        "success": True,
        "action": action,
        "like_count": like_count
    })


# Company Details API endpoint
@app.post("/api/company-details")
async def save_company_details(
    request: Request,
    company_name: str = Form(...),
    registration_number: str = Form(...),
    gst_number: str = Form(...),
    pan_number: str = Form(...),
    industry_sector: str = Form(...),
    year_established: int = Form(...),
    employee_count: str = Form(...),
    registered_address: str = Form(...),
    operational_address: str = Form(""),
    phone_number: str = Form(...),
    email_address: str = Form(...),
    website_url: str = Form(""),
    key_services: str = Form(...),
    specialization_areas: str = Form(...),
    previous_govt_experience: str = Form(""),
    managing_director: str = Form(...),
    technical_head: str = Form(""),
    compliance_officer: str = Form(""),
    company_type: str = Form(""),
    msme_registration: str = Form(""),
    udyam_registration: str = Form(""),
    tan_number: str = Form(""),
    esi_registration: str = Form(""),
    pf_registration: str = Form(""),
    labour_license_number: str = Form(""),
    turnover_year_labels: List[str] = Form([]),
    turnover_year_values: List[str] = Form([]),
    auditor_name: str = Form(""),
    credit_rating: str = Form(""),
    net_worth: str = Form(""),
    credit_facility: str = Form(""),
    bank_name: Optional[str] = Form(None),
    account_number: Optional[str] = Form(None),
    ifsc_code: Optional[str] = Form(None),
    account_holder_name: Optional[str] = Form(None),
    certification_names: List[str] = Form([]),
    remove_certification_ids: List[str] = Form([]),
    certification_files: List[UploadFile] = File([]),
    db: Session = Depends(get_db)
):
    """Save or update company details."""

    def ensure_list(value):
        if value is None:
            return []
        if isinstance(value, list):
            return value
        return [value]

    current_user = get_current_user(request, db)
    if not current_user:
        return JSONResponse(
            status_code=401,
            content={"error": "Authentication required"}
        )

    turnover_year_labels = ensure_list(turnover_year_labels)
    turnover_year_values = ensure_list(turnover_year_values)
    certification_names = ensure_list(certification_names)
    remove_certification_ids = [int(c_id) for c_id in ensure_list(remove_certification_ids) if str(c_id).strip()]
    certification_files = ensure_list(certification_files)

    # Debug: Log raw form data received
    logger.info(f"Raw turnover_year_labels received: {turnover_year_labels}")
    logger.info(f"Raw turnover_year_values received: {turnover_year_values}")

    # Normalize industry sector JSON
    try:
        if industry_sector.startswith('[') and industry_sector.endswith(']'):
            json.loads(industry_sector)
    except (json.JSONDecodeError, AttributeError):
        industry_sector = json.dumps([industry_sector]) if industry_sector else json.dumps([])

    # Build structured payloads
    legal_details = {
        "company_type": company_type.strip(),
        "msme_registration": msme_registration.strip(),
        "udyam_registration": udyam_registration.strip(),
        "tan_number": tan_number.strip(),
        "esi_registration": esi_registration.strip(),
        "pf_registration": pf_registration.strip(),
        "labour_license_number": labour_license_number.strip(),
    }
    legal_details = {k: v for k, v in legal_details.items() if v}

    turnover_history = []
    for index, (label, amount) in enumerate(zip(turnover_year_labels, turnover_year_values)):
        label = str(label).strip()
        amount_raw = str(amount).strip()
        if not label and not amount_raw:
            continue

        if not label:
            label = f"Year {index + 1}"

        numeric_amount: Optional[float] = None
        if amount_raw:
            try:
                numeric_amount = float(amount_raw.replace(',', '').replace('₹', '').strip())
            except ValueError:
                numeric_amount = None

        turnover_history.append({
            "label": label,
            "raw_amount": amount_raw,
            "amount": numeric_amount
        })

    financial_details = {
        "turnover_history": turnover_history,
        "auditor_name": auditor_name.strip() or None,
        "credit_rating": credit_rating.strip() or None,
        "net_worth": net_worth.strip() or None,
        "credit_facility": credit_facility.strip() or None,
    }
    financial_details = {k: v for k, v in financial_details.items() if v}

    # Determine annual turnover placeholder from turnover history
    annual_turnover_value = "N/A"
    if turnover_history:
        first_entry = turnover_history[0]
        annual_turnover_value = first_entry.get("raw_amount") or (
            str(first_entry.get("amount")) if first_entry.get("amount") is not None else "N/A"
        )

    company = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()
    is_new_company = company is None

    if is_new_company:
        company = CompanyDB(
            user_id=current_user.id,
            created_at=datetime.utcnow()
        )
        db.add(company)

    # Update general details
    company.company_name = company_name
    company.registration_number = registration_number
    company.gst_number = gst_number
    company.pan_number = pan_number
    company.industry_sector = industry_sector
    company.year_established = year_established
    company.annual_turnover = annual_turnover_value
    company.employee_count = employee_count
    company.registered_address = registered_address
    company.operational_address = operational_address
    company.phone_number = phone_number
    company.email_address = email_address
    company.website_url = website_url
    company.key_services = key_services
    company.specialization_areas = specialization_areas
    company.previous_govt_experience = previous_govt_experience
    company.certifications = None
    company.managing_director = managing_director
    company.technical_head = technical_head
    company.compliance_officer = compliance_officer
    company.is_complete = True
    company.updated_at = datetime.utcnow()

    # Banking details (retain existing columns)
    if bank_name is not None:
        company.bank_name = bank_name
    if account_number is not None:
        company.account_number = account_number
    if ifsc_code is not None:
        company.ifsc_code = ifsc_code
    if account_holder_name is not None:
        company.account_holder_name = account_holder_name

    # Structured sections
    company.legal_details = legal_details
    company.financial_details = financial_details

    # Debug: Log what's being saved
    logger.info(f"Saving financial_details for user {current_user.id}: {financial_details}")
    logger.info(f"Turnover history records: {len(turnover_history)}")

    # Handle certification removals
    if remove_certification_ids and company.company_certifications:
        for cert in list(company.company_certifications):
            if cert.id in remove_certification_ids:
                try:
                    file_path = Path(cert.file_path)
                    if file_path.exists():
                        file_path.unlink()
                except Exception as exc:
                    logger.warning(f"Failed to delete certification file {cert.file_path}: {exc}")
                db.delete(cert)

    # Handle new certifications
    upload_root = COMPANY_CERTIFICATIONS_DIR / str(current_user.id)
    upload_root.mkdir(parents=True, exist_ok=True)

    for idx, name in enumerate(certification_names):
        name = str(name).strip()
        file_obj = certification_files[idx] if idx < len(certification_files) else None
        has_file = file_obj and getattr(file_obj, "filename", "")

        if not name and not has_file:
            continue

        if not has_file:
            logger.warning(f"Skipping certification '{name}' because no file was provided.")
            continue

        original_filename = Path(file_obj.filename).name  # type: ignore
        safe_original = "".join(ch for ch in original_filename if ch.isalnum() or ch in (" ", ".", "_", "-")).strip()
        if not safe_original:
            safe_original = f"certificate_{uuid.uuid4().hex}.pdf"

        unique_filename = f"{uuid.uuid4().hex}_{safe_original}"
        destination = upload_root / unique_filename

        try:
            content = await file_obj.read()  # type: ignore
            with open(destination, "wb") as out_file:
                out_file.write(content)

            new_cert = CompanyCertificateDB(
                company=company,
                name=name or safe_original,
                file_path=str(destination),
                uploaded_at=datetime.utcnow()
            )
            db.add(new_cert)
        except Exception as exc:
            logger.error(f"Failed to process certification upload '{original_filename}': {exc}")

    db.commit()
    db.refresh(company)

    return JSONResponse(
        status_code=200,
        content={"message": "Company details saved successfully", "redirect": "/dashboard"}
    )


@app.get("/company-certifications/{cert_id}/download")
async def download_company_certification(cert_id: int, request: Request, db: Session = Depends(get_db)):
    """Download a company certification document."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    certification = db.query(CompanyCertificateDB).join(CompanyDB).filter(
        CompanyCertificateDB.id == cert_id,
        CompanyDB.user_id == current_user.id
    ).first()

    if not certification:
        raise HTTPException(status_code=404, detail="Certification not found")

    file_path = Path(certification.file_path)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Certification file not found")

    return FileResponse(str(file_path), filename=file_path.name)

# Employee Management API endpoints
@app.post("/api/employees")
async def add_employee(
    request: Request,
    name: str = Form(""),
    email: str = Form(...),
    role: str = Form(""),
    team: str = Form(""),
    password: str = Form(...),
    is_bd: bool = Form(False),
    profile_picture: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    """Add a new employee to the organization."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user's company
    company = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()
    if not company:
        raise HTTPException(status_code=400, detail="Please complete company details first")

    # Get or create company code for this user
    company_code = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).first()
    if not company_code:
        # Create a company code based on company name
        code = company.company_name[:3].upper() + secrets.token_hex(3).upper()
        company_code = CompanyCodeDB(
            user_id=current_user.id,
            company_name=company.company_name,
            company_code=code,
            created_at=datetime.utcnow()
        )
        db.add(company_code)
        db.commit()
        db.refresh(company_code)

    # Check if employee email already exists
    existing_employee = db.query(EmployeeDB).filter(EmployeeDB.email == email).first()
    if existing_employee:
        raise HTTPException(status_code=400, detail="Employee with this email already exists")

    # Handle profile picture upload
    profile_picture_path = None
    if profile_picture and profile_picture.filename:
        # Create uploads directory if it doesn't exist
        uploads_dir = Path("static/uploads/employee_profiles")
        uploads_dir.mkdir(parents=True, exist_ok=True)

        # Generate unique filename
        file_extension = profile_picture.filename.split('.')[-1]
        unique_filename = f"{uuid.uuid4()}.{file_extension}"
        file_path = uploads_dir / unique_filename

        # Save the file
        with open(file_path, "wb") as buffer:
            content = await profile_picture.read()
            buffer.write(content)

        profile_picture_path = f"/static/uploads/employee_profiles/{unique_filename}"

    # Create employee with the provided password
    employee_id = str(uuid.uuid4())

    employee = EmployeeDB(
        id=employee_id,
        email=email,
        name=name or email.split('@')[0],
        company_code_id=company_code.id,
        password_hash=hash_password(password),
        role=role,
        team=team,
        profile_picture=profile_picture_path,
        is_bd=is_bd,
        is_active=True,
        created_at=datetime.utcnow()
    )

    db.add(employee)
    db.commit()
    db.refresh(employee)

    return JSONResponse(
        status_code=200,
        content={
            "message": "Employee added successfully",
            "employee_id": employee.id,
            "password": password  # Return the password so admin knows what it is
        }
    )

@app.get("/api/employees")
async def get_employees(
    request: Request,
    db: Session = Depends(get_db)
):
    """Get all employees for the current user's organization."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user's company code
    company_code = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).first()
    if not company_code:
        return JSONResponse(status_code=200, content=[])

    # Get all employees for this company
    employees = db.query(EmployeeDB).filter(
        EmployeeDB.company_code_id == company_code.id
    ).all()

    # Convert to list of dicts
    employee_list = []
    for emp in employees:
        employee_list.append({
            "id": emp.id,
            "name": emp.name,
            "email": emp.email,
            "role": emp.role,
            "team": emp.team,
            "profile_picture": emp.profile_picture,
            "is_active": emp.is_active,
            "created_at": emp.created_at.isoformat() if emp.created_at else None
        })

    return JSONResponse(status_code=200, content=employee_list)

@app.delete("/api/employees/{employee_id}")
async def delete_employee(
    request: Request,
    employee_id: str,
    db: Session = Depends(get_db)
):
    """Delete an employee from the organization."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user's company code
    company_code = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).first()
    if not company_code:
        raise HTTPException(status_code=404, detail="Company not found")

    # Find the employee and verify they belong to this company
    employee = db.query(EmployeeDB).filter(
        EmployeeDB.id == employee_id,
        EmployeeDB.company_code_id == company_code.id
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found or does not belong to your organization")

    # Delete related records first to avoid foreign key constraint errors

    # Delete tender assignments
    db.query(TenderAssignmentDB).filter(TenderAssignmentDB.employee_id == employee_id).delete()

    # Delete task comments
    db.query(TaskCommentDB).filter(TaskCommentDB.employee_id == employee_id).delete()

    # Delete tender messages
    db.query(TenderMessageDB).filter(TenderMessageDB.employee_id == employee_id).delete()

    # Delete tasks where employee is the completed_by
    db.query(TaskDB).filter(TaskDB.completed_by == employee_id).update({"completed_by": None})

    # Now delete the employee
    db.delete(employee)
    db.commit()

    return JSONResponse(status_code=200, content={"message": "Employee deleted successfully"})

# Favorites API endpoints
@app.post("/api/favorites/{tender_id}")
@require_company_details
async def add_favorite(
    request: Request,
    tender_id: str,
    db: Session = Depends(get_db)
):
    """Add tender to favorites."""
    from core.dependencies import get_id_for_tender_management, get_user_id_for_queries

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user_id for database storage (company owner's user_id for BD employees)
    user_id_for_storage, _ = get_user_id_for_queries(request, db)
    if not user_id_for_storage:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if tender exists
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if already favorited by this entity
    # For BD employees, check by worked_by fields; for users, check by user_id
    if entity_type == 'bd_employee':
        existing_favorite = db.query(FavoriteDB).filter(
            and_(
                FavoriteDB.user_id == user_id_for_storage,
                FavoriteDB.tender_id == tender_id,
                FavoriteDB.worked_by_type == 'bd_employee',
                FavoriteDB.worked_by_name == entity.name
            )
        ).first()
    else:
        existing_favorite = db.query(FavoriteDB).filter(
            and_(FavoriteDB.user_id == user_id_for_storage, FavoriteDB.tender_id == tender_id)
        ).first()

    if existing_favorite:
        raise HTTPException(status_code=400, detail="Already favorited")

    # Get notes from request body if present (optional - supports JSON or form data)
    notes = ""
    try:
        content_type = request.headers.get("content-type", "").lower()
        if "application/json" in content_type:
            body = await request.json()
            notes = body.get("notes", "")
        elif "multipart/form-data" in content_type or "application/x-www-form-urlencoded" in content_type:
            # Only try to parse form data if content-type indicates it
            # Skip if it's an empty FormData to avoid multipart boundary errors
            if request.headers.get("content-length", "0") != "0":
                form_data = await request.form()
                notes = form_data.get("notes", "")
    except Exception:
        # If parsing fails, default to empty string
        notes = ""

    # Create favorite - use user_id_for_storage (company owner's user_id for BD employees)
    # but track who actually favorited it using worked_by fields
    favorite = FavoriteDB(
        user_id=user_id_for_storage,
        tender_id=tender_id,
        notes=notes,
        worked_by_name=entity.name,
        worked_by_type=entity_type,
        created_at=datetime.utcnow()
    )

    db.add(favorite)
    db.commit()

    # Create deadline notifications for this tender (use entity_id for notifications)
    create_deadline_notifications(entity_id, tender_id, db)

    return {"message": "Added to favorites"}

@app.delete("/api/favorites/{tender_id}")
async def remove_favorite(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Remove tender from favorites."""
    from core.dependencies import get_id_for_tender_management, get_user_id_for_queries

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user_id for database queries (company owner's user_id for BD employees)
    user_id_for_query, _ = get_user_id_for_queries(request, db)
    if not user_id_for_query:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find and delete favorite - for BD employees, check by worked_by fields
    if entity_type == 'bd_employee':
        favorite = db.query(FavoriteDB).filter(
            and_(
                FavoriteDB.user_id == user_id_for_query,
                FavoriteDB.tender_id == tender_id,
                FavoriteDB.worked_by_type == 'bd_employee',
                FavoriteDB.worked_by_name == entity.name
            )
        ).first()
    else:
        favorite = db.query(FavoriteDB).filter(
            and_(FavoriteDB.user_id == user_id_for_query, FavoriteDB.tender_id == tender_id)
        ).first()

    if not favorite:
        raise HTTPException(status_code=404, detail="Favorite not found")

    db.delete(favorite)

    # Delete associated notifications (use entity_id for notifications)
    db.query(NotificationDB).filter(
        NotificationDB.user_id == entity_id,
        NotificationDB.tender_id == tender_id
    ).delete()

    db.commit()

    return {"message": "Removed from favorites"}

@app.put("/api/favorites/{favorite_id}")
async def update_favorite_data(
    request: Request,
    favorite_id: int,
    user_filled_data: str = Form(...),
    db: Session = Depends(get_db)
):
    """Update user-filled data for a favorite."""
    from core.dependencies import get_id_for_tender_management, get_user_id_for_queries

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user_id for database queries (company owner's user_id for BD employees)
    user_id_for_query, _ = get_user_id_for_queries(request, db)
    if not user_id_for_query:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the favorite - for BD employees, check by worked_by fields
    if entity_type == 'bd_employee':
        favorite = db.query(FavoriteDB).filter(
            and_(
                FavoriteDB.id == favorite_id,
                FavoriteDB.user_id == user_id_for_query,
                FavoriteDB.worked_by_type == 'bd_employee',
                FavoriteDB.worked_by_name == entity.name
            )
        ).first()
    else:
        favorite = db.query(FavoriteDB).filter(
            and_(FavoriteDB.id == favorite_id, FavoriteDB.user_id == user_id_for_query)
        ).first()

    if not favorite:
        raise HTTPException(status_code=404, detail="Favorite not found")

    try:
        # Parse the JSON data
        filled_data = json.loads(user_filled_data)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON data")

    # Update the favorite
    favorite.user_filled_data = filled_data  # type: ignore
    setattr(favorite, 'status', "submitted")
    db.commit()

    return {"message": "Favorite data updated successfully"}

@app.post("/api/favorites/{favorite_id}/update-field")
async def update_favorite_field(
    request: Request,
    favorite_id: int,
    field_name: str = Form(...),
    field_value: str = Form(...),
    db: Session = Depends(get_db)
):
    """Update a single user-filled field for a favorite with auto-save."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the favorite
    favorite = db.query(FavoriteDB).filter(
        and_(FavoriteDB.id == favorite_id, FavoriteDB.user_id == entity_id)
    ).first()

    if not favorite:
        raise HTTPException(status_code=404, detail="Favorite not found")

    # Get existing user_filled_data or initialize empty dict
    user_data = favorite.user_filled_data if favorite.user_filled_data else {}

    # Update the specific field
    user_data[field_name] = field_value

    # Save back to database - force SQLAlchemy to detect JSON change
    favorite.user_filled_data = user_data  # type: ignore
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(favorite, "user_filled_data")
    db.commit()
    db.refresh(favorite)

    return {"success": True, "message": "Field updated successfully", "field_name": field_name, "field_value": field_value}

@app.post("/api/favorites/{favorite_id}/submit")
async def submit_favorite(request: Request, favorite_id: int, db: Session = Depends(get_db)):
    """Submit a favorite (mark as completed)."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the favorite
    favorite = db.query(FavoriteDB).filter(
        and_(FavoriteDB.id == favorite_id, FavoriteDB.user_id == entity_id)
    ).first()

    if not favorite:
        raise HTTPException(status_code=404, detail="Favorite not found")

    setattr(favorite, 'status', "submitted")
    db.commit()

    return {"message": "Favorite submitted successfully"}

@app.post("/api/tenders/shortlist")
async def shortlist_tender(
    request: Request,
    tender_id: str = Form(...),
    reason: str = Form(...),
    db: Session = Depends(get_db)
):
    """Shortlist a tender with reason."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if tender exists
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if already shortlisted
    existing = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.user_id == entity_id,
            ShortlistedTenderDB.tender_id == tender_id
        )
    ).first()

    if existing:
        # Update reason if already exists
        existing.reason = reason
        existing.created_at = datetime.utcnow()
    else:
        # Create new shortlist entry
        shortlisted = ShortlistedTenderDB(
            user_id=entity_id,
            tender_id=tender_id,
            reason=reason,
            worked_by_name=entity.name,
            worked_by_type=entity_type
        )
        db.add(shortlisted)

    # Remove from favorites if it's there
    favorite = db.query(FavoriteDB).filter(
        and_(
            FavoriteDB.user_id == entity_id,
            FavoriteDB.tender_id == tender_id
        )
    ).first()

    if favorite:
        db.delete(favorite)

    db.commit()

    # Create deadline notifications for this tender
    create_deadline_notifications(entity_id, tender_id, db)

    return {"message": "Tender shortlisted successfully", "success": True}

@app.post("/api/tenders/reject")
async def reject_tender(
    request: Request,
    tender_id: str = Form(...),
    reason: str = Form(...),
    db: Session = Depends(get_db)
):
    """Reject a tender with reason."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if tender exists
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if already rejected
    existing = db.query(RejectedTenderDB).filter(
        and_(
            RejectedTenderDB.user_id == current_user.id,
            RejectedTenderDB.tender_id == tender_id
        )
    ).first()

    if existing:
        # Update reason if already exists
        existing.reason = reason
        existing.created_at = datetime.utcnow()
    else:
        # Create new reject entry
        rejected = RejectedTenderDB(
            user_id=current_user.id,
            tender_id=tender_id,
            reason=reason
        )
        db.add(rejected)

    # Remove from favorites if it's there
    favorite = db.query(FavoriteDB).filter(
        and_(
            FavoriteDB.user_id == current_user.id,
            FavoriteDB.tender_id == tender_id
        )
    ).first()

    if favorite:
        db.delete(favorite)

    # Delete associated notifications
    db.query(NotificationDB).filter(
        NotificationDB.user_id == current_user.id,
        NotificationDB.tender_id == tender_id
    ).delete()

    db.commit()

    return {"message": "Tender rejected successfully", "success": True}

@app.get("/api/tenders/shortlist/{shortlist_id}/progress")
async def get_shortlist_progress(request: Request, shortlist_id: int, db: Session = Depends(get_db)):
    """Get progress data for a shortlisted tender."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the shortlisted tender
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == entity_id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    return {
        "success": True,
        "progress_data": shortlisted.progress_data or {}
    }

@app.post("/api/tenders/shortlist/{shortlist_id}/progress")
async def update_shortlist_progress(
    request: Request,
    shortlist_id: int,
    db: Session = Depends(get_db)
):
    """Update progress data for a shortlisted tender."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Parse JSON body
    body = await request.json()
    progress_data = body.get('progress_data', {})
    is_awarded = body.get('is_awarded', False)

    # Find the shortlisted tender
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == entity_id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    if is_awarded:
        # Get the tender
        tender = db.query(TenderDB).filter(TenderDB.id == shortlisted.tender_id).first()
        if tender is None:
            raise HTTPException(status_code=404, detail="Tender not found")

        finalize_tender_award_for_user(db, tender, entity_id)
        db.commit()
        return {"success": True, "message": "Tender marked as awarded"}

    # Update progress data when not awarding yet
    shortlisted.progress_data = progress_data
    shortlisted.updated_at = datetime.utcnow()

    db.commit()

    return {"success": True, "message": "Progress updated successfully"}

@app.get("/api/tenders/shortlist/{shortlist_id}/progress-steps")
async def get_shortlist_progress_steps(request: Request, shortlist_id: int, db: Session = Depends(get_db)):
    """Get 6-step progress data for a shortlisted tender."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the shortlisted tender
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == entity_id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    progress_data = shortlisted.progress_data or {}
    step_timestamps = progress_data.get('step_timestamps', {})

    # Get the associated tender to extract deadline dates
    tender = db.query(TenderDB).filter(TenderDB.id == shortlisted.tender_id).first()

    # Extract automatic deadlines from TenderDB critical_dates JSONB
    step_deadlines = {
        "step1": None,  # Clarification End Date
        "step2": None,  # Bid Submission End Date
        "step3": None,  # Bid Opening Date
        "step4": progress_data.get('step4_deadline'),  # Manual deadline
        "step5": progress_data.get('step5_deadline'),  # Manual deadline
        "step6": None   # No deadline for this step
    }

    if tender:
        critical_dates = tender.critical_dates or {}

        # Step 1: Clarification End Date
        if 'Clarification End Date' in critical_dates:
            step_deadlines['step1'] = critical_dates['Clarification End Date']

        # Step 2: Bid Submission End Date (check both deadline column and critical_dates)
        if tender.deadline:
            step_deadlines['step2'] = tender.deadline.isoformat()
        elif 'Bid Submission End Date' in critical_dates:
            step_deadlines['step2'] = critical_dates['Bid Submission End Date']

        # Step 3: Bid Opening Date
        if 'Bid Opening Date' in critical_dates:
            step_deadlines['step3'] = critical_dates['Bid Opening Date']

    return {
        "success": True,
        "step1": progress_data.get('step1', ''),
        "step2": progress_data.get('step2', ''),
        "step3": progress_data.get('step3', ''),
        "step4": progress_data.get('step4', ''),
        "step5": progress_data.get('step5', ''),
        "step6": progress_data.get('step6', ''),
        "step_timestamps": {
            "step1": step_timestamps.get('step1'),
            "step2": step_timestamps.get('step2'),
            "step3": step_timestamps.get('step3'),
            "step4": step_timestamps.get('step4'),
            "step5": step_timestamps.get('step5'),
            "step6": step_timestamps.get('step6'),
        },
        "step_deadlines": step_deadlines,
        "last_updated": shortlisted.updated_at.isoformat() if shortlisted.updated_at else None
    }

@app.put("/api/tenders/shortlist/{shortlist_id}/progress-steps")
async def update_shortlist_progress_steps(
    request: Request,
    shortlist_id: int,
    db: Session = Depends(get_db)
):
    """Update 6-step progress data for a shortlisted tender."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Parse JSON body
    body = await request.json()

    # Find the shortlisted tender
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == entity_id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    # Get existing progress data
    progress_data = shortlisted.progress_data or {}

    # Update only the provided steps
    now_iso = datetime.utcnow().isoformat()
    step_timestamps = progress_data.get('step_timestamps', {})

    for key in ['step1', 'step2', 'step3', 'step4', 'step5', 'step6']:
        if key in body:
            progress_data[key] = body[key]
            step_timestamps[key] = now_iso

    def step_has_value(val) -> bool:
        return val is not None and str(val).strip() != ''

    for idx in range(2, 7):
        current_val = progress_data.get(f'step{idx}')
        previous_val = progress_data.get(f'step{idx - 1}')
        if step_has_value(current_val) and not step_has_value(previous_val):
            raise HTTPException(
                status_code=400,
                detail=f"Complete step {idx - 1} before updating step {idx}."
            )

    progress_data['step_timestamps'] = step_timestamps

    # Save progress data - force SQLAlchemy to detect JSON change
    shortlisted.progress_data = progress_data
    shortlisted.updated_at = datetime.utcnow()

    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(shortlisted, "progress_data")

    db.commit()
    db.refresh(shortlisted)

    # AUTO-TASK CREATION LOGIC
    REMOVAL_OPTIONS = [
        'Cancelled', 'Tender Cancelled', 'Opened & Not Qualified',
        'Opened & Lost', 'Opened & Not Eligible', 'No'
    ]

    # Get tracking data from progress_data
    progress_data = shortlisted.progress_data or {}
    tasks_created_for_stages = progress_data.get('tasks_created_for_stages', [])
    now = datetime.utcnow()

    for step_num in range(1, 7):
        step_key = f'step{step_num}'

        # Check if this step was updated in this request
        if step_key in body:
            step_value = body[step_key]

            # Create tasks if:
            # 1. Step has value
            # 2. Not a removal option
            # 3. Not already created for this stage
            if (step_value and
                str(step_value).strip() != '' and
                step_value not in REMOVAL_OPTIONS and
                step_num not in tasks_created_for_stages):

                # Create tasks for this stage
                tasks_count = auto_create_tasks_for_stage(
                    db=db,
                    shortlist_id=shortlist_id,
                    stage_number=step_num,
                    user_id=entity_id,
                    trigger_date=now
                )

                if tasks_count > 0:
                    # Mark this stage as having tasks created
                    tasks_created_for_stages.append(step_num)

                    # Update progress_data
                    progress_data['tasks_created_for_stages'] = tasks_created_for_stages
                    shortlisted.progress_data = progress_data

                    from sqlalchemy.orm.attributes import flag_modified
                    flag_modified(shortlisted, "progress_data")
                    db.commit()

    return {
        "success": True,
        "message": "Progress updated successfully",
        "last_updated": shortlisted.updated_at.isoformat() if shortlisted.updated_at else None,
        "step_timestamps": step_timestamps
    }

@app.put("/api/tenders/shortlist/{shortlist_id}/deadline")
async def update_manual_deadline(
    request: Request,
    shortlist_id: int,
    db: Session = Depends(get_db)
):
    """Update manual deadline for steps 4 or 5 only."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Parse JSON body
    body = await request.json()
    step = body.get('step')
    deadline = body.get('deadline')

    # Validate step number
    if step not in [4, 5]:
        raise HTTPException(status_code=400, detail="Manual deadlines only allowed for steps 4 and 5")

    # Validate deadline is provided
    if not deadline:
        raise HTTPException(status_code=400, detail="Deadline is required")

    # Validate date format
    try:
        datetime.strptime(deadline, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    # Get shortlist record
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == current_user.id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    # Update progress_data
    progress_data = shortlisted.progress_data or {}
    progress_data[f'step{step}_deadline'] = deadline
    shortlisted.progress_data = progress_data
    shortlisted.updated_at = datetime.utcnow()

    # Force SQLAlchemy to detect JSON change
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(shortlisted, "progress_data")

    db.commit()
    db.refresh(shortlisted)

    return {
        "success": True,
        "message": f"Deadline for step {step} updated successfully",
        "deadline": deadline
    }

@app.delete("/api/tenders/shortlist/{shortlist_id}/remove")
async def remove_shortlisted_tender(
    request: Request,
    shortlist_id: int,
    db: Session = Depends(get_db)
):
    """Remove a tender from shortlist (for removal options)."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Parse JSON body for removal reason
    body = await request.json()
    removal_reason = body.get('removal_reason', 'User removed')

    # Find the shortlisted tender
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == entity_id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    # Delete the shortlisted entry
    db.delete(shortlisted)

    # Delete associated deadline notifications
    db.query(NotificationDB).filter(
        NotificationDB.user_id == entity_id,
        NotificationDB.tender_id == shortlisted.tender_id,
        NotificationDB.notification_type.like('deadline_%')
    ).delete(synchronize_session=False)

    db.commit()

    return {"success": True, "message": f"Tender removed from shortlist: {removal_reason}"}

@app.post("/api/tenders/shortlist/{shortlist_id}/kill")
async def kill_shortlisted_tender(
    request: Request,
    shortlist_id: int,
    db: Session = Depends(get_db)
):
    """Kill a shortlisted tender and move it to dump area."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Parse JSON body
    body = await request.json()
    kill_reason = body.get('kill_reason', '')
    kill_stage = body.get('kill_stage', '')
    progress_data = body.get('progress_data', {})

    # Find the shortlisted tender
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == entity_id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    # Create dumped tender entry
    dumped = DumpedTenderDB(
        user_id=entity_id,
        tender_id=shortlisted.tender_id,
        shortlist_reason=shortlisted.reason,
        kill_reason=kill_reason,
        kill_stage=kill_stage,
        progress_data=progress_data,
        created_at=shortlisted.created_at,
        killed_at=datetime.utcnow()
    )
    db.add(dumped)

    # Remove from shortlisted tenders
    db.delete(shortlisted)

    # Delete associated notifications
    db.query(NotificationDB).filter(
        NotificationDB.user_id == entity_id,
        NotificationDB.tender_id == shortlisted.tender_id
    ).delete()

    db.commit()

    return {"success": True, "message": "Tender moved to dump area"}


# Stage Document Upload/Management Endpoints

@app.post("/api/tenders/shortlist/upload-document")
async def upload_stage_document(
    request: Request,
    file: UploadFile = File(...),
    step_number: int = Form(...),
    shortlist_id: int = Form(...),
    title: str = Form(...),
    notes: str = Form(""),
    db: Session = Depends(get_db)
):
    """Upload a document for a specific tender progress step."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Verify shortlist belongs to user
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == current_user.id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    # Validate file size (10MB max)
    content = await file.read()
    file_size = len(content)

    if file_size > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be less than 10MB")

    # Create upload directory if it doesn't exist
    upload_dir = Path("uploads/stage_documents")
    upload_dir.mkdir(parents=True, exist_ok=True)

    # Generate unique filename
    file_extension = Path(file.filename).suffix if file.filename else ''
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_path = upload_dir / unique_filename

    # Save file
    with open(file_path, "wb") as f:
        f.write(content)

    # Create database entry
    stage_doc = StageDocumentDB(
        shortlist_id=shortlist_id,
        step_number=step_number,
        title=title,
        filename=file.filename or unique_filename,
        file_path=str(file_path),
        file_size=file_size,
        notes=notes,
        uploaded_at=datetime.utcnow()
    )

    db.add(stage_doc)
    db.commit()
    db.refresh(stage_doc)

    return {"success": True, "message": "Document uploaded successfully", "document_id": stage_doc.id}


@app.get("/api/tenders/shortlist/{shortlist_id}/documents")
async def get_stage_documents(
    request: Request,
    shortlist_id: int,
    db: Session = Depends(get_db)
):
    """Get all documents for a shortlisted tender."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Verify shortlist belongs to user
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == current_user.id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    # Get all documents for this shortlist
    documents = db.query(StageDocumentDB).filter(
        StageDocumentDB.shortlist_id == shortlist_id
    ).order_by(StageDocumentDB.uploaded_at.desc()).all()

    # Format documents for response
    docs_list = []
    for doc in documents:
        docs_list.append({
            'id': doc.id,
            'step_number': doc.step_number,
            'title': doc.title,
            'filename': doc.filename,
            'file_size': doc.file_size,
            'notes': doc.notes,
            'uploaded_at': doc.uploaded_at.isoformat() if doc.uploaded_at else None
        })

    return {"success": True, "documents": docs_list}


@app.get("/api/tenders/shortlist/document/{document_id}/download")
async def download_stage_document(
    request: Request,
    document_id: int,
    db: Session = Depends(get_db)
):
    """Download a stage document."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get document
    document = db.query(StageDocumentDB).filter(
        StageDocumentDB.id == document_id
    ).first()

    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Verify access (document belongs to user's shortlist)
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == document.shortlist_id,
            ShortlistedTenderDB.user_id == current_user.id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=403, detail="Access denied")

    # Check if file exists
    file_path = Path(document.file_path)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on server")

    # Return file
    return FileResponse(
        path=str(file_path),
        filename=document.filename,
        media_type='application/octet-stream'
    )


@app.delete("/api/tenders/shortlist/document/{document_id}")
async def delete_stage_document(
    request: Request,
    document_id: int,
    db: Session = Depends(get_db)
):
    """Delete a stage document."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get document
    document = db.query(StageDocumentDB).filter(
        StageDocumentDB.id == document_id
    ).first()

    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Verify access (document belongs to user's shortlist)
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == document.shortlist_id,
            ShortlistedTenderDB.user_id == current_user.id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=403, detail="Access denied")

    # Delete file from filesystem
    file_path = Path(document.file_path)
    if file_path.exists():
        try:
            file_path.unlink()
        except Exception as e:
            logger.error(f"Error deleting file: {e}")

    # Delete database entry
    db.delete(document)
    db.commit()

    return {"success": True, "message": "Document deleted successfully"}


# Employee Assignment Endpoints

@app.get("/api/employees/list")
async def get_employees_list(
    request: Request,
    db: Session = Depends(get_db)
):
    """Get list of all employees for the current user's company."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user's company code
    company_code = db.query(CompanyCodeDB).filter(
        CompanyCodeDB.user_id == current_user.id
    ).first()

    if not company_code:
        return {"success": True, "employees": []}

    # Get all employees for this company
    employees = db.query(EmployeeDB).filter(
        and_(
            EmployeeDB.company_code_id == company_code.id,
            EmployeeDB.is_active == True
        )
    ).all()

    employees_list = [
        {
            'id': emp.id,
            'name': emp.name,
            'email': emp.email,
            'role': emp.role,
            'team': emp.team
        }
        for emp in employees
    ]

    return {"success": True, "employees": employees_list}


@app.post("/api/tenders/shortlist/{shortlist_id}/assign-employee")
async def assign_employee_to_step(
    request: Request,
    shortlist_id: int,
    body: dict,
    db: Session = Depends(get_db)
):
    """Assign an employee to a specific progress step."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Verify shortlist belongs to user
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == entity_id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    step_number = body.get('step_number')
    employee_id = body.get('employee_id')

    if step_number is None or employee_id is None:
        raise HTTPException(status_code=400, detail="step_number and employee_id are required")

    try:
        step_number_int = int(step_number)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="step_number must be between 1 and 6")

    if step_number_int < 1 or step_number_int > 6:
        raise HTTPException(status_code=400, detail="step_number must be between 1 and 6")

    # Get progress data
    progress_data = shortlisted.progress_data or {}

    previous_key = f'step{step_number_int - 1}'
    if step_number_int > 1:
        previous_value = progress_data.get(previous_key)
        if previous_value is None or str(previous_value).strip() == '':
            raise HTTPException(
                status_code=400,
                detail=f"Complete step {step_number_int - 1} before assigning employees to this step."
            )

    # Get or create employee assignments
    if 'employee_assignments' not in progress_data:
        progress_data['employee_assignments'] = {}

    step_key = f'step{step_number_int}'
    if step_key not in progress_data['employee_assignments']:
        progress_data['employee_assignments'][step_key] = []

    # Add employee if not already assigned
    if employee_id not in progress_data['employee_assignments'][step_key]:
        progress_data['employee_assignments'][step_key].append(employee_id)

    # Update database - force SQLAlchemy to detect JSON change
    shortlisted.progress_data = progress_data
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(shortlisted, "progress_data")
    db.commit()
    db.refresh(shortlisted)

    return {"success": True, "message": "Employee assigned successfully"}


@app.post("/api/tenders/shortlist/{shortlist_id}/unassign-employee")
async def unassign_employee_from_step(
    request: Request,
    shortlist_id: int,
    body: dict,
    db: Session = Depends(get_db)
):
    """Unassign an employee from a specific progress step."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Verify shortlist belongs to user
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == entity_id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    step_number = body.get('step_number')
    employee_id = body.get('employee_id')

    if step_number is None or employee_id is None:
        raise HTTPException(status_code=400, detail="step_number and employee_id are required")

    try:
        step_number_int = int(step_number)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="step_number must be between 1 and 6")

    if step_number_int < 1 or step_number_int > 6:
        raise HTTPException(status_code=400, detail="step_number must be between 1 and 6")

    # Get progress data
    progress_data = shortlisted.progress_data or {}

    if step_number_int > 1:
        previous_value = progress_data.get(f'step{step_number_int - 1}')
        if previous_value is None or str(previous_value).strip() == '':
            raise HTTPException(
                status_code=400,
                detail=f"Complete step {step_number_int - 1} before modifying employees on this step."
            )

    # Remove employee from assignments
    if 'employee_assignments' in progress_data:
        step_key = f'step{step_number_int}'
        if step_key in progress_data['employee_assignments']:
            if employee_id in progress_data['employee_assignments'][step_key]:
                progress_data['employee_assignments'][step_key].remove(employee_id)

    # Update database - force SQLAlchemy to detect JSON change
    shortlisted.progress_data = progress_data
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(shortlisted, "progress_data")
    db.commit()
    db.refresh(shortlisted)

    return {"success": True, "message": "Employee unassigned successfully"}


@app.get("/api/tenders/shortlist/{shortlist_id}/assigned-employees")
async def get_assigned_employees(
    request: Request,
    shortlist_id: int,
    db: Session = Depends(get_db)
):
    """Get all employees assigned to each step of a shortlisted tender."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Verify shortlist belongs to user
    shortlisted = db.query(ShortlistedTenderDB).filter(
        and_(
            ShortlistedTenderDB.id == shortlist_id,
            ShortlistedTenderDB.user_id == entity_id
        )
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    # Get progress data
    progress_data = shortlisted.progress_data or {}
    employee_assignments = progress_data.get('employee_assignments', {})

    # Build response with employee details
    assignments = {}
    for step_key in ['step1', 'step2', 'step3', 'step4', 'step5', 'step6']:
        employee_ids = employee_assignments.get(step_key, [])

        if employee_ids:
            employees = db.query(EmployeeDB).filter(
                EmployeeDB.id.in_(employee_ids)
            ).all()

            assignments[step_key] = [
                {
                    'id': emp.id,
                    'name': emp.name,
                    'email': emp.email,
                    'role': emp.role,
                    'team': emp.team
                }
                for emp in employees
            ]
        else:
            assignments[step_key] = []

    return {"success": True, "assignments": assignments}


@app.get("/api/favorites/export/csv")
async def export_favorites_csv(request: Request, db: Session = Depends(get_db)):
    """Export all user favorites to CSV."""
    from core.dependencies import get_id_for_tender_management, get_all_bd_employee_ids_for_company

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get all favorites with tender data
    # Admin sees all BD employee favorites + their own
    if entity_type == 'user':
        bd_employee_ids = get_all_bd_employee_ids_for_company(entity_id, db)
        all_ids = [entity_id] + bd_employee_ids
        favorites = db.query(FavoriteDB).filter(FavoriteDB.user_id.in_(all_ids)).all()
    else:
        # BD employee sees only their own
        favorites = db.query(FavoriteDB).filter(FavoriteDB.user_id == entity_id).all()

    # Prepare CSV data
    import csv
    import io
    from fastapi.responses import StreamingResponse

    output = io.StringIO()
    writer = csv.writer(output)

    # Write headers
    headers = [
        "Date of Tender Favourited",
        "Date of Tender Publish",
        "Tender Details",
        "Tender Documents",
        "Tender Portal",
        "Tender ID",
        "Tender Stage",
        "Type of Tender",
        "State",
        "Client",
        "EMD",
        "Est. Consultancy Fee",
        "Pre-Bid Date",
        "Submission Date",
        "PQ Eligibility",
        "JV Eligibility"
    ]
    writer.writerow(headers)

    # Write data rows
    for fav in favorites:
        if not fav.tender:
            continue

        row = []

        # Date of Tender Favourited
        row.append(fav.created_at.strftime('%Y-%m-%d %H:%M:%S') if fav.created_at else '') # type: ignore

        # Date of Tender Publish
        row.append(fav.tender.published_at.strftime('%Y-%m-%d') if fav.tender.published_at else '')

        # Tender Details
        row.append(fav.tender.summary or fav.tender.title or '')

        # Tender Documents
        documents = fav.tender.tender_documents
        if isinstance(documents, dict) and documents:
            doc_links = []
            for key, value in documents.items():
                if isinstance(value, str) and (value.startswith('http') or 'pdf' in value.lower()):
                    doc_links.append(value)
            row.append('; '.join(doc_links))
        else:
            row.append('')

        # Tender Portal
        row.append(fav.tender.source_url or '')

        # Tender ID
        row.append(fav.tender.tender_id or '')

        # Get user filled data if available
        user_data = fav.user_filled_data or {}

        # Tender Stage - Use user data first, then original data
        stage = user_data.get('Tender Stage', '')
        if not stage:
            critical_dates = fav.tender.critical_dates
            if isinstance(critical_dates, dict):
                stage = critical_dates.get('Tender Stage', '')
        row.append(stage)

        # Type of Tender - Use user data first, then original data
        tender_type = user_data.get('Type of Tender', '')
        if not tender_type:
            tender_type = fav.tender.tender_type or ''
        row.append(tender_type)

        # State - Use user data first, then original data
        state = user_data.get('State', '')
        if not state:
            state = fav.tender.state or ''
        row.append(state)

        # Client - Use user data first, then original data
        client = user_data.get('Client', '')
        if not client:
            client = fav.tender.authority or ''
        row.append(client)

        # EMD - Use user data first, then original data
        emd = user_data.get('EMD', '')
        if not emd:
            emd_details = fav.tender.emd_fee_details
            if isinstance(emd_details, dict):
                emd = str(emd_details.get('Amount', ''))
        row.append(emd)

        # Est. Consultancy Fee - Use user data first, then original data
        fee = user_data.get('Est. Consultancy Fee', '')
        if not fee:
            tender_fee_details = fav.tender.tender_fee_details
            if isinstance(tender_fee_details, dict):
                fee = str(tender_fee_details.get('Amount', ''))
        row.append(fee)

        # Pre-Bid Date - Use user data first, then original data
        pre_bid = user_data.get('Pre-Bid Date', '')
        if not pre_bid:
            critical_dates = fav.tender.critical_dates
            if isinstance(critical_dates, dict):
                pre_bid = critical_dates.get('Pre Bid Meeting Date', '')
        row.append(pre_bid)

        # Submission Date - Use user data first, then original data
        submission = user_data.get('Submission Date', '')
        if not submission:
            critical_dates = fav.tender.critical_dates
            if isinstance(critical_dates, dict):
                submission = critical_dates.get('Bid Submission End Date', '')
        row.append(submission)

        # PQ Eligibility - Use user data first, then original data
        pq = user_data.get('PQ Eligibility', '')
        if not pq:
            additional_fields = fav.tender.additional_fields
            if isinstance(additional_fields, dict):
                pq = additional_fields.get('PQ Eligibility', '')
        row.append(pq)

        # JV Eligibility - Use user data first, then original data
        jv = user_data.get('JV Eligibility', '')
        if not jv:
            additional_fields = fav.tender.additional_fields
            if isinstance(additional_fields, dict):
                jv = additional_fields.get('JV Eligibility', '')
        row.append(jv)

        writer.writerow(row)

    output.seek(0)

    # Return CSV file
    def generate():
        yield output.getvalue()

    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=favorites_export.csv"}
    )

# Custom Cards API endpoints
@app.get("/api/custom-cards")
@require_company_details
async def get_custom_cards(request: Request, db: Session = Depends(get_db)):
    """Get user's custom cards."""
    from core.dependencies import get_id_for_custom_cards, get_user_id_for_queries

    card_owner_id, entity, entity_type = get_id_for_custom_cards(request, db)
    if not card_owner_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user_id for database queries (company owner's user_id for BD employees)
    user_id_for_query, _ = get_user_id_for_queries(request, db)
    if not user_id_for_query:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Convert URLs to state names for display
    from core.cache import URL_TO_STATE

    # Query cards - for BD employees, we need to get all cards for the company
    # and filter client-side, or we can add a worked_by field later
    # For now, get all cards for the user_id (includes both admin and BD employee cards)
    cards = db.query(CustomCardDB).filter(CustomCardDB.user_id == user_id_for_query).all()
    return {"cards": [
        {
            "id": card.id,
            "card_name": card.card_name,
            "core_search_terms": card.core_search_terms,
            "state": card.state,
            "source": URL_TO_STATE.get(card.source, card.source) if card.source else None,  # Convert URL to state name
            "tender_type": card.tender_type,
            "sector": card.sector,
            "sub_sector": card.sub_sector,
            "work_type": card.work_type,
            "created_at": card.created_at.isoformat() if card.created_at else None, # type: ignore
            "updated_at": card.updated_at.isoformat() if card.updated_at else None # type: ignore
        } for card in cards
    ]}

@app.post("/api/custom-cards")
@require_company_details
async def create_custom_card(
    request: Request,
    card_name: str = Form(...),
    core_search_terms: str = Form(...),
    state: str = Form(""),
    sources: str = Form("[]"),  # Changed to receive JSON string
    tender_type: str = Form(""),
    sector: str = Form(""),
    sub_sector: str = Form(""),
    work_type: str = Form(""),
    db: Session = Depends(get_db)
):
    """Create a new custom card."""
    from core.dependencies import get_id_for_custom_cards, get_user_id_for_queries

    card_owner_id, entity, entity_type = get_id_for_custom_cards(request, db)
    if not card_owner_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user_id for database storage (company owner's user_id for BD employees)
    user_id_for_storage, _ = get_user_id_for_queries(request, db)
    if not user_id_for_storage:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if card name already exists for this entity
    # For BD employees, check by user_id_for_storage and card_owner_id logic
    if entity_type == 'bd_employee':
        # BD employees can have cards with same name as admin (separate cards)
        # But we still need to check within the company's user_id
        existing_card = db.query(CustomCardDB).filter(
            and_(CustomCardDB.user_id == user_id_for_storage, CustomCardDB.card_name == card_name)
        ).first()
    else:
        existing_card = db.query(CustomCardDB).filter(
            and_(CustomCardDB.user_id == user_id_for_storage, CustomCardDB.card_name == card_name)
        ).first()

    if existing_card:
        raise HTTPException(status_code=400, detail="Card name already exists")

    # Parse sources JSON
    try:
        sources_list = json.loads(sources) if sources else []
    except json.JSONDecodeError:
        sources_list = []

    # Convert source names to URLs for backend storage if needed
    from core.cache import STATE_TO_URL
    sources_for_db = [STATE_TO_URL.get(src, src) for src in sources_list] if sources_list else None

    # Create new card - use user_id_for_storage (company owner's user_id for BD employees)
    new_card = CustomCardDB(
        user_id=user_id_for_storage,
        card_name=card_name,
        core_search_terms=core_search_terms,
        state=state if state else None,
        sources=sources_for_db,  # Store as JSON array
        tender_type=tender_type if tender_type else None,
        sector=sector if sector else None,
        sub_sector=sub_sector if sub_sector else None,
        work_type=work_type if work_type else None,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow()
    )

    db.add(new_card)
    db.commit()
    db.refresh(new_card)

    return {"message": "Custom card created", "card_id": new_card.id}

@app.put("/api/custom-cards/{card_id}")
async def update_custom_card(
    request: Request,
    card_id: int,
    card_name: str = Form(...),
    core_search_terms: str = Form(...),
    state: str = Form(""),
    source: str = Form(""),
    tender_type: str = Form(""),
    db: Session = Depends(get_db)
):
    """Update a custom card."""
    from core.dependencies import get_id_for_custom_cards, get_user_id_for_queries

    card_owner_id, entity, entity_type = get_id_for_custom_cards(request, db)
    if not card_owner_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user_id for database queries (company owner's user_id for BD employees)
    user_id_for_query, _ = get_user_id_for_queries(request, db)
    if not user_id_for_query:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the card - query by user_id_for_query
    card = db.query(CustomCardDB).filter(
        and_(CustomCardDB.id == card_id, CustomCardDB.user_id == user_id_for_query)
    ).first()

    if not card:
        raise HTTPException(status_code=404, detail="Custom card not found")

    # Check if new card name conflicts with existing cards
    if card_name != card.card_name:
        existing_card = db.query(CustomCardDB).filter(
            and_(CustomCardDB.user_id == user_id_for_query, CustomCardDB.card_name == card_name)
        ).first()
        if existing_card:
            raise HTTPException(status_code=400, detail="Card name already exists")

    # Update card
    card.card_name = card_name # type: ignore
    card.core_search_terms = core_search_terms # type: ignore
    card.state = state if state else None # type: ignore
    card.source = source if source else None # type: ignore
    card.tender_type = tender_type if tender_type else None # type: ignore
    card.updated_at = datetime.utcnow() # type: ignore

    db.commit()

    return {"message": "Custom card updated"}

# Project Management Routes
@app.get("/add_project", response_class=HTMLResponse)
@require_company_details
async def add_project_page(request: Request, db: Session = Depends(get_db)):
    """Show the add project form."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Get user's company details to retrieve industry sectors
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()
    user_industry_sectors = []
    user_sectors_data = []
    if company_details and company_details.industry_sector:
        if company_details.industry_sector.startswith('[{'):
            # New format: [{"sector": "...", "subsectors": [...]}]
            try:
                sectors_data = json.loads(company_details.industry_sector)
                user_industry_sectors = [s['sector'] for s in sectors_data if 'sector' in s]
                user_sectors_data = sectors_data
            except (json.JSONDecodeError, KeyError):
                user_industry_sectors = [company_details.industry_sector]
        elif company_details.industry_sector.startswith('['):
            try:
                user_industry_sectors = json.loads(company_details.industry_sector)
            except json.JSONDecodeError:
                user_industry_sectors = [company_details.industry_sector]
        else:
            user_industry_sectors = [company_details.industry_sector]

    return templates.TemplateResponse("add_project.html", {
        "request": request,
        "current_user": current_user,
        "user_industry_sectors": user_industry_sectors,
        "user_sectors_data": user_sectors_data,
        "now": datetime.now,
        "selected_font": get_active_font()
    })

@app.post("/add_project")
@require_company_details
async def submit_project(
    request: Request,
    # Section 1: Project Details
    project_name: str = Form(...),
    project_description: str = Form(""),
    complete_scope_of_work: str = Form(""),
    client_name: str = Form(""),
    financing_authority: str = Form(""),
    sector: str = Form(""),
    sub_sector: str = Form(""),
    consultancy_fee: str = Form(""),  # Changed to str to handle empty values
    project_cost: str = Form(""),     # Changed to str to handle empty values
    start_date: Optional[str] = Form(None),
    end_date: Optional[str] = Form(None),
    ongoing: str = Form("false"),
    jv_partner: str = Form(""),
    country: str = Form("India"),
    states: str = Form(""),
    cities: str = Form(""),
    # Section 2: Services Rendered (will be parsed from form data)
    # Section 3: Documents (file uploads)
    db: Session = Depends(get_db)
):
    """Handle project submission."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Parse form data manually to handle arrays and complex data
    form_data = await request.form()

    # Convert numeric fields from strings to float or None
    parsed_consultancy_fee = None
    if consultancy_fee and consultancy_fee.strip():
        try:
            parsed_consultancy_fee = float(consultancy_fee)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid consultancy fee value")

    parsed_project_cost = None
    if project_cost and project_cost.strip():
        try:
            parsed_project_cost = float(project_cost)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid project cost value")

    # Parse dates
    parsed_start_date = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    parsed_end_date = None
    if end_date and ongoing != "true":
        parsed_end_date = datetime.strptime(end_date, "%Y-%m-%d")

        # Validate that end_date is not in the future
        today = datetime.utcnow().date()
        if parsed_end_date.date() > today:
            raise HTTPException(
                status_code=400,
                detail="Project completion date cannot be in the future. Please select today's date or earlier."
            )

    # Calculate project duration in months
    project_duration_months = None
    if parsed_start_date:
        end_for_calc = parsed_end_date if parsed_end_date else datetime.utcnow()
        if parsed_start_date <= end_for_calc:
            diff = end_for_calc - parsed_start_date
            project_duration_months = round(diff.days / 30.44)  # Approximate months

    # Parse states and cities as lists
    states_list = [s.strip() for s in states.split(',') if s.strip()] if states else []
    cities_list = [c.strip() for c in cities.split(',') if c.strip()] if cities else []

    financing_authority_entries = []
    if financing_authority:
        financing_authority_entries = [entry.strip() for entry in financing_authority.split(',') if entry.strip()]
    financing_authority_value = ", ".join(financing_authority_entries) if financing_authority_entries else "Financing Not Required"

    # Parse services rendered
    services_rendered = {}
    for key in PROJECT_SERVICE_KEYS:
        service_value = form_data.get(f'services[{key}]', '')
        if service_value:
            services_rendered[key] = service_value

    # Handle document uploads
    documents = {}

    # Ensure uploads directory exists
    os.makedirs("uploads/projects", exist_ok=True)

    for doc_type in PROJECT_DOCUMENT_TYPES:
        doc_files = form_data.getlist(f'documents[{doc_type}][]')
        file_entries = []
        for file in doc_files:
            if hasattr(file, 'filename') and file.filename:  # It's a file upload
                try:
                    # Generate unique filename
                    file_extension = os.path.splitext(file.filename)[1]
                    unique_filename = f"{uuid.uuid4()}{file_extension}"
                    file_path = f"uploads/projects/{unique_filename}"

                    with open(file_path, "wb") as buffer:
                        buffer.write(await file.read())

                    # Store both file path and original filename
                    file_entries.append({
                        "file_path": file_path,
                        "original_filename": file.filename
                    })
                except Exception as e:
                    logger.error(f"Error uploading {doc_type} file: {e}")
                    continue

        if file_entries:
            documents[doc_type] = file_entries

    # Generate project_id
    try:
        project_id_value = generate_project_id(current_user.id, db)
    except Exception as e:
        logger.error(f"Error generating project_id: {e}")
        # Fallback: generate ID even if there's an error
        project_id_value = generate_project_id(current_user.id, db)

    # Create new project entry
    new_project = ProjectDB(
        user_id=current_user.id,
        project_name=project_name,
        project_description=project_description,
        complete_scope_of_work=complete_scope_of_work,
        client_name=client_name,
        sector=sector,
        sub_sector=sub_sector,
        consultancy_fee=parsed_consultancy_fee,
        project_cost=parsed_project_cost,
        start_date=parsed_start_date,
        end_date=parsed_end_date,
        project_duration_months=project_duration_months,
        financing_authority=financing_authority_value,
        jv_partner=jv_partner,
        country=country,
        states=states_list,
        cities=cities_list,
        services_rendered=services_rendered,
        documents=documents,
        project_id=project_id_value
    )

    db.add(new_project)
    db.commit()
    db.refresh(new_project)

    return RedirectResponse(url="/projects?show_all=true", status_code=302)


@app.get("/edit_project/{project_id}", response_class=HTMLResponse)
@enforce_test_quarantine
@require_company_details
async def edit_project_page(
    request: Request, 
    project_id: int, 
    return_url: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Show the edit project form with pre-filled data."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    # Get the project
    project = db.query(ProjectDB).filter(
        ProjectDB.id == project_id,
        ProjectDB.user_id == current_user.id
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get user's company details to retrieve industry sectors
    company_details = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()
    user_industry_sectors = []
    user_sectors_data = []
    if company_details and company_details.industry_sector:
        if company_details.industry_sector.startswith('[{'):
            # New format: [{"sector": "...", "subsectors": [...]}]
            try:
                sectors_data = json.loads(company_details.industry_sector)
                user_industry_sectors = [s['sector'] for s in sectors_data if 'sector' in s]
                user_sectors_data = sectors_data
            except (json.JSONDecodeError, KeyError):
                user_industry_sectors = [company_details.industry_sector]
        elif company_details.industry_sector.startswith('['):
            try:
                user_industry_sectors = json.loads(company_details.industry_sector)
            except json.JSONDecodeError:
                user_industry_sectors = [company_details.industry_sector]
        else:
            user_industry_sectors = [company_details.industry_sector]

    return templates.TemplateResponse("edit_project.html", {
        "request": request,
        "current_user": current_user,
        "project": project,
        "user_industry_sectors": user_industry_sectors,
        "user_sectors_data": user_sectors_data,
        "now": datetime.now,
        "return_url": return_url,
        "selected_font": get_active_font()
    })


@app.post("/edit_project/{project_id}")
@require_company_details
async def update_project(
    request: Request,
    project_id: int,
    # Section 1: Project Details
    project_name: str = Form(...),
    project_description: str = Form(""),
    complete_scope_of_work: str = Form(""),
    client_name: str = Form(""),
    financing_authority: str = Form(""),
    sector: str = Form(""),
    sub_sector: str = Form(""),
    consultancy_fee: str = Form(""),  # Changed to str to handle empty values
    project_cost: str = Form(""),     # Changed to str to handle empty values
    start_date: Optional[str] = Form(None),
    end_date: Optional[str] = Form(None),
    ongoing: str = Form("false"),
    jv_partner: str = Form(""),
    country: str = Form("India"),
    states: str = Form(""),
    cities: str = Form(""),
    return_url: Optional[str] = Form(None),
    # Section 2: Services Rendered (will be parsed from form data)
    # Section 3: Documents (file uploads)
    db: Session = Depends(get_db)
):
    """Handle project update."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get the project
    project = db.query(ProjectDB).filter(
        ProjectDB.id == project_id,
        ProjectDB.user_id == current_user.id
    ).first()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Parse form data manually to handle arrays and complex data
    form_data = await request.form()

    # Convert numeric fields from strings to float or None
    parsed_consultancy_fee = None
    if consultancy_fee and consultancy_fee.strip():
        try:
            parsed_consultancy_fee = float(consultancy_fee)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid consultancy fee value")

    parsed_project_cost = None
    if project_cost and project_cost.strip():
        try:
            parsed_project_cost = float(project_cost)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid project cost value")

    # Parse dates
    parsed_start_date = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    parsed_end_date = None
    if end_date and ongoing != "true":
        parsed_end_date = datetime.strptime(end_date, "%Y-%m-%d")

        # Validate that end_date is not in the future
        today = datetime.utcnow().date()
        if parsed_end_date.date() > today:
            raise HTTPException(
                status_code=400,
                detail="Project completion date cannot be in the future. Please select today's date or earlier."
            )

    # Calculate project duration in months
    project_duration_months = None
    if parsed_start_date:
        end_for_calc = parsed_end_date if parsed_end_date else datetime.utcnow()
        if parsed_start_date <= end_for_calc:
            diff = end_for_calc - parsed_start_date
            project_duration_months = round(diff.days / 30.44)  # Approximate months

    # Parse states and cities as lists
    states_list = [s.strip() for s in states.split(',') if s.strip()] if states else []
    cities_list = [c.strip() for c in cities.split(',') if c.strip()] if cities else []

    financing_authority_entries = []
    if financing_authority:
        financing_authority_entries = [entry.strip() for entry in financing_authority.split(',') if entry.strip()]
    financing_authority_value = ", ".join(financing_authority_entries) if financing_authority_entries else "Financing Not Required"

    # Parse services rendered
    services_rendered = {}
    for key in PROJECT_SERVICE_KEYS:
        service_value = form_data.get(f'services[{key}]', '')
        if service_value:
            services_rendered[key] = service_value

    # Handle document uploads (append to existing documents)
    existing_documents = project.documents if project.documents else {}

    # Ensure uploads directory exists
    os.makedirs("uploads/projects", exist_ok=True)

    logger.info(f"[EDIT] Processing document uploads for project {project_id}")

    for doc_type in PROJECT_DOCUMENT_TYPES:
        doc_files = form_data.getlist(f'documents[{doc_type}][]')
        file_entries = []

        logger.info(f"[EDIT] Found {len(doc_files)} files for document type: {doc_type}")

        for file in doc_files:
            if hasattr(file, 'filename') and file.filename:  # It's a file upload
                try:
                    # Generate unique filename
                    file_extension = os.path.splitext(file.filename)[1]
                    unique_filename = f"{uuid.uuid4()}{file_extension}"
                    file_path = f"uploads/projects/{unique_filename}"

                    with open(file_path, "wb") as buffer:
                        buffer.write(await file.read())

                    # Store both file path and original filename
                    file_entries.append({
                        "file_path": file_path,
                        "original_filename": file.filename
                    })
                    logger.info(f"[EDIT] Successfully uploaded: {file.filename} -> {file_path}")
                except Exception as e:
                    logger.error(f"[EDIT] Error uploading {doc_type} file {file.filename}: {e}")
                    continue

        # Append new files to existing documents for this category
        if file_entries:
            if doc_type not in existing_documents:
                existing_documents[doc_type] = []
            existing_documents[doc_type].extend(file_entries)
            logger.info(f"[EDIT] Added {len(file_entries)} files to {doc_type}")

    # Generate project_id if it doesn't exist (for legacy projects)
    if not project.project_id:
        try:
            project.project_id = generate_project_id(current_user.id, db)
            logger.info(f"[EDIT] Generated project_id: {project.project_id} for project {project_id}")
        except Exception as e:
            logger.error(f"[EDIT] Error generating project_id: {e}")
            # Continue without project_id - not critical for edit operation

    # Update project fields
    project.project_name = project_name
    project.project_description = project_description
    project.complete_scope_of_work = complete_scope_of_work
    project.client_name = client_name
    project.sector = sector
    project.sub_sector = sub_sector
    project.consultancy_fee = parsed_consultancy_fee
    project.project_cost = parsed_project_cost
    project.start_date = parsed_start_date
    project.end_date = parsed_end_date
    project.project_duration_months = project_duration_months
    project.financing_authority = financing_authority_value
    project.jv_partner = jv_partner
    project.country = country
    project.states = states_list
    project.cities = cities_list
    project.services_rendered = services_rendered
    project.documents = existing_documents

    # CRITICAL: Flag the JSONB column as modified so SQLAlchemy tracks the change
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(project, "documents")

    logger.info(f"[EDIT] Final document count: {sum(len(v) for v in existing_documents.values()) if existing_documents else 0}")
    logger.info(f"[EDIT] ✓ Flagged 'documents' field as modified for SQLAlchemy tracking")

    # If project was auto-generated and is now being edited, mark as completed_by_user
    if project.is_auto_generated and project.completion_status == "incomplete":
        project.completion_status = "completed_by_user"

    try:
        db.commit()
        db.refresh(project)
        logger.info(f"[EDIT] Successfully updated project {project_id}")
    except Exception as e:
        db.rollback()
        logger.error(f"[EDIT] Failed to update project {project_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update project: {str(e)}")

    # Redirect back to project detail, preserving return_url
    redirect_url = f"/project/{project_id}"
    if return_url:
        redirect_url += f"?return_url={return_url}"
    return RedirectResponse(url=redirect_url, status_code=302)


@app.delete("/project/{project_id}/document/{doc_type}/{doc_index}")
@require_company_details
async def delete_project_document(
    request: Request,
    project_id: int,
    doc_type: str,
    doc_index: int,
    db: Session = Depends(get_db)
):
    """Delete a specific document from a project."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    logger.info(f"[DELETE] Deleting document: project={project_id}, type={doc_type}, index={doc_index}")

    # Get project and verify ownership
    project = db.query(ProjectDB).filter(
        ProjectDB.id == project_id,
        ProjectDB.user_id == current_user.id
    ).first()

    if not project:
        logger.error(f"[DELETE] Project {project_id} not found")
        raise HTTPException(status_code=404, detail="Project not found")

    # Check if project has documents
    if not project.documents or doc_type not in project.documents:
        logger.error(f"[DELETE] Document type '{doc_type}' not found in project {project_id}")
        raise HTTPException(status_code=404, detail=f"Document type '{doc_type}' not found")

    # Get the documents for this type
    docs = project.documents[doc_type]
    if not isinstance(docs, list):
        docs = [docs]

    # Check if index is valid
    if doc_index < 0 or doc_index >= len(docs):
        logger.error(f"[DELETE] Invalid document index {doc_index} for type '{doc_type}'")
        raise HTTPException(status_code=404, detail="Document not found")

    # Get the document to delete
    doc_to_delete = docs[doc_index]

    # Extract file path
    if isinstance(doc_to_delete, dict):
        file_path = doc_to_delete.get('file_path', '')
        filename = doc_to_delete.get('original_filename', 'unknown')
    else:
        file_path = doc_to_delete
        filename = os.path.basename(file_path)

    logger.info(f"[DELETE] Deleting file: {filename} at {file_path}")

    # Delete file from filesystem
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"[DELETE] ✓ File deleted from disk: {file_path}")
        else:
            logger.warning(f"[DELETE] ⚠ File not found on disk: {file_path}")
    except Exception as e:
        logger.error(f"[DELETE] Error deleting file from disk: {e}")
        # Continue with database deletion even if file deletion fails

    # Remove document from list
    docs.pop(doc_index)

    # Update project documents
    if len(docs) == 0:
        # Remove the category if no documents left
        del project.documents[doc_type]
        logger.info(f"[DELETE] ✓ Removed empty category '{doc_type}'")
    else:
        # Update the category with remaining documents
        project.documents[doc_type] = docs

    # CRITICAL: Flag as modified for SQLAlchemy
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(project, "documents")

    logger.info(f"[DELETE] ✓ Flagged 'documents' field as modified")

    # Commit changes
    try:
        db.commit()
        db.refresh(project)
        logger.info(f"[DELETE] ✓ Successfully deleted document from project {project_id}")

        # Verify deletion
        verify_project = db.query(ProjectDB).filter(ProjectDB.id == project_id).first()
        if verify_project and verify_project.documents:
            remaining_count = sum(len(v) if isinstance(v, list) else 1 for v in verify_project.documents.values())
            logger.info(f"[DELETE] ✓ VERIFICATION: {remaining_count} documents remaining in database")
        else:
            logger.info(f"[DELETE] ✓ VERIFICATION: No documents remaining in project")

        return {"success": True, "message": f"Document '{filename}' deleted successfully"}

    except Exception as e:
        db.rollback()
        logger.error(f"[DELETE] ✗ Failed to delete document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete document: {str(e)}")


@app.post("/projects/upload")
@require_company_details
async def upload_projects_bulk(
    request: Request,
    excel_file: UploadFile = File(...),
    zip_file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    """Bulk upload projects from an Excel file. Attachments can be added later via edit functionality."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Validate Excel file
    excel_filename = excel_file.filename or ""
    excel_extension = Path(excel_filename).suffix.lower()
    if excel_extension != ".xlsx":
        params = urlencode({
            "show_all": "true",
            "upload_status": "error",
            "imported": 0,
            "skipped": 0,
            "failed": 1,
            "upload_message": "Invalid Excel file. Please upload a .xlsx file."
        })
        return RedirectResponse(url=f"/projects?{params}", status_code=302)

    # Read Excel file contents
    excel_bytes = await excel_file.read()

    # Validate file sizes
    MAX_EXCEL_SIZE = 10 * 1024 * 1024  # 10 MB

    if len(excel_bytes) > MAX_EXCEL_SIZE:
        params = urlencode({
            "show_all": "true",
            "upload_status": "error",
            "imported": 0,
            "skipped": 0,
            "failed": 1,
            "upload_message": "Excel file is too large. Maximum size is 10 MB."
        })
        return RedirectResponse(url=f"/projects?{params}", status_code=302)

    if not excel_bytes:
        params = urlencode({
            "show_all": "true",
            "upload_status": "error",
            "imported": 0,
            "skipped": 0,
            "failed": 1,
            "upload_message": "The uploaded Excel file is empty."
        })
        return RedirectResponse(url=f"/projects?{params}", status_code=302)

    # Extract zip file and get filename mapping
    import zipfile
    import shutil
    from datetime import datetime as dt_module

    def extract_zip_attachments(zip_data: bytes, user_id: str) -> tuple[Dict[str, str], str]:
        """
        Extract zip file and return a mapping of filename -> saved_file_path.

        Args:
            zip_data: The zip file content as bytes
            user_id: User ID for directory isolation

        Returns:
            Tuple of (filename_to_path_mapping, extraction_directory)

        Raises:
            ValueError: If zip is invalid or extraction fails
        """
        # Create unique extraction directory
        upload_timestamp = dt_module.utcnow().strftime("%Y%m%d_%H%M%S_%f")
        extraction_dir = f"uploads/projects/{user_id}/{upload_timestamp}"
        os.makedirs(extraction_dir, exist_ok=True)

        filename_mapping: Dict[str, str] = {}

        try:
            # Create a BytesIO object from zip bytes
            zip_buffer = BytesIO(zip_data)

            # Open and validate zip file
            with zipfile.ZipFile(zip_buffer, 'r') as zip_ref:
                # Test zip integrity
                bad_file = zip_ref.testzip()
                if bad_file:
                    raise ValueError(f"Corrupted file in zip: {bad_file}")

                # Extract all files
                for zip_info in zip_ref.infolist():
                    # Skip directories
                    if zip_info.is_dir():
                        continue

                    # Get original filename (handle nested paths)
                    original_filename = os.path.basename(zip_info.filename)

                    # Sanitize filename to prevent path traversal
                    safe_filename = original_filename.replace("../", "").replace("..\\", "")
                    if not safe_filename:
                        continue

                    # Extract file
                    try:
                        extracted_data = zip_ref.read(zip_info.filename)

                        # Save file with sanitized name
                        file_path = os.path.join(extraction_dir, safe_filename)

                        # Handle duplicate filenames by appending counter
                        base_name, ext = os.path.splitext(safe_filename)
                        counter = 1
                        while os.path.exists(file_path):
                            safe_filename = f"{base_name}_{counter}{ext}"
                            file_path = os.path.join(extraction_dir, safe_filename)
                            counter += 1

                        with open(file_path, 'wb') as f:
                            f.write(extracted_data)

                        # Store mapping: original filename -> saved path
                        filename_mapping[original_filename] = file_path

                    except Exception as e:
                        logger.warning(f"Failed to extract file {zip_info.filename}: {e}")
                        continue

            if not filename_mapping:
                raise ValueError("No valid files found in zip archive")

            logger.info(f"Extracted {len(filename_mapping)} files from zip to {extraction_dir}")
            return filename_mapping, extraction_dir

        except zipfile.BadZipFile:
            # Clean up on failure
            if os.path.exists(extraction_dir):
                shutil.rmtree(extraction_dir)
            raise ValueError("Invalid or corrupted zip file")
        except Exception as e:
            # Clean up on failure
            if os.path.exists(extraction_dir):
                shutil.rmtree(extraction_dir)
            raise ValueError(f"Failed to extract zip file: {str(e)}")

    # Extract attachments from zip file (if provided)
    attachment_files_map: Dict[str, str] = {}
    extraction_directory: Optional[str] = None
    files_extracted_count = 0
    
    if zip_file is not None:
        # Validate Zip file
        zip_filename = zip_file.filename or ""
        zip_extension = Path(zip_filename).suffix.lower()
        if zip_extension != ".zip":
            params = urlencode({
                "show_all": "true",
                "upload_status": "error",
                "imported": 0,
                "skipped": 0,
                "failed": 1,
                "upload_message": "Invalid Zip file. Please upload a .zip file."
            })
            return RedirectResponse(url=f"/projects?{params}", status_code=302)

        # Read zip file contents
        zip_bytes = await zip_file.read()

        # Validate zip file size
        MAX_ZIP_SIZE = 500 * 1024 * 1024   # 500 MB
        if len(zip_bytes) > MAX_ZIP_SIZE:
            params = urlencode({
                "show_all": "true",
                "upload_status": "error",
                "imported": 0,
                "skipped": 0,
                "failed": 1,
                "upload_message": "Zip file is too large. Maximum size is 500 MB."
            })
            return RedirectResponse(url=f"/projects?{params}", status_code=302)

        if not zip_bytes:
            params = urlencode({
                "show_all": "true",
                "upload_status": "error",
                "imported": 0,
                "skipped": 0,
                "failed": 1,
                "upload_message": "The uploaded Zip file is empty."
            })
            return RedirectResponse(url=f"/projects?{params}", status_code=302)

        # Extract attachments from zip file
        try:
            attachment_files_map, extraction_directory = extract_zip_attachments(zip_bytes, current_user.id)
            files_extracted_count = len(attachment_files_map)
            logger.info(f"Successfully extracted {files_extracted_count} files for user {current_user.id}")
        except ValueError as exc:
            params = urlencode({
                "show_all": "true",
                "upload_status": "error",
                "imported": 0,
                "skipped": 0,
                "failed": 1,
                "upload_message": str(exc)
            })
            return RedirectResponse(url=f"/projects?{params}", status_code=302)
        except Exception as exc:
            logger.error("Failed to extract zip file", exc_info=exc)
            params = urlencode({
                "show_all": "true",
                "upload_status": "error",
                "imported": 0,
                "skipped": 0,
                "failed": 1,
                "upload_message": "Unable to extract attachments from zip file. Please check the file and try again."
            })
            return RedirectResponse(url=f"/projects?{params}", status_code=302)

    def normalize_header(header: Any) -> str:
        if header is None:
            return ""
        header_str = str(header).strip().lower()
        header_str = re.sub(r'[^a-z0-9]+', '_', header_str)
        return header_str.strip('_')

    def extract_rows_from_xlsx(data: bytes) -> List[tuple[int, Dict[str, Any]]]:
        workbook = load_workbook(filename=BytesIO(data), data_only=True)
        sheet = workbook.active
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            raise ValueError("The uploaded file is missing a header row.")

        normalized_headers = [normalize_header(cell) for cell in header_cells]

        # Debug: Log found headers with their column indices
        logger.info(f"Excel upload - Total columns: {len(normalized_headers)}")
        for idx, header in enumerate(normalized_headers):
            if header in ["project_name", "project_description", "complete_scope_of_work", "scope_of_work", "client_name"]:
                logger.info(f"  Column {idx}: '{header}' (original: '{header_cells[idx]}')")

        if "project_description" in normalized_headers:
            desc_idx = normalized_headers.index("project_description")
            logger.info(f"✓ project_description column found at index {desc_idx}")
        else:
            logger.warning("✗ project_description column NOT found")

        if "complete_scope_of_work" in normalized_headers:
            scope_idx = normalized_headers.index("complete_scope_of_work")
            logger.info(f"✓ complete_scope_of_work column found at index {scope_idx}")
        elif "scope_of_work" in normalized_headers:
            scope_idx = normalized_headers.index("scope_of_work")
            logger.info(f"✓ scope_of_work column found at index {scope_idx}")
        else:
            logger.warning("✗ Neither complete_scope_of_work nor scope_of_work column found")

        rows: List[tuple[int, Dict[str, Any]]] = []

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            if all(
                (value is None) or (isinstance(value, str) and not value.strip())
                for value in row
            ):
                continue

            row_data: Dict[str, Any] = {}
            for idx, value in enumerate(row):
                if idx >= len(normalized_headers):
                    continue
                key = normalized_headers[idx]
                if not key:
                    continue
                row_data[key] = value

                # Debug: Log project_name, project_description, and scope fields
                if key in ["project_name", "project_description", "complete_scope_of_work", "scope_of_work"]:
                    value_preview = str(value)[:100] if value else "EMPTY"
                    logger.info(f"Row {row_index}, Col {idx} [{key}]: {value_preview}")

            if row_data:
                rows.append((row_index, row_data))

        return rows

    def extract_rows_from_csv(data: bytes) -> List[tuple[int, Dict[str, Any]]]:
        text_stream = StringIO(data.decode("utf-8-sig"))
        reader = csv.reader(text_stream)
        try:
            header_row = next(reader)
        except StopIteration:
            raise ValueError("The uploaded file is empty.")

        normalized_headers = [normalize_header(cell) for cell in header_row]
        rows: List[tuple[int, Dict[str, Any]]] = []

        for row_index, row_values in enumerate(reader, start=2):
            if not row_values:
                continue
            if all(not (value or "").strip() for value in row_values):
                continue

            row_data: Dict[str, Any] = {}
            for idx, value in enumerate(row_values):
                if idx >= len(normalized_headers):
                    continue
                key = normalized_headers[idx]
                if not key:
                    continue
                row_data[key] = value.strip() if isinstance(value, str) else value

            if row_data:
                rows.append((row_index, row_data))

        return rows

    try:
        # Only process xlsx files now (no CSV support with attachments)
        extracted_rows = extract_rows_from_xlsx(excel_bytes)
    except ValueError as exc:
        params = urlencode({
            "show_all": "true",
            "upload_status": "error",
            "imported": 0,
            "skipped": 0,
            "failed": 1,
            "upload_message": str(exc)
        })
        return RedirectResponse(url=f"/projects?{params}", status_code=302)
    except Exception as exc:
        logger.error("Failed to read uploaded project file", exc_info=exc)
        params = urlencode({
            "show_all": "true",
            "upload_status": "error",
            "imported": 0,
            "skipped": 0,
            "failed": 1,
            "upload_message": "Unable to read the uploaded file. Please check the template and try again."
        })
        return RedirectResponse(url=f"/projects?{params}", status_code=302)

    if not extracted_rows:
        params = urlencode({
            "show_all": "true",
            "upload_status": "error",
            "imported": 0,
            "skipped": 0,
            "failed": 1,
            "upload_message": "No project rows were found in the uploaded file."
        })
        return RedirectResponse(url=f"/projects?{params}", status_code=302)

    def parse_float(value: Any) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        value_str = str(value).strip()
        if not value_str:
            return None
        value_str = value_str.replace(",", "")
        value_str = re.sub(r"[^\d\.\-]", "", value_str)
        if not value_str or value_str in {"-", "."}:
            return None
        try:
            return float(value_str)
        except ValueError:
            return None

    def parse_date(value: Any) -> Optional[datetime]:
        if value in (None, "", " "):
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        if isinstance(value, (int, float)):
            # Excel serial date (assuming 1899-12-30 origin)
            try:
                return datetime(1899, 12, 30) + timedelta(days=float(value))
            except Exception:
                return None
        value_str = str(value).strip()
        if not value_str:
            return None
        date_formats = [
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%b-%Y",
            "%d %b %Y",
            "%b %d %Y",
            "%Y/%m/%d",
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(value_str, fmt)
            except ValueError:
                continue
        return None

    def parse_bool(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        if value is None:
            return False
        value_str = str(value).strip().lower()
        return value_str in {"true", "yes", "1", "y", "ongoing"}

    def parse_list(value: Any) -> List[str]:
        if value is None:
            return []
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        value_str = str(value)
        if not value_str.strip():
            return []
        items = re.split(r"[,;\n]+", value_str)
        return [item.strip() for item in items if item.strip()]

    document_types = PROJECT_DOCUMENT_TYPES
    document_header_variants = {
        variant
        for doc_type in document_types
        for variant in (
            doc_type,
            f"{doc_type}_files",
            f"{doc_type}_names"
        )
    }

    def parse_document_value(value: Any) -> List[str]:
        if value is None:
            return []
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        if isinstance(value, dict):
            collected: List[str] = []
            for item_value in value.values():
                collected.extend(parse_document_value(item_value))
            return collected
        value_str = str(value).strip()
        if not value_str:
            return []
        try:
            parsed_json = json.loads(value_str)
        except (json.JSONDecodeError, TypeError):
            parsed_json = None
        if parsed_json is not None:
            return parse_document_value(parsed_json)
        return [item.strip() for item in re.split(r"[,;\n]+", value_str) if item.strip()]

    existing_name_rows = db.query(ProjectDB.project_name).filter(ProjectDB.user_id == current_user.id).all()
    existing_names = {
        (name or "").strip().lower()
        for (name,) in existing_name_rows
        if name
    }

    service_keys = set(PROJECT_SERVICE_KEYS)

    successes = 0
    duplicates = 0
    updated = 0  # Track projects updated with new column data
    failures = 0
    corrected = 0  # Track rows with auto-corrected date issues
    error_messages: List[str] = []
    warning_messages: List[str] = []  # Track auto-corrections
    new_projects: List[ProjectDB] = []
    all_unmatched_files: set = set()  # Track unique unmatched filenames
    generated_project_ids: Set[str] = set()  # Track project IDs generated in this bulk upload session

    for row_number, row in extracted_rows:
        project_name_raw = row.get("project_name")
        project_name = str(project_name_raw).strip() if project_name_raw is not None else ""
        if not project_name:
            failures += 1
            error_messages.append(f"Row {row_number}: Missing project_name.")
            continue

        normalized_name = project_name.lower()
        if normalized_name in existing_names:
            # Check if there's new data in complete_scope_of_work or scope_of_work to update
            new_scope = str(row.get("complete_scope_of_work") or row.get("scope_of_work") or "").strip()
            if new_scope:
                # Find the existing project and update complete_scope_of_work if it's empty
                existing_project = db.query(ProjectDB).filter(
                    ProjectDB.user_id == current_user.id,
                    func.lower(ProjectDB.project_name) == normalized_name
                ).first()
                if existing_project and not existing_project.complete_scope_of_work:
                    existing_project.complete_scope_of_work = new_scope
                    updated += 1
                    logger.info(f"Updated complete_scope_of_work for existing project: {project_name}")
            duplicates += 1
            continue

        project_description = str(row.get("project_description") or "").strip()
        # Check both 'complete_scope_of_work' and 'scope_of_work' column names
        complete_scope_of_work = str(row.get("complete_scope_of_work") or row.get("scope_of_work") or "").strip()

        # Debug logging for description and scope during upload
        logger.info(f"Uploading project '{project_name}' - Description: {len(project_description)} chars, Scope: {len(complete_scope_of_work)} chars")

        client_name = str(row.get("client_name") or "").strip()
        sector = str(row.get("sector") or "").strip()
        sub_sector = str(row.get("sub_sector") or "").strip()
        consultancy_fee = parse_float(row.get("consultancy_fee"))
        project_cost = parse_float(row.get("project_cost"))
        start_date = parse_date(row.get("start_date"))
        end_date = parse_date(row.get("end_date"))
        ongoing = parse_bool(row.get("ongoing"))

        # Auto-correct: flip dates if end_date is before start_date
        if start_date and end_date and not ongoing and end_date < start_date:
            start_date, end_date = end_date, start_date  # Flip the dates
            corrected += 1
            warning_messages.append(f"Row {row_number} ({project_name}): Start and end dates were flipped.")

        if ongoing:
            end_date = None

        states_list = parse_list(row.get("states"))
        cities_list = parse_list(row.get("cities"))
        jv_partner = str(row.get("jv_partner") or "").strip()
        country = str(row.get("country") or "").strip() or "India"
        financing_authority_raw = str(row.get("financing_authority") or "").strip()
        financing_authority_entries = [item.strip() for item in financing_authority_raw.split(",") if item.strip()]
        financing_authority_value = ", ".join(financing_authority_entries) if financing_authority_entries else "Financing Not Required"

        project_duration_months: Optional[int] = None
        if start_date:
            duration_end = end_date if end_date else datetime.utcnow()
            if duration_end >= start_date:
                duration_days = (duration_end - start_date).days
                project_duration_months = round(duration_days / 30.44)

        services_rendered: Dict[str, Any] = {}
        services_value = row.get("services_rendered")
        if services_value:
            if isinstance(services_value, dict):
                for key, value in services_value.items():
                    if value and str(value).strip():
                        services_rendered[str(key)] = str(value).strip()
            elif isinstance(services_value, list):
                for idx, value in enumerate(services_value, start=1):
                    if value and str(value).strip():
                        services_rendered[f"service_{idx}"] = str(value).strip()
            elif isinstance(services_value, str):
                text = services_value.strip()
                if text:
                    try:
                        parsed = json.loads(text)
                        if isinstance(parsed, dict):
                            for key, value in parsed.items():
                                if value and str(value).strip():
                                    services_rendered[str(key)] = str(value).strip()
                        elif isinstance(parsed, list):
                            for idx, value in enumerate(parsed, start=1):
                                if value and str(value).strip():
                                    services_rendered[f"service_{idx}"] = str(value).strip()
                        else:
                            services_rendered["details"] = text
                    except json.JSONDecodeError:
                        items = [item.strip() for item in re.split(r"[,;\n]+", text) if item.strip()]
                        if len(items) == 1:
                            services_rendered["details"] = items[0]
                        else:
                            for idx, item in enumerate(items, start=1):
                                services_rendered[f"service_{idx}"] = item

        documents_map: Dict[str, List[str]] = {}
        for doc_type in document_types:
            raw_value = (
                row.get(f"{doc_type}_files")
                or row.get(f"{doc_type}_names")
                or row.get(doc_type)
            )
            parsed_files = parse_document_value(raw_value)
            if parsed_files:
                documents_map[doc_type] = list(dict.fromkeys(parsed_files))

        # Match filenames from Excel to actual files in zip
        resolved_documents_map: Dict[str, List[str]] = {}

        for doc_type, filenames in documents_map.items():
            resolved_paths = []

            for filename in filenames:
                filename_clean = filename.strip()
                if not filename_clean:
                    continue

                # Try exact match first
                if filename_clean in attachment_files_map:
                    resolved_paths.append(attachment_files_map[filename_clean])
                else:
                    # Try case-insensitive match
                    filename_lower = filename_clean.lower()
                    found = False
                    for zip_filename, zip_path in attachment_files_map.items():
                        if zip_filename.lower() == filename_lower:
                            resolved_paths.append(zip_path)
                            found = True
                            break

                    if not found:
                        # Log warning but continue (don't fail entire upload)
                        logger.warning(
                            f"Row {row_number}: File '{filename_clean}' "
                            f"not found in zip for doc_type '{doc_type}'"
                        )
                        all_unmatched_files.add(filename_clean)  # Track unmatched file

            if resolved_paths:
                resolved_documents_map[doc_type] = resolved_paths

        for key, value in row.items():
            if key in {"project_name", "project_description", "complete_scope_of_work", "client_name", "sector", "sub_sector",
                       "consultancy_fee", "project_cost", "start_date", "end_date", "ongoing",
                       "states", "cities", "jv_partner", "country", "services_rendered", "financing_authority"} or key in document_header_variants:
                continue
            if key in service_keys or key.startswith("services_") or key.startswith("service_"):
                service_key = key
                if key.startswith("services_"):
                    service_key = key.replace("services_", "", 1)
                elif key.startswith("service_"):
                    service_key = key.replace("service_", "", 1)
                if value and str(value).strip():
                    services_rendered[service_key] = str(value).strip()

        # Extract project_id from Excel row (check for id, project_id, ID, PROJECT_ID)
        project_id_from_excel = None
        for id_key in ["id", "project_id", "ID", "PROJECT_ID"]:
            if id_key in row and row[id_key]:
                project_id_from_excel = str(row[id_key]).strip()
                if project_id_from_excel:
                    break
        
        # Generate project_id if not provided in Excel
        try:
            if project_id_from_excel:
                # Validate format before using
                if re.match(r'^[A-Z]{2,10}-\d{6,}$', project_id_from_excel):
                    # Check uniqueness in database and in-memory set
                    existing = db.query(ProjectDB).filter(
                        ProjectDB.user_id == current_user.id,
                        ProjectDB.project_id == project_id_from_excel
                    ).first()
                    if existing or project_id_from_excel in generated_project_ids:
                        # Duplicate found, generate new one
                        project_id_value = generate_project_id(current_user.id, db, in_memory_project_ids=generated_project_ids)
                        logger.warning(f"Row {row_number}: Duplicate project_id '{project_id_from_excel}', generated new: {project_id_value}")
                    else:
                        project_id_value = project_id_from_excel
                else:
                    # Invalid format, generate new one
                    project_id_value = generate_project_id(current_user.id, db, in_memory_project_ids=generated_project_ids)
                    logger.warning(f"Row {row_number}: Invalid project_id format '{project_id_from_excel}', generated new: {project_id_value}")
            else:
                # No project_id in Excel, generate one
                project_id_value = generate_project_id(current_user.id, db, in_memory_project_ids=generated_project_ids)
        except Exception as e:
            logger.error(f"Row {row_number}: Error generating project_id: {e}")
            # Fallback: generate ID even if there's an error
            project_id_value = generate_project_id(current_user.id, db, in_memory_project_ids=generated_project_ids)
        
        # Add generated project_id to tracking set to ensure uniqueness within this bulk upload
        generated_project_ids.add(project_id_value)

        new_project = ProjectDB(
            user_id=current_user.id,
            project_name=project_name,
            project_description=project_description or None,
            complete_scope_of_work=complete_scope_of_work or None,
            client_name=client_name or None,
            sector=sector or None,
            sub_sector=sub_sector or None,
            consultancy_fee=consultancy_fee,
            project_cost=project_cost,
            start_date=start_date,
            end_date=end_date,
            project_duration_months=project_duration_months,
            financing_authority=financing_authority_value,
            jv_partner=jv_partner or None,
            country=country,
            states=states_list,
            cities=cities_list,
            services_rendered=services_rendered,
            documents=resolved_documents_map,  # Use resolved file paths from zip
            project_id=project_id_value
        )

        new_projects.append(new_project)
        successes += 1
        existing_names.add(normalized_name)

    if new_projects:
        try:
            db.add_all(new_projects)
            db.commit()
        except Exception as exc:
            logger.error("Failed to save uploaded projects", exc_info=exc)
            db.rollback()
            failures += successes
            successes = 0
            error_messages.append("Unable to save uploaded projects. No new projects were created.")
            # Clean up extracted files on database failure
            if extraction_directory and os.path.exists(extraction_directory):
                try:
                    shutil.rmtree(extraction_directory)
                    logger.info(f"Cleaned up extraction directory after database failure: {extraction_directory}")
                except Exception as cleanup_exc:
                    logger.error(f"Failed to cleanup extraction directory: {cleanup_exc}")

    # Commit any updates to existing projects (e.g., complete_scope_of_work)
    if updated > 0:
        try:
            db.commit()
            logger.info(f"Updated {updated} existing projects with new column data")
        except Exception as exc:
            logger.error("Failed to save updated projects", exc_info=exc)
            db.rollback()
            updated = 0

    status_value = "success" if (successes > 0 or updated > 0) else "error"
    params_dict = {
        "show_all": "true",
        "upload_status": status_value,
        "imported": successes,
        "updated": updated,
        "skipped": duplicates - updated,  # Adjusted: don't count updated as skipped
        "failed": failures,
        "corrected": corrected,
        "files_processed": files_extracted_count,
        "files_unmatched": len(all_unmatched_files)
    }
    if error_messages:
        params_dict["upload_message"] = error_messages[0][:200]
    if warning_messages:
        params_dict["warning_message"] = "; ".join(warning_messages[:3])  # Show up to 3 warnings

    redirect_url = f"/projects?{urlencode(params_dict)}"
    return RedirectResponse(url=redirect_url, status_code=302)


@app.get("/projects/upload/template")
@require_company_details
async def download_projects_template(
    request: Request,
    db: Session = Depends(get_db)
):
    """Provide a spreadsheet template for bulk project uploads."""
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Projects"

    base_headers = [
        "project_name",
        "project_description",
        "complete_scope_of_work",  # Also accepts 'scope_of_work'
        "client_name",
        "financing_authority",
        "sector",
        "sub_sector",
        "consultancy_fee",
        "project_cost",
        "start_date",
        "end_date",
        "ongoing",
        "country",
        "states",
        "cities",
        "jv_partner"
    ]
    service_headers = [f"services_{key}" for key in PROJECT_SERVICE_KEYS]
    document_headers = [f"{doc_type}_files" for doc_type in PROJECT_DOCUMENT_TYPES]
    column_headers = base_headers + service_headers + document_headers

    sheet.append(column_headers)
    sheet.freeze_panes = "A2"

    sample_data = {
        "project_name": "Urban Water Supply Modernisation",
        "project_description": "Comprehensive upgrade of water supply infrastructure covering distribution, treatment, and monitoring systems.",
        "complete_scope_of_work": "The scope of work includes: (1) Detailed engineering design for 250 km pipeline network, (2) Preparation of DPR and feasibility studies, (3) GIS mapping and asset management system, (4) Construction supervision and quality assurance, (5) Capacity building and training for O&M staff, (6) Environmental and social impact assessment, (7) Final project completion and handover documentation.",
        "client_name": "Delhi Jal Board",
        "financing_authority": "World Bank, State Finance Department",
        "sector": "Water Supply",
        "sub_sector": "Distribution Network",
        "consultancy_fee": "8500000",
        "project_cost": "325000000",
        "start_date": "2021-04-01",
        "end_date": "2023-03-31",
        "ongoing": "No",
        "country": "India",
        "states": "Delhi, Haryana",
        "cities": "New Delhi, Gurugram",
        "jv_partner": "XYZ Engineering Pvt Ltd",
        "services_design_engineering": "Detailed hydraulic design deliverables",
        "services_dpr_feasibility": "Prepared techno-economic feasibility report",
        "services_gis_data": "Digitised distribution network maps",
        "services_pmc": "Planning, monitoring and supervision activities",
        "services_pmu": "Implementation support and capacity building",
        "services_advisory_capacity": "Capacity building workshops for local staff",
        "services_survey_investigations": "Topographical survey and site investigations",
        "services_environmental_social": "Environmental compliance monitoring",
        "tender_documents_files": "RFP.pdf, BOQ.xlsx",
        "technical_proposal_files": "Technical_Proposal.pdf",
        "financial_proposal_files": "Financial_Bid.pdf",
        "work_order_files": "Work_Order.pdf",
        "deliverables_files": "Inception_Report.pdf, Final_Report.pdf",
        "completion_certificate_files": "Completion_Certificate.pdf",
        "invoices_receipts_files": "Invoice1.pdf, Receipt1.pdf",
        "other_documents_files": "Site_Photos.zip"
    }
    sample_row = [sample_data.get(header, "") for header in column_headers]
    sheet.append(sample_row)
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = 18
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column].width = adjusted_width

    instructions_sheet = workbook.create_sheet(title="Instructions")
    instructions_sheet.append(["How to use this template"])
    instructions_lines = [
        "1. Fill one row per completed project. Leave optional columns blank if not applicable.",
        "2. Use YYYY-MM-DD format for dates. Set 'Yes' in the ongoing column for active projects.",
        "3. Separate multiple states, cities, financing authorities, and service notes with commas.",
        "4. Attachment columns should list the exact filenames you will upload (comma-separated).",
        "5. You can extend the template with additional rows as needed before uploading."
    ]
    for line in instructions_lines:
        instructions_sheet.append([line])
    instructions_sheet.column_dimensions["A"].width = 120

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = "past_projects_template.xlsx"
    response_headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=response_headers
    )


@app.post("/api/projects/delete-all")
@require_company_details
async def delete_all_projects(
    request: Request,
    db: Session = Depends(get_db)
):
    """Delete all projects for the current user with confirmation."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    try:
        # Parse request body
        body = await request.json()
        confirmation_text = body.get("confirmation_text", "")

        # Validate confirmation text - must be exactly "DELETE" in all caps
        if confirmation_text != "DELETE":
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "error": "Invalid confirmation text. Please type DELETE (all caps) to confirm."
                }
            )

        # Count projects before deletion
        project_count = db.query(ProjectDB).filter(
            ProjectDB.user_id == current_user.id
        ).count()

        if project_count == 0:
            return JSONResponse(
                status_code=200,
                content={
                    "success": True,
                    "deleted_count": 0,
                    "message": "No projects to delete."
                }
            )

        # Delete all projects for the current user
        db.query(ProjectDB).filter(
            ProjectDB.user_id == current_user.id
        ).delete(synchronize_session=False)

        db.commit()

        return JSONResponse(
            status_code=200,
            content={
                "success": True,
                "deleted_count": project_count,
                "message": f"Successfully deleted {project_count} project(s)."
            }
        )

    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting all projects for user {current_user.id}: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "error": "An error occurred while deleting projects. Please try again."
            }
        )


@app.post("/api/certificates/delete-all")
@require_company_details
async def delete_all_completion_certificates(
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Delete ALL completion certificates from ALL sources for the current user with confirmation.
    This includes:
    1. CertificateDB table (new system with AI extraction)
    2. ProjectDB.documents['completion_certificate'] (old system)
    3. All associated files, vectors, and FAISS indices
    """
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    try:
        # Parse request body
        body = await request.json()
        confirmation_text = body.get("confirmation_text", "")

        # Validate confirmation text - must be exactly "DELETE" in all caps
        if confirmation_text != "DELETE":
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "error": "Invalid confirmation text. Please type DELETE (all caps) to confirm."
                }
            )

        # ============================================================
        # PART 1: Delete from NEW CertificateDB table
        # ============================================================
        
        # Get all certificates from CertificateDB for this user
        certificates_from_db = db.query(CertificateDB).filter(
            CertificateDB.user_id == current_user.id
        ).all()

        certificatedb_count = len(certificates_from_db)
        certificatedb_deleted = 0
        certificatedb_files_deleted = 0
        
        logger.info(f"Found {certificatedb_count} certificate(s) in CertificateDB for user {current_user.id}")

        # Delete each certificate with full cleanup (vectors, FAISS indices, files)
        for certificate in certificates_from_db:
            try:
                # Use the certificate processor's cleanup method
                certificate_processor.delete_certificate_with_cleanup(
                    certificate_id=certificate.id,
                    user_id=current_user.id
                )
                certificatedb_deleted += 1
                certificatedb_files_deleted += 1
                logger.info(f"Deleted certificate {certificate.id} ({certificate.project_name})")
            except Exception as exc:
                logger.error(f"Failed to delete certificate {certificate.id}: {exc}")
                # Continue with other certificates even if one fails

        # ============================================================
        # PART 2: Delete from OLD ProjectDB.documents system
        # ============================================================
        
        # Get all projects for the current user
        projects = db.query(ProjectDB).filter(
            ProjectDB.user_id == current_user.id
        ).all()

        projectdb_cert_count = 0
        projectdb_affected_projects = 0
        projectdb_files_deleted = 0

        # Iterate through all projects and delete completion certificates
        for project in projects:
            if project.documents and 'completion_certificate' in project.documents:
                certificates = project.documents['completion_certificate']

                if certificates and len(certificates) > 0:
                    projectdb_affected_projects += 1
                    projectdb_cert_count += len(certificates)

                    # Delete the actual files from disk
                    for cert_path in certificates:
                        try:
                            if os.path.exists(cert_path):
                                os.remove(cert_path)
                                projectdb_files_deleted += 1
                                logger.info(f"Deleted old certificate file: {cert_path}")
                        except Exception as exc:
                            logger.warning(f"Failed to delete certificate file {cert_path}: {exc}")

                    # Remove completion_certificate from the documents JSON
                    documents_copy = dict(project.documents)
                    if 'completion_certificate' in documents_copy:
                        del documents_copy['completion_certificate']
                    project.documents = documents_copy

        # Commit all project changes
        db.commit()

        # ============================================================
        # PART 3: Calculate totals and return response
        # ============================================================
        
        total_deleted = certificatedb_deleted + projectdb_cert_count
        total_files_deleted = certificatedb_files_deleted + projectdb_files_deleted

        if total_deleted == 0:
            return JSONResponse(
                status_code=200,
                content={
                    "success": True,
                    "deleted_count": 0,
                    "affected_projects": 0,
                    "message": "No completion certificates found to delete."
                }
            )

        # Build detailed message
        message_parts = []
        if certificatedb_deleted > 0:
            message_parts.append(f"{certificatedb_deleted} certificate(s) from Certificate Vault")
        if projectdb_cert_count > 0:
            message_parts.append(f"{projectdb_cert_count} certificate(s) from {projectdb_affected_projects} project(s)")
        
        message = f"Successfully deleted {', '.join(message_parts)}."

        logger.info(f"Delete all certificates completed for user {current_user.id}: {message}")

        return JSONResponse(
            status_code=200,
            content={
                "success": True,
                "deleted_count": total_deleted,
                "certificatedb_count": certificatedb_deleted,
                "projectdb_count": projectdb_cert_count,
                "affected_projects": projectdb_affected_projects,
                "deleted_files": total_files_deleted,
                "message": message
            }
        )

    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting all completion certificates for user {current_user.id}: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "error": "An error occurred while deleting completion certificates. Please try again."
            }
        )


@app.delete("/api/custom-cards/{card_id}")
async def delete_custom_card(request: Request, card_id: int, db: Session = Depends(get_db)):
    """Delete a custom card."""
    from core.dependencies import get_id_for_custom_cards, get_user_id_for_queries

    card_owner_id, entity, entity_type = get_id_for_custom_cards(request, db)
    if not card_owner_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user_id for database queries (company owner's user_id for BD employees)
    user_id_for_query, _ = get_user_id_for_queries(request, db)
    if not user_id_for_query:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find and delete card - query by user_id_for_query
    card = db.query(CustomCardDB).filter(
        and_(CustomCardDB.id == card_id, CustomCardDB.user_id == user_id_for_query)
    ).first()

    if not card:
        raise HTTPException(status_code=404, detail="Custom card not found")

    db.delete(card)
    db.commit()

    return {"message": "Custom card deleted"}

@app.post("/api/custom-cards/{card_id}/search")
async def search_with_custom_card(
    request: Request,
    card_id: int,
    sector: str = Form(""),
    sub_sector: str = Form(""),
    work_type: str = Form(""),
    search: str = Form(""),
    category: str = Form(""),
    state: str = Form(""),
    source: str = Form(""),
    min_value: str = Form(""),
    max_value: str = Form(""),
    db: Session = Depends(get_db)
):
    """Search tenders using custom card with dynamic filters."""
    from core.dependencies import get_id_for_custom_cards, get_user_id_for_queries

    card_owner_id, entity, entity_type = get_id_for_custom_cards(request, db)
    if not card_owner_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get user_id for database queries (company owner's user_id for BD employees)
    user_id_for_query, _ = get_user_id_for_queries(request, db)
    if not user_id_for_query:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if user has complete company details (skip for BD employees)
    if entity_type == 'user' and not user_has_complete_company_details(entity.id, db):
        raise HTTPException(status_code=403, detail="Complete company details required")

    # Find the card - query by user_id_for_query
    card = db.query(CustomCardDB).filter(
        and_(CustomCardDB.id == card_id, CustomCardDB.user_id == user_id_for_query)
    ).first()

    if not card:
        raise HTTPException(status_code=404, detail="Custom card not found")

    # Build search query using card criteria + dynamic filters
    query = db.query(TenderDB)

    # Filter out expired tenders and awarded tenders
    # Awarded tenders should only appear in tender_management page
    now = datetime.utcnow()
    query = query.filter(
        TenderDB.awarded == False,  # Exclude awarded tenders from search
        or_(
            TenderDB.deadline == None,
            TenderDB.deadline >= now
        )
    )

    # Apply core search terms from card (using OR logic for broader results)
    if card.core_search_terms:
        # Split on commas and strip whitespace from each keyword
        search_terms = [term.strip() for term in card.core_search_terms.split(',') if term.strip()]

        # Build OR conditions for all terms across ALL tender fields
        term_conditions = []
        for term in search_terms:
            search_term = f"%{term}%"
            term_conditions.append(
                or_(
                    TenderDB.title.ilike(search_term),
                    TenderDB.summary.ilike(search_term),
                    TenderDB.authority.ilike(search_term),
                    TenderDB.category.ilike(search_term),
                    TenderDB.organisation_chain.ilike(search_term),
                    TenderDB.tender_reference_number.ilike(search_term),
                    TenderDB.tender_id.ilike(search_term),
                    TenderDB.tender_type.ilike(search_term),
                    TenderDB.tender_category.ilike(search_term),
                    TenderDB.state.ilike(search_term),
                    TenderDB.source.ilike(search_term),
                    # Search in JSON fields using PostgreSQL JSON operators
                    TenderDB.work_item_details.cast(String).ilike(search_term),
                    TenderDB.tender_inviting_authority.cast(String).ilike(search_term),
                    # Cast tags array to string for searching
                    TenderDB.tags.cast(String).ilike(search_term)
                )
            )
        # Combine all term conditions with OR (any keyword match = include tender)
        if term_conditions:
            query = query.filter(or_(*term_conditions))

    # Apply card filters
    if card.state:
        query = query.filter(TenderDB.state == card.state)
    if card.source:
        query = query.filter(TenderDB.source == card.source)
    if card.tender_type:
        query = query.filter(TenderDB.tender_type == card.tender_type)

    # Apply additional search filter (comprehensive search across all fields)
    if search and search.strip():
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                TenderDB.title.ilike(search_term),
                TenderDB.summary.ilike(search_term),
                TenderDB.authority.ilike(search_term),
                TenderDB.category.ilike(search_term),
                TenderDB.organisation_chain.ilike(search_term),
                TenderDB.tender_reference_number.ilike(search_term),
                TenderDB.tender_id.ilike(search_term),
                TenderDB.tender_type.ilike(search_term),
                TenderDB.tender_category.ilike(search_term),
                TenderDB.state.ilike(search_term),
                TenderDB.source.ilike(search_term),
                TenderDB.work_item_details.cast(String).ilike(search_term),
                TenderDB.tender_inviting_authority.cast(String).ilike(search_term),
                TenderDB.tags.cast(String).ilike(search_term)
            )
        )

    # Apply dynamic filters (override card defaults if provided)
    # Handle empty values as 'All' - no filtering needed
    if sector and sector.strip() and sector != "All":
        # For now, search in category field - you may need to adjust based on your data structure
        query = query.filter(TenderDB.category.ilike(f"%{sector}%"))
    elif card.sector and card.sector != "All":
        query = query.filter(TenderDB.category.ilike(f"%{card.sector}%"))

    if sub_sector and sub_sector.strip() and sub_sector != "All":
        query = query.filter(TenderDB.category.ilike(f"%{sub_sector}%"))
    elif card.sub_sector and card.sub_sector != "All":
        query = query.filter(TenderDB.category.ilike(f"%{card.sub_sector}%"))

    if work_type and work_type.strip() and work_type != "All":
        query = query.filter(TenderDB.tender_type.ilike(f"%{work_type}%"))
    elif card.work_type and card.work_type != "All":
        query = query.filter(TenderDB.tender_type.ilike(f"%{card.work_type}%"))

    # Apply category filter
    if category and category.strip():
        query = query.filter(TenderDB.category.ilike(f"%{category}%"))

    # Apply state filter
    if state and state.strip():
        query = query.filter(TenderDB.state == state)

    # Apply source filter
    if source and source.strip():
        query = query.filter(TenderDB.source == source)

    # Apply value range filters
    if min_value and min_value.strip():
        try:
            min_val = float(min_value)
            query = query.filter(TenderDB.estimated_value >= min_val)
        except ValueError:
            pass

    if max_value and max_value.strip():
        try:
            max_val = float(max_value)
            query = query.filter(TenderDB.estimated_value <= max_val)
        except ValueError:
            pass

    # Execute query with pagination
    page = int(request.query_params.get('page', 1))
    per_page = 20
    skip = (page - 1) * per_page

    total_tenders = query.count()
    tenders = query.order_by(desc(TenderDB.published_at)).offset(skip).limit(per_page).all()

    # Get seen tender IDs for the current entity
    seen_tender_ids = set()
    if entity_type == 'user':
        seen_records = db.query(SeenTenderDB).filter(SeenTenderDB.user_id == entity_id).all()
        seen_tender_ids = {record.tender_id for record in seen_records}
    elif entity_type == 'bd_employee':
        seen_records = db.query(SeenTenderDB).filter(SeenTenderDB.employee_id == entity_id).all()
        seen_tender_ids = {record.tender_id for record in seen_records}

    # Pagination info
    total_pages = (total_tenders + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages

    return {
        "tenders": [
            {
                "id": tender.id,
                "title": tender.title,
                "authority": tender.authority,
                "state": tender.state,
                "category": tender.category,
                "tender_type": tender.tender_type,
                "source": tender.source,
                "estimated_value": tender.estimated_value,
                "summary": tender.summary,
                "published_at": tender.published_at.isoformat() if tender.published_at else None,
                "deadline": tender.deadline.isoformat() if tender.deadline else None,
                "is_seen": tender.id in seen_tender_ids,
                "metadata": {
                    "work_item_details": tender.work_item_details or {},
                    "critical_dates": tender.critical_dates or {},
                    "tender_documents": tender.tender_documents or {},
                    "tender_inviting_authority": tender.tender_inviting_authority or {}
                }
            } for tender in tenders
        ],
        "total_tenders": total_tenders,
        "page": page,
        "total_pages": total_pages,
        "has_prev": has_prev,
        "has_next": has_next,
        "card": {
            "id": card.id,
            "card_name": card.card_name,
            "core_search_terms": card.core_search_terms
        }
    }

@app.get("/api/custom-cards/{card_id}/tenders")
def get_custom_card_tenders(
    card_id: int,
    skip: int = 0,
    limit: int = 20,
    request: Request = None, # type: ignore
    db: Session = Depends(get_db)
):
    """Get tenders matching a custom card's criteria."""
    from core.dependencies import get_id_for_custom_cards

    card_owner_id, entity, entity_type = get_id_for_custom_cards(request, db)
    if not card_owner_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the card
    card = db.query(CustomCardDB).filter(
        and_(CustomCardDB.id == card_id, CustomCardDB.user_id == card_owner_id)
    ).first()

    if not card:
        raise HTTPException(status_code=404, detail="Custom card not found")

    # Build query based on card criteria
    query = db.query(TenderDB)

    # Filter out expired tenders (unless awarded)
    now = datetime.utcnow()
    query = query.filter(
        or_(
            TenderDB.deadline == None,
            TenderDB.deadline >= now,
            TenderDB.awarded == True
        )
    )

    # Core search terms are mandatory - if no terms provided, return empty
    if not card.core_search_terms or not card.core_search_terms.strip(): # type: ignore
        return {
            "tenders": [],
            "total": 0,
            "card": {
                "id": card.id,
                "card_name": card.card_name,
                "core_search_terms": card.core_search_terms,
                "state": card.state,
                "tender_type": card.tender_type
            }
        }

    # Apply core search terms (comma-separated, OR logic)
    core_terms = [term.strip() for term in card.core_search_terms.split(',') if term.strip()] # type: ignore
    if core_terms:
        core_conditions = []
        for term in core_terms:
            term_pattern = f"%{term}%"
            core_conditions.append(
                or_(
                    TenderDB.title.ilike(term_pattern),
                    TenderDB.summary.ilike(term_pattern)
                )
            )
        query = query.filter(or_(*core_conditions))

    # Apply state filter (comma-separated, OR logic)
    if card.state and card.state.strip(): # type: ignore
        state_terms = [state.strip() for state in card.state.split(',') if state.strip()] # type: ignore
        if state_terms:
            state_conditions = [TenderDB.state == state for state in state_terms]
            query = query.filter(or_(*state_conditions))

    # Apply source filter (comma-separated, OR logic)
    if card.source and card.source.strip(): # type: ignore
        source_terms = [source.strip() for source in card.source.split(',') if source.strip()] # type: ignore
        if source_terms:
            source_conditions = [TenderDB.source == source for source in source_terms]
            query = query.filter(or_(*source_conditions))

    # Apply tender type filter (comma-separated, OR logic)
    if card.tender_type and card.tender_type.strip(): # type: ignore
        type_terms = [t_type.strip() for t_type in card.tender_type.split(',') if t_type.strip()] # type: ignore
        if type_terms:
            type_conditions = [TenderDB.tender_type == t_type for t_type in type_terms]
            query = query.filter(or_(*type_conditions))

    # Get total count
    total_tenders = query.count()

    # Apply pagination and ordering
    tenders = query.order_by(desc(TenderDB.published_at)).offset(skip).limit(limit).all()

    return {
        "tenders": tenders, #[tender.to_frontend_format() for tender in tenders],
        "total": total_tenders,
        "card": {
            "id": card.id,
            "card_name": card.card_name,
            "core_search_terms": card.core_search_terms,
            "state": card.state,
            "tender_type": card.tender_type
        }
    }

# User profile API endpoints
@app.put("/api/user/profile")
@require_company_details
async def update_profile(
    request: Request,
    name: str = Form(...),
    company: str = Form(""),
    role: str = Form(""),
    db: Session = Depends(get_db)
):
    """Update user profile."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Update user info
    current_user.name = name # type: ignore
    current_user.company = company # type: ignore
    current_user.role = role # type: ignore

    db.commit()

    return RedirectResponse(url="/home", status_code=302)

# Company Code API endpoints
@app.get("/api/company-codes")
async def get_company_codes(request: Request, db: Session = Depends(get_db)):
    """Get user's company codes."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()

    # Add employee count for each company code
    codes_with_count = []
    for code in company_codes:
        employee_count = db.query(EmployeeDB).filter(EmployeeDB.company_code_id == code.id).count()
        codes_with_count.append({
            "id": code.id,
            "company_name": code.company_name,
            "company_code": code.company_code,
            "created_at": code.created_at.isoformat(),
            "employee_count": employee_count
        })

    return {"company_codes": codes_with_count}

@app.post("/api/company-codes")
async def create_company_code(
    request: Request,
    company_name: str = Form(...),
    company_code: str = Form(...),
    db: Session = Depends(get_db)
):
    """Create a new company code."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if company code already exists
    existing_code = db.query(CompanyCodeDB).filter(CompanyCodeDB.company_code == company_code.upper()).first()
    if existing_code:
        raise HTTPException(status_code=400, detail="Company code already exists")

    # Create new company code
    new_code = CompanyCodeDB(
        user_id=current_user.id,
        company_name=company_name,
        company_code=company_code.upper(),
        created_at=datetime.utcnow()
    )

    db.add(new_code)
    db.commit()
    db.refresh(new_code)

    return {"message": "Company code created", "company_code_id": new_code.id}

# API endpoints for frontend
@app.get("/api/tenders")
async def get_tenders(
    skip: int = 0,
    limit: int = 20,
    search: str = "",
    category: str = "",
    state: str = "",
    db: Session = Depends(get_db)
):
    """Get tenders with pagination and filters."""
    query = db.query(TenderDB)

    # Filter out expired tenders (unless awarded)
    now = datetime.utcnow()
    query = query.filter(
        or_(
            TenderDB.deadline == None,
            TenderDB.deadline >= now,
            TenderDB.awarded == True
        )
    )

    # Apply filters
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                TenderDB.title.ilike(search_term),
                TenderDB.authority.ilike(search_term),
                TenderDB.summary.ilike(search_term)
            )
        )
    
    if category:
        query = query.filter(TenderDB.category == category)
    
    if state:
        query = query.filter(TenderDB.state == state)
    
    tenders = query.order_by(desc(TenderDB.published_at)).offset(skip).limit(limit).all()
    
    return {
        "tenders": [tender.to_frontend_format() for tender in tenders],
        "total": query.count()
    }

@app.get("/api/tenders/recommended")
async def get_recommended_tenders(
    request: Request,
    sector: str = "",
    search: str = "",
    state: str = "",
    category: str = "",
    source: str = "",
    min_value: Optional[float] = None,
    max_value: Optional[float] = None,
    page: int = 1,
    per_page: int = 20,
    db: Session = Depends(get_db)
):
    """
    Get recommended tenders based on user's certificates and projects.

    Enhanced Scoring System:
    - Hybrid Filter: keyword OR sector OR location (changed to OR logic for looser matching)
    - Semantic AI: 15% weight using OpenAI embeddings for meaning-based matching
    - Text Matching: 35% weight for keyword/phrase overlap
    - Sector: 20% weight for category matching
    - Services: 15% weight for service type matching
    - Location: 10% weight for geographic matching
    - Authority: 5% weight for client matching

    Returns tenders in two categories (very loose thresholds for maximum coverage):
    - more_reco: Score > 65% (Highly Recommended) - lowered from 70%
    - less_reco: Score 30-65% (Other Recommendations) - lowered from 45%
    """
    # Get current user or BD employee
    from core.dependencies import get_current_user_or_bd_employee
    entity, entity_type = get_current_user_or_bd_employee(request, db)
    if not entity:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # For recommendation scoring, we need user_id (company owner's ID for BD employees)
    if entity_type == 'user':
        current_user = entity
        user_id_for_scoring = current_user.id
    else:
        # BD employee - get company owner's user_id
        from core.dependencies import get_user_id_for_queries
        user_id_for_scoring, _ = get_user_id_for_queries(request, db)
        if not user_id_for_scoring:
            raise HTTPException(status_code=401, detail="Company owner not found")
        # Get the actual user for scoring
        current_user = db.query(UserDB).filter(UserDB.id == user_id_for_scoring).first()
        if not current_user:
            raise HTTPException(status_code=401, detail="Company owner not found")

    try:
        import time
        start_time = time.time()

        # Initialize embedding manager for semantic similarity
        embedding_manager = None
        user_profile_embedding = None

        try:
            # Try to initialize with OpenAI (requires OPENAI_API_KEY env var)
            embedding_manager = TenderEmbeddingManager(db)
            logger.info("TenderEmbeddingManager initialized successfully")

            # Generate user profile embedding once (cached for reuse)
            user_profile_embedding = embedding_manager.get_user_profile_embedding(user_id_for_scoring)
            if user_profile_embedding is not None:
                logger.info(f"User profile embedding generated: {len(user_profile_embedding)} dimensions")
            else:
                logger.warning("Failed to generate user profile embedding (no project data)")
        except Exception as e:
            logger.warning(f"Could not initialize semantic similarity (OpenAI unavailable): {e}")
            # Continue without semantic similarity (uses keyword matching only)

        # Initialize recommendation scorer
        scorer = TenderRecommendationScorer(
            db,
            current_user.id,
            embedding_manager=embedding_manager,
            user_profile_embedding=user_profile_embedding
        )

        # Build user profile from certificates and projects
        user_profile = scorer.build_user_profile()

        # Check if user has any expertise data
        if not any([
            user_profile.get("sectors"),
            user_profile.get("keywords"),
            user_profile.get("services")
        ]):
            return {
                "more_reco": [],
                "less_reco": [],
                "total_more": 0,
                "total_less": 0,
                "message": "No certificates or projects found. Add your expertise to get recommendations.",
                "performance": {
                    "total_time_ms": int((time.time() - start_time) * 1000),
                    "semantic_enabled": False
                }
            }

        # Query tenders with filters (same as regular search)
        query = db.query(TenderDB)

        # Filter out expired tenders
        now = datetime.utcnow()
        query = query.filter(
            or_(
                TenderDB.deadline == None,
                TenderDB.deadline >= now,
                TenderDB.awarded == False
            )
        )

        # Apply sector filter if provided
        if sector:
            query = query.filter(
                or_(
                    TenderDB.category.ilike(f"%{sector}%"),
                    TenderDB.tender_type.ilike(f"%{sector}%"),
                    TenderDB.tender_category.ilike(f"%{sector}%")
                )
            )

        # Apply other filters
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    TenderDB.title.ilike(search_term),
                    TenderDB.authority.ilike(search_term),
                    TenderDB.summary.ilike(search_term),
                    cast(TenderDB.work_item_details, String).ilike(search_term)
                )
            )

        if category:
            query = query.filter(TenderDB.category == category)

        if state:
            query = query.filter(TenderDB.state == state)

        if source:
            query = query.filter(TenderDB.source == source)

        if min_value is not None:
            query = query.filter(TenderDB.estimated_value >= min_value)

        if max_value is not None:
            query = query.filter(TenderDB.estimated_value <= max_value)

        # Get all matching tenders (limit to prevent performance issues)
        tenders = query.order_by(desc(TenderDB.published_at)).limit(1000).all()

        # Score each tender (with debug logging)
        scored_tenders = []
        for tender in tenders:
            score, metadata = scorer.score_tender(tender)
            # Debug logging for petroleum tenders
            if 'petroleum' in tender.title.lower() or 'oil' in tender.title.lower() or 'iocl' in tender.authority.lower():
                logger.info(f"PETROLEUM TENDER SCORED: '{tender.title[:60]}' -> {score:.1f}% | Breakdown: {metadata}")
            if score >= 30:  # Lowered from 45 to 30 for maximum coverage
                scored_tenders.append((tender, score, metadata))

        # Categorize results
        categorized = scorer.categorize_results(scored_tenders)

        # Paginate results within each category
        # For simplicity, we'll return all results and let frontend handle pagination
        # In production, you might want to implement server-side pagination

        # Get seen tender IDs for the current entity
        seen_tender_ids = set()
        if entity_type == 'user':
            seen_records = db.query(SeenTenderDB).filter(SeenTenderDB.user_id == entity.id).all()
            seen_tender_ids = {record.tender_id for record in seen_records}
        elif entity_type == 'bd_employee':
            seen_records = db.query(SeenTenderDB).filter(SeenTenderDB.employee_id == entity.id).all()
            seen_tender_ids = {record.tender_id for record in seen_records}

        # Convert tenders to frontend format
        more_reco_formatted = []
        for item in categorized["more_reco"]:
            tender_data = item["tender"].to_frontend_format()
            tender_data["recommendation_score"] = item["score"]
            tender_data["match_reasons"] = item["match_reasons"]
            tender_data["semantic_similarity"] = item.get("semantic_similarity", 0)
            tender_data["is_seen"] = tender_data["id"] in seen_tender_ids
            more_reco_formatted.append(tender_data)

        less_reco_formatted = []
        for item in categorized["less_reco"]:
            tender_data = item["tender"].to_frontend_format()
            tender_data["recommendation_score"] = item["score"]
            tender_data["match_reasons"] = item["match_reasons"]
            tender_data["semantic_similarity"] = item.get("semantic_similarity", 0)
            tender_data["is_seen"] = tender_data["id"] in seen_tender_ids
            less_reco_formatted.append(tender_data)

        # Calculate performance metrics
        total_time_ms = int((time.time() - start_time) * 1000)

        # Get cache statistics if available
        cache_stats = {}
        if embedding_manager:
            cache_stats = embedding_manager.get_cache_stats()

        performance = {
            "total_time_ms": total_time_ms,
            "tenders_evaluated": len(tenders),
            "tenders_passed_filter": len(scored_tenders),
            "semantic_enabled": embedding_manager is not None and user_profile_embedding is not None,
            "cache_stats": cache_stats
        }

        logger.info(f"Recommendations generated in {total_time_ms}ms: {categorized['total_more']} highly recommended, {categorized['total_less']} other recommendations")
        if cache_stats:
            logger.info(f"Cache performance: {cache_stats.get('cache_hit_rate', 0):.1f}% hit rate, {cache_stats.get('api_calls', 0)} API calls")

        return {
            "more_reco": more_reco_formatted,
            "less_reco": less_reco_formatted,
            "total_more": categorized["total_more"],
            "total_less": categorized["total_less"],
            "message": None,
            "performance": performance
        }

    except Exception as e:
        logger.error(f"Error generating recommendations: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generating recommendations: {str(e)}")

@app.get("/api/tenders/recommended/debug")
async def debug_recommendations(
    request: Request,
    sample_size: int = 10,
    db: Session = Depends(get_db)
):
    """
    Debug endpoint to show user profile and sample tender matches.
    Helps diagnose why recommendations are/aren't working.
    """
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        # Build user profile
        scorer = TenderRecommendationScorer(db, current_user.id)
        profile = scorer.build_user_profile()

        # Sample tenders for testing
        sample_tenders = db.query(TenderDB).filter(
            or_(
                TenderDB.deadline == None,
                TenderDB.deadline >= datetime.utcnow()
            )
        ).order_by(desc(TenderDB.published_at)).limit(sample_size).all()

        # Score each sample tender
        debug_results = []
        for tender in sample_tenders:
            score, metadata = scorer.score_tender(tender)
            debug_results.append({
                "tender_id": tender.id,
                "title": tender.title,
                "category": tender.category,
                "tender_type": tender.tender_type,
                "total_score": score,
                "sector_score": metadata["sector_match"],
                "scope_score": metadata["scope_match"],
                "services_score": metadata["services_match"],
                "location_score": metadata["location_match"],
                "match_reasons": metadata["match_reasons"],
                "passed_threshold": score >= 50
            })

        return {
            "user_profile": {
                "total_sectors": len(profile["sectors"]),
                "sectors": list(profile["sectors"])[:20],
                "total_keywords": len(profile["keywords"]),
                "keywords_sample": list(profile["keywords"])[:50],
                "total_phrases": len(profile["phrases"]),
                "phrases_sample": list(profile["phrases"])[:30],
                "total_services": len(profile["services"]),
                "services": list(profile["services"])[:20],
                "total_locations": len(profile["locations"]),
                "locations": list(profile["locations"]),
                "total_clients": len(profile["clients"]),
            },
            "matching_config": {
                "fuzzy_threshold": scorer.FUZZY_THRESHOLD,
                "fuzzy_strong_threshold": scorer.FUZZY_STRONG_THRESHOLD,
                "minimum_score": 50,
                "weights": {
                    "sector": scorer.WEIGHT_SECTOR,
                    "scope": scorer.WEIGHT_SCOPE_TEXT,
                    "services": scorer.WEIGHT_SERVICES,
                    "location": scorer.WEIGHT_LOCATION,
                    "authority": scorer.WEIGHT_AUTHORITY
                }
            },
            "sample_matches": debug_results,
            "summary": {
                "total_sampled": len(sample_tenders),
                "passed_threshold": sum(1 for r in debug_results if r["passed_threshold"]),
                "failed_threshold": sum(1 for r in debug_results if not r["passed_threshold"])
            }
        }

    except Exception as e:
        logger.error(f"Error in debug endpoint: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Debug failed: {str(e)}")

@app.get("/api/tender/{tender_id}")
async def get_tender(tender_id: str, request: Request, db: Session = Depends(get_db)):
    """Get single tender details."""
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if tender has expired
    current_user = get_current_user(request, db)
    if current_user and tender.deadline:
        now = datetime.utcnow()
        if tender.deadline < now:
            # Check if user has this tender favorited or shortlisted
            is_favorited = db.query(FavoriteDB).filter(
                FavoriteDB.user_id == current_user.id,
                FavoriteDB.tender_id == tender_id
            ).first()

            is_shortlisted = db.query(ShortlistedTenderDB).filter(
                ShortlistedTenderDB.user_id == current_user.id,
                ShortlistedTenderDB.tender_id == tender_id
            ).first()

            if is_favorited or is_shortlisted:
                raise HTTPException(
                    status_code=410,
                    detail="This tender has expired. Deadline has passed and access is no longer available."
                )

    return tender.to_frontend_format()

@app.get("/debug/sync", response_class=HTMLResponse)
async def debug_sync_page(request: Request):
    """Debug page for checking and fixing deliverable sync issues."""
    return templates.TemplateResponse("debug_sync.html", {
        "request": request,
        "selected_font": get_active_font()
    })

@app.get("/api/debug/project/{project_id}/documents")
async def debug_project_documents(request: Request, project_id: int, db: Session = Depends(get_db)):
    """Check what documents are actually stored in the project."""
    project = db.query(ProjectDB).filter(ProjectDB.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return {
        "project_id": project.id,
        "project_name": project.project_name,
        "documents": project.documents,
        "documents_type": str(type(project.documents)),
        "task_deliverables_count": len(project.documents.get("Task Deliverables", [])) if project.documents else 0,
        "all_categories": list(project.documents.keys()) if project.documents and isinstance(project.documents, dict) else []
    }

@app.post("/api/debug/project/{project_id}/cleanup-duplicates")
async def cleanup_project_duplicates(request: Request, project_id: int, db: Session = Depends(get_db)):
    """Remove duplicate deliverables from a project."""
    project = db.query(ProjectDB).filter(ProjectDB.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    if not project.documents:
        return {
            "message": "No documents found",
            "removed_count": 0
        }
    
    project_docs = project.documents if isinstance(project.documents, dict) else {}
    deliverable_category = "Task Deliverables"
    
    if deliverable_category not in project_docs:
        return {
            "message": "No task deliverables found",
            "removed_count": 0
        }
    
    deliverables = project_docs[deliverable_category]
    original_count = len(deliverables)
    
    # Remove duplicates
    new_deliverables = []
    seen_task_file_ids = set()
    removed_items = []
    
    for doc in deliverables:
        if isinstance(doc, dict):
            task_file_id = doc.get('task_file_id')
            if task_file_id:
                if task_file_id in seen_task_file_ids:
                    # Duplicate found
                    removed_items.append({
                        "filename": doc.get('original_filename', 'unknown'),
                        "reason": "Duplicate task_file_id"
                    })
                    continue
                
                seen_task_file_ids.add(task_file_id)
                new_deliverables.append(doc)
            else:
                # No task_file_id - check if it has complete metadata
                has_metadata = doc.get('uploaded_by') and doc.get('task_title')
                
                if has_metadata:
                    new_deliverables.append(doc)
                else:
                    # Missing metadata - likely a duplicate with system-generated name
                    removed_items.append({
                        "filename": doc.get('file_path', 'unknown'),
                        "reason": "Missing metadata (uploaded_by, task_title)"
                    })
        else:
            # Old format (string path), keep as-is
            new_deliverables.append(doc)
    
    removed_count = original_count - len(new_deliverables)
    
    if removed_count > 0:
        project_docs[deliverable_category] = new_deliverables
        project.documents = project_docs
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(project, "documents")
        db.commit()
    
    return {
        "message": f"Removed {removed_count} duplicate entries",
        "original_count": original_count,
        "new_count": len(new_deliverables),
        "removed_items": removed_items
    }

@app.get("/api/debug/tender/{tender_id}/sync-status")
async def debug_tender_sync_status(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Debug endpoint to check tender-project-deliverable sync status."""
    # Try to get current user, but allow access even without authentication for debugging
    current_user = get_current_user(request, db)
    
    # Find tender
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    
    # Check project
    project = db.query(ProjectDB).filter(ProjectDB.source_tender_id == tender.id).first()
    
    # Check assignments and tasks
    assignments = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.tender_id == tender.id).all()
    
    task_files = []
    for assignment in assignments:
        tasks = db.query(TaskDB).filter(TaskDB.assignment_id == assignment.id).all()
        for task in tasks:
            files = db.query(TaskFileDB).filter(TaskFileDB.task_id == task.id).all()
            for f in files:
                employee = db.query(EmployeeDB).filter(EmployeeDB.id == f.employee_id).first()
                task_files.append({
                    "file_id": f.id,
                    "filename": f.filename,
                    "task_title": task.title,
                    "uploaded_by": employee.name if employee else "Unknown",
                    "uploaded_at": f.created_at.isoformat() if f.created_at else None
                })
    
    # Check project deliverables
    project_deliverables = []
    synced_file_ids = set()
    if project and project.documents:
        deliverables = project.documents.get("Task Deliverables", [])
        for d in deliverables:
            if isinstance(d, dict):
                project_deliverables.append(d)
                if "task_file_id" in d:
                    synced_file_ids.add(d["task_file_id"])
    
    # Identify missing files
    missing_files = [f for f in task_files if f["file_id"] not in synced_file_ids]
    
    return {
        "tender": {
            "id": tender.id,
            "title": tender.title,
            "awarded": tender.awarded,
            "awarded_by": tender.awarded_by
        },
        "project": {
            "exists": project is not None,
            "id": project.id if project else None,
            "name": project.project_name if project else None,
            "source_tender_id": project.source_tender_id if project else None
        },
        "task_files_count": len(task_files),
        "task_files": task_files,
        "project_deliverables_count": len(project_deliverables),
        "project_deliverables": project_deliverables,
        "missing_files_count": len(missing_files),
        "missing_files": missing_files,
        "status": "OK" if len(missing_files) == 0 else "MISSING_FILES"
    }

@app.get("/api/debug/tender/{tender_id}/force-sync")
@app.post("/api/debug/tender/{tender_id}/force-sync")
async def debug_force_sync_deliverables(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Force sync all deliverables for a tender to its project."""
    # Try to get current user, but allow access even without authentication for debugging
    current_user = get_current_user(request, db)
    
    import uuid
    
    # Find tender
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    
    if not tender.awarded:
        raise HTTPException(status_code=400, detail="Tender is not awarded")
    
    # Find or create project
    project = db.query(ProjectDB).filter(ProjectDB.source_tender_id == tender.id).first()
    
    if not project and tender.title:
        # Try to find by name
        project = db.query(ProjectDB).filter(
            and_(
                ProjectDB.user_id == tender.awarded_by,
                ProjectDB.project_name.ilike(f"%{tender.title[:50]}%")
            )
        ).first()
        
        if project:
            project.source_tender_id = tender.id
            db.commit()
    
    if not project:
        raise HTTPException(status_code=404, detail="No project found for this tender")
    
    # Get all task files
    assignments = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.tender_id == tender.id).all()
    
    synced_count = 0
    skipped_count = 0
    
    for assignment in assignments:
        tasks = db.query(TaskDB).filter(TaskDB.assignment_id == assignment.id).all()
        for task in tasks:
            files = db.query(TaskFileDB).filter(TaskFileDB.task_id == task.id).all()
            
            for task_file in files:
                # Check if already synced
                project_docs = project.documents or {}
                if not isinstance(project_docs, dict):
                    project_docs = {}
                
                deliverable_category = "Task Deliverables"
                if deliverable_category not in project_docs:
                    project_docs[deliverable_category] = []
                
                already_synced = any(
                    isinstance(doc, dict) and doc.get('task_file_id') == task_file.id
                    for doc in project_docs[deliverable_category]
                )
                
                if already_synced:
                    skipped_count += 1
                    continue
                
                # Sync the file
                deliverables_dir = "project_deliverables"
                os.makedirs(deliverables_dir, exist_ok=True)
                
                timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                unique_filename = f"{timestamp}_{uuid.uuid4().hex[:8]}_{task_file.filename}"
                file_path = os.path.join(deliverables_dir, unique_filename)
                
                with open(file_path, "wb") as f:
                    f.write(task_file.file_data)
                
                employee = db.query(EmployeeDB).filter(EmployeeDB.id == task_file.employee_id).first()
                
                deliverable_info = {
                    "file_path": file_path,
                    "original_filename": task_file.filename,
                    "uploaded_by": employee.name if employee else "Unknown",
                    "uploaded_at": task_file.created_at.isoformat() if task_file.created_at else datetime.utcnow().isoformat(),
                    "task_title": task.title,
                    "task_file_id": task_file.id,
                    "file_size": task_file.file_size,
                    "description": task_file.description
                }
                
                project_docs[deliverable_category].append(deliverable_info)
                synced_count += 1
    
    # Update project documents and mark as modified for SQLAlchemy
    from sqlalchemy.orm.attributes import flag_modified
    project.documents = project_docs
    flag_modified(project, "documents")
    
    db.commit()
    db.refresh(project)
    
    return {
        "message": "Sync completed",
        "tender_id": tender.id,
        "project_id": project.id,
        "synced_count": synced_count,
        "skipped_count": skipped_count,
        "total_deliverables": len(project.documents.get("Task Deliverables", [])) if project.documents else 0
    }

@app.post("/api/tender/{tender_id}/award")
async def award_tender(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Mark a tender as awarded for employee task assignment."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the tender
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if tender is already awarded
    if tender.awarded is True:
        raise HTTPException(status_code=400, detail="Tender is already awarded")

    # Set worked_by metadata before finalizing
    tender.worked_by_name = entity.name
    tender.worked_by_type = entity_type

    # Finalize award and auto-create project
    project = finalize_tender_award_for_user(db, tender, entity_id)

    db.commit()

    # Build response with project info if created
    response = {
        "message": "Tender marked as awarded successfully",
        "tender_id": tender_id
    }

    if project:
        response["project_id"] = project.id
        response["project_name"] = project.project_name
        response["project_created"] = True
    else:
        response["project_created"] = False

    return response


@app.delete("/api/tender/{tender_id}/award")
async def delete_awarded_tender(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Delete an awarded tender. Only available from the awarded tenders view."""
    from core.dependencies import get_id_for_tender_management

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    if tender.awarded is not True:
        raise HTTPException(status_code=400, detail="Tender is not marked as awarded")

    # BD employees can only delete their own awarded tenders
    # Admins can delete any awarded tender from their company
    if entity_type == 'bd_employee' and tender.awarded_by != entity_id:
        raise HTTPException(status_code=403, detail="You are not allowed to delete this awarded tender")

    try:
        db.delete(tender)
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error(f"Failed to delete awarded tender {tender_id}: {exc}")
        raise HTTPException(status_code=500, detail="Failed to delete awarded tender") from exc

    return {"message": "Awarded tender deleted successfully", "tender_id": tender_id}

# Notification API endpoints
@app.get("/api/notifications")
async def get_notifications(request: Request, db: Session = Depends(get_db)):
    """Get all unread notifications for the current user or BD employee."""
    from core.dependencies import get_id_for_tender_management, get_all_bd_employee_ids_for_company

    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Determine IDs to query based on user type
    if entity_type == 'user':
        # Admin: Get all BD employee IDs + their own ID
        bd_employee_ids = get_all_bd_employee_ids_for_company(entity_id, db)
        all_ids = [entity_id] + bd_employee_ids
    else:
        # BD Employee: Only their own ID
        all_ids = [entity_id]

    # Get all unread notifications for the entity(ies), ordered by creation date (newest first)
    notifications = db.query(NotificationDB).filter(
        NotificationDB.user_id.in_(all_ids),
        NotificationDB.is_read == False
    ).order_by(desc(NotificationDB.created_at)).all()

    # Format notifications for frontend
    notification_list = []
    for notif in notifications:
        # Determine notification icon and title based on type
        icon = "⏰"
        title = 'Tender deadline approaching'

        if notif.notification_type == 'tender_expired':
            icon = "❌"
            title = 'Tender expired and removed'
        elif "10 days" in notif.message or "10d" in notif.notification_type:
            icon = "📅"
        elif "7 days" in notif.message or "7d" in notif.notification_type:
            icon = "⏰"
        elif "5 days" in notif.message or "5d" in notif.notification_type:
            icon = "⚠️"
        elif "2 days" in notif.message or "2d" in notif.notification_type:
            icon = "🚨"
        elif "today" in notif.notification_type or "expires today" in notif.message.lower():
            icon = "🔥"

        # Calculate time ago
        time_diff = datetime.utcnow() - notif.created_at
        if time_diff.days > 0:
            time_ago = f"{time_diff.days} day{'s' if time_diff.days > 1 else ''} ago"
        elif time_diff.seconds >= 3600:
            hours = time_diff.seconds // 3600
            time_ago = f"{hours} hour{'s' if hours > 1 else ''} ago"
        else:
            minutes = max(1, time_diff.seconds // 60)
            time_ago = f"{minutes} minute{'s' if minutes > 1 else ''} ago"

        notification_list.append({
            'id': notif.id,
            'icon': icon,
            'title': title,
            'message': notif.message,
            'time_ago': time_ago,
            'tender_id': notif.tender_id,
            'is_read': notif.is_read
        })

    return {
        'notifications': notification_list,
        'count': len(notification_list)
    }

@app.post("/api/notifications/{notification_id}/mark-read")
async def mark_notification_read(request: Request, notification_id: int, db: Session = Depends(get_db)):
    """Mark a specific notification as read."""
    from core.dependencies import get_id_for_tender_management, get_all_bd_employee_ids_for_company
    
    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Determine IDs to query based on user type
    if entity_type == 'user':
        # Admin: Can mark notifications for all BD employees + their own
        bd_employee_ids = get_all_bd_employee_ids_for_company(entity_id, db)
        all_ids = [entity_id] + bd_employee_ids
    else:
        # BD Employee: Only their own ID
        all_ids = [entity_id]

    # Find the notification
    notification = db.query(NotificationDB).filter(
        NotificationDB.id == notification_id,
        NotificationDB.user_id.in_(all_ids)
    ).first()

    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")

    # Mark as read
    notification.is_read = True  # type: ignore
    db.commit()

    return {"message": "Notification marked as read"}

@app.post("/api/notifications/mark-all-read")
async def mark_all_notifications_read(request: Request, db: Session = Depends(get_db)):
    """Mark all notifications as read for the current user or BD employee."""
    from core.dependencies import get_id_for_tender_management, get_all_bd_employee_ids_for_company
    
    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Determine IDs to query based on user type
    if entity_type == 'user':
        # Admin: Can mark all notifications for BD employees + their own
        bd_employee_ids = get_all_bd_employee_ids_for_company(entity_id, db)
        all_ids = [entity_id] + bd_employee_ids
    else:
        # BD Employee: Only their own ID
        all_ids = [entity_id]

    # Update all unread notifications
    db.query(NotificationDB).filter(
        NotificationDB.user_id.in_(all_ids),
        NotificationDB.is_read == False
    ).update({"is_read": True})

    db.commit()

    return {"message": "All notifications marked as read"}

@app.get("/api/expert/notifications")
@require_expert_login
async def get_expert_notifications(
    request: Request,
    db: Session = Depends(get_db),
    limit: int = 20
):
    """Fetch notifications for currently logged-in expert."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")

    notifications = db.query(ExpertNotificationDB).filter(
        ExpertNotificationDB.expert_id == current_expert.id
    ).order_by(desc(ExpertNotificationDB.created_at)).limit(limit).all()

    unread_count = db.query(ExpertNotificationDB).filter(
        ExpertNotificationDB.expert_id == current_expert.id,
        ExpertNotificationDB.is_read == False
    ).count()

    def format_time_ago(created_at: datetime) -> str:
        diff = datetime.utcnow() - created_at
        if diff.days > 0:
            return f"{diff.days} day{'s' if diff.days != 1 else ''} ago"
        hours = diff.seconds // 3600
        if hours > 0:
            return f"{hours} hour{'s' if hours != 1 else ''} ago"
        minutes = max(1, diff.seconds // 60)
        return f"{minutes} minute{'s' if minutes != 1 else ''} ago"

    formatted = []
    for notif in notifications:
        formatted.append({
            "id": notif.id,
            "title": notif.title,
            "message": notif.message,
            "time_ago": format_time_ago(notif.created_at),
            "is_read": notif.is_read,
            "link": notif.link,
            "type": notif.notification_type
        })

    return {
        "notifications": formatted,
        "count": unread_count
    }

@app.post("/api/expert/notifications/{notification_id}/mark-read")
@require_expert_login
async def mark_expert_notification_read(
    request: Request,
    notification_id: str,
    db: Session = Depends(get_db)
):
    """Mark a specific expert notification as read."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")

    notification = db.query(ExpertNotificationDB).filter(
        ExpertNotificationDB.id == notification_id,
        ExpertNotificationDB.expert_id == current_expert.id
    ).first()

    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")

    notification.is_read = True
    db.commit()

    return {"message": "Notification marked as read"}

@app.post("/api/expert/notifications/mark-all-read")
@require_expert_login
async def mark_all_expert_notifications_read(
    request: Request,
    db: Session = Depends(get_db)
):
    """Mark every notification for the expert as read."""
    current_expert = get_current_expert(request, db)
    if not current_expert:
        raise HTTPException(status_code=401, detail="Authentication required")

    db.query(ExpertNotificationDB).filter(
        ExpertNotificationDB.expert_id == current_expert.id,
        ExpertNotificationDB.is_read == False
    ).update({"is_read": True})
    db.commit()

    return {"message": "All notifications marked as read"}

@app.post("/api/notifications/check-deadlines")
async def check_deadline_notifications_endpoint(db: Session = Depends(get_db)):
    """
    Check all favorited and shortlisted tenders for deadline notifications.
    This endpoint can be called by a cron job or background scheduler.
    """
    check_all_deadline_notifications(db)
    return {"message": "Deadline notifications checked and created"}


# ============================================================
# REMINDER ENDPOINTS
# ============================================================

@app.post("/api/reminders")
async def create_reminder(request: Request, db: Session = Depends(get_db)):
    """Create a new reminder for a tender."""
    from core.dependencies import get_id_for_tender_management
    
    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    try:
        data = await request.json()
        tender_id = data.get('tender_id')
        tender_title = data.get('tender_title')
        reminder_datetime_str = data.get('reminder_datetime')
        note = data.get('note', '')

        # Validate required fields
        if not tender_id or not reminder_datetime_str:
            raise HTTPException(status_code=400, detail="Missing required fields")

        # Parse datetime
        from datetime import datetime
        try:
            reminder_datetime = datetime.fromisoformat(reminder_datetime_str)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid datetime format")

        # Check if reminder is in the future
        if reminder_datetime <= datetime.utcnow():
            raise HTTPException(status_code=400, detail="Reminder must be in the future")

        # Check if tender exists
        tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
        if not tender:
            raise HTTPException(status_code=404, detail="Tender not found")

        # Create reminder
        from database import ReminderDB
        reminder = ReminderDB(
            user_id=entity_id,
            tender_id=tender_id,
            reminder_datetime=reminder_datetime,
            title=tender_title or tender.title,
            note=note,
            worked_by_name=entity.name,
            worked_by_type=entity_type
        )

        db.add(reminder)
        db.commit()
        db.refresh(reminder)

        return {
            "success": True,
            "message": "Reminder created successfully",
            "reminder": {
                "id": reminder.id,
                "tender_id": reminder.tender_id,
                "reminder_datetime": reminder.reminder_datetime.isoformat(),
                "title": reminder.title,
                "note": reminder.note
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating reminder: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create reminder")


@app.get("/api/reminders")
async def get_reminders(request: Request, db: Session = Depends(get_db)):
    """Get all reminders for the current user or BD employee."""
    from core.dependencies import get_id_for_tender_management, get_all_bd_employee_ids_for_company
    
    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    try:
        from database import ReminderDB
        from datetime import datetime

        # Determine IDs to query based on user type
        if entity_type == 'user':
            # Admin: Get all BD employee IDs + their own ID
            bd_employee_ids = get_all_bd_employee_ids_for_company(entity_id, db)
            all_ids = [entity_id] + bd_employee_ids
        else:
            # BD Employee: Only their own ID
            all_ids = [entity_id]

        # Get all pending reminders for the entity(ies)
        reminders = db.query(ReminderDB).filter(
            ReminderDB.user_id.in_(all_ids),
            ReminderDB.is_dismissed == False
        ).order_by(ReminderDB.reminder_datetime).all()

        result = []
        for reminder in reminders:
            result.append({
                "id": reminder.id,
                "tender_id": reminder.tender_id,
                "title": reminder.title,
                "reminder_datetime": reminder.reminder_datetime.isoformat(),
                "note": reminder.note,
                "is_triggered": reminder.is_triggered,
                "created_at": reminder.created_at.isoformat(),
                "worked_by_name": reminder.worked_by_name,
                "worked_by_type": reminder.worked_by_type
            })

        return {"reminders": result}

    except Exception as e:
        logger.error(f"Error fetching reminders: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch reminders")


@app.post("/api/reminders/{reminder_id}/dismiss")
async def dismiss_reminder(request: Request, reminder_id: int, db: Session = Depends(get_db)):
    """Dismiss a reminder."""
    from core.dependencies import get_id_for_tender_management
    
    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    try:
        from database import ReminderDB

        reminder = db.query(ReminderDB).filter(
            ReminderDB.id == reminder_id,
            ReminderDB.user_id == entity_id
        ).first()

        if not reminder:
            raise HTTPException(status_code=404, detail="Reminder not found")

        reminder.is_dismissed = True
        db.commit()

        return {"success": True, "message": "Reminder dismissed"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error dismissing reminder: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to dismiss reminder")


@app.delete("/api/reminders/{reminder_id}")
async def delete_reminder(request: Request, reminder_id: int, db: Session = Depends(get_db)):
    """Delete a reminder."""
    from core.dependencies import get_id_for_tender_management
    
    entity_id, entity, entity_type = get_id_for_tender_management(request, db)
    if not entity_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    try:
        from database import ReminderDB

        reminder = db.query(ReminderDB).filter(
            ReminderDB.id == reminder_id,
            ReminderDB.user_id == entity_id
        ).first()

        if not reminder:
            raise HTTPException(status_code=404, detail="Reminder not found")

        db.delete(reminder)
        db.commit()

        return {"success": True, "message": "Reminder deleted"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting reminder: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete reminder")


@app.post("/api/reminders/check-triggered")
async def check_triggered_reminders(db: Session = Depends(get_db)):
    """
    Check for reminders that should be triggered and create notifications.
    This endpoint should be called by a background scheduler every minute.
    """
    try:
        from database import ReminderDB, NotificationDB
        from datetime import datetime

        # Get reminders that should be triggered (past their time and not yet triggered)
        now = datetime.utcnow()
        reminders_to_trigger = db.query(ReminderDB).filter(
            ReminderDB.reminder_datetime <= now,
            ReminderDB.is_triggered == False,
            ReminderDB.is_dismissed == False
        ).all()

        notifications_created = 0
        for reminder in reminders_to_trigger:
            # Create notification for this reminder
            notification = NotificationDB(
                user_id=reminder.user_id,
                tender_id=reminder.tender_id,
                notification_type='reminder',
                message=f"Reminder: {reminder.title}",
                tender_title=reminder.title,
                is_read=False
            )
            db.add(notification)

            # Mark reminder as triggered
            reminder.is_triggered = True
            notifications_created += 1

        db.commit()

        logger.info(f"Checked reminders, created {notifications_created} notifications")
        return {
            "success": True,
            "message": f"Created {notifications_created} reminder notifications"
        }

    except Exception as e:
        logger.error(f"Error checking triggered reminders: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to check reminders")


# ============================================================
# AI SUMMARY ENDPOINT
# ============================================================



@app.get("/api/tender/{tender_id}/employee-tasks")
async def get_tender_employee_tasks(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Get employee tasks for a specific tender."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the tender
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Get all assignments for this tender
    assignments = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.tender_id == tender_id).all()

    employee_tasks = []
    for assignment in assignments:
        if not assignment.employee:
            continue

        # Get tasks for this assignment
        tasks = db.query(TaskDB).filter(TaskDB.assignment_id == assignment.id).all()

        # Format tasks for frontend
        formatted_tasks = []
        for task in tasks:
            formatted_tasks.append({
                'id': task.id,
                'title': task.title,
                'description': task.description,
                'priority': task.priority,
                'status': task.status,
                'deadline': task.deadline.isoformat() if task.deadline else None, # type: ignore
                'created_at': task.created_at.isoformat() if task.created_at else None, # type: ignore
                'completed_at': task.completed_at.isoformat() if task.completed_at else None # type: ignore
            })

        employee_tasks.append({
            'employee_id': assignment.employee.id,
            'employee_name': assignment.employee.name,
            'role': assignment.role,
            'tasks': formatted_tasks
        })

    return {"employee_tasks": employee_tasks}

# Employee Task Management API endpoints
@app.post("/api/employee/tasks/{task_id}/complete")
async def complete_task(request: Request, task_id: int, db: Session = Depends(get_db)):
    """Mark a task as completed."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the task
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Check if employee is assigned to this task
    assignment = db.query(TenderAssignmentDB).filter(
        and_(TenderAssignmentDB.id == task.assignment_id, TenderAssignmentDB.employee_id == current_employee.id)
    ).first()

    if not assignment:
        raise HTTPException(status_code=403, detail="Not authorized to complete this task")

    # Update task status
    task.status = 'completed'  # type: ignore
    task.completed_at = datetime.utcnow()  # type: ignore
    task.completed_by = current_employee.id  # type: ignore

    db.commit()

    return {"message": "Task marked as completed"}

@app.post("/api/employee/tasks/{task_id}/comment")
async def add_task_comment(
    request: Request,
    task_id: int,
    comment: str = Form(...),
    db: Session = Depends(get_db)
):
    """Add a comment to a task."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the task
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Check if employee is assigned to this task
    assignment = db.query(TenderAssignmentDB).filter(
        and_(TenderAssignmentDB.id == task.assignment_id, TenderAssignmentDB.employee_id == current_employee.id)
    ).first()

    if not assignment:
        raise HTTPException(status_code=403, detail="Not authorized to comment on this task")

    # Create comment
    new_comment = TaskCommentDB(
        task_id=task_id,
        employee_id=current_employee.id,
        comment=comment,
        created_at=datetime.utcnow()
    )

    db.add(new_comment)
    db.commit()

    return {"message": "Comment added successfully"}

@app.post("/api/employee/messages")
async def send_message(
    request: Request,
    assignment_id: int = Form(...),
    message: str = Form(...),
    db: Session = Depends(get_db)
):
    """Send a message in the tender chat."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if employee is assigned to this tender
    assignment = db.query(TenderAssignmentDB).filter(
        and_(TenderAssignmentDB.id == assignment_id, TenderAssignmentDB.employee_id == current_employee.id)
    ).first()

    if not assignment:
        raise HTTPException(status_code=403, detail="Not authorized to send messages for this tender")

    # Create message
    timestamp = datetime.utcnow()
    new_message = TenderMessageDB(
        assignment_id=assignment_id,
        employee_id=current_employee.id,
        message=message,
        created_at=timestamp
    )

    db.add(new_message)
    db.commit()

    # Broadcast message via WebSocket
    await ws_manager.broadcast_chat_message(assignment.tender_id, {
        "sender_type": "employee",
        "sender_name": current_employee.name,
        "message": message,
        "timestamp": timestamp.isoformat(),
        "formatted_time": timestamp.strftime('%d %b, %I:%M %p')
    })

    return {"message": "Message sent successfully"}

@app.get("/api/employee/messages/{assignment_id}")
async def get_messages(
    request: Request,
    assignment_id: int,
    since: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get messages for a tender assignment, optionally since a specific timestamp."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if employee is assigned to this tender
    assignment = db.query(TenderAssignmentDB).filter(
        and_(TenderAssignmentDB.id == assignment_id, TenderAssignmentDB.employee_id == current_employee.id)
    ).first()

    if not assignment:
        raise HTTPException(status_code=403, detail="Not authorized to view messages for this tender")

    # Build query
    query = db.query(TenderMessageDB).filter(TenderMessageDB.assignment_id == assignment_id)

    # Filter by timestamp if provided
    if since:
        try:
            since_dt = datetime.fromisoformat(since.replace('Z', '+00:00'))
            query = query.filter(TenderMessageDB.created_at > since_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid timestamp format")

    # Get messages with robust ordering
    messages = query.order_by(TenderMessageDB.created_at.asc(), TenderMessageDB.id.asc()).all()

    # Format messages for frontend
    formatted_messages = []
    for message in messages:
        formatted_messages.append({
            'id': message.id,
            'employee_id': message.employee_id,
            'employee_name': message.employee.name if message.employee else 'Unknown',
            'message': message.message,
            'created_at': message.created_at.isoformat() if message.created_at else None    # type: ignore
        })

    return {"messages": formatted_messages}

@app.get("/api/employee/messages/tender/{tender_id}")
async def get_tender_messages(
    request: Request,
    tender_id: str,
    since: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get chat history for a tender for both employees and managers."""
    current_user = get_current_user(request, db)
    current_employee = get_current_employee(request, db)

    if not current_user and not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Ensure tender exists
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    assignment_query = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.tender_id == tender_id)

    assignment_ids: List[int] = []

    if current_employee:
        # Verify employee is mapped to this tender
        employee_assignment = assignment_query.filter(TenderAssignmentDB.employee_id == current_employee.id).first()
        if not employee_assignment:
            raise HTTPException(status_code=403, detail="Not authorized for this tender")
        # Employees can see the entire tender conversation
        assignment_ids = [assignment.id for assignment in assignment_query.all()]
    else:
        # Manager view — restrict to their company employees
        company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
        company_code_ids = [code.id for code in company_codes]

        if not company_code_ids:
            raise HTTPException(status_code=403, detail="No employees assigned to this tender")

        assignments = assignment_query.join(EmployeeDB).filter(
            EmployeeDB.company_code_id.in_(company_code_ids)
        ).all()

        assignment_ids = [assignment.id for assignment in assignments]

    if not assignment_ids:
        return {"messages": []}

    query = db.query(TenderMessageDB).filter(TenderMessageDB.assignment_id.in_(assignment_ids))

    # Filter by timestamp if provided
    if since:
        try:
            since_dt = datetime.fromisoformat(since.replace('Z', '+00:00'))
            query = query.filter(TenderMessageDB.created_at > since_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid timestamp format")

    # Get messages with robust ordering
    messages = query.order_by(TenderMessageDB.created_at.asc(), TenderMessageDB.id.asc()).all()

    # Format messages for frontend
    formatted_messages = []
    for message in messages:
        # Determine sender name
        if message.employee:
            sender_name = message.employee.name
        else:
            assignment_owner = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == message.assignment_id).first()
            if assignment_owner and assignment_owner.assigned_by:
                manager = db.query(UserDB).filter(UserDB.id == assignment_owner.assigned_by).first()
                sender_name = f"{manager.name} (Manager)" if manager else 'Manager'
            else:
                sender_name = 'Manager'

        formatted_messages.append({
            'id': message.id,
            'employee_id': message.employee_id,
            'employee_name': sender_name,
            'message': message.message,
            'created_at': message.created_at.isoformat() if message.created_at else None    # type: ignore
        })

    return {"messages": formatted_messages}

# Employee Task Assignment API endpoints
@app.post("/api/employee/assignments")
async def create_assignment(
    request: Request,
    tender_id: str = Form(...),
    employee_id: str = Form(...),
    role: str = Form(...),
    priority: str = Form("medium"),
    db: Session = Depends(get_db)
):
    """Create a new tender assignment for an employee."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if tender exists
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if employee exists and belongs to user's company
    employee = db.query(EmployeeDB).filter(EmployeeDB.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Check if employee belongs to user's company
    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
    employee_company_codes = [code.id for code in company_codes]
    if employee.company_code_id not in employee_company_codes:
        raise HTTPException(status_code=403, detail="Employee does not belong to your company")

    # Check if assignment already exists
    existing_assignment = db.query(TenderAssignmentDB).filter(
        and_(TenderAssignmentDB.tender_id == tender_id, TenderAssignmentDB.employee_id == employee_id)
    ).first()

    if existing_assignment:
        raise HTTPException(status_code=400, detail="Employee is already assigned to this tender")

    # Create assignment
    new_assignment = TenderAssignmentDB(
        tender_id=tender_id,
        employee_id=employee_id,
        role=role,
        assigned_by=current_user.id,
        priority=priority,
        assigned_at=datetime.utcnow()
    )

    db.add(new_assignment)
    db.commit()
    db.refresh(new_assignment)

    return {"message": "Assignment created successfully", "assignment_id": new_assignment.id}

@app.post("/api/employee/tasks")
async def create_task(
    request: Request,
    assignment_id: int = Form(...),
    title: str = Form(...),
    description: str = Form(""),
    priority: str = Form("medium"),
    estimated_hours: Optional[float] = Form(None),
    deadline: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Create a new task for an assignment."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if assignment exists and belongs to user's company
    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")

    # Check if assignment belongs to user's company
    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
    employee_company_codes = [code.id for code in company_codes]
    if assignment.employee.company_code_id not in employee_company_codes:
        raise HTTPException(status_code=403, detail="Assignment does not belong to your company")

    # Parse deadline
    parsed_deadline = None
    if deadline:
        try:
            parsed_deadline = datetime.fromisoformat(deadline.replace('Z', '+00:00'))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid deadline format")

    # Create task
    new_task = TaskDB(
        assignment_id=assignment_id,
        title=title,
        description=description,
        priority=priority,
        estimated_hours=estimated_hours,
        deadline=parsed_deadline,
        created_at=datetime.utcnow()
    )

    db.add(new_task)
    db.commit()
    db.refresh(new_task)

    return {"message": "Task created successfully", "task_id": new_task.id}


# ==================== Task Template Management ====================

@app.get("/api/task-templates")
async def get_task_templates(
    request: Request,
    stage: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Get all task templates for the current user (admin-only)."""
    from core.dependencies import get_current_user

    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    query = db.query(StageTaskTemplateDB).filter(
        StageTaskTemplateDB.user_id == current_user.id,
        StageTaskTemplateDB.parent_template_id == None  # Only main tasks
    )

    if stage is not None:
        if stage < 1 or stage > 6:
            raise HTTPException(status_code=400, detail="Stage must be 1-6")
        query = query.filter(StageTaskTemplateDB.stage_number == stage)

    templates = query.order_by(
        StageTaskTemplateDB.stage_number,
        StageTaskTemplateDB.task_order
    ).all()

    # Format response with subtasks
    result = []
    for template in templates:
        subtasks = db.query(StageTaskTemplateDB).filter(
            StageTaskTemplateDB.parent_template_id == template.id
        ).order_by(StageTaskTemplateDB.task_order).all()

        result.append({
            "id": template.id,
            "stage_number": template.stage_number,
            "task_title": template.task_title,
            "task_description": template.task_description,
            "priority": template.priority,
            "estimated_hours": template.estimated_hours,
            "deadline_days": template.deadline_days,
            "task_order": template.task_order,
            "subtasks": [{
                "id": st.id,
                "task_title": st.task_title,
                "task_description": st.task_description,
                "priority": st.priority,
                "estimated_hours": st.estimated_hours,
                "deadline_days": st.deadline_days,
                "task_order": st.task_order
            } for st in subtasks]
        })

    return {"success": True, "templates": result}


@app.post("/api/task-templates")
async def create_task_template(request: Request, db: Session = Depends(get_db)):
    """Create a new task template (admin-only)."""
    from core.dependencies import get_current_user

    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    body = await request.json()

    # Validate required fields
    required_fields = ['stage_number', 'task_title', 'deadline_days']
    for field in required_fields:
        if field not in body:
            raise HTTPException(status_code=400, detail=f"Missing required field: {field}")

    stage_number = body['stage_number']
    if stage_number < 1 or stage_number > 6:
        raise HTTPException(status_code=400, detail="Stage must be 1-6")

    # Validate parent_template_id if subtask
    parent_template_id = body.get('parent_template_id')
    if parent_template_id:
        parent = db.query(StageTaskTemplateDB).filter(
            StageTaskTemplateDB.id == parent_template_id,
            StageTaskTemplateDB.user_id == current_user.id,
            StageTaskTemplateDB.parent_template_id == None  # Can't nest subtasks
        ).first()
        if not parent:
            raise HTTPException(status_code=400, detail="Invalid parent template")

    # Get max task_order for this stage
    max_order = db.query(func.max(StageTaskTemplateDB.task_order)).filter(
        StageTaskTemplateDB.user_id == current_user.id,
        StageTaskTemplateDB.stage_number == stage_number,
        StageTaskTemplateDB.parent_template_id == parent_template_id
    ).scalar() or 0

    # Create template
    new_template = StageTaskTemplateDB(
        user_id=current_user.id,
        stage_number=stage_number,
        task_title=body['task_title'],
        task_description=body.get('task_description', ''),
        priority=body.get('priority', 'medium'),
        estimated_hours=body.get('estimated_hours'),
        deadline_days=body['deadline_days'],
        parent_template_id=parent_template_id,
        is_subtask=parent_template_id is not None,
        task_order=max_order + 1
    )

    db.add(new_template)
    db.commit()
    db.refresh(new_template)

    return {
        "success": True,
        "message": "Task template created",
        "template_id": new_template.id
    }


@app.put("/api/task-templates/{template_id}")
async def update_task_template(
    request: Request,
    template_id: int,
    db: Session = Depends(get_db)
):
    """Update an existing task template (admin-only)."""
    from core.dependencies import get_current_user

    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find template
    template = db.query(StageTaskTemplateDB).filter(
        StageTaskTemplateDB.id == template_id,
        StageTaskTemplateDB.user_id == current_user.id
    ).first()

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    body = await request.json()

    # Update fields
    if 'task_title' in body:
        template.task_title = body['task_title']
    if 'task_description' in body:
        template.task_description = body['task_description']
    if 'priority' in body:
        template.priority = body['priority']
    if 'estimated_hours' in body:
        template.estimated_hours = body['estimated_hours']
    if 'deadline_days' in body:
        template.deadline_days = body['deadline_days']
    if 'task_order' in body:
        template.task_order = body['task_order']

    template.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(template)

    return {
        "success": True,
        "message": "Task template updated"
    }


@app.delete("/api/task-templates/{template_id}")
async def delete_task_template(
    request: Request,
    template_id: int,
    db: Session = Depends(get_db)
):
    """Delete a task template and its subtasks via cascade (admin-only)."""
    from core.dependencies import get_current_user

    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find template
    template = db.query(StageTaskTemplateDB).filter(
        StageTaskTemplateDB.id == template_id,
        StageTaskTemplateDB.user_id == current_user.id
    ).first()

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    db.delete(template)
    db.commit()

    return {
        "success": True,
        "message": "Task template deleted"
    }


def auto_create_tasks_for_stage(
    db: Session,
    shortlist_id: int,
    stage_number: int,
    user_id: str,
    trigger_date: datetime
):
    """
    Automatically create tasks for all assigned employees based on templates.

    Args:
        db: Database session
        shortlist_id: ShortlistedTenderDB ID
        stage_number: Stage number (1-6)
        user_id: User ID who owns the templates
        trigger_date: Date when stage was changed (used to calculate deadlines)

    Returns:
        Number of tasks created
    """
    from sqlalchemy.orm import joinedload

    # Get shortlisted tender with tender info
    shortlisted = db.query(ShortlistedTenderDB).options(
        joinedload(ShortlistedTenderDB.tender)
    ).filter(ShortlistedTenderDB.id == shortlist_id).first()

    if not shortlisted:
        return 0

    tender_id = shortlisted.tender_id

    # Get all assignments for this tender
    assignments = db.query(TenderAssignmentDB).filter(
        TenderAssignmentDB.tender_id == tender_id
    ).all()

    if not assignments:
        return 0

    # Get task templates for this stage (main tasks only)
    templates = db.query(StageTaskTemplateDB).filter(
        StageTaskTemplateDB.user_id == user_id,
        StageTaskTemplateDB.stage_number == stage_number,
        StageTaskTemplateDB.parent_template_id == None
    ).order_by(StageTaskTemplateDB.task_order).all()

    if not templates:
        return 0

    tasks_created = 0

    # Create tasks for each assignment
    for assignment in assignments:
        for template in templates:
            # Calculate deadline
            deadline = trigger_date + timedelta(days=template.deadline_days)

            # Create main task
            new_task = TaskDB(
                assignment_id=assignment.id,
                title=template.task_title,
                description=template.task_description,
                priority=template.priority,
                estimated_hours=template.estimated_hours,
                deadline=deadline,
                status='pending',
                created_at=datetime.utcnow()
            )

            db.add(new_task)
            db.flush()  # Get task ID for subtasks
            tasks_created += 1

            # Get subtasks for this template
            subtask_templates = db.query(StageTaskTemplateDB).filter(
                StageTaskTemplateDB.parent_template_id == template.id
            ).order_by(StageTaskTemplateDB.task_order).all()

            # Create subtasks
            for subtask_template in subtask_templates:
                subtask_deadline = trigger_date + timedelta(days=subtask_template.deadline_days)

                new_subtask = TaskDB(
                    assignment_id=assignment.id,
                    title=subtask_template.task_title,
                    description=subtask_template.task_description,
                    priority=subtask_template.priority,
                    estimated_hours=subtask_template.estimated_hours,
                    deadline=subtask_deadline,
                    status='pending',
                    parent_task_id=new_task.id,
                    is_subtask=True,
                    subtask_order=subtask_template.task_order,
                    created_at=datetime.utcnow()
                )

                db.add(new_subtask)
                tasks_created += 1

    db.commit()
    return tasks_created


@app.post("/api/tenders/assignments/{assignment_id}/create-pending-tasks")
async def create_pending_tasks_for_late_assignment(
    request: Request,
    assignment_id: int,
    db: Session = Depends(get_db)
):
    """Create tasks for incomplete stages when an employee is assigned late."""
    from core.dependencies import get_current_user

    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get assignment
    assignment = db.query(TenderAssignmentDB).filter(
        TenderAssignmentDB.id == assignment_id
    ).first()

    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")

    # Get shortlisted tender
    shortlisted = db.query(ShortlistedTenderDB).filter(
        ShortlistedTenderDB.tender_id == assignment.tender_id,
        ShortlistedTenderDB.user_id == current_user.id
    ).first()

    if not shortlisted:
        raise HTTPException(status_code=404, detail="Shortlisted tender not found")

    progress_data = shortlisted.progress_data or {}

    REMOVAL_OPTIONS = [
        'Cancelled', 'Tender Cancelled', 'Opened & Not Qualified',
        'Opened & Lost', 'Opened & Not Eligible', 'No'
    ]

    total_created = 0
    now = datetime.utcnow()

    # Check stages 1-6
    for stage_num in range(1, 7):
        step_key = f'step{stage_num}'
        step_value = progress_data.get(step_key, '')

        # Create tasks if:
        # 1. Stage has been started (has a value)
        # 2. Value is not a removal option
        # 3. Stage is not complete (stage 6 with "Yes" is complete)
        is_complete = (stage_num == 6 and step_value == 'Yes')

        if (step_value and
            str(step_value).strip() != '' and
            step_value not in REMOVAL_OPTIONS and
            not is_complete):

            # Get templates for this stage
            templates = db.query(StageTaskTemplateDB).filter(
                StageTaskTemplateDB.user_id == current_user.id,
                StageTaskTemplateDB.stage_number == stage_num,
                StageTaskTemplateDB.parent_template_id == None
            ).order_by(StageTaskTemplateDB.task_order).all()

            # Create tasks for this assignment
            for template in templates:
                deadline = now + timedelta(days=template.deadline_days)

                new_task = TaskDB(
                    assignment_id=assignment_id,
                    title=template.task_title,
                    description=template.task_description,
                    priority=template.priority,
                    estimated_hours=template.estimated_hours,
                    deadline=deadline,
                    status='pending',
                    created_at=now
                )

                db.add(new_task)
                db.flush()
                total_created += 1

                # Create subtasks
                subtask_templates = db.query(StageTaskTemplateDB).filter(
                    StageTaskTemplateDB.parent_template_id == template.id
                ).order_by(StageTaskTemplateDB.task_order).all()

                for subtask_template in subtask_templates:
                    subtask_deadline = now + timedelta(days=subtask_template.deadline_days)

                    new_subtask = TaskDB(
                        assignment_id=assignment_id,
                        title=subtask_template.task_title,
                        description=subtask_template.task_description,
                        priority=subtask_template.priority,
                        estimated_hours=subtask_template.estimated_hours,
                        deadline=subtask_deadline,
                        status='pending',
                        parent_task_id=new_task.id,
                        is_subtask=True,
                        subtask_order=subtask_template.task_order,
                        created_at=now
                    )

                    db.add(new_subtask)
                    total_created += 1

    db.commit()

    return {
        "success": True,
        "message": f"Created {total_created} tasks for late assignment",
        "tasks_created": total_created
    }


@app.post("/api/tasks/subtask")
async def create_subtask(
    request: Request,
    parent_task_id: int = Form(...),
    title: str = Form(...),
    description: str = Form(""),
    deadline: Optional[str] = Form(None),
    priority: str = Form("medium"),
    db: Session = Depends(get_db)
):
    """Create a subtask under a parent task (1-level deep only).
    Subtasks automatically inherit the employee assignment from the parent task."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if parent task exists
    parent_task = db.query(TaskDB).filter(TaskDB.id == parent_task_id).first()
    if not parent_task:
        raise HTTPException(status_code=404, detail="Parent task not found")

    # Prevent creating subtask of a subtask (1-level deep only)
    if parent_task.is_subtask:
        raise HTTPException(status_code=400, detail="Cannot create subtask of a subtask. Only 1-level hierarchy allowed.")

    # Verify assignment belongs to user's company
    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == parent_task.assignment_id).first()
    if not assignment or assignment.assigned_by != current_user.id:
        raise HTTPException(status_code=403, detail="Unauthorized")

    # Parse deadline
    parsed_deadline = None
    if deadline and deadline.strip():
        try:
            parsed_deadline = datetime.fromisoformat(deadline.replace('Z', '+00:00'))
        except ValueError:
            parsed_deadline = parent_task.deadline  # Default to parent deadline

    # Automatically inherit employee assignment from parent task
    assigned_employee_id = assignment.employee_id

    # Get next subtask order
    max_order = db.query(func.max(TaskDB.subtask_order)).filter(TaskDB.parent_task_id == parent_task_id).scalar() or 0

    # Create subtask
    new_subtask = TaskDB(
        assignment_id=parent_task.assignment_id,
        title=title,
        description=description,
        priority=priority,
        deadline=parsed_deadline,
        parent_task_id=parent_task_id,
        is_subtask=True,
        subtask_order=max_order + 1,
        created_at=datetime.utcnow()
    )

    db.add(new_subtask)
    db.commit()
    db.refresh(new_subtask)

    # Create notification for assigned employee
    notification = EmployeeNotificationDB(
        employee_id=assigned_employee_id,
        notification_type="subtask_assigned",
        title="New Subtask Assigned",
        message=f"You've been assigned subtask: {title}",
        related_task_id=new_subtask.id,
        related_tender_id=assignment.tender_id,
        created_at=datetime.utcnow()
    )
    db.add(notification)
    db.commit()

    return {"message": "Subtask created successfully", "subtask_id": new_subtask.id}


@app.put("/api/tasks/{task_id}")
async def update_task(
    request: Request,
    task_id: int,
    title: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    priority: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    deadline: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Update an existing task. User or assigned employee can update."""
    current_user = get_current_user(request, db)
    current_employee = get_current_employee(request, db)

    if not current_user and not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get task
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Verify authorization
    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == task.assignment_id).first()
    if current_user:
        if assignment.assigned_by != current_user.id:
            raise HTTPException(status_code=403, detail="Unauthorized")
    elif current_employee:
        if assignment.employee_id != current_employee.id:
            raise HTTPException(status_code=403, detail="Unauthorized")

    # Update fields
    if title:
        task.title = title
    if description is not None:
        task.description = description
    if priority:
        task.priority = priority
    if status:
        task.status = status
        if status == "completed":
            task.completed_at = datetime.utcnow()
            if current_employee:
                task.completed_by = current_employee.id
    if deadline:
        try:
            task.deadline = datetime.fromisoformat(deadline.replace('Z', '+00:00'))
        except ValueError:
            pass

    db.commit()
    db.refresh(task)

    # Notify employee of update (if updated by manager)
    if current_user:
        notification = EmployeeNotificationDB(
            employee_id=assignment.employee_id,
            notification_type="task_updated",
            title="Task Updated",
            message=f"Task '{task.title}' has been updated",
            related_task_id=task.id,
            related_tender_id=assignment.tender_id,
            created_at=datetime.utcnow()
        )
        db.add(notification)
        db.commit()

    return {"message": "Task updated successfully", "task_id": task.id}


@app.delete("/api/tasks/{task_id}")
async def delete_task(
    request: Request,
    task_id: int,
    db: Session = Depends(get_db)
):
    """Delete a task and all its subtasks (cascade). Manager only."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get task
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Verify authorization
    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == task.assignment_id).first()
    if assignment.assigned_by != current_user.id:
        raise HTTPException(status_code=403, detail="Unauthorized")

    # Get affected employees before deletion
    affected_employees = set()
    affected_employees.add(assignment.employee_id)
    for subtask in task.subtasks:
        affected_employees.add(assignment.employee_id)

    task_title = task.title

    # Delete task (cascade deletes subtasks, concerns, etc.)
    db.delete(task)
    db.commit()

    # Notify affected employees
    for employee_id in affected_employees:
        notification = EmployeeNotificationDB(
            employee_id=employee_id,
            notification_type="task_deleted",
            title="Task Deleted",
            message=f"Task '{task_title}' and its subtasks have been removed",
            related_tender_id=assignment.tender_id,
            created_at=datetime.utcnow()
        )
        db.add(notification)
    db.commit()

    return {"message": "Task deleted successfully"}


@app.get("/api/tasks/{task_id}/subtasks")
async def get_task_subtasks(
    request: Request,
    task_id: int,
    db: Session = Depends(get_db)
):
    """Get all subtasks for a parent task."""
    current_user = get_current_user(request, db)
    current_employee = get_current_employee(request, db)

    if not current_user and not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get task with subtasks
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Get subtasks ordered by subtask_order
    subtasks = db.query(TaskDB).filter(
        TaskDB.parent_task_id == task_id
    ).order_by(TaskDB.subtask_order).all()

    subtasks_data = []
    for subtask in subtasks:
        assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == subtask.assignment_id).first()
        employee = db.query(EmployeeDB).filter(EmployeeDB.id == assignment.employee_id).first()

        subtasks_data.append({
            "id": subtask.id,
            "title": subtask.title,
            "description": subtask.description,
            "status": subtask.status,
            "priority": subtask.priority,
            "deadline": subtask.deadline.isoformat() if subtask.deadline else None,
            "completed_at": subtask.completed_at.isoformat() if subtask.completed_at else None,
            "employee": {
                "id": employee.id,
                "name": employee.name,
                "email": employee.email
            } if employee else None
        })

    return {"subtasks": subtasks_data}


@app.post("/api/tasks/{task_id}/concerns")
async def create_task_concern(
    request: Request,
    task_id: int,
    concern_type: str = Form(...),
    title: str = Form(...),
    description: str = Form(...),
    priority: str = Form("medium"),
    db: Session = Depends(get_db)
):
    """Employee raises a concern about a specific task."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Employee authentication required")

    # Verify task exists and employee is assigned to it
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == task.assignment_id).first()
    if assignment.employee_id != current_employee.id:
        raise HTTPException(status_code=403, detail="You are not assigned to this task")

    # Create concern
    concern = TaskConcernDB(
        task_id=task_id,
        employee_id=current_employee.id,
        concern_type=concern_type,
        title=title,
        description=description,
        priority=priority,
        status="open",
        created_at=datetime.utcnow()
    )

    db.add(concern)
    db.commit()
    db.refresh(concern)

    # Notify manager
    manager = db.query(UserDB).filter(UserDB.id == assignment.assigned_by).first()
    if manager:
        manager_notification = NotificationDB(
            user_id=manager.id,
            tender_id=assignment.tender_id,
            notification_type="concern_raised",
            message=f"{current_employee.name} raised a {concern_type} concern on task '{task.title}'",
            tender_title=task.title,
            is_read=False,
            created_at=datetime.utcnow()
        )
        db.add(manager_notification)
        db.commit()

    return {"message": "Concern raised successfully", "concern_id": concern.id}


@app.post("/api/tasks/{task_id}/files")
async def upload_task_file(
    request: Request,
    task_id: int,
    file: UploadFile = File(...),
    description: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Upload a deliverable file for a task. Employees can upload files they've completed."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Employee authentication required")

    # Verify task exists and employee is assigned to it
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == task.assignment_id).first()
    if assignment.employee_id != current_employee.id:
        raise HTTPException(status_code=403, detail="You are not assigned to this task")

    # Read file data
    file_data = await file.read()
    file_size = len(file_data)

    # Determine MIME type
    mime_type = file.content_type or "application/octet-stream"

    # Create task file record
    task_file = TaskFileDB(
        task_id=task_id,
        employee_id=current_employee.id,
        filename=file.filename,
        mime_type=mime_type,
        file_size=file_size,
        file_data=file_data,
        description=description,
        created_at=datetime.utcnow()
    )

    db.add(task_file)
    db.commit()
    db.refresh(task_file)

    # Notify manager
    manager = db.query(UserDB).filter(UserDB.id == assignment.assigned_by).first()
    if manager:
        try:
            # Check if notification already exists for this tender and type
            existing_notification = db.query(NotificationDB).filter(
                NotificationDB.user_id == manager.id,
                NotificationDB.tender_id == assignment.tender_id,
                NotificationDB.notification_type == "file_uploaded"
            ).first()
            
            if existing_notification:
                # Update existing notification with latest upload info
                existing_notification.message = f"{current_employee.name} uploaded deliverable '{file.filename}' for task '{task.title}'"
                existing_notification.tender_title = task.title
                existing_notification.is_read = False
                existing_notification.created_at = datetime.utcnow()
            else:
                # Create new notification
                manager_notification = NotificationDB(
                    user_id=manager.id,
                    tender_id=assignment.tender_id,
                    notification_type="file_uploaded",
                    message=f"{current_employee.name} uploaded deliverable '{file.filename}' for task '{task.title}'",
                    tender_title=task.title,
                    is_read=False,
                    created_at=datetime.utcnow()
                )
                db.add(manager_notification)

            db.commit()
        except Exception as e:
            # Don't let notification failure block file upload
            logger.warning(f"Failed to create/update notification: {e}")
            db.rollback()

    # =====================================================================
    # SYNC DELIVERABLE TO PAST PROJECT (if tender is awarded)
    # =====================================================================
    sync_success = False
    sync_message = ""
    
    try:
        # Check if the tender is awarded
        tender = db.query(TenderDB).filter(TenderDB.id == assignment.tender_id).first()
        
        if not tender:
            sync_message = "Tender not found"
            logger.warning(f"Tender not found for assignment {assignment.id}")
        elif not tender.awarded:
            sync_message = f"Tender not awarded (awarded={tender.awarded})"
            logger.info(f"Tender {tender.id} is not awarded. Deliverable uploaded to task only.")
        else:
            logger.info(f"✓ Tender {tender.id} is awarded, looking for linked project...")
            
            # Find the corresponding project by source_tender_id
            project = db.query(ProjectDB).filter(
                ProjectDB.source_tender_id == tender.id
            ).first()
            
            # If no project found by source_tender_id, try to find by name and link it
            if not project and tender.title:
                logger.info(f"  No project found by source_tender_id, searching by name...")
                project = db.query(ProjectDB).filter(
                    and_(
                        ProjectDB.user_id == tender.awarded_by,
                        ProjectDB.project_name.ilike(f"%{tender.title[:50]}%"),
                        ProjectDB.source_tender_id == None
                    )
                ).first()
                
                if project:
                    logger.info(f"  Found matching project by name, linking it...")
                    project.source_tender_id = tender.id
                    db.commit()
                    logger.info(f"  ✓ Linked project {project.id} to tender {tender.id}")
            
            if project:
                logger.info(f"✓ Found project {project.id} ('{project.project_name[:50]}'), syncing deliverable...")
                
                # Get or initialize project documents
                project_docs = project.documents or {}
                if not isinstance(project_docs, dict):
                    project_docs = {}
                
                # Add to "Task Deliverables" category
                deliverable_category = "Task Deliverables"
                if deliverable_category not in project_docs:
                    project_docs[deliverable_category] = []
                
                # Check if this file is already synced BEFORE creating file on disk
                already_synced = any(
                    isinstance(doc, dict) and doc.get('task_file_id') == task_file.id
                    for doc in project_docs[deliverable_category]
                )
                
                if already_synced:
                    logger.info(f"  ℹ Deliverable already synced (task_file_id={task_file.id})")
                    sync_success = True
                    sync_message = "Already synced to project"
                else:
                    # Save the deliverable file to filesystem
                    deliverables_dir = "project_deliverables"
                    os.makedirs(deliverables_dir, exist_ok=True)
                    
                    # Create a unique filename to avoid collisions
                    import uuid
                    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                    unique_filename = f"{timestamp}_{uuid.uuid4().hex[:8]}_{file.filename}"
                    file_path = os.path.join(deliverables_dir, unique_filename)
                    
                    # Write file to disk
                    with open(file_path, "wb") as f:
                        f.write(file_data)
                    
                    # Add file metadata
                    deliverable_info = {
                        "file_path": file_path,
                        "original_filename": file.filename,
                        "uploaded_by": current_employee.name,
                        "uploaded_at": datetime.utcnow().isoformat(),
                        "task_title": task.title,
                        "task_file_id": task_file.id,
                        "file_size": file_size,
                        "description": description
                    }
                    project_docs[deliverable_category].append(deliverable_info)
                    
                    # Update project documents and mark as modified for SQLAlchemy
                    from sqlalchemy.orm.attributes import flag_modified
                    project.documents = project_docs
                    flag_modified(project, "documents")
                    db.commit()
                    
                    sync_success = True
                    sync_message = f"Synced to project '{project.project_name}'"
                    logger.info(
                        f"✓✓✓ SUCCESSFULLY SYNCED deliverable '{file.filename}' from task '{task.title}' "
                        f"to project '{project.project_name}' (Project ID: {project.id})"
                    )
            else:
                sync_message = "No project found for awarded tender"
                logger.error(
                    f"✗ CRITICAL: No project found for awarded tender {tender.id} ('{tender.title[:50]}'). "
                    f"Deliverable '{file.filename}' uploaded to task but NOT synced to project. "
                    f"Run diagnose_and_fix_project_links.py to resolve this issue."
                )
                
    except Exception as e:
        sync_message = f"Sync failed: {str(e)}"
        logger.error(f"✗ Failed to sync deliverable to project: {e}", exc_info=True)
        # Don't rollback - the task file upload should succeed even if sync fails

    response = {
        "message": "File uploaded successfully",
        "file_id": task_file.id,
        "filename": file.filename,
        "sync_status": sync_message,
        "synced_to_project": sync_success
    }
    
    return response


@app.get("/api/tasks/{task_id}/files/{file_id}/download")
async def download_task_file(
    request: Request,
    task_id: int,
    file_id: int,
    db: Session = Depends(get_db)
):
    """Download a task deliverable file. Both employee and manager can download."""
    current_user = get_current_user(request, db)
    current_employee = get_current_employee(request, db)

    if not current_user and not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Verify task exists
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Get file
    task_file = db.query(TaskFileDB).filter(
        and_(TaskFileDB.id == file_id, TaskFileDB.task_id == task_id)
    ).first()

    if not task_file:
        raise HTTPException(status_code=404, detail="File not found")

    # Verify authorization
    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == task.assignment_id).first()
    if current_employee:
        # Employee can only download their own files or files from their assigned tasks
        if assignment.employee_id != current_employee.id and task_file.employee_id != current_employee.id:
            raise HTTPException(status_code=403, detail="Unauthorized")
    elif current_user:
        # Manager can download if they assigned the task
        if assignment.assigned_by != current_user.id:
            raise HTTPException(status_code=403, detail="Unauthorized")

    # Return file
    return Response(
        content=task_file.file_data,
        media_type=task_file.mime_type,
        headers={
            "Content-Disposition": f'attachment; filename="{task_file.filename}"',
            "Cache-Control": "public, max-age=3600"
        }
    )


@app.get("/api/tasks/{task_id}/files")
async def list_task_files(
    request: Request,
    task_id: int,
    db: Session = Depends(get_db)
):
    """List all files for a task."""
    current_user = get_current_user(request, db)
    current_employee = get_current_employee(request, db)

    if not current_user and not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Verify task exists
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Verify authorization
    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == task.assignment_id).first()
    if current_employee:
        if assignment.employee_id != current_employee.id:
            raise HTTPException(status_code=403, detail="Unauthorized")
    elif current_user:
        if assignment.assigned_by != current_user.id:
            raise HTTPException(status_code=403, detail="Unauthorized")

    # Get files
    files = db.query(TaskFileDB).filter(TaskFileDB.task_id == task_id).order_by(TaskFileDB.created_at.desc()).all()

    files_data = []
    for file_obj in files:
        files_data.append({
            "id": file_obj.id,
            "filename": file_obj.filename,
            "mime_type": file_obj.mime_type,
            "file_size": file_obj.file_size,
            "description": file_obj.description,
            "created_at": file_obj.created_at.isoformat() if file_obj.created_at else None,
            "uploaded_by": file_obj.employee.name if file_obj.employee else "Unknown"
        })

    return {"files": files_data}


@app.post("/api/tasks/{task_id}/progress")
async def add_task_progress_update(
    request: Request,
    task_id: int,
    update_text: str = Form(...),
    db: Session = Depends(get_db)
):
    """Add a progress update/work log to a task. Automatically changes task status to 'in_progress' if it's the first update."""
    # Authenticate employee
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Validate update_text length
    if not update_text or len(update_text.strip()) < 1:
        raise HTTPException(status_code=400, detail="Update text cannot be empty")
    if len(update_text) > 2000:
        raise HTTPException(status_code=400, detail="Update text too long (max 2000 characters)")

    # Find the task
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Check if employee is assigned to this task's tender
    assignment = db.query(TenderAssignmentDB).filter(
        and_(
            TenderAssignmentDB.id == task.assignment_id,
            TenderAssignmentDB.employee_id == current_employee.id
        )
    ).first()

    if not assignment:
        raise HTTPException(status_code=403, detail="Not authorized to update this task")

    # Check if this is the first progress update for this task
    existing_updates_count = db.query(TaskProgressUpdateDB).filter(
        TaskProgressUpdateDB.task_id == task_id
    ).count()

    # If first update and task is pending, change to in_progress
    if existing_updates_count == 0 and task.status == 'pending':
        task.status = 'in_progress'  # type: ignore

    # Create progress update
    progress_update = TaskProgressUpdateDB(
        task_id=task_id,
        employee_id=current_employee.id,
        update_text=update_text.strip(),
        created_at=datetime.utcnow()
    )

    db.add(progress_update)

    try:
        db.commit()
        db.refresh(progress_update)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to add progress update: {str(e)}")

    # Return the created update with employee info
    return {
        "message": "Progress update added successfully",
        "update": {
            "id": progress_update.id,
            "task_id": progress_update.task_id,
            "employee_id": progress_update.employee_id,
            "employee_name": current_employee.name,
            "employee_avatar": current_employee.name[0].upper() if current_employee.name else "?",
            "update_text": progress_update.update_text,
            "created_at": progress_update.created_at.isoformat(),
            "formatted_date": progress_update.created_at.strftime('%d %b %Y, %I:%M %p'),
            "is_edited": progress_update.is_edited,
            "is_current_user": True
        },
        "task_status": task.status
    }


@app.get("/api/tasks/{task_id}/progress")
async def get_task_progress_updates(
    request: Request,
    task_id: int,
    order: str = "desc",
    db: Session = Depends(get_db)
):
    """Get all progress updates for a task. Returns updates with employee information."""
    # Authenticate employee
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the task
    task = db.query(TaskDB).filter(TaskDB.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Check if employee is assigned to this task's tender
    assignment = db.query(TenderAssignmentDB).filter(
        and_(
            TenderAssignmentDB.id == task.assignment_id,
            TenderAssignmentDB.employee_id == current_employee.id
        )
    ).first()

    if not assignment:
        raise HTTPException(status_code=403, detail="Not authorized to view this task")

    # Query progress updates with employee info
    query = db.query(TaskProgressUpdateDB, EmployeeDB).join(
        EmployeeDB, TaskProgressUpdateDB.employee_id == EmployeeDB.id
    ).filter(TaskProgressUpdateDB.task_id == task_id)

    # Apply ordering
    if order.lower() == "asc":
        query = query.order_by(TaskProgressUpdateDB.created_at.asc())
    else:
        query = query.order_by(TaskProgressUpdateDB.created_at.desc())

    results = query.all()

    # Format response
    updates = []
    for update, employee in results:
        updates.append({
            "id": update.id,
            "task_id": update.task_id,
            "employee_id": update.employee_id,
            "employee_name": employee.name,
            "employee_avatar": employee.name[0].upper() if employee.name else "?",
            "update_text": update.update_text,
            "created_at": update.created_at.isoformat(),
            "formatted_date": update.created_at.strftime('%d %b %Y, %I:%M %p'),
            "is_edited": update.is_edited,
            "is_current_user": update.employee_id == current_employee.id
        })

    return {
        "task_id": task_id,
        "total_updates": len(updates),
        "updates": updates
    }


@app.delete("/api/tasks/{task_id}/progress/{update_id}")
async def delete_task_progress_update(
    request: Request,
    task_id: int,
    update_id: int,
    db: Session = Depends(get_db)
):
    """Delete a progress update. Only the creator can delete their own updates."""
    # Authenticate employee
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the progress update
    update = db.query(TaskProgressUpdateDB).filter(
        and_(
            TaskProgressUpdateDB.id == update_id,
            TaskProgressUpdateDB.task_id == task_id
        )
    ).first()

    if not update:
        raise HTTPException(status_code=404, detail="Progress update not found")

    # Check if current employee is the creator
    if update.employee_id != current_employee.id:
        raise HTTPException(status_code=403, detail="You can only delete your own progress updates")

    try:
        db.delete(update)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete progress update: {str(e)}")

    return {"message": "Progress update deleted successfully"}


@app.get("/api/tender/{tender_id}/concerns")
async def get_tender_concerns(
    request: Request,
    tender_id: str,
    status_filter: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get all concerns for a tender. User or assigned employee can access."""
    current_user = get_current_user(request, db)
    current_employee = get_current_employee(request, db)

    if not current_user and not current_employee:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get all assignments for this tender
    assignments = db.query(TenderAssignmentDB).filter(
        TenderAssignmentDB.tender_id == tender_id
    ).all()

    if not assignments:
        return {"concerns": []}

    # Verify access
    if current_user:
        # Manager must be the one who assigned
        if not any(a.assigned_by == current_user.id for a in assignments):
            raise HTTPException(status_code=403, detail="Unauthorized")
    elif current_employee:
        # Employee must be assigned to this tender
        if not any(a.employee_id == current_employee.id for a in assignments):
            raise HTTPException(status_code=403, detail="Unauthorized")

    # Get all task IDs for these assignments
    task_ids = []
    for assignment in assignments:
        tasks = db.query(TaskDB).filter(TaskDB.assignment_id == assignment.id).all()
        task_ids.extend([t.id for t in tasks])

    # Get concerns for these tasks
    query = db.query(TaskConcernDB).filter(TaskConcernDB.task_id.in_(task_ids))

    if status_filter:
        query = query.filter(TaskConcernDB.status == status_filter)

    concerns = query.order_by(TaskConcernDB.created_at.desc()).all()

    concerns_data = []
    for concern in concerns:
        task = db.query(TaskDB).filter(TaskDB.id == concern.task_id).first()
        employee = db.query(EmployeeDB).filter(EmployeeDB.id == concern.employee_id).first()

        concerns_data.append({
            "id": concern.id,
            "task_id": concern.task_id,
            "task_title": task.title if task else "Unknown",
            "concern_type": concern.concern_type,
            "title": concern.title,
            "description": concern.description,
            "status": concern.status,
            "priority": concern.priority,
            "raised_by": {
                "id": employee.id,
                "name": employee.name,
                "email": employee.email
            } if employee else None,
            "created_at": concern.created_at.isoformat() if concern.created_at else None,
            "resolved_at": concern.resolved_at.isoformat() if concern.resolved_at else None,
            "resolution_notes": concern.resolution_notes
        })

    return {"concerns": concerns_data}


@app.put("/api/concerns/{concern_id}/resolve")
async def resolve_concern(
    request: Request,
    concern_id: int,
    resolution_notes: str = Form(...),
    db: Session = Depends(get_db)
):
    """Manager resolves a concern."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get concern
    concern = db.query(TaskConcernDB).filter(TaskConcernDB.id == concern_id).first()
    if not concern:
        raise HTTPException(status_code=404, detail="Concern not found")

    # Verify authorization
    task = db.query(TaskDB).filter(TaskDB.id == concern.task_id).first()
    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == task.assignment_id).first()
    if assignment.assigned_by != current_user.id:
        raise HTTPException(status_code=403, detail="Unauthorized")

    # Resolve concern
    concern.status = "resolved"
    concern.resolved_by = current_user.id
    concern.resolved_at = datetime.utcnow()
    concern.resolution_notes = resolution_notes

    db.commit()

    # Notify employee
    notification = EmployeeNotificationDB(
        employee_id=concern.employee_id,
        notification_type="concern_resolved",
        title="Concern Resolved",
        message=f"Your concern '{concern.title}' has been resolved",
        related_task_id=concern.task_id,
        related_tender_id=assignment.tender_id,
        created_at=datetime.utcnow()
    )
    db.add(notification)
    db.commit()

    return {"message": "Concern resolved successfully"}


@app.get("/api/tender/{tender_id}/suggested-employees")
async def get_suggested_employees(
    request: Request,
    tender_id: str,
    db: Session = Depends(get_db)
):
    """Get suggested employees based on previous assignments during shortlisted phase."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if this tender was shortlisted before being awarded
    shortlisted_record = db.query(ShortlistedTenderDB).filter(
        ShortlistedTenderDB.tender_id == tender_id,
        ShortlistedTenderDB.user_id == current_user.id
    ).first()

    suggested_employees = []

    if shortlisted_record and shortlisted_record.progress_data:
        progress_data = shortlisted_record.progress_data
        employee_ids = set()

        # Extract employee IDs from all steps
        for step_num in range(1, 7):
            step_key = f'step{step_num}_employees'
            if step_key in progress_data:
                step_employees = progress_data[step_key]
                if isinstance(step_employees, list):
                    employee_ids.update(step_employees)

        # Get employee details
        if employee_ids:
            employees = db.query(EmployeeDB).filter(EmployeeDB.id.in_(list(employee_ids))).all()

            for emp in employees:
                # Count previous assignments
                assignment_count = db.query(TenderAssignmentDB).filter(
                    TenderAssignmentDB.employee_id == emp.id,
                    TenderAssignmentDB.assigned_by == current_user.id
                ).count()

                suggested_employees.append({
                    "id": emp.id,
                    "name": emp.name,
                    "email": emp.email,
                    "team": emp.team,
                    "role": emp.role,
                    "previous_assignments": assignment_count,
                    "worked_on_shortlist": True
                })

    return {"suggested_employees": suggested_employees}


@app.get("/api/employee/notifications")
async def get_employee_notifications(
    request: Request,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """Get employee notifications."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Employee authentication required")

    # Get notifications (unread first, then recent read)
    notifications = db.query(EmployeeNotificationDB).filter(
        EmployeeNotificationDB.employee_id == current_employee.id
    ).order_by(
        EmployeeNotificationDB.is_read.asc(),
        EmployeeNotificationDB.created_at.desc()
    ).limit(limit).all()

    notifications_data = []
    for notif in notifications:
        notifications_data.append({
            "id": notif.id,
            "notification_type": notif.notification_type,
            "title": notif.title,
            "message": notif.message,
            "is_read": notif.is_read,
            "created_at": notif.created_at.isoformat() if notif.created_at else None,
            "related_task_id": notif.related_task_id,
            "related_tender_id": notif.related_tender_id
        })

    return {"notifications": notifications_data}


@app.put("/api/employee/notifications/{notification_id}/read")
async def mark_notification_read(
    request: Request,
    notification_id: int,
    db: Session = Depends(get_db)
):
    """Mark employee notification as read."""
    current_employee = get_current_employee(request, db)
    if not current_employee:
        raise HTTPException(status_code=401, detail="Employee authentication required")

    notification = db.query(EmployeeNotificationDB).filter(
        EmployeeNotificationDB.id == notification_id,
        EmployeeNotificationDB.employee_id == current_employee.id
    ).first()

    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")

    notification.is_read = True
    db.commit()

    return {"message": "Notification marked as read"}


@app.post("/api/tasks/create")
async def create_task_for_tender(
    request: Request,
    tender_id: str = Form(...),
    assigned_employee_id: str = Form(...),
    title: str = Form(...),
    description: str = Form(""),
    priority: str = Form("medium"),
    estimated_hours: Optional[str] = Form(None),
    deadline: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Create a new task for a tender assignment."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Validate and parse estimated_hours
    parsed_hours = None
    if estimated_hours and estimated_hours.strip():
        try:
            parsed_hours = float(estimated_hours)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid estimated hours: {estimated_hours}")

    # Find or create assignment for this employee and tender
    assignment = db.query(TenderAssignmentDB).filter(
        and_(
            TenderAssignmentDB.tender_id == tender_id,
            TenderAssignmentDB.employee_id == assigned_employee_id
        )
    ).first()

    if not assignment:
        # Create a new assignment if it doesn't exist
        assignment = TenderAssignmentDB(
            tender_id=tender_id,
            employee_id=assigned_employee_id,
            role="Team Member",  # Default role
            assigned_by=current_user.id,
            priority="medium"
        )
        db.add(assignment)
        db.flush()

    # Parse deadline
    parsed_deadline = None
    if deadline and deadline.strip():
        try:
            # Handle datetime-local format (YYYY-MM-DDTHH:MM)
            parsed_deadline = datetime.fromisoformat(deadline.replace('Z', '+00:00'))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid deadline format: {deadline}")

    # Create task
    new_task = TaskDB(
        assignment_id=assignment.id,
        title=title,
        description=description,
        priority=priority,
        estimated_hours=parsed_hours,
        deadline=parsed_deadline,
        created_at=datetime.utcnow()
    )

    db.add(new_task)
    db.commit()
    db.refresh(new_task)

    # Notify employee
    notification = EmployeeNotificationDB(
        employee_id=assigned_employee_id,
        notification_type="task_assigned",
        title="New Task Assigned",
        message=f"You have been assigned a new task: {title}",
        related_task_id=new_task.id,
        related_tender_id=tender_id
    )
    db.add(notification)
    db.commit()

    return {"task_id": new_task.id, "message": "Task created successfully"}


@app.post("/api/tender/message")
async def send_tender_message(
    request: Request,
    tender_id: str = Form(...),
    message: str = Form(...),
    db: Session = Depends(get_db)
):
    """Send a message to the tender team chat."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get all assignments for this tender
    assignments = db.query(TenderAssignmentDB).filter(
        TenderAssignmentDB.tender_id == tender_id
    ).all()

    if not assignments:
        raise HTTPException(status_code=400, detail="Please add team members before using chat")

    # Get company name from CompanyDB
    company = db.query(CompanyDB).filter(CompanyDB.user_id == current_user.id).first()
    sender_name = company.company_name if company and company.company_name else current_user.email

    # Send message to all assignments (manager message, so employee_id is NULL)
    timestamp = datetime.utcnow()
    for assignment in assignments:
        tender_message = TenderMessageDB(
            assignment_id=assignment.id,
            employee_id=None,  # NULL means manager sent it
            message=message,
            created_at=timestamp
        )
        db.add(tender_message)

    db.commit()

    # Broadcast message via WebSocket
    try:
        await ws_manager.broadcast_chat_message(tender_id, {
            "sender_type": "manager",
            "sender_name": sender_name,
            "message": message,
            "timestamp": timestamp.isoformat(),
            "formatted_time": timestamp.strftime('%d %b, %I:%M %p')
        })
    except Exception as e:
        logger.warning(f"Failed to broadcast message via WebSocket: {e}")
        # Don't fail the request if WebSocket broadcast fails

    return {"message": "Message sent successfully"}


@app.post("/api/tender/assign")
async def assign_employee_to_tender(
    request: Request,
    tender_id: str = Form(...),
    employee_id: str = Form(...),
    role: str = Form(...),
    priority: str = Form("medium"),
    db: Session = Depends(get_db)
):
    """Assign an employee to a tender."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if tender exists
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if employee exists and belongs to user's company
    employee = db.query(EmployeeDB).filter(EmployeeDB.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Verify employee belongs to user's company
    user_company_codes = db.query(CompanyCodeDB).filter(
        CompanyCodeDB.user_id == current_user.id
    ).all()
    company_code_ids = [cc.id for cc in user_company_codes]

    if employee.company_code_id not in company_code_ids:
        raise HTTPException(status_code=403, detail="Employee does not belong to your company")

    # Check if assignment already exists
    existing_assignment = db.query(TenderAssignmentDB).filter(
        and_(
            TenderAssignmentDB.tender_id == tender_id,
            TenderAssignmentDB.employee_id == employee_id
        )
    ).first()

    if existing_assignment:
        return {"message": "Employee already assigned to this tender", "assignment_id": existing_assignment.id}

    # Create new assignment
    new_assignment = TenderAssignmentDB(
        tender_id=tender_id,
        employee_id=employee_id,
        role=role,
        assigned_by=current_user.id,
        priority=priority,
        assigned_at=datetime.utcnow()
    )

    db.add(new_assignment)
    db.commit()
    db.refresh(new_assignment)

    # Notify employee
    notification = EmployeeNotificationDB(
        employee_id=employee_id,
        notification_type="tender_assigned",
        title="Assigned to Tender",
        message=f"You have been assigned to tender: {tender.title}",
        related_tender_id=tender_id,
        created_at=datetime.utcnow()
    )
    db.add(notification)
    db.commit()

    return {"message": "Employee assigned successfully", "assignment_id": new_assignment.id}


@app.post("/api/employee/assignments-with-tasks")
async def create_assignment_with_tasks(
    request: Request,
    tender_id: str = Form(...),
    employee_id: str = Form(...),
    role: str = Form(...),
    priority: str = Form("medium"),
    tasks: str = Form(...),
    db: Session = Depends(get_db)
):
    """Create a new tender assignment with multiple tasks."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if tender exists
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Check if employee exists and belongs to user's company
    employee = db.query(EmployeeDB).filter(EmployeeDB.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Check if employee belongs to user's company
    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
    employee_company_codes = [code.id for code in company_codes]
    if employee.company_code_id not in employee_company_codes:
        raise HTTPException(status_code=403, detail="Employee does not belong to your company")

    # Check if assignment already exists
    existing_assignment = db.query(TenderAssignmentDB).filter(
        and_(TenderAssignmentDB.tender_id == tender_id, TenderAssignmentDB.employee_id == employee_id)
    ).first()

    if existing_assignment:
        raise HTTPException(status_code=400, detail="Employee is already assigned to this tender")

    # Parse tasks
    try:
        tasks_data = json.loads(tasks)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid tasks format")

    # Create assignment
    new_assignment = TenderAssignmentDB(
        tender_id=tender_id,
        employee_id=employee_id,
        role=role,
        assigned_by=current_user.id,
        priority=priority,
        assigned_at=datetime.utcnow()
    )

    db.add(new_assignment)
    db.commit()
    db.refresh(new_assignment)

    # Create tasks
    created_tasks = []
    for task_data in tasks_data:
        # Parse deadline
        parsed_deadline = None
        if task_data.get('deadline'):
            try:
                parsed_deadline = datetime.fromisoformat(task_data['deadline'].replace('Z', '+00:00'))
            except ValueError:
                continue  # Skip invalid deadlines

        new_task = TaskDB(
            assignment_id=new_assignment.id,
            title=task_data['title'],
            description=task_data.get('description', ''),
            priority=task_data.get('priority', 'medium'),
            deadline=parsed_deadline,
            created_at=datetime.utcnow()
        )

        db.add(new_task)
        created_tasks.append(new_task)

    db.commit()

    return {
        "message": "Assignment and tasks created successfully",
        "assignment_id": new_assignment.id,
        "tasks_count": len(created_tasks)
    }

@app.get("/api/tender/{tender_id}/assignments")
async def get_tender_assignments(request: Request, tender_id: str, db: Session = Depends(get_db)):
    """Get all employee assignments for a specific tender."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Check if tender exists
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Get user's company codes to filter assignments
    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
    company_code_ids = [code.id for code in company_codes]

    # Get assignments for this tender where employees belong to user's company
    assignments = db.query(TenderAssignmentDB).join(EmployeeDB).filter(
        and_(TenderAssignmentDB.tender_id == tender_id, EmployeeDB.company_code_id.in_(company_code_ids))
    ).all()

    # Format assignments for frontend
    formatted_assignments = []
    for assignment in assignments:
        if not assignment.employee:
            continue

        # Get task count for this assignment
        task_count = db.query(TaskDB).filter(TaskDB.assignment_id == assignment.id).count()
        completed_tasks = db.query(TaskDB).filter(
            and_(TaskDB.assignment_id == assignment.id, TaskDB.status == 'completed')
        ).count()

        formatted_assignments.append({
            'id': assignment.id,
            'employee_id': assignment.employee.id,
            'employee_name': assignment.employee.name,
            'employee_email': assignment.employee.email,
            'role': assignment.role,
            'priority': assignment.priority,
            'assigned_at': assignment.assigned_at.isoformat() if assignment.assigned_at else None, # type: ignore
            'task_count': task_count,
            'completed_tasks': completed_tasks
        })

    return {"assignments": formatted_assignments}

@app.delete("/api/employee/assignments/{assignment_id}")
async def unassign_employee(request: Request, assignment_id: int, db: Session = Depends(get_db)):
    """Un-assign an employee from a tender."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Find the assignment
    assignment = db.query(TenderAssignmentDB).filter(TenderAssignmentDB.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")

    # Check if assignment belongs to user's company
    company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == current_user.id).all()
    employee_company_codes = [code.id for code in company_codes]
    if assignment.employee.company_code_id not in employee_company_codes:
        raise HTTPException(status_code=403, detail="Assignment does not belong to your company")

    # Delete the assignment (this will cascade delete tasks and messages)
    db.delete(assignment)
    db.commit()

    return {"message": "Employee un-assigned successfully"}

# Certificate API endpoints
# NOTE: Certificate upload endpoint moved to api/routes/certificates.py for bulk upload support
# The old single-file upload endpoint is now replaced with a more powerful bulk upload system
# that supports multiple files, folders, zip files, and duplicate detection.
# See api/routes/certificates.py for the new implementation.

# @app.post("/api/certificates/upload")
# @require_company_details
# async def upload_certificate(
#     request: Request,
#     file: UploadFile = File(...),
#     db: Session = Depends(get_db)
# ):
#     """OLD ENDPOINT - Replaced by bulk upload endpoint in api/routes/certificates.py"""
#     pass

@app.get("/api/certificates")
@require_company_details
async def get_certificates(
    request: Request,
    search: str = "",
    client: str = "",
    location: str = "",
    page: int = 1,
    per_page: int = 10,
    db: Session = Depends(get_db)
):
    """Get user's certificates with search and filtering."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Build query
    query = db.query(CertificateDB).filter(
        and_(CertificateDB.user_id == current_user.id, CertificateDB.processing_status == "completed")
    )

    # Apply filters
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                CertificateDB.project_name.ilike(search_term),
                CertificateDB.extracted_text.ilike(search_term)
            )
        )

    if client:
        client_term = f"%{client}%"
        query = query.filter(CertificateDB.client_name.ilike(client_term))

    if location:
        location_term = f"%{location}%"
        query = query.filter(CertificateDB.location.ilike(location_term))

    # Get total count
    total_certificates = query.count()

    # Apply pagination
    certificates = query.order_by(desc(CertificateDB.created_at)).offset((page - 1) * per_page).limit(per_page).all()

    # Format results
    results = []
    for cert in certificates:
        results.append({
            "id": cert.id,
            "project_name": cert.project_name,
            "client_name": cert.client_name,
            "completion_date": cert.completion_date.isoformat() if cert.completion_date else None,
            "project_value": cert.project_value,
            "services_rendered": cert.services_rendered,
            "location": cert.location,
            "original_filename": cert.original_filename,
            "created_at": cert.created_at.isoformat() if cert.created_at else None,
            "processed_at": cert.processed_at.isoformat() if cert.processed_at else None
        })

    return {
        "certificates": results,
        "total": total_certificates,
        "page": page,
        "per_page": per_page,
        "total_pages": (total_certificates + per_page - 1) // per_page
    }

@app.post("/api/certificates/search")
@require_company_details
async def search_certificates_post(
    request: Request,
    db: Session = Depends(get_db)
):
    """Search certificates with powerful keyword matching and filtering."""
    current_user = get_current_user(request, db)
    if not current_user:
        logger.error("❌ Certificate search: User not authenticated")
        raise HTTPException(status_code=401, detail="Authentication required")

    logger.info(f"🔍 Certificate search: User {current_user.email} (ID: {current_user.id})")

    # Parse request body
    try:
        body = await request.json()
        query = body.get("query", "").strip()
        filter_type = body.get("filter", "all")
        limit = body.get("limit", 50)
        logger.info(f"🔍 Search query: '{query}', filter: {filter_type}, limit: {limit}")
    except:
        query = ""
        filter_type = "all"
        limit = 50

    # Build base query
    base_query = db.query(CertificateDB).filter(
        and_(
            CertificateDB.user_id == current_user.id,
            CertificateDB.processing_status == "completed"
        )
    )

    # If no search query, return all certificates for the user
    if not query:
        all_certificates = base_query.order_by(desc(CertificateDB.created_at)).limit(limit).all()
        logger.info(f"✅ Returning {len(all_certificates)} certificates (no search query)")

        # Format results
        results = []
        for cert in all_certificates:
            results.append({
                "id": cert.id,
                "project_name": cert.project_name,
                "client_name": cert.client_name,
                "completion_date": cert.completion_date.isoformat() if cert.completion_date else None,
                "project_value": cert.project_value,
                "services_rendered": cert.services_rendered,
                "location": cert.location,
                "original_filename": cert.original_filename,
                "created_at": cert.created_at.isoformat() if cert.created_at else None
            })

        return {"certificates": results}
    else:
        # Powerful keyword search with different strategies
        search_term = f"%{query}%"

        if filter_type == "all":
            # UNIVERSAL SEARCH - Search EVERY field in the certificate database
            # This searches across 20+ fields including text, JSONB arrays, dates, and numbers
            logger.info(f"🔍 Performing universal search across all certificate fields")
            filtered_query = base_query.filter(
                or_(
                    # Core text fields
                    CertificateDB.project_name.ilike(search_term),
                    CertificateDB.client_name.ilike(search_term),
                    CertificateDB.location.ilike(search_term),
                    CertificateDB.original_filename.ilike(search_term),

                    # Large text fields - Full certificate content
                    CertificateDB.extracted_text.ilike(search_term),  # Full OCR text
                    CertificateDB.verbatim_certificate.ilike(search_term),  # Full certificate text
                    CertificateDB.scope_of_work.ilike(search_term),  # Project scope
                    CertificateDB.issuing_authority_details.ilike(search_term),  # Authority info
                    CertificateDB.performance_remarks.ilike(search_term),  # Feedback/remarks
                    CertificateDB.signing_authority_details.ilike(search_term),  # Signatories

                    # Reference and metadata fields
                    CertificateDB.certificate_number.ilike(search_term),  # Certificate IDs
                    CertificateDB.role_lead_jv.ilike(search_term),  # Lead/JV role
                    CertificateDB.funding_agency.ilike(search_term),  # Funding sources
                    CertificateDB.duration.ilike(search_term),  # Project duration

                    # Financial fields (search as text)
                    CertificateDB.consultancy_fee_inr.ilike(search_term),  # Fee amounts
                    CertificateDB.project_value_inr.ilike(search_term),  # Project value text
                    cast(CertificateDB.project_value, String).ilike(search_term),  # Numeric value as text

                    # JSONB array fields (cast to text for searching)
                    cast(CertificateDB.services_rendered, String).ilike(search_term),  # Services array
                    cast(CertificateDB.sectors, String).ilike(search_term),  # Sectors array
                    cast(CertificateDB.sub_sectors, String).ilike(search_term),  # Sub-sectors array
                    cast(CertificateDB.jv_partners, String).ilike(search_term),  # JV partners array
                    cast(CertificateDB.metrics, String).ilike(search_term),  # Metrics JSON

                    # Date fields (cast to text for date string matching)
                    cast(CertificateDB.completion_date, String).ilike(search_term),  # Completion dates
                    cast(CertificateDB.start_date, String).ilike(search_term),  # Start dates
                    cast(CertificateDB.end_date, String).ilike(search_term)  # End dates
                )
            )
        elif filter_type == "project_name":
            filtered_query = base_query.filter(CertificateDB.project_name.ilike(search_term))
        elif filter_type == "client_name":
            filtered_query = base_query.filter(CertificateDB.client_name.ilike(search_term))
        elif filter_type == "location":
            filtered_query = base_query.filter(CertificateDB.location.ilike(search_term))
        elif filter_type == "services":
            # Search in services_rendered array (stored as JSON)
            filtered_query = base_query.filter(
                cast(CertificateDB.services_rendered, String).ilike(search_term)
            )
        else:
            filtered_query = base_query

        # Also try vector search for semantic similarity
        try:
            vector_results = certificate_processor.search_certificates(
                query=query,
                user_id=current_user.id,
                limit=limit//2  # Get half from vector search
            )
            vector_cert_ids = [result['certificate'].id for result in vector_results]
            logger.info(f"✅ Vector search returned {len(vector_cert_ids)} certificates")
        except Exception as vector_error:
            logger.warning(f"⚠️ Vector search failed: {vector_error}")
            vector_cert_ids = []

        # Execute the filtered query
        certificates = filtered_query.order_by(desc(CertificateDB.created_at)).limit(limit).all()
        logger.info(f"✅ Text search returned {len(certificates)} certificates (query: '{query}')")

    # Format results
    results = []
    for cert in certificates:
        results.append({
            "id": cert.id,
            "project_name": cert.project_name,
            "client_name": cert.client_name,
            "completion_date": cert.completion_date.isoformat() if cert.completion_date else None,
            "project_value": cert.project_value,
            "services_rendered": cert.services_rendered,
            "location": cert.location,
            "original_filename": cert.original_filename,
            "created_at": cert.created_at.isoformat() if cert.created_at else None
        })

    logger.info(f"✅ Returning {len(results)} certificates (with search query)")
    return {"certificates": results}

@app.get("/api/certificates/debug-auth")
async def debug_auth_status(
    request: Request,
    db: Session = Depends(get_db)
):
    """Debug endpoint to check authentication status."""
    current_user = get_current_user(request, db)

    if not current_user:
        return {
            "authenticated": False,
            "message": "No user logged in",
            "session_token": request.cookies.get('session_token', 'Not found')
        }

    # Count certificates for this user
    cert_count = db.query(CertificateDB).filter(
        and_(
            CertificateDB.user_id == current_user.id,
            CertificateDB.processing_status == "completed"
        )
    ).count()

    return {
        "authenticated": True,
        "user_id": current_user.id,
        "email": current_user.email,
        "name": current_user.name,
        "certificate_count": cert_count
    }

@app.get("/api/certificates/batch-pdf-test")
async def test_batch_pdf_endpoint():
    """Simple test endpoint to verify routing works."""
    print("🔥 TEST ENDPOINT HIT!")
    return {"status": "success", "message": "Test endpoint works!"}

@app.get("/api/certificates/batch-pdf")
async def generate_certificates_batch_pdf(
    request: Request,
    ids: str = "",
    db: Session = Depends(get_db)
):
    """Generate a combined PDF from multiple certificates."""
    print("\n" + "=" * 80)
    print("🔥 BATCH PDF ENDPOINT HIT!")
    print(f"Request URL: {request.url}")
    print(f"IDs parameter received: {ids[:100] if ids else 'EMPTY'}...")
    print("=" * 80 + "\n")

    try:
        print("BATCH PDF GENERATION STARTED")
        print(f"Raw IDs parameter (first 200 chars): {ids[:200] if ids else 'EMPTY'}...")

        current_user = get_current_user(request, db)
        if not current_user:
            print("ERROR: No current user found - authentication failed")
            raise HTTPException(status_code=401, detail="Authentication required")

        print(f"Current user: {current_user.id} ({current_user.email})")

        # Parse certificate IDs
        cert_ids = [cert_id.strip() for cert_id in ids.split(',') if cert_id.strip()]
        print(f"Parsed {len(cert_ids)} certificate IDs")

        if not cert_ids:
            print("ERROR: No certificate IDs after parsing")
            raise HTTPException(status_code=400, detail="No certificate IDs provided")

        # Limit to top 10 certificates
        original_count = len(cert_ids)
        cert_ids = cert_ids[:10]
        print(f"Limited from {original_count} to {len(cert_ids)} certificates (top 10)")
        print(f"Certificate IDs to fetch: {cert_ids}")

        # Get certificates and preserve the order
        print(f"Querying database for certificates with user_id={current_user.id}")
        certificates = db.query(CertificateDB).filter(
            and_(
                CertificateDB.id.in_(cert_ids),
                CertificateDB.user_id == current_user.id
            )
        ).all()

        print(f"Database returned {len(certificates)} certificates")

        if certificates:
            print("Certificate details:")
            for cert in certificates:
                print(f"  - ID: {cert.id}, Project: {cert.project_name}, Client: {cert.client_name}")

        if not certificates:
            print(f"ERROR: No certificates found in database for user {current_user.id}")
            print(f"ERROR: Searched for IDs: {cert_ids}")

            # Try to find if certificates exist at all
            all_certs = db.query(CertificateDB).filter(CertificateDB.id.in_(cert_ids)).all()
            if all_certs:
                print(f"ERROR: Found {len(all_certs)} certificates but they belong to different users:")
                for cert in all_certs:
                    print(f"  - ID: {cert.id}, User: {cert.user_id}")
            else:
                print("ERROR: Certificates with these IDs do not exist in database at all")

            raise HTTPException(status_code=404, detail="No certificates found")

        # Sort certificates to match the order of cert_ids (maintains similarity score order)
        cert_dict = {cert.id: cert for cert in certificates}
        certificates = [cert_dict[cert_id] for cert_id in cert_ids if cert_id in cert_dict]
        print(f"Sorted {len(certificates)} certificates in order")

    except HTTPException:
        raise
    except Exception as e:
        print(f"EXCEPTION in batch-pdf endpoint: {str(e)}")
        print(f"Exception type: {type(e).__name__}")
        import traceback
        print(f"Traceback:\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

    # Generate combined PDF using ReportLab
    print("Starting PDF generation with ReportLab")
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        import io
        from datetime import datetime
        print("All PDF imports successful")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.75*inch, bottomMargin=0.75*inch)
        print("PDF document created")

        elements = []
        styles = getSampleStyleSheet()
    except Exception as e:
        print(f"ERROR during PDF setup: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error setting up PDF: {str(e)}")

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=8,
        fontName='Helvetica-Bold'
    )

    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#6b7280'),
        alignment=TA_CENTER,
        spaceAfter=6
    )

    # Add each certificate (one per page)
    print(f"Adding {len(certificates)} certificates to PDF")
    for idx, cert in enumerate(certificates, 1):
        try:
            print(f"Processing certificate {idx}/{len(certificates)}: {cert.id}")
            # Certificate header with number
            elements.append(Paragraph(f"<b>Certificate {idx} of {len(certificates)}</b>", heading_style))
            elements.append(Spacer(1, 0.15*inch))

            # Project name as title
            if cert.project_name:
                print(f"  Project name: {cert.project_name}")
                project_title_style = ParagraphStyle(
                    'ProjectTitle',
                    parent=styles['Heading3'],
                    fontSize=12,
                    textColor=colors.HexColor('#111827'),
                    spaceAfter=10,
                    fontName='Helvetica-Bold'
                )
                elements.append(Paragraph(cert.project_name, project_title_style))
        except Exception as e:
            print(f"ERROR processing certificate {idx}: {str(e)}")
            import traceback
            print(traceback.format_exc())
            continue

        # Certificate details table
        data = []

        # Basic information
        if cert.client_name:
            data.append(['Client Name:', cert.client_name])

        if cert.location:
            data.append(['Location:', cert.location])

        if cert.project_value:
            data.append(['Contract Value:', f"₹{cert.project_value:,.2f}"])

        if cert.completion_date:
            data.append(['Completion Date:', cert.completion_date.strftime('%B %d, %Y')])

        # Add services rendered if available
        if cert.services_rendered:
            services = cert.services_rendered if isinstance(cert.services_rendered, list) else [cert.services_rendered]
            services_text = ", ".join(services) if len(services) <= 10 else ", ".join(services[:10]) + f" (+{len(services) - 10} more)"
            data.append(['Services/Work Done:', services_text])

        # Add extracted text if available (as work description)
        if cert.extracted_text and len(cert.extracted_text.strip()) > 0:
            # Truncate to first 500 characters
            text_preview = cert.extracted_text[:500] + "..." if len(cert.extracted_text) > 500 else cert.extracted_text
            work_desc = Paragraph(text_preview, styles['Normal'])
            data.append(['Description:', work_desc])

        # Only create table if we have data
        if data:
            table = Table(data, colWidths=[2*inch, 4.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#eff6ff')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e40af')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(table)

        # Add page break except for last certificate
        if idx < len(certificates):
            elements.append(PageBreak())

    # Build PDF
    try:
        print(f"Building PDF with {len(elements)} elements")
        doc.build(elements)
        buffer.seek(0)
        pdf_size = len(buffer.getvalue())
        print(f"PDF built successfully, size: {pdf_size} bytes")

        # Return PDF
        from fastapi.responses import StreamingResponse
        filename = f"similar_certificates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        print(f"Returning PDF as StreamingResponse: {filename}")
        print("=" * 80 + "\n")

        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        print(f"ERROR building PDF: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error building PDF: {str(e)}")

@app.get("/api/certificates/{certificate_id}")
async def get_certificate(
    request: Request,
    certificate_id: str,
    db: Session = Depends(get_db)
):
    """Get detailed certificate information."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    certificate = db.query(CertificateDB).filter(
        and_(CertificateDB.id == certificate_id, CertificateDB.user_id == current_user.id)
    ).first()

    if not certificate:
        raise HTTPException(status_code=404, detail="Certificate not found")

    return {
        "id": certificate.id,
        "project_name": certificate.project_name,
        "client_name": certificate.client_name,
        "completion_date": certificate.completion_date.isoformat() if certificate.completion_date else None,
        "project_value": certificate.project_value,
        "services_rendered": certificate.services_rendered,
        "location": certificate.location,
        "original_filename": certificate.original_filename,
        "extracted_text": certificate.extracted_text,
        "processing_status": certificate.processing_status,
        "processing_error": certificate.processing_error,
        "created_at": certificate.created_at.isoformat() if certificate.created_at else None,
        "processed_at": certificate.processed_at.isoformat() if certificate.processed_at else None
    }

@app.get("/api/certificates/{certificate_id}/file")
async def get_certificate_file(
    request: Request,
    certificate_id: str,
    db: Session = Depends(get_db)
):
    """View certificate file (for preview in browser)."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    certificate = db.query(CertificateDB).filter(
        and_(CertificateDB.id == certificate_id, CertificateDB.user_id == current_user.id)
    ).first()

    if not certificate:
        raise HTTPException(status_code=404, detail="Certificate not found")

    # Try S3 first, then fall back to local file
    if certificate.s3_key:
        try:
            from s3_utils import get_presigned_url, download_from_s3
            # Generate presigned URL for 1 hour access
            presigned_url = get_presigned_url(certificate.s3_key, expiration=3600)
            if presigned_url:
                # Redirect to presigned URL for direct access
                return RedirectResponse(url=presigned_url, status_code=302)
            else:
                # If presigned URL fails, download and serve
                success, file_data, content_type = download_from_s3(certificate.s3_key)
                if success:
                    return Response(
                        content=file_data,
                        media_type=content_type,
                        headers={"Content-Disposition": "inline"}
                    )
        except Exception as e:
            logger.error(f"Error accessing S3 file: {e}")
            # Fall through to local file check

    # Fall back to local file if S3 is not available
    file_path = certificate.file_path
    if file_path and os.path.exists(file_path):
        # Determine media type based on file extension
        file_extension = os.path.splitext(file_path)[1].lower()
        media_types = {
            '.pdf': 'application/pdf',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png'
        }
        media_type = media_types.get(file_extension, 'application/octet-stream')

        # Return file for inline viewing (not download)
        return FileResponse(
            path=file_path,
            media_type=media_type,
            headers={
                "Content-Disposition": "inline"
            }
        )

    raise HTTPException(status_code=404, detail="Certificate file not found")

@app.get("/api/certificates/{certificate_id}/download")
async def download_certificate_file(
    request: Request,
    certificate_id: str,
    db: Session = Depends(get_db)
):
    """Download certificate file."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    certificate = db.query(CertificateDB).filter(
        and_(CertificateDB.id == certificate_id, CertificateDB.user_id == current_user.id)
    ).first()

    if not certificate:
        raise HTTPException(status_code=404, detail="Certificate not found")

    # Try S3 first, then fall back to local file
    if certificate.s3_key:
        try:
            from s3_utils import get_presigned_url, download_from_s3
            # Generate presigned URL for download (1 hour access)
            presigned_url = get_presigned_url(
                certificate.s3_key,
                expiration=3600,
                filename=certificate.original_filename or "certificate"
            )
            if presigned_url:
                # Redirect to presigned URL for direct download
                return RedirectResponse(url=presigned_url, status_code=302)
            else:
                # If presigned URL fails, download and serve
                success, file_data, content_type = download_from_s3(certificate.s3_key)
                if success:
                    return Response(
                        content=file_data,
                        media_type=content_type,
                        headers={
                            "Content-Disposition": f'attachment; filename="{certificate.original_filename or "certificate"}"'
                        }
                    )
        except Exception as e:
            logger.error(f"Error accessing S3 file: {e}")
            # Fall through to local file check

    # Fall back to local file if S3 is not available
    file_path = certificate.file_path
    if file_path and os.path.exists(file_path):
        # Determine media type based on file extension
        file_extension = os.path.splitext(file_path)[1].lower()
        media_types = {
            '.pdf': 'application/pdf',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png'
        }
        media_type = media_types.get(file_extension, 'application/octet-stream')
        
        return FileResponse(
            path=file_path,
            media_type=media_type,
            filename=certificate.original_filename or "certificate",
            headers={"Content-Disposition": f'attachment; filename="{certificate.original_filename or "certificate"}"'}
        )

    raise HTTPException(status_code=404, detail="Certificate file not found")


# ============================================================
# CERTIFICATE MATCHING PLACEHOLDER ROUTE
# ============================================================

SIMILAR_CERTIFICATES_DISABLED_MSG = (
    "Certificate similarity search is temporarily unavailable while we redesign this workflow."
)

@app.get("/api/tenders/{tender_id}/similar-certificates")
async def find_similar_certificates(
    request: Request,
    tender_id: str,
    db: Session = Depends(get_db)
):
    """Return a placeholder response until the new matcher ships."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # We still check that the tender exists to mimic old behavior
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    return {
        'tender_id': tender_id,
        'tender_title': tender.title,
        'certificates': [],
        'count': 0,
        'message': SIMILAR_CERTIFICATES_DISABLED_MSG,
    }

@app.get("/api/certificates/analytics")
async def get_certificate_analytics(
    request: Request,
    db: Session = Depends(get_db)
):
    """Get certificate analytics and statistics."""
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get basic counts
    total_certificates = db.query(CertificateDB).filter(
        and_(CertificateDB.user_id == current_user.id, CertificateDB.processing_status == "completed")
    ).count()

    # Get date ranges
    certificates = db.query(CertificateDB).filter(
        and_(CertificateDB.user_id == current_user.id, CertificateDB.processing_status == "completed")
    ).all()

    # Calculate analytics
    total_value = sum(cert.project_value for cert in certificates if cert.project_value)
    locations = {}
    clients = {}
    services = {}

    for cert in certificates:
        # Count locations
        if cert.location:
            locations[cert.location] = locations.get(cert.location, 0) + 1

        # Count clients
        if cert.client_name:
            clients[cert.client_name] = clients.get(cert.client_name, 0) + 1

        # Count services
        if cert.services_rendered:
            for service in cert.services_rendered:
                services[service] = services.get(service, 0) + 1

    # Get completion date ranges
    completion_dates = [cert.completion_date for cert in certificates if cert.completion_date]
    date_range = None
    if completion_dates:
        min_date = min(completion_dates)
        max_date = max(completion_dates)
        date_range = {
            "earliest": min_date.isoformat(),
            "latest": max_date.isoformat()
        }

    return {
        "total_certificates": total_certificates,
        "total_project_value": total_value,
        "date_range": date_range,
        "top_locations": sorted(locations.items(), key=lambda x: x[1], reverse=True)[:10],
        "top_clients": sorted(clients.items(), key=lambda x: x[1], reverse=True)[:10],
        "top_services": sorted(services.items(), key=lambda x: x[1], reverse=True)[:10]
    }

@app.get("/api/certificates/manual-clause/filter-options")
async def get_manual_clause_filter_options(
    request: Request,
    db: Session = Depends(get_db)
):
    """Get all unique filter options from user's certificates for manual clause filtering."""
    from core.dependencies import get_current_user_or_bd_employee

    # Check if user or BD employee is logged in
    entity, entity_type = get_current_user_or_bd_employee(request, db)
    if not entity:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get the user_id for queries
    user_id_for_queries = None
    if entity_type == 'user':
        user_id_for_queries = entity.id
    else:  # BD employee
        company_code = db.query(CompanyCodeDB).filter(CompanyCodeDB.id == entity.company_code_id).first()
        if company_code:
            user_id_for_queries = company_code.user_id

    if not user_id_for_queries:
        raise HTTPException(status_code=401, detail="Authentication required")

    # Get all completed certificates for this user
    certificates = db.query(CertificateDB).filter(
        and_(CertificateDB.user_id == user_id_for_queries, CertificateDB.processing_status == "completed")
    ).all()

    # Extract unique values for each filter
    clients = set()
    locations = set()
    services = set()
    sectors = set()
    sub_sectors = set()
    funding_agencies = set()
    roles = set()
    durations = set()
    certificate_numbers = set()

    for cert in certificates:
        if cert.client_name:
            clients.add(cert.client_name)
        if cert.location:
            locations.add(cert.location)
        if cert.services_rendered and isinstance(cert.services_rendered, list):
            services.update(cert.services_rendered)
        if cert.sector:
            sectors.add(cert.sector)
        if cert.sub_sector:
            sub_sectors.add(cert.sub_sector)
        if cert.funding_agency:
            funding_agencies.add(cert.funding_agency)
        if cert.role:
            roles.add(cert.role)
        if cert.duration:
            durations.add(cert.duration)
        if cert.certificate_number:
            certificate_numbers.add(cert.certificate_number)

    # Convert sets to sorted lists
    return {
        "clients": sorted(list(clients)),
        "locations": sorted(list(locations)),
        "services": sorted(list(services)),
        "sectors": sorted(list(sectors)),
        "sub_sectors": sorted(list(sub_sectors)),
        "funding_agencies": sorted(list(funding_agencies)),
        "roles": sorted(list(roles)),
        "durations": sorted(list(durations)),
        "certificate_numbers": sorted(list(certificate_numbers))
    }

# WebSocket Connection Manager for Real-time Chat
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, List[WebSocket]] = defaultdict(list)
        self.employee_sessions: Dict[WebSocket, str] = {}

    async def connect(self, websocket: WebSocket, tender_id: str, employee_id: str, already_accepted: bool = False):
        if not already_accepted:
            await websocket.accept()
        self.active_connections[tender_id].append(websocket)
        self.employee_sessions[websocket] = employee_id

    def disconnect(self, websocket: WebSocket, tender_id: str):
        if websocket in self.active_connections[tender_id]:
            self.active_connections[tender_id].remove(websocket)
        if websocket in self.employee_sessions:
            del self.employee_sessions[websocket]

    async def broadcast_to_tender(self, tender_id: str, message: dict, exclude_websocket: Optional[WebSocket] = None):
        for connection in self.active_connections[tender_id]:
            if connection != exclude_websocket:
                try:
                    await connection.send_json(message)
                except Exception as e:
                    logger.error(f"Error sending message to websocket: {e}")
                    # Remove broken connection
                    if connection in self.active_connections[tender_id]:
                        self.active_connections[tender_id].remove(connection)
                    if connection in self.employee_sessions:
                        del self.employee_sessions[connection]

manager = ConnectionManager()

@app.websocket("/ws/chat/{tender_id}")
async def chat_websocket(websocket: WebSocket, tender_id: str):
    """WebSocket endpoint for real-time chat."""
    # Create database session
    db = SessionLocal()

    try:
        # Get token from query parameters
        token = websocket.query_params.get('token')

        if not token:
            # Fallback to cookies if no token in query params
            session_token = websocket.cookies.get('employee_session_token')
            user_type = 'employee'

            if not session_token:
                session_token = websocket.cookies.get('session_token')
                user_type = 'manager'

            if not session_token:
                await websocket.close(code=1008, reason="Authentication required")
                return
        else:
            session_token = token
            # Determine user type based on session data - will be determined later
            user_type = None

        # Validate session - check both employee and manager sessions
        session_data = user_sessions.get(session_token)
        if not session_data or session_data['expires_at'] < datetime.utcnow():
            await websocket.close(code=1008, reason="Invalid or expired session")
            return

        user_id = session_data['user_id']

        # Determine user type if not already set
        if user_type is None:
            # Check if this is an employee session by looking up the user in EmployeeDB
            employee = db.query(EmployeeDB).filter(EmployeeDB.id == user_id).first()
            user_type = 'employee' if employee else 'manager'
            logger.info(f"Determined user type: {user_type} for user {user_id}")

        # Get company codes for managers (needed later)
        company_code_ids = []
        if user_type == 'manager':
            company_codes = db.query(CompanyCodeDB).filter(CompanyCodeDB.user_id == user_id).all()
            company_code_ids = [code.id for code in company_codes]

        # Check authorization based on user type
        assignment = None
        if user_type == 'employee':
            # Check if employee is assigned to this tender
            assignment = db.query(TenderAssignmentDB).filter(
                and_(TenderAssignmentDB.tender_id == tender_id, TenderAssignmentDB.employee_id == user_id)
            ).first()

            if not assignment:
                await websocket.close(code=1008, reason="Not authorized for this tender")
                return
        else:
            # For managers, check if they have any employees assigned to this tender
            if not company_code_ids:
                await websocket.close(code=1008, reason="Not authorized for this tender")
                return

            # Check if any employees from manager's company are assigned to this tender
            assignment_count = db.query(TenderAssignmentDB).join(EmployeeDB).filter(
                and_(TenderAssignmentDB.tender_id == tender_id, EmployeeDB.company_code_id.in_(company_code_ids))
            ).count()

            if assignment_count == 0:
                await websocket.close(code=1008, reason="Not authorized for this tender")
                return

        # Connect to chat room
        await manager.connect(websocket, tender_id, user_id)

        try:
            while True:
                data = await websocket.receive_json()

                if data.get('type') == 'message':
                    message_text = data.get('message', '').strip()
                    if message_text:
                        # Create a new database session for each message to ensure isolation
                        message_db = SessionLocal()
                        try:
                            # For managers, we need to use an assignment that belongs to their company
                            if user_type == 'manager':
                                # Get the first assignment for this tender that belongs to the manager's company
                                first_assignment = message_db.query(TenderAssignmentDB).join(EmployeeDB).filter(
                                    and_(TenderAssignmentDB.tender_id == tender_id, EmployeeDB.company_code_id.in_(company_code_ids))
                                ).first()
                                if first_assignment:
                                    assignment_id = first_assignment.id
                                else:
                                    # No assignments from manager's company, skip
                                    continue
                            else:
                                if not assignment:
                                    continue
                                assignment_id = assignment.id

                            # Save message to database with proper transaction handling
                            new_message = TenderMessageDB(
                                assignment_id=assignment_id,
                                employee_id=user_id if user_type == 'employee' else None,  # NULL for manager messages
                                message=message_text,
                                created_at=datetime.utcnow()
                            )
                            message_db.add(new_message)
                            message_db.commit()
                            message_db.refresh(new_message)

                            # Get user/employee details for broadcast
                            if user_type == 'employee':
                                sender = message_db.query(EmployeeDB).filter(EmployeeDB.id == user_id).first()
                                sender_name = sender.name if sender else 'Unknown'
                            else:
                                # For manager
                                sender = message_db.query(UserDB).filter(UserDB.id == user_id).first()
                                sender_name = f"{sender.name} (Manager)" if sender else 'Manager'

                            # Broadcast message to all connected clients in this tender
                            message_data = {
                                'type': 'new_message',
                                'message': {
                                    'id': new_message.id,
                                    'employee_id': new_message.employee_id,  # Use the actual employee_id from DB
                                    'employee_name': sender_name,
                                    'message': message_text,
                                    'created_at': new_message.created_at.isoformat() if new_message.created_at is not None else None
                                }
                            }

                            await manager.broadcast_to_tender(tender_id, message_data, websocket)

                        except Exception as msg_error:
                            logger.error(f"Error saving message to database: {msg_error}")
                            message_db.rollback()
                            # Send error back to client
                            await websocket.send_json({
                                'type': 'error',
                                'message': 'Failed to save message. Please try again.'
                            })
                        finally:
                            message_db.close()

        except WebSocketDisconnect:
            manager.disconnect(websocket, tender_id)
        except Exception as e:
            logger.error(f"WebSocket error: {e}")
            manager.disconnect(websocket, tender_id)
    except Exception as setup_error:
        logger.error(f"WebSocket setup error: {setup_error}")
        await websocket.close(code=1011, reason="Internal server error")
    finally:
        db.close()


# ==========================================
# TENDER RESPONSE PREPARATION ROUTES
# ==========================================

@app.get("/prepare-response/{tender_id}", response_class=HTMLResponse)
async def prepare_response_page(
    tender_id: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: UserDB = Depends(get_current_user)
):
    """Render the prepare response page for a tender."""
    # Get tender details
    tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    # Get all existing responses for this tender
    responses = db.query(TenderResponseDB).filter(
        and_(
            TenderResponseDB.tender_id == tender_id,
            TenderResponseDB.user_id == current_user.id
        )
    ).order_by(desc(TenderResponseDB.created_at)).all()

    # Get documents for each response
    for response in responses:
        response.documents = db.query(ResponseDocumentDB).filter(
            ResponseDocumentDB.response_id == response.id
        ).order_by(ResponseDocumentDB.display_order).all()

    return templates.TemplateResponse("prepare_response.html", {
        "request": request,
        "tender": tender,
        "responses": responses,
        "current_user": current_user,
        "selected_font": get_active_font()
    })


@app.post("/api/response/upload-document")
async def upload_response_document(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserDB = Depends(get_current_user)
):
    """Upload a document for the response (temporary, not yet saved to a response)."""
    try:
        # Create upload directory if it doesn't exist
        upload_dir = Path("uploads/response_documents")
        upload_dir.mkdir(parents=True, exist_ok=True)

        # Generate unique filename
        file_ext = Path(file.filename).suffix
        unique_filename = f"{uuid.uuid4()}{file_ext}"
        file_path = upload_dir / unique_filename

        # Save file
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)

        # Get file size
        file_size = len(content)

        return JSONResponse({
            "success": True,
            "filename": file.filename,
            "unique_filename": unique_filename,
            "file_path": str(file_path),
            "file_size": file_size,
            "file_type": file.content_type
        })
    except Exception as e:
        logger.error(f"Error uploading document: {e}")
        raise HTTPException(status_code=500, detail="Failed to upload document")


@app.post("/api/response/finalize")
async def finalize_response(
    tender_id: str = Form(...),
    response_name: str = Form(...),
    response_type: str = Form(...),
    remarks: str = Form(None),
    documents: str = Form(...),  # JSON string of documents
    signature: UploadFile = File(None),
    stamp: UploadFile = File(None),
    db: Session = Depends(get_db),
    current_user: UserDB = Depends(get_current_user)
):
    """Finalize and save a tender response with all documents."""
    try:
        # Parse documents JSON
        docs_list = json.loads(documents)

        # Handle signature upload
        signature_path = None
        if signature and signature.filename:
            sig_dir = Path("uploads/signatures")
            sig_dir.mkdir(parents=True, exist_ok=True)
            sig_ext = Path(signature.filename).suffix
            sig_filename = f"{uuid.uuid4()}{sig_ext}"
            signature_path = sig_dir / sig_filename
            with open(signature_path, "wb") as buffer:
                buffer.write(await signature.read())
            signature_path = str(signature_path)

        # Handle stamp upload
        stamp_path = None
        if stamp and stamp.filename:
            stamp_dir = Path("uploads/signatures")
            stamp_dir.mkdir(parents=True, exist_ok=True)
            stamp_ext = Path(stamp.filename).suffix
            stamp_filename = f"{uuid.uuid4()}{stamp_ext}"
            stamp_path = stamp_dir / stamp_filename
            with open(stamp_path, "wb") as buffer:
                buffer.write(await stamp.read())
            stamp_path = str(stamp_path)

        # Create response record
        response = TenderResponseDB(
            user_id=current_user.id,
            tender_id=tender_id,
            response_name=response_name,
            response_type=response_type,
            remarks=remarks,
            signature_path=signature_path,
            stamp_path=stamp_path,
            is_finalized=True,
            finalized_at=datetime.utcnow()
        )
        db.add(response)
        db.flush()  # Get the response ID

        # Add documents to response
        for idx, doc in enumerate(docs_list):
            response_doc = ResponseDocumentDB(
                response_id=response.id,
                document_name=doc['document_name'],
                filename=doc['filename'],
                file_path=doc['file_path'],
                file_size=doc.get('file_size', 0),
                file_type=doc.get('file_type', ''),
                display_order=idx
            )
            db.add(response_doc)

        db.commit()

        # Generate PDF (we'll do this next)
        pdf_path = await generate_response_pdf(response.id, db)
        response.pdf_path = pdf_path
        response.pdf_filename = f"{response_name}.pdf"
        db.commit()

        return JSONResponse({
            "success": True,
            "response_id": response.id,
            "pdf_path": pdf_path
        })
    except Exception as e:
        db.rollback()
        logger.error(f"Error finalizing response: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to finalize response: {str(e)}")


async def generate_response_pdf(response_id: int, db: Session) -> str:
    """Generate a PDF from a tender response."""
    try:
        # Import PDF libraries
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Image as RLImage, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib import colors
        from PyPDF2 import PdfMerger

        # Get response and documents
        response = db.query(TenderResponseDB).filter(TenderResponseDB.id == response_id).first()
        if not response:
            raise Exception("Response not found")

        documents = db.query(ResponseDocumentDB).filter(
            ResponseDocumentDB.response_id == response_id
        ).order_by(ResponseDocumentDB.display_order).all()

        # Create PDF directory
        pdf_dir = Path("uploads/response_pdfs")
        pdf_dir.mkdir(parents=True, exist_ok=True)

        # Generate cover page PDF
        cover_pdf_path = pdf_dir / f"cover_{response_id}.pdf"
        doc = SimpleDocTemplate(str(cover_pdf_path), pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4f46e5'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#374151'),
            spaceAfter=12,
            spaceBefore=20
        )

        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12
        )

        # Add title
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(response.response_name, title_style))
        story.append(Spacer(1, 0.3*inch))

        # Add response details
        story.append(Paragraph(f"<b>Response Type:</b> {response.response_type}", normal_style))
        story.append(Paragraph(f"<b>Date:</b> {response.finalized_at.strftime('%d %B %Y')}", normal_style))

        if response.remarks:
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph("<b>Remarks:</b>", heading_style))
            story.append(Paragraph(response.remarks, normal_style))

        # Add document list
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("<b>Included Documents:</b>", heading_style))

        for idx, doc in enumerate(documents, 1):
            story.append(Paragraph(f"{idx}. {doc.document_name}", normal_style))

        # Add signature and stamp if available
        story.append(Spacer(1, 0.5*inch))

        sig_stamp_data = []
        if response.signature_path and Path(response.signature_path).exists():
            sig_stamp_data.append(["Signature:", RLImage(response.signature_path, width=2*inch, height=1*inch)])
        if response.stamp_path and Path(response.stamp_path).exists():
            sig_stamp_data.append(["Stamp:", RLImage(response.stamp_path, width=2*inch, height=1*inch)])

        if sig_stamp_data:
            sig_table = Table(sig_stamp_data, colWidths=[1.5*inch, 2.5*inch])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ]))
            story.append(sig_table)

        # Build cover page
        doc.build(story)

        # Merge with uploaded documents
        merger = PdfMerger()
        merger.append(str(cover_pdf_path))

        for doc in documents:
            if Path(doc.file_path).exists() and doc.file_path.lower().endswith('.pdf'):
                try:
                    merger.append(doc.file_path)
                except Exception as e:
                    logger.warning(f"Could not merge PDF {doc.file_path}: {e}")

        # Final PDF path
        final_pdf_path = pdf_dir / f"response_{response_id}.pdf"
        merger.write(str(final_pdf_path))
        merger.close()

        # Clean up cover page
        cover_pdf_path.unlink()

        return str(final_pdf_path)

    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        raise


@app.get("/api/response/{response_id}/download")
async def download_response_pdf(
    response_id: int,
    db: Session = Depends(get_db),
    current_user: UserDB = Depends(get_current_user)
):
    """Download a generated response PDF."""
    response = db.query(TenderResponseDB).filter(
        and_(
            TenderResponseDB.id == response_id,
            TenderResponseDB.user_id == current_user.id
        )
    ).first()

    if not response or not response.pdf_path:
        raise HTTPException(status_code=404, detail="Response not found")

    if not Path(response.pdf_path).exists():
        raise HTTPException(status_code=404, detail="PDF file not found")

    return FileResponse(
        path=response.pdf_path,
        filename=response.pdf_filename,
        media_type='application/pdf'
    )


@app.delete("/api/response/document/{filename}")
async def delete_temp_document(
    filename: str,
    db: Session = Depends(get_db),
    current_user: UserDB = Depends(get_current_user)
):
    """Delete a temporarily uploaded document."""
    try:
        file_path = Path("uploads/response_documents") / filename
        if file_path.exists():
            file_path.unlink()
        return JSONResponse({"success": True})
    except Exception as e:
        logger.error(f"Error deleting document: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete document")


# ==================== WebSocket Endpoints ====================

@app.websocket("/ws/tender/{tender_id}")
async def websocket_tender_endpoint(
    websocket: WebSocket,
    tender_id: str,
    db: Session = Depends(get_db)
):
    """
    WebSocket endpoint for real-time updates in tender workspace.
    Supports both manager and employee connections for instant chat and task updates.

    Connection format:
    - Client sends initial message: {"type": "auth", "user_type": "manager|employee", "user_id": "...", "user_name": "..."}
    - Server responds with: {"type": "connected", "message": "Connected successfully"}
    - Then real-time events flow bidirectionally
    """
    connection_established = False
    user_type = None
    user_id = None
    user_name = None

    try:
        # Accept connection temporarily
        await websocket.accept()

        # Wait for authentication message
        auth_message = await websocket.receive_json()

        if auth_message.get("type") != "auth":
            await websocket.send_json({"type": "error", "message": "Authentication required"})
            await websocket.close()
            return

        user_type = auth_message.get("user_type")  # "manager" or "employee"
        user_id = auth_message.get("user_id")
        user_name = auth_message.get("user_name")

        if not all([user_type, user_id, user_name]):
            await websocket.send_json({"type": "error", "message": "Invalid authentication data"})
            await websocket.close()
            return

        # Verify access to tender
        if user_type == "manager":
            # Verify user is the manager who assigned this tender
            tender = db.query(TenderDB).filter(TenderDB.id == tender_id).first()
            if not tender or tender.awarded_by != user_id:
                await websocket.send_json({"type": "error", "message": "Unauthorized access"})
                await websocket.close()
                return
        elif user_type == "employee":
            # Verify employee is assigned to this tender
            assignment = db.query(TenderAssignmentDB).filter(
                and_(
                    TenderAssignmentDB.tender_id == tender_id,
                    TenderAssignmentDB.employee_id == user_id
                )
            ).first()
            if not assignment:
                await websocket.send_json({"type": "error", "message": "Unauthorized access"})
                await websocket.close()
                return
        else:
            await websocket.send_json({"type": "error", "message": "Invalid user type"})
            await websocket.close()
            return

        # Connection is now authenticated - add to manager (already accepted above)
        await ws_manager.connect(websocket, tender_id, user_type, user_id, user_name, already_accepted=True)
        connection_established = True

        # Send confirmation
        await websocket.send_json({
            "type": "connected",
            "message": "Connected successfully",
            "tender_id": tender_id,
            "room_size": ws_manager.get_room_size(tender_id)
        })

        # Listen for messages
        while True:
            try:
                data = await websocket.receive_json()

                # Handle ping/pong for keep-alive
                if data.get("type") == "ping":
                    await websocket.send_json({"type": "pong"})
                    continue

                # Echo back any other messages (client-side events handled via HTTP endpoints)
                logger.info(f"Received WebSocket message from {user_type} {user_name}: {data.get('type')}")

            except WebSocketDisconnect:
                break
            except Exception as e:
                logger.error(f"Error in WebSocket loop: {e}")
                break

    except WebSocketDisconnect:
        logger.info(f"WebSocket disconnected: {user_type} {user_name}")
    except Exception as e:
        logger.error(f"WebSocket error: {e}")
    finally:
        if connection_established:
            ws_manager.disconnect(websocket)


# ==================== Test Certificate Endpoints (Complete Isolation) ====================

# Import certificate helper functions
from api.routes.certificates import (
    get_distinct_values, get_distinct_json_values, calculate_fee_ranges,
    parse_consultancy_fee, format_currency_range, count_active_filters,
    calculate_compliance, get_compliance_levels, extract_json_from_response,
    search_certificate_by_keyword, calculate_keyword_matches
)
from core.openai_wrapper import OpenAIWrapper, OpenAIServiceError

def extract_test_token_from_request(request: Request, body: Optional[dict] = None) -> Optional[str]:
    """Extract test_token from query params or JSON body."""
    # Try query params first
    test_token = request.query_params.get('test_token')
    if test_token:
        return test_token
    # Try JSON body if provided
    if body:
        return body.get('test_token')
    return None

@app.get("/public-test/api/certificates/filter-options")
async def test_get_certificate_filter_options(
    request: Request,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Get filter options for certificate search."""
    test_user = validate_test_token_and_get_user(test_token, db)
    
    logger.info(f"📊 [TEST] Fetching filter options for test user {test_user.email}")
    
    try:
        # Get distinct values for categorical filters
        clients = get_distinct_values(db, 'client_name', test_user.id)
        locations = get_distinct_values(db, 'location', test_user.id)
        funding_agencies = get_distinct_values(db, 'funding_agency', test_user.id)
        
        # Get distinct values from JSONB fields
        services = get_distinct_json_values(db, 'services_rendered', test_user.id)
        sectors = get_distinct_json_values(db, 'sectors', test_user.id)
        sub_sectors = get_distinct_json_values(db, 'sub_sectors', test_user.id)
        
        # Calculate consultancy fee ranges (8 equal divisions)
        fee_ranges = calculate_fee_ranges(db, test_user.id)
        
        # Get project value min/max
        value_stats = db.query(
            func.min(CertificateDB.project_value),
            func.max(CertificateDB.project_value)
        ).filter(
            CertificateDB.user_id == test_user.id,
            CertificateDB.project_value.isnot(None)
        ).first()
        
        # Get date ranges
        date_stats = db.query(
            func.min(CertificateDB.completion_date),
            func.max(CertificateDB.completion_date)
        ).filter(
            CertificateDB.user_id == test_user.id,
            CertificateDB.completion_date.isnot(None)
        ).first()
        
        response = {
            "success": True,
            "clients": clients,
            "locations": locations,
            "services": services,
            "sectors": sectors,
            "sub_sectors": sub_sectors,
            "funding_agencies": funding_agencies,
            "roles": ["Lead Consultant", "JV Partner", "Consortium"],
            "processing_statuses": ["completed", "processing", "failed", "duplicate"],
            "ranges": {
                "consultancy_fee": fee_ranges,
                "project_value": {
                    "min": float(value_stats[0]) if value_stats[0] else 0,
                    "max": float(value_stats[1]) if value_stats[1] else 0
                },
                "completion_date": {
                    "min": date_stats[0].isoformat() if date_stats[0] else None,
                    "max": date_stats[1].isoformat() if date_stats[1] else None
                }
            }
        }
        
        logger.info(f"✅ [TEST] Filter options generated: {len(clients)} clients, {len(locations)} locations, {len(services)} services")
        
        return response
        
    except Exception as e:
        logger.error(f"❌ [TEST] Error fetching filter options: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching filter options: {str(e)}")


@app.get("/public-test/api/certificates/manual-clause/filter-options")
async def test_get_manual_clause_filter_options(
    request: Request,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Get manual clause filter options."""
    test_user = validate_test_token_and_get_user(test_token, db)
    
    logger.info(f"📊 [TEST] Fetching manual clause filter options for test user {test_user.email}")
    
    try:
        # Get distinct values for categorical filters
        clients = get_distinct_values(db, 'client_name', test_user.id)
        locations = get_distinct_values(db, 'location', test_user.id)
        funding_agencies = get_distinct_values(db, 'funding_agency', test_user.id)
        durations = get_distinct_values(db, 'duration', test_user.id)
        certificate_numbers = get_distinct_values(db, 'certificate_number', test_user.id)
        roles = get_distinct_values(db, 'role_lead_jv', test_user.id)
        
        # If no roles in DB, use default options
        if not roles:
            roles = ["Lead Consultant", "JV Partner", "Consortium", "Solo"]
        
        # Get distinct values from JSONB fields
        services = get_distinct_json_values(db, 'services_rendered', test_user.id)
        sectors = get_distinct_json_values(db, 'sectors', test_user.id)
        sub_sectors = get_distinct_json_values(db, 'sub_sectors', test_user.id)
        
        # Get consultancy fee min/max (for range slider)
        fee_stats = db.query(
            func.min(CertificateDB.consultancy_fee_numeric),
            func.max(CertificateDB.consultancy_fee_numeric)
        ).filter(
            CertificateDB.user_id == test_user.id,
            CertificateDB.consultancy_fee_numeric.isnot(None)
        ).first()
        
        # Get project value min/max
        value_stats = db.query(
            func.min(CertificateDB.project_value),
            func.max(CertificateDB.project_value)
        ).filter(
            CertificateDB.user_id == test_user.id,
            CertificateDB.project_value.isnot(None)
        ).first()
        
        # Get date ranges
        date_stats = db.query(
            func.min(CertificateDB.completion_date),
            func.max(CertificateDB.completion_date)
        ).filter(
            CertificateDB.user_id == test_user.id,
            CertificateDB.completion_date.isnot(None)
        ).first()
        
        response = {
            "success": True,
            "clients": clients,
            "locations": locations,
            "services": services,
            "sectors": sectors,
            "sub_sectors": sub_sectors,
            "funding_agencies": funding_agencies,
            "roles": roles,
            "durations": durations,
            "certificate_numbers": certificate_numbers,
            "consultancy_fee_range": {
                "min": float(fee_stats[0]) if fee_stats[0] else 0,
                "max": float(fee_stats[1]) if fee_stats[1] else 0
            },
            "project_value_range": {
                "min": float(value_stats[0]) if value_stats[0] else 0,
                "max": float(value_stats[1]) if value_stats[1] else 0
            },
            "completion_date_range": {
                "min": date_stats[0].isoformat() if date_stats[0] else None,
                "max": date_stats[1].isoformat() if date_stats[1] else None
            }
        }
        
        logger.info(f"✅ [TEST] Manual clause filter options generated: {len(clients)} clients, {len(locations)} locations, {len(services)} services")
        
        return response
        
    except Exception as e:
        logger.error(f"❌ [TEST] Error fetching manual clause filter options: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching filter options: {str(e)}")


@app.post("/public-test/api/certificates/search")
async def test_search_certificates(
    request: Request,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Search certificates with universal search across all fields."""
    test_user = validate_test_token_and_get_user(test_token, db)
    
    logger.info(f"🔍 [TEST] Certificate search: Test user {test_user.email} (ID: {test_user.id})")
    
    # Parse request body with pagination support
    try:
        body = await request.json()
        query = body.get("query", "").strip()
        filter_type = body.get("filter", "all")
        page = body.get("page", 1)
        per_page = body.get("per_page", 30)
        filters = body.get("filters", {})
        
        # Ensure valid pagination values
        page = max(1, int(page))
        per_page = max(1, min(100, int(per_page)))
        
        logger.info(f"🔍 [TEST] Search query: '{query}', filter: {filter_type}, page: {page}, per_page: {per_page}")
    except Exception as e:
        logger.error(f"❌ [TEST] Failed to parse request body: {e}")
        query = ""
        filter_type = "all"
        page = 1
        per_page = 30
        filters = {}
    
    # Build base query - filter by test user ID
    filter_conditions = [
        CertificateDB.user_id == test_user.id,
        CertificateDB.processing_status == "completed"
    ]
    
    # Apply filters if provided (same logic as regular endpoint)
    filters_applied = False
    if filters:
        filters_applied = True
        
        if filters.get("clients"):
            filter_conditions.append(CertificateDB.client_name.in_(filters["clients"]))
        
        if filters.get("locations"):
            filter_conditions.append(CertificateDB.location.in_(filters["locations"]))
        
        if filters.get("consultancy_fee_range"):
            fee_range = filters["consultancy_fee_range"]
            min_fee = fee_range.get("min")
            max_fee = fee_range.get("max")
            if min_fee is not None:
                filter_conditions.append(CertificateDB.consultancy_fee_numeric >= min_fee)
            if max_fee is not None:
                filter_conditions.append(CertificateDB.consultancy_fee_numeric <= max_fee)
        
        if filters.get("project_value_range"):
            pv_range = filters["project_value_range"]
            if pv_range.get("min") is not None:
                filter_conditions.append(CertificateDB.project_value >= pv_range["min"])
            if pv_range.get("max") is not None:
                filter_conditions.append(CertificateDB.project_value <= pv_range["max"])
        
        if filters.get("completion_date_range"):
            date_range = filters["completion_date_range"]
            if date_range.get("start"):
                try:
                    start_date = datetime.fromisoformat(date_range["start"].replace("Z", "+00:00"))
                    filter_conditions.append(CertificateDB.completion_date >= start_date)
                except Exception:
                    pass
            if date_range.get("end"):
                try:
                    end_date = datetime.fromisoformat(date_range["end"].replace("Z", "+00:00"))
                    filter_conditions.append(CertificateDB.completion_date <= end_date)
                except Exception:
                    pass
        
        if filters.get("services"):
            services_conditions = []
            for service in filters["services"]:
                services_conditions.append(
                    CertificateDB.services_rendered.op('@>')(func.jsonb_build_array(service))
                )
            if services_conditions:
                filter_conditions.append(or_(*services_conditions))
        
        if filters.get("sectors"):
            sectors_conditions = []
            for sector in filters["sectors"]:
                sectors_conditions.append(
                    CertificateDB.sectors.op('@>')(func.jsonb_build_array(sector))
                )
            if sectors_conditions:
                filter_conditions.append(or_(*sectors_conditions))
        
        if filters.get("sub_sectors"):
            subsectors_conditions = []
            for subsector in filters["sub_sectors"]:
                subsectors_conditions.append(
                    CertificateDB.sub_sectors.op('@>')(func.jsonb_build_array(subsector))
                )
            if subsectors_conditions:
                filter_conditions.append(or_(*subsectors_conditions))
        
        if filters.get("funding_agencies"):
            filter_conditions.append(CertificateDB.funding_agency.in_(filters["funding_agencies"]))
    
    base_query = db.query(CertificateDB).filter(and_(*filter_conditions))
    
    # If no search query, return all certificates with pagination
    if not query:
        total_count = base_query.count()
        total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1
        offset = (page - 1) * per_page
        
        all_certificates = base_query.order_by(desc(CertificateDB.created_at)).offset(offset).limit(per_page).all()
        
        results = []
        for cert in all_certificates:
            results.append({
                "id": cert.id,
                "project_name": cert.project_name,
                "client_name": cert.client_name,
                "completion_date": cert.completion_date.isoformat() if cert.completion_date else None,
                "project_value": cert.project_value,
                "services_rendered": cert.services_rendered,
                "location": cert.location,
                "original_filename": cert.original_filename,
                "created_at": cert.created_at.isoformat() if cert.created_at else None
            })
        
        return {
            "certificates": results,
            "total": total_count,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1,
            "filters_applied": filters_applied
        }
    
    # UNIVERSAL SEARCH - Search EVERY field
    search_term = f"%{query}%"
    
    filtered_query = base_query.filter(
        or_(
            CertificateDB.project_name.ilike(search_term),
            CertificateDB.client_name.ilike(search_term),
            CertificateDB.location.ilike(search_term),
            CertificateDB.original_filename.ilike(search_term),
            CertificateDB.extracted_text.ilike(search_term),
            CertificateDB.verbatim_certificate.ilike(search_term),
            CertificateDB.scope_of_work.ilike(search_term),
            CertificateDB.certificate_number.ilike(search_term),
            CertificateDB.role_lead_jv.ilike(search_term),
            CertificateDB.funding_agency.ilike(search_term),
            CertificateDB.duration.ilike(search_term),
            CertificateDB.consultancy_fee_inr.ilike(search_term),
            CertificateDB.project_value_inr.ilike(search_term),
            cast(CertificateDB.project_value, String).ilike(search_term),
            cast(CertificateDB.services_rendered, String).ilike(search_term),
            cast(CertificateDB.sectors, String).ilike(search_term),
            cast(CertificateDB.sub_sectors, String).ilike(search_term),
            cast(CertificateDB.completion_date, String).ilike(search_term),
            cast(CertificateDB.start_date, String).ilike(search_term),
            cast(CertificateDB.end_date, String).ilike(search_term)
        )
    )
    
    total_count = filtered_query.count()
    total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1
    offset = (page - 1) * per_page
    
    certificates = filtered_query.order_by(desc(CertificateDB.created_at)).offset(offset).limit(per_page).all()
    
    logger.info(f"✅ [TEST] Universal search returned {len(certificates)} of {total_count} certificates for query: '{query}' (page {page}/{total_pages})")
    
    results = []
    for cert in certificates:
        results.append({
            "id": cert.id,
            "project_name": cert.project_name,
            "client_name": cert.client_name,
            "completion_date": cert.completion_date.isoformat() if cert.completion_date else None,
            "project_value": cert.project_value,
            "services_rendered": cert.services_rendered,
            "location": cert.location,
            "original_filename": cert.original_filename,
            "created_at": cert.created_at.isoformat() if cert.created_at else None
        })
    
    return {
        "certificates": results,
        "total": total_count,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "has_next": page < total_pages,
        "has_prev": page > 1,
        "filters_applied": filters_applied
    }


@app.post("/public-test/api/certificates/manual-clause/search")
async def test_manual_clause_search(
    request: Request,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Manual clause search with compliance scoring."""
    test_user = validate_test_token_and_get_user(test_token, db)
    
    try:
        # Parse request body
        body = await request.json()
        filters = body.get('filters', {})
        
        # Count active filters
        total_filters = count_active_filters(filters)
        
        if total_filters == 0:
            return {
                "success": False,
                "error": "At least one filter must be applied",
                "results": [],
                "compliance_summary": {
                    "total_filters_applied": 0,
                    "compliance_levels": []
                }
            }
        
        logger.info(f"🔍 [TEST] MANUAL CLAUSE SEARCH - Test user: {test_user.email}, Total Filters: {total_filters}")
        
        # Fetch ALL completed certificates for the test user
        certificates = db.query(CertificateDB).filter(
            CertificateDB.user_id == test_user.id,
            CertificateDB.processing_status == "completed"
        ).all()
        
        logger.info(f"📊 [TEST] Found {len(certificates)} total certificates to evaluate")
        
        # Calculate compliance for each certificate
        results = []
        compliance_counts = {}
        
        for cert in certificates:
            compliance = calculate_compliance(cert, filters)
            
            # Only include certificates with at least 1 filter met
            if compliance['filters_met'] > 0:
                level = compliance['filters_met']
                compliance_counts[level] = compliance_counts.get(level, 0) + 1
                
                results.append({
                    'certificate': {
                        'id': cert.id,
                        'project_name': cert.project_name,
                        'client_name': cert.client_name,
                        'location': cert.location,
                        'services_rendered': cert.services_rendered,
                        'sectors': cert.sectors,
                        'sub_sectors': cert.sub_sectors,
                        'funding_agency': cert.funding_agency,
                        'role_lead_jv': cert.role_lead_jv,
                        'duration': cert.duration,
                        'certificate_number': cert.certificate_number,
                        'consultancy_fee_numeric': cert.consultancy_fee_numeric,
                        'consultancy_fee_inr': cert.consultancy_fee_inr,
                        'project_value': cert.project_value,
                        'project_value_inr': cert.project_value_inr,
                        'completion_date': cert.completion_date.isoformat() if cert.completion_date else None,
                        'created_at': cert.created_at.isoformat() if cert.created_at else None,
                    },
                    'compliance': compliance
                })
        
        # Sort results by compliance (highest first)
        results.sort(key=lambda x: (-x['compliance']['filters_met'], x['certificate']['created_at']), reverse=False)
        
        # Build compliance summary
        compliance_levels = get_compliance_levels(total_filters)
        compliance_summary = {
            'total_filters_applied': total_filters,
            'compliance_levels': [
                {
                    'level': f"{level}/{total_filters}",
                    'count': compliance_counts.get(level, 0)
                }
                for level in compliance_levels
            ]
        }
        
        logger.info(f"✅ [TEST] Search complete: {len(results)} matching certificates found")
        
        return {
            "success": True,
            "results": results,
            "compliance_summary": compliance_summary
        }
        
    except Exception as e:
        logger.error(f"❌ [TEST] Error in manual clause search: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Search error: {str(e)}")


@app.post("/public-test/api/certificates/ai-extract-clauses")
async def test_ai_extract_clauses(
    request: Request,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Extract tender clauses from text/images using GPT-4o."""
    test_user = validate_test_token_and_get_user(test_token, db)
    
    try:
        data = await request.json()
        text_input = (data.get('text') or '').strip()
        images = data.get('images') or []
        
        if not text_input and not images:
            raise HTTPException(status_code=400, detail="Please provide text or images")
        
        # Get OpenAI API key
        api_key = os.getenv('OPENAI_API_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="OpenAI API key not configured")
        
        ai_client = OpenAIWrapper(api_key)
        
        # Clause extraction prompt (same as regular endpoint)
        system_prompt = """You are an expert at analyzing tender documents and extracting key requirements.

Analyze the provided text/images and extract the following information.

CRITICAL: Your response must be ONLY a valid JSON object, with no markdown formatting, no code blocks, no explanatory text before or after. Start with { and end with }.

Required JSON structure:
{
  "location": "Extracted location/city/state (single string)",
  "sectors": ["Sector 1", "Sector 2"],
  "sub_sectors": ["Sub-sector 1"],
  "services": ["Service 1", "Service 2"],
  "client_type": "Type of client organization",
  "funding_agency": "Funding agency if mentioned",
  "project_value_min": number or null,
  "project_value_max": number or null,
  "consultancy_fee_min": number or null,
  "consultancy_fee_max": number or null,
  "duration": "Project duration as mentioned",
  "role_requirement": "Lead/JV/Consortium/etc",
  "completion_date_start": "YYYY-MM-DD or null",
  "completion_date_end": "YYYY-MM-DD or null"
}

Extract only information explicitly stated. Use null for missing fields.
DO NOT wrap your response in ```json or ``` blocks. Return ONLY the raw JSON object."""
        
        # Build messages for GPT-4o
        messages = [{"role": "system", "content": system_prompt}]
        
        if images and len(images) > 0:
            content_parts = []
            if text_input:
                content_parts.append({"type": "text", "text": f"Text input: {text_input}"})
            else:
                content_parts.append({"type": "text", "text": "Analyze these tender document images and extract the clause information."})
            
            for img_base64 in images[:5]:
                content_parts.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/jpeg;base64,{img_base64}",
                        "detail": "high"
                    }
                })
            messages.append({"role": "user", "content": content_parts})
        else:
            messages.append({"role": "user", "content": text_input})
        
        logger.info(f"🤖 [TEST] Extracting clauses using GPT-4o for test user {test_user.email}")
        
        # Call GPT-4o
        response = ai_client.chat_completion(
            messages=messages,
            model="gpt-4o",
            max_completion_tokens=2000,
            temperature=0.1,
            timeout=60.0
        )
        
        # Parse JSON response
        clauses = extract_json_from_response(response)
        
        logger.info(f"✅ [TEST] Clause extraction complete: {list(clauses.keys())}")
        
        return {"success": True, "clauses": clauses}
        
    except json.JSONDecodeError as e:
        logger.error(f"❌ [TEST] Failed to parse AI response as JSON: {e}")
        raise HTTPException(status_code=500, detail="AI returned invalid JSON response")
    except OpenAIServiceError as e:
        logger.error(f"❌ [TEST] OpenAI API error: {e}")
        raise HTTPException(status_code=500, detail=f"AI service error: {str(e)}")
    except Exception as e:
        logger.error(f"❌ [TEST] Error in clause extraction: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Extraction error: {str(e)}")


@app.post("/public-test/api/certificates/ai-extract-conditions")
async def test_ai_extract_conditions(
    request: Request,
    test_token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Test endpoint: Extract tender requirements/conditions directly as filter values."""
    test_user = validate_test_token_and_get_user(test_token, db)
    
    try:
        data = await request.json()
        text_input = (data.get('text') or '').strip()
        images = data.get('images') or []
        
        if not text_input and not images:
            raise HTTPException(status_code=400, detail="Please provide text or images")
        
        # Get OpenAI API key
        api_key = os.getenv('OPENAI_API_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="OpenAI API key not configured")
        
        ai_client = OpenAIWrapper(api_key)
        
        # Condition extraction prompt (same as regular endpoint)
        system_prompt = """You are analyzing a TENDER DOCUMENT to extract the ELIGIBILITY CRITERIA and REQUIREMENTS that companies must meet to qualify.

Your goal: Extract CONDITIONS that applicants must satisfy, not general project descriptions.

Focus on these requirement types:

1. **Location Requirements**: Where must the company have prior experience? Extract ALL mentioned locations (cities, states, regions, countries).

2. **Sector/Domain Experience**: What sectors/industries? (Water Supply, Healthcare, Transport, Urban Development, Energy, Education, etc.)

3. **Services Required**: What type of work must they have done? (Feasibility Study, DPR Preparation, PMC, Project Management, Detailed Engineering, Supervision, etc.)

4. **Client Type**: Who should prior clients be? (Government, PSU, Private Sector, Municipal Corporation, International Organization, State Government, Central Government, etc.)

5. **Funding Agency**: Any specific funding body experience required? (World Bank, ADB, JICA, DFID, KfW, etc.)

6. **Financial Criteria**:
   - Minimum consultancy fee from past similar projects (in INR)
   - Minimum project value handled (in INR)

7. **Duration**: Required project duration experience (e.g., "12 months", "18 months", "2 years")

8. **Role**: What role experience needed? (Lead Consultant, JV Partner, Consortium Member, Solo Consultant)

9. **Completion Date**: When should past projects have been completed?
   - After date: Projects completed after this date qualify
   - Before date: Projects completed before this date qualify

10. **Certificate Date**: How recent should experience certificates be?
    - After date: Certificates issued after this date are valid

EXTRACTION RULES:
- Extract MULTIPLE values per field if tender mentions multiple options
- Use EXACT TERMS from tender text (don't paraphrase or simplify)
- For "similar" or "related" work, extract the specific examples given
- For "OR" conditions ("Water Supply OR Sanitation"), extract both as separate items
- For ranges: extract min/max values; use null if unbounded
- For dates: use "YYYY-MM-DD" format or null if not specified
- If tender says "preference for", still extract it as a condition

CRITICAL OUTPUT RULES:
- Return ONLY valid JSON, no markdown, no code blocks, no explanation
- ONLY include fields that are EXPLICITLY mentioned in the tender
- DO NOT include fields with empty arrays [] or null values
- If a field is not mentioned in the tender, OMIT it entirely from the JSON
- For ranges: only include if at least one bound (min or max) is specified
- For dates: only include if at least one date (after or before) is mentioned

DO NOT wrap response in ```json blocks. Return raw JSON only."""
        
        # Build messages for GPT-4o
        messages = [{"role": "system", "content": system_prompt}]
        
        if images and len(images) > 0:
            content_parts = []
            if text_input:
                content_parts.append({
                    "type": "text",
                    "text": f"Tender Document Text:\n\n{text_input}\n\nExtract the eligibility requirements and conditions from this tender."
                })
            else:
                content_parts.append({
                    "type": "text",
                    "text": "Analyze these tender document images and extract the eligibility requirements and conditions that companies must meet."
                })
            
            for img_base64 in images[:5]:
                content_parts.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/jpeg;base64,{img_base64}",
                        "detail": "high"
                    }
                })
            messages.append({"role": "user", "content": content_parts})
        else:
            messages.append({
                "role": "user",
                "content": f"Tender Document Text:\n\n{text_input}\n\nExtract the eligibility requirements and conditions from this tender."
            })
        
        logger.info(f"🤖 [TEST] Extracting tender CONDITIONS using GPT-4o for test user {test_user.email}")
        
        # Call GPT-4o
        response = ai_client.chat_completion(
            messages=messages,
            model="gpt-4o",
            max_completion_tokens=2500,
            temperature=0.1,
            timeout=60.0
        )
        
        # Parse JSON response
        conditions = extract_json_from_response(response)
        
        logger.info(f"✅ [TEST] Condition extraction complete: {list(conditions.keys())}")
        
        return {
            "success": True,
            "filters": conditions,
            "raw_extraction": conditions
        }
        
    except json.JSONDecodeError as e:
        logger.error(f"❌ [TEST] Failed to parse AI response as JSON: {e}")
        raise HTTPException(status_code=500, detail="AI returned invalid JSON response")
    except OpenAIServiceError as e:
        logger.error(f"❌ [TEST] OpenAI API error: {e}")
        raise HTTPException(status_code=500, detail=f"AI service error: {str(e)}")
    except Exception as e:
        logger.error(f"❌ [TEST] Error in condition extraction: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Extraction error: {str(e)}")


# Health check endpoint
@app.get("/api/health")
async def health_check():
    """Health check endpoint."""
    from core.redis_client import is_redis_available, get_redis_client
    from certificate_queue import is_redis_available as cert_queue_redis_available, get_queue_status
    
    # Check database connection
    db_status = "ok"
    try:
        db = SessionLocal()
        db.execute(text("SELECT 1"))
        db.close()
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    # Check Redis connection (for sessions)
    redis_session_status = "ok" if is_redis_available() else "unavailable"
    redis_session_url = os.getenv('REDIS_URL', 'redis://localhost:6379/0')
    
    # Check Redis connection (for certificate queue)
    redis_queue_status = "ok" if cert_queue_redis_available() else "unavailable"
    queue_status = get_queue_status()
    
    return {
        "status": "ok" if db_status == "ok" else "degraded",
        "database": db_status,
        "redis_session": {
            "status": redis_session_status,
            "url": redis_session_url.split('@')[-1] if '@' in redis_session_url else redis_session_url  # Hide credentials
        },
        "redis_queue": {
            "status": redis_queue_status,
            "queue_size": queue_status.get('queue_size', 0),
            "processing_count": queue_status.get('processing_count', 0),
            "active_workers": queue_status.get('active_workers', 0)
        },
        "timestamp": datetime.utcnow().isoformat()
    }

@app.get("/api/redis/status")
async def redis_status_endpoint(request: Request, db: Session = Depends(get_db)):
    """Check Redis connection status for certificate processing."""
    from core.redis_client import is_redis_available, get_redis_client
    from certificate_queue import is_redis_available as cert_queue_redis_available, get_queue_status, get_redis_connection
    
    current_user = get_current_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Test session Redis
    session_redis_ok = False
    session_redis_error = None
    try:
        session_client = get_redis_client()
        if session_client:
            session_client.ping()
            session_redis_ok = True
    except Exception as e:
        session_redis_error = str(e)
    
    # Test certificate queue Redis
    queue_redis_ok = False
    queue_redis_error = None
    queue_info = {}
    try:
        queue_redis_ok = cert_queue_redis_available()
        if queue_redis_ok:
            queue_info = get_queue_status()
            # Try to get connection and test
            try:
                r = get_redis_connection()
                r.ping()
            except Exception as e:
                queue_redis_error = str(e)
                queue_redis_ok = False
    except Exception as e:
        queue_redis_error = str(e)
    
    return {
        "session_redis": {
            "connected": session_redis_ok,
            "error": session_redis_error,
            "url": os.getenv('REDIS_URL', 'redis://localhost:6379/0').split('@')[-1] if '@' in os.getenv('REDIS_URL', '') else os.getenv('REDIS_URL', 'redis://localhost:6379/0')
        },
        "queue_redis": {
            "connected": queue_redis_ok,
            "error": queue_redis_error,
            "queue_status": queue_info
        }
    }

# Cleanup task (run periodically)
@app.get("/api/admin/cleanup")
async def cleanup_old_data(db: Session = Depends(get_db)):
    """Cleanup old tenders and orphaned records (admin endpoint)."""
    tender_count = cleanup_old_tenders(db, days_old=60)
    orphaned_counts = cleanup_orphaned_records(db)
    total_orphaned = sum(orphaned_counts.values())

    return {
        "message": f"Cleaned up {tender_count} old tenders and {total_orphaned} orphaned records",
        "old_tenders_cleaned": tender_count,
        "orphaned_records_cleaned": orphaned_counts
    }

@app.post("/api/admin/backfill-project-ids")
async def backfill_project_ids(db: Session = Depends(get_db)):
    """
    Backfill project IDs for all existing projects that don't have one.
    Projects are processed ordered by creation date, grouped by user.
    """
    try:
        # Query all projects without project_id, ordered by created_at
        projects_without_id = db.query(ProjectDB).filter(
            ProjectDB.project_id.is_(None)
        ).order_by(ProjectDB.user_id, ProjectDB.created_at).all()
        
        if not projects_without_id:
            return {
                "success": True,
                "message": "All projects already have IDs",
                "updated": 0
            }
        
        updated_count = 0
        errors = []
        
        # Group by user_id to process sequentially per user
        current_user_id = None
        for project in projects_without_id:
            try:
                # Generate project_id for this project
                project_id_value = generate_project_id(project.user_id, db)
                project.project_id = project_id_value
                updated_count += 1
                
                # Commit in batches of 50 to avoid long transactions
                if updated_count % 50 == 0:
                    db.commit()
                    logger.info(f"Backfilled {updated_count} project IDs so far...")
                
            except Exception as e:
                error_msg = f"Error generating ID for project {project.id} (user {project.user_id}): {str(e)}"
                logger.error(error_msg)
                errors.append(error_msg)
                continue
        
        # Final commit for remaining projects
        db.commit()
        
        result = {
            "success": True,
            "message": f"Backfilled project IDs for {updated_count} projects",
            "updated": updated_count,
            "total_processed": len(projects_without_id)
        }
        
        if errors:
            result["errors"] = errors[:10]  # Limit error messages
            result["error_count"] = len(errors)
        
        logger.info(f"Backfill completed: {updated_count} projects updated")
        return result
        
    except Exception as e:
        db.rollback()
        logger.error(f"Backfill failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Backfill failed: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=5001, reload=True)
